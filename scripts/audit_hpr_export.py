"""Reconcile an OSIP holdings export against its source workbook and dividends."""

from __future__ import annotations

import argparse
from datetime import date
from decimal import Decimal, InvalidOperation
import json
from pathlib import Path

from openpyxl import load_workbook

from osip_dashboard.ingestion import parse_osip_workbook
from osip_dashboard.services.dividends import load_dividend_history, lot_dividend_contribution
from osip_dashboard.services.holdings_export import (
    estimated_coupon_income_kzt,
    estimated_paid_coupon_income_native,
)
from osip_dashboard.services.hpr import hpr_amount, hpr_percent


TOLERANCE = Decimal("0.02")


def _estimated_hpr_income(lot: object, contribution: object, *, report_date: date, usd_rate: Decimal | None) -> tuple[Decimal, Decimal]:
    """Mirror the holdings export's coupon-plus-dividend HPR inputs.

    Coupon income is the estimated paid amount (gross 360-day estimate less
    current accrued coupon where applicable); dividends are the validated
    ex-date/pay-date contribution. Missing components contribute zero, just
    as they do in ``create_holdings_xlsx``.
    """
    dividend_kzt = getattr(contribution, "kzt_amount", None)
    dividend_native = getattr(contribution, "native_amount", None)
    matched_count = getattr(contribution, "matched_count", 0)
    coupon_native = estimated_paid_coupon_income_native(lot, report_date)
    coupon_kzt = estimated_coupon_income_kzt(lot, coupon_native)
    kzt_income = (dividend_kzt or Decimal("0")) + (coupon_kzt or Decimal("0"))

    dividend_usd: Decimal | None = Decimal("0")
    currency = str(getattr(lot, "instrument_currency", "") or "").upper()
    if matched_count:
        if currency == "USD":
            dividend_usd = dividend_native
        elif currency == "KZT" and usd_rate and dividend_kzt is not None:
            dividend_usd = dividend_kzt / usd_rate
        else:
            dividend_usd = None
    coupon_usd: Decimal | None = None
    if coupon_native is not None:
        if currency == "USD":
            coupon_usd = coupon_native
        elif currency == "KZT" and usd_rate:
            coupon_usd = coupon_native / usd_rate
    usd_income = (dividend_usd or Decimal("0")) + (coupon_usd or Decimal("0"))
    return kzt_income, usd_income


def _decimal(value: object) -> Decimal | None:
    if value in (None, "", "Недоступно"):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _same(actual: object, expected: Decimal | None) -> bool:
    if expected is None:
        return actual in (None, "Недоступно")
    parsed = _decimal(actual)
    return parsed is not None and abs(parsed - expected) <= TOLERANCE


def audit(source_path: Path, export_path: Path, *, portfolio_code: str, dividends_path: Path, as_of: date) -> dict[str, object]:
    snapshot = parse_osip_workbook(source_path, portfolio_code=portfolio_code, source_name=source_path.name)
    source_lots = sorted(
        snapshot.positions,
        key=lambda lot: (lot.security_code or "", lot.isin or "", lot.source.row_number if lot.source else 0),
    )
    workbook = load_workbook(export_path, data_only=True, read_only=True)
    if "Позиции по лотам" not in workbook.sheetnames:
        raise ValueError("Export is missing the Позиции по лотам sheet")
    sheet = workbook["Позиции по лотам"]
    header_row = next((row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Код инструмента"), None)
    total_row = next((row for row in range((header_row or 0) + 1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Итого"), None)
    if header_row is None or total_row is None:
        raise ValueError("Export has no identifiable lot table or total row")
    export_rows = list(range(header_row + 1, total_row))
    columns = {
        sheet.cell(header_row, column).value: column
        for column in range(1, sheet.max_column + 1)
    }
    required_columns = {
        "Эмитент", "Стоимость покупки, ₸", "Дата погашения облигации", "Стоимость открытия", "HPR (расч.), KZT",
        "HPR (расч.), FX", "HPR (расч.), KZT, %", "HPR (расч.), FX, %",
        "Текущая балансовая цена за единицу (источник OSIP)",
    }
    missing_columns = sorted(required_columns - columns.keys())
    if missing_columns:
        raise ValueError(f"Export is missing explicit HPR columns: {', '.join(missing_columns)}")
    issuer_column = columns.get("Эмитент")
    if issuer_column is None or columns.get("Дата погашения облигации") != issuer_column + 1:
        raise ValueError("Дата погашения облигации must immediately follow Эмитент")
    if len(source_lots) != len(export_rows):
        raise ValueError(f"Lot count mismatch: source={len(source_lots)} export={len(export_rows)}")

    history = load_dividend_history(str(dividends_path))
    usd_rates = sorted(
        lot.report_fx_rate
        for lot in source_lots
        if lot.instrument_currency == "USD"
        and lot.report_fx_rate is not None
        and lot.report_fx_rate > 0
    )
    usd_rate = usd_rates[len(usd_rates) // 2] if usd_rates else None
    if usd_rates and usd_rates[-1] / usd_rates[0] - Decimal("1") > Decimal("0.01"):
        # Match holdings_export._snapshot_usd_rates: a mixed-rate workbook
        # must not produce a guessed USD amount from a median rate.
        usd_rate = None
    issues: list[dict[str, object]] = []
    dividend_total = Decimal("0")
    matched_lots = 0
    for lot, row in zip(source_lots, export_rows):
        contribution = lot_dividend_contribution(lot, history=history, current_date=as_of)
        dividend_total += contribution.kzt_amount or Decimal("0")
        matched_lots += bool(contribution.matched_count)
        # Keep this reconciliation aligned with the holdings exporter. HPR
        # includes validated dividends plus the estimated paid-coupon amount.
        estimated_kzt_income, estimated_usd_income = _estimated_hpr_income(
            lot, contribution, report_date=snapshot.report_date, usd_rate=usd_rate
        )
        purchase_kzt = lot.purchase_amount_kzt
        carrying_kzt = lot.derived_carrying_value_kzt
        expected_hpr_kzt = hpr_amount(purchase_kzt, carrying_kzt, estimated_kzt_income)
        expected_hpr_percent = hpr_percent(purchase_kzt, carrying_kzt, estimated_kzt_income)
        if lot.instrument_currency == "USD":
            purchase_usd = lot.purchase_amount_native
            carrying_usd = lot.carrying_amount_native
        elif lot.instrument_currency == "KZT" and usd_rate:
            purchase_usd = purchase_kzt / usd_rate if purchase_kzt is not None else None
            carrying_usd = carrying_kzt / usd_rate if carrying_kzt is not None else None
        else:
            purchase_usd = carrying_usd = None
        expected_hpr_usd = hpr_amount(purchase_usd, carrying_usd, estimated_usd_income)
        expected_hpr_usd_percent = hpr_percent(purchase_usd, carrying_usd, estimated_usd_income)
        if lot.instrument_currency == "KZT":
            # The common USD conversion factor cancels in the percentage
            # ratio, so this remains auditable even when USD amounts are
            # unavailable because no agreed USD/KZT rate exists.
            expected_hpr_usd_percent = expected_hpr_percent
        checks = {
            "purchase KZT": (sheet.cell(row, columns["Стоимость покупки, ₸"]).value, purchase_kzt),
            "MV KZT": (sheet.cell(row, columns["Текущая стоимость, KZT"]).value, carrying_kzt),
            "purchase native": (sheet.cell(row, columns["Стоимость открытия"]).value, lot.purchase_amount_native),
            "carrying price native": (
                sheet.cell(row, columns["Текущая балансовая цена за единицу (источник OSIP)"]).value,
                lot.carrying_price_native,
            ),
            "HPR KZT": (sheet.cell(row, columns["HPR (расч.), KZT"]).value, expected_hpr_kzt),
            "HPR FX": (sheet.cell(row, columns["HPR (расч.), FX"]).value, expected_hpr_usd),
            "HPR KZT %": (sheet.cell(row, columns["HPR (расч.), KZT, %"]).value, expected_hpr_percent),
            "HPR FX %": (sheet.cell(row, columns["HPR (расч.), FX, %"]).value, expected_hpr_usd_percent),
        }
        for field, (actual, expected) in checks.items():
            if not _same(actual, expected):
                issues.append({"row": row, "instrument": lot.security_code, "field": field, "actual": actual, "expected": str(expected) if expected is not None else "Недоступно"})
        actual_maturity = sheet.cell(row, columns["Дата погашения облигации"]).value
        if hasattr(actual_maturity, "date"):
            actual_maturity = actual_maturity.date()
        expected_maturity = lot.close_date or "Недоступно"
        if actual_maturity != expected_maturity:
            issues.append({"row": row, "instrument": lot.security_code, "field": "maturity date", "actual": actual_maturity, "expected": expected_maturity})
    workbook.close()
    return {
        "source": str(source_path),
        "export": str(export_path),
        "portfolio": portfolio_code,
        "as_of": as_of.isoformat(),
        "source_lots": len(source_lots),
        "export_lots": len(export_rows),
        "matched_dividend_lots": matched_lots,
        "dividend_total_kzt": str(dividend_total),
        "issues": issues,
        "passed": not issues,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Reconcile a holdings export's HPR values to the OSIP source and Bloomberg dividends.")
    parser.add_argument("source", type=Path, help="Source OSIP .xls workbook")
    parser.add_argument("export", type=Path, help="Generated holdings .xlsx export")
    parser.add_argument("--portfolio-code", required=True, help="Portfolio code assigned to the source workbook")
    parser.add_argument("--dividends", type=Path, required=True, help="Bloomberg dividend dictionary .xlsx")
    parser.add_argument("--as-of", type=date.fromisoformat, default=date.today(), help="Current-date cutoff for received dividends (YYYY-MM-DD)")
    args = parser.parse_args()
    print(json.dumps(audit(args.source, args.export, portfolio_code=args.portfolio_code, dividends_path=args.dividends, as_of=args.as_of), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
