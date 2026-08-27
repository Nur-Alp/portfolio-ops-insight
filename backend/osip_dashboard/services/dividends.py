"""Read the supplied dividend history and calculate received lot income."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from functools import lru_cache
import hashlib
import io
import json
import os
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook


_DEFAULT_PATH = Path(__file__).resolve().parents[3] / "sources" / "dividends.xlsx"
_US_TAX_FACTOR = Decimal("0.85")
_STALE_AFTER_DAYS = 35
_override_path: Path | None = None
_metadata_path: Path | None = None


class DividendValidationError(ValueError):
    """The uploaded Bloomberg dividend history is missing usable records."""


@dataclass(frozen=True)
class DividendRecord:
    ticker: str
    dividend: Decimal
    declaration_date: date | None
    ex_date: date | None
    record_date: date | None
    pay_date: date | None
    dividend_type: str
    tickers_list: tuple[str, ...] = ()


@dataclass(frozen=True)
class DividendContribution:
    """A lot's received dividend, in native currency and KZT when possible."""

    native_amount: Decimal = Decimal("0")
    kzt_amount: Decimal | None = Decimal("0")
    matched_count: int = 0
    unavailable: bool = False


@dataclass(frozen=True)
class DividendDataStatus:
    """Coverage and freshness metadata for the dividend reference file."""

    freshness: str
    source_filename: str | None
    source_sha256: str | None
    source_date: date | None
    uploaded_at: datetime | None
    latest_ex_date: date | None
    latest_pay_date: date | None
    future_pay_count: int
    row_count: int
    ticker_count: int
    stale_after_days: int = _STALE_AFTER_DAYS


def configure_dividend_data_root(root: Path | None) -> None:
    """Configure runtime storage for uploaded Bloomberg dividend histories."""
    global _override_path, _metadata_path
    _override_path = (root / "dividends.xlsx") if root is not None else None
    _metadata_path = (root / "dividends.metadata.json") if root is not None else None
    load_dividend_history.cache_clear()


def dividend_source_path() -> Path:
    configured = os.getenv("OSIP_DIVIDENDS_FILE")
    if configured:
        return Path(configured)
    if _override_path is not None and _override_path.exists():
        return _override_path
    return _DEFAULT_PATH


def normalize_ticker(value: Any) -> str:
    """Normalize Bloomberg-style tickers while preserving the market token."""
    text = re.sub(r"\s+", " ", str(value or "").strip().upper())
    if text.endswith(" EQUITY"):
        text = text[:-7].rstrip()
    return text


def is_us_listed(ticker: Any) -> bool:
    """Return true only when US is a standalone ticker token."""
    return "US" in normalize_ticker(ticker).split()


def _header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError):
        return None


def _date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    text = str(value).strip()
    for parser in (date.fromisoformat,):
        try:
            return parser(text[:10])
        except ValueError:
            pass
    return None


def _split_tickers(value: Any) -> tuple[str, ...]:
    if not value:
        return ()
    return tuple(
        normalize_ticker(part)
        for part in re.split(r"[,;|\n]", str(value))
        if normalize_ticker(part)
    )


def _date_from_filename(filename: str | None) -> date | None:
    if not filename:
        return None
    match = re.search(r"(?<!\d)(\d{1,2})[.\-_](\d{1,2})[.\-_](\d{2,4})(?!\d)", filename)
    if not match:
        return None
    day, month, year = (int(value) for value in match.groups())
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _parse_workbook_object(workbook: Any) -> DividendHistory:
    records: list[DividendRecord] = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue
            header_map = {_header_key(value): index for index, value in enumerate(header)}

            def column(*names: str) -> int | None:
                return next((header_map[name] for name in names if name in header_map), None)

            ticker_column = column("ticker", "id")
            dividend_column = column("dividend")
            ex_column = column("exdate", "exdividenddate")
            pay_column = column("paydate", "payable", "payabledate")
            if ticker_column is None or dividend_column is None or ex_column is None or pay_column is None:
                continue
            declaration_column = column("declarationdate", "declaration")
            record_column = column("recorddate", "record")
            type_column = column("dividendtype", "type")
            aliases_column = column("tickerslist", "tickerlist", "aliases")
            for row in rows:
                ticker = row[ticker_column] if ticker_column < len(row) else None
                dividend = _decimal(row[dividend_column] if dividend_column < len(row) else None)
                ex_date = _date(row[ex_column] if ex_column < len(row) else None)
                pay_date = _date(row[pay_column] if pay_column < len(row) else None)
                if not ticker or dividend is None or ex_date is None or pay_date is None:
                    continue
                records.append(
                    DividendRecord(
                        ticker=str(ticker).strip(),
                        dividend=dividend,
                        declaration_date=_date(row[declaration_column]) if declaration_column is not None and declaration_column < len(row) else None,
                        ex_date=ex_date,
                        record_date=_date(row[record_column]) if record_column is not None and record_column < len(row) else None,
                        pay_date=pay_date,
                        dividend_type=str(row[type_column] or "").strip() if type_column is not None and type_column < len(row) else "",
                        tickers_list=_split_tickers(row[aliases_column]) if aliases_column is not None and aliases_column < len(row) else (),
                    )
                )
    finally:
        workbook.close()
    return DividendHistory(records)


class DividendHistory:
    def __init__(self, records: Iterable[DividendRecord] = ()) -> None:
        self.records = tuple(records)
        by_ticker: dict[str, list[DividendRecord]] = {}
        for record in self.records:
            keys = tuple(dict.fromkeys((normalize_ticker(record.ticker), *record.tickers_list)))
            for key in keys:
                if key:
                    by_ticker.setdefault(key, []).append(record)
        self._by_ticker = {key: tuple(value) for key, value in by_ticker.items()}

    def for_ticker(self, ticker: Any) -> tuple[DividendRecord, ...]:
        return self._by_ticker.get(normalize_ticker(ticker), ())

    def received(
        self,
        ticker: Any,
        *,
        purchase_date: date | None,
        current_date: date,
    ) -> tuple[DividendRecord, ...]:
        if purchase_date is None:
            return ()
        return tuple(
            record
            for record in self.for_ticker(ticker)
            if record.ex_date is not None
            and record.pay_date is not None
            and record.ex_date > purchase_date
            and record.pay_date < current_date
        )


def _parse_workbook(path: Path) -> DividendHistory:
    if not path.exists():
        return DividendHistory()
    workbook = load_workbook(path, read_only=True, data_only=True)
    return _parse_workbook_object(workbook)


@lru_cache(maxsize=8)
def load_dividend_history(path: str | None = None) -> DividendHistory:
    """Load the current uploaded history, falling back to the packaged file."""
    selected = Path(path) if path else dividend_source_path()
    return _parse_workbook(selected)


def _metadata() -> dict[str, Any]:
    if _metadata_path is None or not _metadata_path.exists():
        return {}
    try:
        return json.loads(_metadata_path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}


def dividend_data_status(*, as_of: date | None = None) -> DividendDataStatus:
    """Return whether the active dividend history is current enough to trust."""
    selected = dividend_source_path()
    if not selected.exists():
        return DividendDataStatus("missing", None, None, None, None, None, None, 0, 0, 0)
    history = load_dividend_history()
    records = history.records
    metadata = _metadata()
    source_date = _date(metadata.get("source_date")) or _date_from_filename(metadata.get("source_filename") or selected.name)
    uploaded_datetime = None
    raw_uploaded = metadata.get("uploaded_at")
    if raw_uploaded:
        try:
            uploaded_datetime = datetime.fromisoformat(str(raw_uploaded))
        except ValueError:
            uploaded_datetime = None
    latest_ex = max((record.ex_date for record in records if record.ex_date), default=None)
    latest_pay = max((record.pay_date for record in records if record.pay_date), default=None)
    today = as_of or date.today()
    future_pay_count = sum(1 for record in records if record.pay_date is not None and record.pay_date >= today)
    if not records:
        freshness = "missing"
    elif source_date is None:
        freshness = "unknown"
    else:
        freshness = "fresh" if (today - source_date).days <= _STALE_AFTER_DAYS else "stale"
    return DividendDataStatus(
        freshness=freshness,
        source_filename=metadata.get("source_filename") or selected.name,
        source_sha256=metadata.get("source_sha256"),
        source_date=source_date,
        uploaded_at=uploaded_datetime,
        latest_ex_date=latest_ex,
        latest_pay_date=latest_pay,
        future_pay_count=future_pay_count,
        row_count=len(records),
        ticker_count=len({normalize_ticker(record.ticker) for record in records}),
    )


def replace_dividend_history(*, filename: str, content: bytes) -> DividendDataStatus:
    """Validate and install a Bloomberg dividend workbook in runtime storage."""
    if not filename.lower().endswith(".xlsx"):
        raise DividendValidationError("Загрузите файл Bloomberg в формате .xlsx")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        history = _parse_workbook_object(workbook)
    except Exception as exc:
        raise DividendValidationError("Не удалось прочитать книгу дивидендов Bloomberg") from exc
    if not history.records:
        raise DividendValidationError("В книге не найдены строки с тикером, дивидендом, ex-date и pay-date")
    if _override_path is None or _metadata_path is None:
        raise DividendValidationError("Хранилище загруженных дивидендов не настроено")
    _override_path.parent.mkdir(parents=True, exist_ok=True)
    _override_path.write_bytes(content)
    source_date = _date_from_filename(filename)
    metadata = {
        "source_filename": filename,
        "source_sha256": hashlib.sha256(content).hexdigest(),
        "source_date": source_date.isoformat() if source_date else None,
        "uploaded_at": datetime.now().astimezone().isoformat(),
        "row_count": len(history.records),
        "ticker_count": len({normalize_ticker(record.ticker) for record in history.records}),
    }
    _metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    load_dividend_history.cache_clear()
    return dividend_data_status()


def lot_dividend_contribution(
    lot: Any,
    *,
    history: DividendHistory | None = None,
    current_date: date | None = None,
) -> DividendContribution:
    """Calculate net received dividends for one lot in native currency/KZT."""
    history = history or load_dividend_history()
    as_of = current_date or date.today()
    records = history.received(
        getattr(lot, "security_code", ""),
        purchase_date=getattr(lot, "purchase_date", None),
        current_date=as_of,
    )
    if not records:
        return DividendContribution()
    quantity = _decimal(getattr(lot, "quantity", None)) or Decimal("0")
    tax_factor = _US_TAX_FACTOR if is_us_listed(getattr(lot, "security_code", "")) else Decimal("1")
    native_amount = sum((record.dividend * quantity * tax_factor for record in records), Decimal("0"))
    currency = str(getattr(lot, "instrument_currency", "") or "").upper()
    if currency == "KZT":
        return DividendContribution(native_amount, native_amount, len(records), False)
    fx_rate = _decimal(getattr(lot, "report_fx_rate", None))
    if fx_rate is not None and fx_rate > 0:
        return DividendContribution(native_amount, native_amount * fx_rate, len(records), False)
    return DividendContribution(native_amount, None, len(records), True)
