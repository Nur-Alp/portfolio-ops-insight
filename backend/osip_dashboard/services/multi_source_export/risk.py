"""Risk-limits summary writer (SOBSTV + TABYS combined view)."""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from osip_dashboard.persistence.models import DatasetVersion

from .shared import _BORDER, _matches_term, _summary_decimal, _write_multi_series_summary_chart


def _write_risk_summary(workbook: Workbook, dataset_by_type: dict[str, DatasetVersion], *, term: str = "") -> None:
    """One "Сводка по лимитам" sheet combining both portfolios' risk-limit lines.

    SOBSTV and TABYS are independently uploaded/versioned (see
    ingestion/multi_source.py's `_parse_risk_sobstv`/`_parse_risk_tabys`
    docstrings), but a risk reviewer wants one cross-portfolio view of
    what's breached, not two separate tabs to cross-reference by hand.
    Counts are recomputed from the term-filtered record set (not the
    precomputed per-dataset summary), same reasoning as corporate finance's
    summary: a search term must narrow the summary the same way it narrows
    the detail sheets.
    """
    sobstv = dataset_by_type.get("risk_limits_sobstv")
    tabys = dataset_by_type.get("risk_limits_tabys")
    if sobstv is None and tabys is None:
        return
    normalized_term = term.strip().casefold()
    records = [
        record for dataset in (sobstv, tabys) if dataset is not None
        for record in dataset.records if _matches_term(record, normalized_term)
    ]
    sheet = workbook.create_sheet("Сводка по лимитам")
    sheet.sheet_view.showGridLines = False
    # No frozen panes: this is a dashboard of stacked tables/charts, read
    # chart-first rather than scrolled as a long table - a frozen column A
    # has nothing to do here. Deliberate, documented exception to the
    # workbook navigation standard (docs/export-column-audit.md), enforced
    # by tests/export_compliance.py's _UNFROZEN_DASHBOARD_SHEETS.
    row = 2
    sheet.cell(row, 1, "Сводка по лимитам риска").font = Font(bold=True, size=14)
    row += 2
    control_records = [record for record in records if record.payload.get("dimension") not in {"exposure_detail", "country_instrument_detail"}]
    breach_count = sum(1 for record in control_records if record.payload.get("signal") == "breach")
    unknown_count = sum(1 for record in control_records if record.payload.get("signal") is None)
    not_applicable_count = sum(1 for record in control_records if record.payload.get("signal") == "not_applicable")
    sheet.cell(row, 1, "Лимитных строк").font = Font(bold=True)
    sheet.cell(row, 2, len(control_records))
    sheet.cell(row + 1, 1, "Превышений лимита").font = Font(bold=True)
    sheet.cell(row + 1, 2, breach_count)
    sheet.cell(row + 2, 1, "Деталей экспозиции").font = Font(bold=True)
    sheet.cell(row + 2, 2, len(records) - len(control_records))
    sheet.cell(row + 3, 1, "Статус не определён").font = Font(bold=True)
    sheet.cell(row + 3, 2, unknown_count)
    sheet.cell(row + 4, 1, "Без применимого лимита").font = Font(bold=True)
    sheet.cell(row + 4, 2, not_applicable_count)
    row += 6

    detail_dimensions = {"exposure_detail", "country_instrument_detail"}

    def signal_bucket(record: Any) -> str:
        if record.payload.get("dimension") in detail_dimensions:
            return "not_applicable"
        signal = str(record.payload.get("signal") or "").strip().casefold()
        if signal in {"ok", "pass", "within", "within_limit"}:
            return "within"
        if signal in {"breach", "exceeded", "limit_exceeded"}:
            return "breach"
        if signal in {"not_applicable", "n/a"}:
            return "not_applicable"
        return "unknown"

    status_mix: dict[str, dict[str, int]] = {}
    portfolio_mix: dict[str, dict[str, int]] = {}
    # The status chart answers a limit-status question, so exclude the two
    # exposure-detail dimensions.  Those rows remain visible in the raw tabs,
    # the separate exposure count, portfolio "Детали" series, and currency
    # exposure chart; including them here made a status total that mixed
    # detail rows with actual limit controls.
    for record in control_records:
        dimension = _RISK_DIMENSION_LABELS.get(record.payload.get("dimension"), record.payload.get("dimension") or "Не указано")
        bucket = signal_bucket(record)
        dimension_counts = status_mix.setdefault(dimension, {"within": 0, "breach": 0, "unknown": 0, "not_applicable": 0})
        dimension_counts[bucket] += 1

    # Detail rows (exposure_detail/country_instrument_detail - informational,
    # no limit of their own) are excluded from this comparison: they have no
    # bearing on limit compliance, and SOBSTV's much larger detail count
    # dwarfed the actually-comparable bars (breaches/unknown/limit lines).
    # The total detail count is still visible separately, in the KPI row at
    # the top of this sheet ("Деталей экспозиции").
    for record in records:
        bucket = signal_bucket(record)
        if bucket == "not_applicable":
            continue
        portfolio = str(record.payload.get("portfolio_code") or "Не указано")
        portfolio_counts = portfolio_mix.setdefault(portfolio, {"limit": 0, "breach": 0, "unknown": 0})
        portfolio_counts["limit"] += 1
        if bucket in {"breach", "unknown"}:
            portfolio_counts[bucket] += 1

    status_rows = [
        [dimension, Decimal(counts["within"]), Decimal(counts["breach"]), Decimal(counts["unknown"]), Decimal(counts["not_applicable"])]
        for dimension, counts in sorted(status_mix.items(), key=lambda pair: -sum(pair[1].values()))
    ]
    if status_rows:
        row = _write_multi_series_summary_chart(
            sheet, row, "Статусы лимитов по измерению",
            ["Измерение", "В пределах лимита", "Превышение", "Не определено", "Без применимого лимита"], status_rows,
            numeric_formats={2: "#,##0", 3: "#,##0", 4: "#,##0", 5: "#,##0"},
            value_col_first=2, value_col_last=5, chart_title="Статусы лимитов по измерению", widths=[30, 18, 16, 16, 24], horizontal=True,
        )

    portfolio_rows = [
        [portfolio, Decimal(counts["limit"]), Decimal(counts["breach"]), Decimal(counts["unknown"])]
        for portfolio, counts in sorted(portfolio_mix.items())
    ]
    if len(portfolio_rows) >= 2:
        row = _write_multi_series_summary_chart(
            sheet, row, "Сравнение портфелей",
            ["Портфель", "Лимитные строки", "Превышения", "Не определено"], portfolio_rows,
            numeric_formats={2: "#,##0", 3: "#,##0", 4: "#,##0"},
            value_col_first=2, value_col_last=4, chart_title="Сравнение портфелей по статусам", widths=[18, 18, 16, 16],
        )

    def optional_decimal(value: Any) -> Decimal | None:
        if value in (None, "", "Недоступно"):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    utilization_rows: list[list[Any]] = []
    for record in control_records:
        payload = record.payload
        # Mirrors _risk_utilization's own basis order in ingestion/multi_source.py
        # - must stay in sync with it. base_label names what limit_pct is a
        # percentage OF (e.g. "Капитал"/capital); when it's set, that's a
        # different denominator than actual_pct's own (total assets), so
        # skip straight to KZT/USD instead of dividing two mismatched
        # percentages (confirmed on real data: a SOBSTV issuer row otherwise
        # produced a nonsense 383% "utilization" for a row the source itself
        # reports as OK).
        basis_pairs = (
            (("KZT", "actual_kzt", "limit_kzt"), ("USD", "actual_usd", "limit_usd"))
            if payload.get("base_label") not in (None, "")
            else (("%", "actual_pct", "limit_pct"), ("KZT", "actual_kzt", "limit_kzt"), ("USD", "actual_usd", "limit_usd"))
        )
        selected: tuple[str, Decimal] | None = None
        for basis, actual_field, limit_field in basis_pairs:
            actual = optional_decimal(payload.get(actual_field))
            limit = optional_decimal(payload.get(limit_field))
            if actual is not None and limit not in (None, Decimal("0")):
                selected = (basis, actual / limit)
                break
        if selected is not None:
            basis, utilization = selected
            portfolio = str(payload.get("portfolio_code") or "Не указано")
            dimension = _RISK_DIMENSION_LABELS.get(payload.get("dimension"), payload.get("dimension") or "Не указано")
            label = str(payload.get("label") or "Не указано")
            full_label = f"{portfolio} · {label}"
            chart_label = full_label if len(full_label) <= 48 else f"{full_label[:45]}…"
            # Signal is the source's own OK/breach determination (see
            # _risk_signal in ingestion), not derived from this ratio - for a
            # row whose % limit is expressed against a different base than
            # its % actual (e.g. an issuer limit set as a % of capital, while
            # the actual is a % of total assets), utilization can read as
            # >100% while the source still reports OK. Showing the signal
            # here prevents that from reading as an unflagged breach.
            # near_breach is a separate stored flag (see _risk_near_breach in
            # ingestion), not a third value of signal - a row can be
            # signal="OK" and near_breach=True at once. Without checking it
            # here, a genuinely near-breach row silently read as plain "В
            # пределах лимита" with no warning at all.
            signal = str(payload.get("signal") or "").strip().casefold()
            signal_label = (
                "Превышение" if signal == "breach"
                else "Близко к превышению" if payload.get("near_breach")
                else "В пределах лимита" if signal == "ok"
                else "Не определено"
            )
            utilization_rows.append([chart_label, full_label, dimension, basis, utilization, signal_label])
    utilization_rows.sort(key=lambda values: values[4], reverse=True)
    utilization_rows = utilization_rows[:10]
    if utilization_rows:
        row = _write_multi_series_summary_chart(
            sheet, row, "Наибольшее использование лимита (топ-10)",
            ["Позиция (график)", "Полное наименование", "Измерение", "База", "Использование лимита", "Сигнал"], utilization_rows,
            numeric_formats={5: "0.0%;[Red](0.0%);-"}, value_col_first=5, value_col_last=5,
            chart_title="Использование лимита, топ-10", widths=[32, 44, 24, 12, 20, 18], chart_number_format="0%", horizontal=True,
            # These are by definition the 10 highest-utilization rows -
            # often clustered close together (e.g. all between 60% and
            # 96%) - so start the axis just below the smallest of them
            # instead of at 0%, making the real spread visible instead of
            # every bar reading as nearly full-length.
            min_value_padding=0.05,
        )

    # Mirrors the website's per-dimension "Использование лимита по
    # выбранному измерению" chart, which is an interactive dropdown there -
    # a static workbook can't replicate a dropdown, so every dimension that
    # actually has utilization data gets its own pair of charts here instead
    # (one per portfolio, skipped where that portfolio has none). "duration"
    # is excluded for the same reason it's excluded from the website's
    # dropdown: it's measured as modified duration against a limit, and
    # ingestion never computes a utilization ratio for it.
    # Reads the stored utilization field (actual/limit, base_label-aware -
    # see _risk_utilization in ingestion), not actual_pct directly, for the
    # same reason the "Использование лимита, топ-10" table above does: a
    # row's actual_pct alone is its share of the portfolio, not of its own
    # limit, and dividing by a mismatched-base limit_pct produces nonsense
    # for some rows (confirmed on real data - a SOBSTV issuer at 86.2% of
    # assets against a 22.5%-of-capital limit is NOT 383% over).
    utilization_dimension_order = [
        "instrument_category", "country", "issuer", "sector", "ifrs", "currency", "fx_position", "instrument_issuer",
    ]
    for dimension in utilization_dimension_order:
        for portfolio in ("SOBSTV", "TABYS"):
            eligible = [
                record for record in control_records
                if record.payload.get("portfolio_code") == portfolio
                and record.payload.get("dimension") == dimension
                and record.payload.get("utilization") not in (None, "")
            ]
            if not eligible:
                continue
            eligible.sort(key=lambda record: _summary_decimal(record.payload.get("utilization")), reverse=True)
            dimension_utilization_rows = [
                [record.payload.get("label") or "Не указано", _summary_decimal(record.payload.get("utilization"))]
                for record in eligible[:8]
            ]
            dimension_label = _RISK_DIMENSION_LABELS.get(dimension, dimension)
            row = _write_multi_series_summary_chart(
                sheet, row, f"Использование лимита по измерению «{dimension_label}» · {portfolio}",
                ["Наименование", "Использование лимита"], dimension_utilization_rows,
                numeric_formats={2: "0.0%;[Red](0.0%);-"}, value_col_first=2, value_col_last=2,
                chart_title=f"Использование лимита: {dimension_label} · {portfolio}", widths=[36, 20],
                chart_number_format="0%", horizontal=True,
            )

    exposure_mix: dict[str, Decimal] = {}
    exposure_counts: dict[str, int] = {}
    for record in records:
        if record.payload.get("dimension") != "exposure_detail":
            continue
        amount_kzt = optional_decimal(record.payload.get("amount_kzt"))
        if amount_kzt is None:
            continue
        currency = str(record.payload.get("currency") or "Не указано")
        exposure_mix[currency] = exposure_mix.get(currency, Decimal("0")) + amount_kzt
        exposure_counts[currency] = exposure_counts.get(currency, 0) + 1
    exposure_rows = [[currency, amount, Decimal(exposure_counts[currency])] for currency, amount in sorted(exposure_mix.items(), key=lambda pair: pair[1], reverse=True)]
    if exposure_rows:
        row = _write_multi_series_summary_chart(
            sheet, row, "Валютная экспозиция по валюте",
            ["Валюта", "Сумма, KZT", "Строк источника"], exposure_rows,
            numeric_formats={2: "#,##0.00;[Red](#,##0.00);-", 3: "#,##0"}, value_col_first=2, value_col_last=2,
            chart_title="Валютная экспозиция, KZT", widths=[18, 22, 18],
        )

    duration_rows: list[list[Any]] = []
    for record in records:
        if record.payload.get("dimension") != "duration":
            continue
        modified = optional_decimal(record.payload.get("modified_duration"))
        limit = optional_decimal(record.payload.get("duration_limit"))
        if modified is None or limit is None:
            continue
        portfolio = str(record.payload.get("portfolio_code") or "Не указано")
        issuer = str(record.payload.get("issuer") or record.payload.get("label") or "Не указано")
        isin = str(record.payload.get("isin") or "")
        label = f"{portfolio} · {issuer}" + (f" · {isin}" if isin else "")
        headroom = optional_decimal(record.payload.get("duration_headroom"))
        chart_label = label if len(label) <= 48 else f"{label[:45]}…"
        duration_rows.append([chart_label, label, modified, limit, headroom if headroom is not None else modified - limit])
    duration_rows.sort(key=lambda values: values[2] / values[3] if values[3] else Decimal("0"), reverse=True)
    duration_rows = duration_rows[:15]
    if duration_rows:
        row = _write_multi_series_summary_chart(
            sheet, row, f"Контроли дюрации (топ-{len(duration_rows)} по использованию)",
            ["Позиция (график)", "Полная позиция", "Модифицированная", "Лимит", "Запас"], duration_rows,
            numeric_formats={3: "0.00;[Red](0.00);-", 4: "0.00;[Red](0.00);-", 5: "0.00;[Red](0.00);-"},
            value_col_first=3, value_col_last=4, chart_title="Модифицированная дюрация против лимита", widths=[32, 44, 20, 16, 16], chart_number_format="0.00", horizontal=True,
        )

    if breach_count:
        breach_rows = [
            [record.payload.get("portfolio_code") or "Недоступно", _RISK_DIMENSION_LABELS.get(record.payload.get("dimension"), record.payload.get("dimension")), record.payload.get("label") or "Недоступно"]
            for record in records if record.payload.get("signal") == "breach"
        ]
        sheet.cell(row, 1, "Превышенные лимиты").font = Font(bold=True)
        header_row = row + 1
        for column, label in enumerate(["Портфель", "Измерение", "Наименование"], 1):
            cell = sheet.cell(header_row, column, label)
            cell.font = Font(bold=True); cell.border = _BORDER
        for offset, values in enumerate(breach_rows, header_row + 1):
            for column, value in enumerate(values, 1):
                cell = sheet.cell(offset, column, value)
                cell.border = _BORDER
        for column, width in enumerate([16, 24, 42], 1):
            sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, width)
    else:
        sheet.cell(row, 1, "Превышений лимита не обнаружено.").font = Font(italic=True, color="666666")


_RISK_DIMENSION_LABELS = {
    "instrument_category": "Класс инструментов",
    "country": "Страна",
    "issuer": "Эмитент",
    "sector": "Отрасль GICS",
    "ifrs": "Классификация МСФО",
    "currency": "Валюта",
    "fx_position": "Открытая валютная позиция",
    "instrument_issuer": "Инструмент одного эмитента",
    "duration": "Дюрация",
    "exposure_detail": "Детализация экспозиции",
    "country_instrument_detail": "Детализация по странам",
}
