"""Cross-domain constants and low-level Excel-writing helpers.

Shared by every ``_write_<domain>_summary`` writer and by
``create_module_xlsx`` itself.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from osip_dashboard.persistence.models import DatasetVersion
from osip_dashboard.services.date_provenance import filename_date_warning
from osip_dashboard.services.excel_charts import chart_series, write_bar_chart, write_line_chart, write_pie_chart
from osip_dashboard.services.fx_rates import FxRate


_BORDER = Border(**{side: Side(style="thin", color="000000") for side in ("left", "right", "top", "bottom")})
# Matches Excel's built-in "Bad" cell style - flags a row without needing to
# guess at execution_status wording, since failure_reason is only ever
# populated by ingestion when a trade actually failed (see multi_source.py).
_FAILED_TRADE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# Source workbooks store these rates as percentage points (for example, 0.4051
# means 0.4051%), not Excel fractions.  Keep the numeric value unchanged and
# quote the sign so Excel displays the source scale without multiplying it.
_POINTS_PERCENT_FORMAT = '0.00"%";[Red](0.00"%");-'
_RISK_COLUMN_WIDTHS = {
    # Risk exports contain large KZT/USD balances.  Header-length sizing
    # leaves these columns at ~14 characters, which makes Excel render valid
    # numeric values as ####.  Keep the source values numeric and give the
    # rendered values enough room for grouping separators and decimals.
    "limit_kzt": 20,
    "actual_kzt": 20,
    "free_limit_kzt": 22,
    "limit_usd": 20,
    "actual_usd": 20,
    "free_limit_usd": 22,
    "amount_native": 20,
    "amount_kzt": 20,
    "fx_rate": 16,
    "duration_limit": 18,
    "modified_duration": 22,
    "duration_headroom": 20,
}
# risk_limits_sobstv/tabys' raw sheet is split one-sheet-per-dimension (see
# create_module_xlsx) rather than one combined table - SOBSTV alone spans 8
# structurally different control sheets in the source (limit %/KZT/USD
# dimensions, duration, and two informational detail dimensions with no
# limit at all), and cramming them into one table made most cells
# "Недоступно" for whichever fields belonged to a different dimension.
_RISK_CONTROL_FIELDS = [
    "portfolio_code", "label", "limit_pct", "limit_kzt", "actual_pct", "actual_kzt", "free_limit_kzt",
    "limit_usd", "actual_usd", "free_limit_usd", "signal", "utilization", "near_breach",
    "near_breach_threshold", "near_breach_policy_version", "base_label",
]
_RISK_DIMENSION_FIELDS: dict[str, list[str]] = {
    "instrument_category": _RISK_CONTROL_FIELDS,
    "country": _RISK_CONTROL_FIELDS,
    "issuer": _RISK_CONTROL_FIELDS,
    "sector": _RISK_CONTROL_FIELDS,
    "ifrs": _RISK_CONTROL_FIELDS,
    "currency": _RISK_CONTROL_FIELDS,
    "fx_position": _RISK_CONTROL_FIELDS,
    "instrument_issuer": _RISK_CONTROL_FIELDS,
    # Duration is measured as modified duration vs. a limit, not a %/KZT/USD
    # basis - none of _RISK_CONTROL_FIELDS' limit columns ever apply to it.
    "duration": ["portfolio_code", "label", "issuer", "isin", "country", "currency", "duration_limit", "modified_duration", "duration_headroom", "signal"],
    # Informational detail, not a limit control - no limit/signal fields.
    "exposure_detail": ["portfolio_code", "label", "currency", "instrument_group", "amount_native", "fx_rate", "amount_kzt"],
    "country_instrument_detail": ["portfolio_code", "label", "country", "currency", "instrument_category", "amount_native", "amount_kzt"],
}
# Same order as displayed on the website's dimension dropdown (see
# DomainCharts.tsx); keeps generated sheet order predictable.
_RISK_DIMENSION_SHEET_ORDER = [
    "instrument_category", "country", "issuer", "sector", "ifrs",
    "currency", "fx_position", "instrument_issuer", "duration",
    "exposure_detail", "country_instrument_detail",
]
_RISK_DIMENSION_SHEET_LABELS = {
    "instrument_category": "Классы инструментов", "country": "Страны", "issuer": "Эмитенты",
    "sector": "Отрасли", "ifrs": "МСФО", "currency": "Валюты", "fx_position": "Валютная позиция",
    "instrument_issuer": "Инстр. одного эмитента", "duration": "Дюрация",
    "exposure_detail": "Расшифровка", "country_instrument_detail": "Детали по странам",
}
_LABELS = {
    "instrument": "Инструмент", "isin": "ISIN", "issuer": "Эмитент", "quantity": "Количество",
    "currency": "Валюта", "purchase_value_kzt": "Стоимость приобретения, KZT", "nav_kzt": "СЧА, KZT",
    "unit_value_kzt": "Стоимость пая, KZT", "trade_date": "Дата сделки", "side": "Сторона",
    "client_name": "Клиент", "account": "Лицевой счёт", "iin": "ИИН", "branch": "Филиал",
    "manager": "Менеджер", "venue": "Торговая площадка", "amount": "Сумма", "execution_status": "Статус исполнения",
    "issuer": "Эмитент", "subject": "Предмет", "placement_raw": "Объём размещения (исходный текст)",
    "demand_raw": "Удовлетворённый спрос (исходный текст)", "fee_received_kzt": "Полученное вознаграждение, KZT",
    "duration_raw": "Срок (исходный текст)", "active": "Действующий", "sheet": "Лист",
    "rows": "Строки", "columns": "Колонки", "formula_error_count": "Ошибки формул",
    "instrument_type": "Тип инструмента", "identifier": "Идентификатор", "market": "Рынок",
    "underlying": "Базовый актив / рейтинг", "counterparty": "Контрагент", "settlement_date": "Дата расчёта",
    "obligation_status": "Статус обязательства", "maturity_date": "Дата погашения", "coupon_payment_date": "Дата выплаты купона",
    "coupon_percent": "Купон, %", "days_to_maturity": "Дней до погашения", "value_kzt": "Стоимость, KZT",
    "client_type": "Тип клиента", "securities_value_kzt": "Стоимость бумаг, KZT", "cash_share": "Доля денег", "income": "Доход",
    "trade_number": "№ сделки", "security_type": "Вид ЦБ", "failure_reason": "Причина неисполнения",
    "cash_kzt": "Денежные средства, KZT", "total_assets_kzt": "Активы всего, KZT",
    "record_type": "Тип записи", "market_value_kzt": "Рыночная стоимость, KZT",
    "date": "Дата", "units": "Паи", "investors": "Инвесторов",
    "source_name": "Клиент (исходное имя)", "normalized_name": "Клиент (нормализовано)",
    "open_date": "Дата открытия", "match_status": "Статус сопоставления",
    "trade_weight_percent": "Вес оборота в валюте, %",
    "holding_weight_percent": "Вес в портфеле, %",
    "citizenship": "Гражданство", "resident": "Резидент", "economic_sector": "Отраслевой сектор",
    "document_type": "Тип документа", "category": "Категория", "agent": "Агент",
    "security_code": "Код ценной бумаги", "nominal": "Номинал", "nominal_currency": "Валюта номинала",
    "market_price": "Рыночная цена", "price_currency": "Валюта цены",
    "portfolio_code": "Портфель", "dimension": "Измерение", "label": "Наименование",
    "limit_pct": "Лимит, %", "limit_kzt": "Лимит, KZT", "actual_pct": "Факт, %", "actual_kzt": "Факт, KZT",
    "free_limit_kzt": "Свободный лимит, KZT", "limit_usd": "Лимит, USD", "actual_usd": "Факт, USD",
    "free_limit_usd": "Свободный лимит, USD", "signal": "Сигнал", "base_label": "База лимита",
    "utilization": "Использование лимита", "near_breach": "Близко к превышению", "near_breach_threshold": "Порог близости к превышению",
    "near_breach_policy_version": "Версия политики порога близости к превышению",
    "duration_limit": "Лимит дюрации", "modified_duration": "Модифицированная дюрация",
    "duration_headroom": "Запас по дюрации", "country": "Страна", "instrument_category": "Категория инструмента",
    "instrument_group": "Группа инструмента", "amount_native": "Сумма в валюте", "fx_rate": "Курс",
    "amount_kzt": "Сумма, KZT",
    "line_code": "Код строки", "line_label": "Наименование статьи", "section": "Раздел",
    "current_period_kzt": "На конец периода, тыс. KZT", "prior_period_kzt": "На начало периода, тыс. KZT",
    "quarter_kzt": "За отчетный квартал, тыс. KZT", "ytd_kzt": "С начала года, тыс. KZT",
    "prior_quarter_kzt": "За аналогичный квартал пред. года, тыс. KZT", "prior_ytd_kzt": "С начала пред. года, тыс. KZT",
    "year_2023_kzt": "2023, тыс. KZT", "year_2024_kzt": "2024, тыс. KZT",
    "budget_9m_2025_kzt": "Бюджет 9М 2025, тыс. KZT", "actual_9m_2025_kzt": "Факт 9М 2025, тыс. KZT",
    "budget_2025_kzt": "Бюджет 2025, тыс. KZT", "oct_2025_kzt": "Окт 2025, тыс. KZT",
    "nov_2025_kzt": "Ноя 2025, тыс. KZT", "dec_2025_kzt": "Дек 2025, тыс. KZT",
    "forecast_2025_kzt": "Прогноз 2025, тыс. KZT", "execution_pct": "% исполнения", "deviation_kzt": "Отклонение, тыс. KZT",
    "category": "Категория", "security_code": "Код ЦБ", "coupon_rate": "Ставка купона/репо",
    "purchase_price": "Цена покупки", "purchase_value_kzt": "Объём покупки, KZT",
    "carrying_value_kzt": "Балансовая стоимость, KZT", "market_value_kzt": "Рыночная стоимость, KZT",
    "reserve_kzt": "Сумма резерва, KZT", "accrued_income_kzt": "Накопленный доход, KZT",
}


def _open_charts_sheet(workbook: Workbook, title: str, description: str) -> tuple[Any, int]:
    """Get-or-append onto the workbook's single "Данные графиков" sheet.

    The brokerage export now bundles the clients summary onto the same
    workbook (see routes/multi_source.py), so a second summary section
    writing its own "Данные графиков" sheet would otherwise collide with
    the first one's (openpyxl silently renames the duplicate rather than
    erroring, which just hides a second, out-of-order chart-data sheet).
    Appending below the existing content keeps everything on one sheet.
    """
    if "Данные графиков" in workbook.sheetnames:
        sheet = workbook["Данные графиков"]
        start_row = sheet.max_row + 3
    else:
        sheet = workbook.create_sheet("Данные графиков")
        sheet.sheet_view.showGridLines = False
        # No frozen panes - see the comment on the module summary sheets
        # (e.g. _write_accounting_summary): this is chart-source data, read
        # chart-first, not a table meant to be scrolled with a pinned label
        # column. Documented exception, enforced by
        # tests/export_compliance.py's _UNFROZEN_DASHBOARD_SHEETS.
        start_row = 1
    sheet.cell(start_row, 1, title).font = Font(bold=True)
    sheet.cell(start_row + 1, 1, description).alignment = Alignment(wrap_text=True)
    sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=4)
    return sheet, start_row + 3


def _matches_term(record: Any, normalized_term: str) -> bool:
    if not normalized_term:
        return True
    return any(normalized_term in str(value).casefold() for value in record.payload.values())


def _payload_date(record: Any, field: str) -> date | None:
    """Read an ISO date from a dataset row without guessing formats."""
    value = record.payload.get(field)
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


def _summary_decimal(value: Any) -> Decimal:
    """Summary dicts store amounts as strings (see ingestion's
    ``_decimal_text``); convert back for charting/number formatting."""
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal("0")


def _security_class_bucket(security_type: str) -> str:
    """Classify a trade's security type as Stocks/Bonds/Other by keyword.

    Mirrors the frontend's ``securityClassBucket`` in DomainCharts.tsx so the
    Excel and web charts agree on the same split. GDR/ADR count as stocks
    (depositary receipts of equity); funds and index instruments fall to
    "Прочее" rather than being guessed into either bucket.
    """
    normalized = security_type.strip().casefold()
    if any(token in normalized for token in ("облигац", "гцб", "евронот")):
        return "Облигации"
    if any(token in normalized for token in ("акци", "gdr", "adr")):
        return "Акции"
    return "Прочее"


def _fx_method_note(rate_date: date | None, currencies: list[str], fx_by_currency: dict[str, FxRate | None]) -> str:
    """Document the KZT-equivalent conversion method directly on the sheet.

    The turnover table aggregates trades across the whole published period,
    not a single day, so there is no one "correct" trade-level rate to apply
    - this converts the period's total using one dated NBK rate per
    currency, which is an approximation for comparison, not an official
    conversion. That trade-off, the reference date, and each resolved (or
    unavailable) rate are spelled out here rather than left implicit.
    """
    if rate_date is None:
        return "Эквивалент в KZT не рассчитан: у набора данных нет бизнес-даты для запроса курса НБК."
    parts = []
    for currency in currencies:
        if currency == "KZT":
            continue
        fx = fx_by_currency.get(currency)
        if fx is None:
            parts.append(f"{currency} = недоступно (лента НБК недоступна)")
        else:
            fallback_note = " (оффлайн-резерв)" if fx.fallback else ""
            parts.append(f"{currency} = {fx.rate:,.4f} KZT на {fx.effective_date.isoformat()}{fallback_note}")
    rates_text = "; ".join(parts) if parts else "нет валют, отличных от KZT"
    return (
        f"Эквивалент в KZT рассчитан по официальному курсу Национального Банка Казахстана (nationalbank.kz) "
        f"на бизнес-дату набора данных {rate_date.isoformat()}, применённому ко всей сумме периода целиком — "
        f"не к каждой сделке по её собственной дате. Это приближение для сопоставления величин, не официальная "
        f"конвертация. Курсы: {rates_text}."
    )


def _write_summary_table_with_chart(
    sheet, start_row: int, section_title: str, headers: list[str], rows: list[list[Any]],
    *, numeric_formats: dict[int, str], chart_column: int, chart_title: str, chart_kind: str,
    widths: list[int], log_scale: bool = False, add_weight: bool = False, render_chart: bool = True,
) -> int:
    """A small bordered table plus a pie or bar chart; returns the next free row."""
    display_headers = list(headers)
    display_rows = [list(values) for values in rows]
    display_formats = dict(numeric_formats)
    display_widths = list(widths)
    if add_weight:
        # The weight is derived from the same rows shown in this table.  This
        # is appropriate for composition/count/fee mixes, but not for event
        # schedules or tables whose values are deductions rather than parts
        # of a whole.
        total = sum(
            (value for value in (row_values[chart_column - 1] for row_values in rows)
             if isinstance(value, Decimal)),
            Decimal("0"),
        )
        display_headers.append("Вес, %")
        for row_values in display_rows:
            value = row_values[chart_column - 1] if chart_column <= len(row_values) else None
            row_values.append(value / total if isinstance(value, Decimal) and total else "Недоступно")
        weight_column = len(display_headers)
        display_formats[weight_column] = "0.0%;[Red](0.0%);-"
        display_widths.append(14)
    sheet.cell(start_row, 1, section_title).font = Font(bold=True, color="000000")
    header_row = start_row + 1
    for column, label in enumerate(display_headers, 1):
        cell = sheet.cell(header_row, column, label)
        cell.font = Font(bold=True, color="000000")
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for offset, values in enumerate(display_rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(offset, column, value)
            cell.border = _BORDER
            if column in display_formats and isinstance(value, Decimal):
                cell.number_format = display_formats[column]
    last_row = header_row + len(display_rows)
    for column, width in enumerate(display_widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, width)
    if rows and render_chart:
        anchor = f"{get_column_letter(len(display_headers) + 2)}{start_row}"
        if chart_kind == "pie":
            labels, values = chart_series(rows, chart_column)
            write_pie_chart(
                sheet, label_col=1, value_col=chart_column, first_row=header_row + 1, last_row=last_row,
                values=values, title=chart_title, anchor=anchor,
            )
        elif chart_kind == "line":
            write_line_chart(
                sheet, label_col=1, value_col=chart_column, header_row=header_row,
                first_row=header_row + 1, last_row=last_row, title=chart_title, anchor=anchor,
            )
        else:
            write_bar_chart(
                sheet, label_col=1, value_col_first=chart_column, value_col_last=chart_column,
                header_row=header_row, first_row=header_row + 1, last_row=last_row,
                title=chart_title, anchor=anchor, log_scale=log_scale,
            )
    # Charts are 9cm tall (see excel_charts.py); a plain "+3 rows" gap left
    # successive charts anchored inside the previous one's still-drawn area
    # - confirmed against real Excel. Leave enough rows for the chart itself,
    # not just the table beneath it.
    return max(last_row + 3, start_row + 18) if rows else last_row + 3


def _write_multi_series_summary_chart(
    sheet, start_row: int, section_title: str, headers: list[str], rows: list[list[Any]],
    *, numeric_formats: dict[int, str], value_col_first: int, value_col_last: int,
    chart_title: str, widths: list[int], chart_number_format: str = "#,##0", horizontal: bool = False,
    render_chart: bool = True, min_value_padding: float | None = None,
) -> int:
    """Write a summary table and a clustered chart with multiple value series."""
    sheet.cell(start_row, 1, section_title).font = Font(bold=True, color="000000")
    header_row = start_row + 1
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, label)
        cell.font = Font(bold=True, color="000000")
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for offset, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(offset, column, value)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column == 1)
            if isinstance(value, Decimal) and column in numeric_formats:
                cell.number_format = numeric_formats[column]
    last_row = header_row + len(rows)
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, width)
    if rows and render_chart:
        anchor = f"{get_column_letter(len(headers) + 2)}{start_row}"
        write_bar_chart(
            sheet, label_col=1, value_col_first=value_col_first, value_col_last=value_col_last,
            header_row=header_row, first_row=header_row + 1, last_row=last_row,
            title=chart_title, anchor=anchor, y_axis_number_format=chart_number_format, horizontal=horizontal,
            row_separator_gridlines=horizontal, min_value_padding=min_value_padding,
        )
    # Keep the next summary section below the full chart frame.  Horizontal
    # charts are intentionally taller than the legacy compact charts, so use a
    # larger footprint and a visible five-row breathing space after the table.
    chart_footprint = 30 if horizontal else 22
    return max(last_row + 5, start_row + chart_footprint) if rows else last_row + 5


def _write_chart_source_pie(
    charts_sheet, charts_row: int, *, title: str, table_name: str,
    headers: tuple[str, str], rows: list[list[Any]], anchor: str, anchor_worksheet,
) -> int:
    """Store a native pie chart's editable values on the final chart-data sheet.

    Presentation sheets retain the full readable table, while every chart
    series in a brokerage export points to the final ``Данные графиков``
    sheet.  This prevents an Excel user changing chart data by accident in a
    visible operational table and keeps all chart input together.
    """
    if not rows:
        return charts_row
    charts_sheet.cell(charts_row, 1, title).font = Font(bold=True)
    header_row = charts_row + 1
    for column, label in enumerate(headers, 1):
        charts_sheet.cell(header_row, column, label).font = Font(bold=True)
    for offset, (label, value) in enumerate(rows, header_row + 1):
        charts_sheet.cell(offset, 1, label)
        charts_sheet.cell(offset, 2, value).number_format = "#,##0.00;[Red](#,##0.00);-"
    last_row = header_row + len(rows)
    table = Table(displayName=table_name, ref=f"A{header_row}:B{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    charts_sheet.add_table(table)
    charts_sheet.column_dimensions["A"].width = max(charts_sheet.column_dimensions["A"].width or 0, 28)
    charts_sheet.column_dimensions["B"].width = max(charts_sheet.column_dimensions["B"].width or 0, 22)
    labels, values = chart_series(rows, chart_column=2)
    write_pie_chart(
        charts_sheet,
        label_col=1,
        value_col=2,
        first_row=header_row + 1,
        last_row=last_row,
        values=values,
        title=title,
        anchor=anchor,
        anchor_worksheet=anchor_worksheet,
    )
    return max(last_row + 3, charts_row + 18)


def _write_full_mix_table(
    sheet, row: int, section_title: str, mix: dict[str, int] | dict[str, Decimal], *,
    charts_sheet, charts_row: int, table_name: str, chart_title: str, log_scale: bool = False,
    value_header: str = "Количество", numeric_format: str = "#,##0", chart_kind: str = "bar",
) -> tuple[int, int]:
    """Write every category in ``mix`` to ``sheet`` (no "Прочее" folding),
    and a separate <2%-grouped chart-source Table to ``charts_sheet``. The
    native chart is anchored on ``sheet`` beside its table. Skips the chart
    entirely (table only) when there's just one category - a single bar has
    nothing to compare itself against and only wastes space (confirmed
    against a real export where every trade shared one execution status).
    Returns (next row on ``sheet``, next row on ``charts_sheet``)."""
    raw_rows = [[label, value if isinstance(value, Decimal) else Decimal(value)] for label, value in sorted(mix.items(), key=lambda pair: pair[1], reverse=True)]
    total_value = sum((values[1] for values in raw_rows), Decimal("0"))
    display_rows = [values + [values[1] / total_value if total_value else "Недоступно"] for values in raw_rows]
    sheet.cell(row, 1, section_title).font = Font(bold=True)
    header_row = row + 1
    for column, label in enumerate(["Категория", value_header, "Вес, %"], 1):
        cell = sheet.cell(header_row, column, label)
        cell.font = Font(bold=True); cell.border = _BORDER
    for offset, values in enumerate(display_rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(offset, column, value)
            cell.border = _BORDER
            if isinstance(value, Decimal):
                cell.number_format = "0.0%;[Red](0.0%);-" if column == 3 else numeric_format
    last_row = header_row + len(raw_rows)
    for column, width in enumerate([32, 16, 14], 1):
        sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, width)

    next_charts_row = charts_row
    if len(raw_rows) >= 2:
        grouped_labels, grouped_values = chart_series(raw_rows, chart_column=2)
        charts_sheet.cell(charts_row, 1, chart_title).font = Font(bold=True)
        charts_header_row = charts_row + 1
        charts_sheet.cell(charts_header_row, 1, "Категория")
        charts_sheet.cell(charts_header_row, 2, value_header)
        for offset, (label, value) in enumerate(zip(grouped_labels, grouped_values)):
            charts_sheet.cell(charts_header_row + 1 + offset, 1, label)
            charts_sheet.cell(charts_header_row + 1 + offset, 2, value).number_format = numeric_format
        charts_last_row = charts_header_row + len(grouped_labels)
        table = Table(displayName=table_name, ref=f"A{charts_header_row}:B{charts_last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        charts_sheet.add_table(table)
        for column, width in (("A", 34), ("B", 14)):
            charts_sheet.column_dimensions[column].width = max(charts_sheet.column_dimensions[column].width or 0, width)
        if chart_kind == "pie":
            write_pie_chart(
                charts_sheet, label_col=1, value_col=2, first_row=charts_header_row + 1, last_row=charts_last_row,
                values=grouped_values, title=chart_title, anchor=f"E{row}", anchor_worksheet=sheet,
            )
        else:
            write_bar_chart(
                charts_sheet, label_col=1, value_col_first=2, value_col_last=2, header_row=charts_header_row,
                first_row=charts_header_row + 1, last_row=charts_last_row, title=chart_title,
                anchor=f"E{row}", anchor_worksheet=sheet, log_scale=log_scale,
            )
        # Charts are 9cm tall. Reserve their full footprint plus a small
        # blank margin before the next table; a plain "+3 rows" gap leaves
        # the next chart visually touching or overlapping the previous one.
        next_charts_row = max(charts_last_row + 3, charts_row + 20)
        # The table itself also needs to clear the chart anchored beside it.
        return max(last_row + 3, row + 20), next_charts_row
    return last_row + 3, next_charts_row


def _title(sheet, dataset: DatasetVersion, *, workbook_report_date: date | None = None) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Portfolio Operations Insight — контролируемая выгрузка"
    sheet["A1"].font = Font(bold=True, size=14)
    metadata = [
        ("Набор", dataset.dataset_type), ("Область", dataset.scope_code),
        ("Бизнес-дата", dataset.business_date or "Недоступно"), ("Дата источника", dataset.source_report_date or "Недоступно"),
        ("Версия", dataset.version), ("Исходный файл", dataset.source_upload.original_filename),
    ]
    for index, (label, value) in enumerate(metadata, 2):
        sheet.cell(index, 1, label).font = Font(bold=True)
        sheet.cell(index, 2, value)
    sheet["D2"] = "Операционные данные из опубликованной версии; не официальный NAV, доходность или бухгалтерская отчётность."
    sheet["D2"].alignment = Alignment(wrap_text=True)
    internal_date = workbook_report_date or dataset.source_report_date or dataset.business_date
    warning = filename_date_warning(dataset.source_upload.original_filename, internal_date)
    if warning:
        sheet.merge_cells("D3:F3")
        sheet["D3"] = warning
        sheet["D3"].font = Font(bold=True, color="9C0006", size=9)
        sheet["D3"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[3].height = 28

def _fields(dataset_type: str, records: list[Any], *, include_client_details: bool = True) -> list[str]:
    # Split into one sheet per dimension (see create_module_xlsx) - each
    # dimension has its own field list in _RISK_DIMENSION_FIELDS rather than
    # one combined list, since e.g. duration's fields never apply to a
    # country-limit row and vice versa.
    risk_dimension_prefix = next(
        (prefix for prefix in ("risk_limits_sobstv_", "risk_limits_tabys_") if dataset_type.startswith(prefix)),
        None,
    )
    if risk_dimension_prefix is not None:
        preferred = _RISK_DIMENSION_FIELDS.get(dataset_type.removeprefix(risk_dimension_prefix), [])
    else:
        preferred = {
            "fund_holdings": ["instrument", "isin", "issuer", "quantity", "currency", "purchase_value_kzt"],
            "fund_valuation": ["nav_kzt", "unit_value_kzt", "currency"],
            "fund_unit_series": ["date", "nav_kzt", "units", "investors", "unit_value_kzt"],
            "brokerage_trade_ledger": [
                # Do not put client identity, account or counterparty fields in a
                # Brokerage export. Those belong to the Clients export and can be
                # requested there by the local client-operations operator.
                "trade_number", "trade_date", "settlement_date", "side", "venue", "instrument", "issuer",
                "security_type", "isin", "quantity", "amount", "currency", "execution_status", "failure_reason",
            ],
            "derivatives_register": ["instrument_type", "identifier", "market", "underlying", "counterparty", "quantity", "amount", "currency", "settlement_date", "obligation_status"],
            # Split into two sheets by record_type (see create_module_xlsx) -
            # each pseudo dataset_type below only ever contains one row shape,
            # so record_type itself and the other shape's fields are omitted
            # rather than showing as "Недоступно" on every row.
            "client_account_snapshot_client": ["client_name", "account", "iin", "citizenship", "resident", "economic_sector", "document_type", "branch", "category", "agent", "manager", "cash_kzt", "total_assets_kzt"],
            "client_account_snapshot_position": ["client_name", "account", "issuer", "security_type", "security_code", "isin", "nominal", "nominal_currency", "quantity", "market_price", "price_currency", "market_value_kzt"],
            # match_status ("exact"/"ambiguous"/"unmatched") is an internal
            # reconciliation signal against this workbook's own client registry -
            # nothing downstream depends on it and this is a local dashboard, not
            # a formal reconciliation process, so it's dropped from the export
            # rather than surfacing an "unmatched" label with no action attached.
            "client_open_dates": ["source_name", "normalized_name", "open_date"],
            "client_maturity_calendar": ["client_name", "manager", "instrument", "isin", "maturity_date", "coupon_payment_date", "coupon_percent", "days_to_maturity", "value_kzt"],
            "client_dashboard_snapshot": ["client_name", "manager", "client_type", "opening_date", "cash_kzt", "securities_value_kzt", "total_assets_kzt", "cash_share", "income", "status"],
            "corporate_finance_register": ["issuer", "subject", "isins", "placement_raw", "demand_raw", "investors", "commission_rate", "fee_received_kzt", "duration_raw", "active"],
            "accounting_landing": ["sheet", "rows", "columns", "formula_error_count", "dates"],
            "accounting_balance_sheet": ["line_code", "line_label", "section", "current_period_kzt", "prior_period_kzt"],
            "accounting_income_statement": ["line_code", "line_label", "quarter_kzt", "ytd_kzt", "prior_quarter_kzt", "prior_ytd_kzt"],
            "accounting_budget": ["section", "line_label", "year_2023_kzt", "year_2024_kzt", "budget_9m_2025_kzt", "actual_9m_2025_kzt", "budget_2025_kzt", "oct_2025_kzt", "nov_2025_kzt", "dec_2025_kzt", "forecast_2025_kzt", "execution_pct", "deviation_kzt"],
            "accounting_portfolio_detail": ["category", "issuer", "isin", "security_code", "security_type", "currency", "coupon_rate", "nominal", "quantity", "purchase_price", "purchase_value_kzt", "carrying_value_kzt", "market_value_kzt", "reserve_kzt", "accrued_income_kzt"],
        }.get(dataset_type, [])
    # record_type is a DatasetRecord column, not a payload key - it would
    # never appear in `available` (built from payload keys only) even though
    # every record has one, silently dropping it from any preferred list that
    # names it (confirmed: client_account_snapshot's "record_type" column
    # never rendered). Treat it as always available instead.
    #
    # A key existing in payload isn't enough - ingestion always sets a
    # column's key even when the source cell was blank (e.g. every
    # client_account_snapshot row has a "branch" key, even when every
    # single one is ""), so a key-only check let fully-blank columns
    # through as if they had real data. Require at least one non-blank
    # value across these exact records instead.
    available = {
        key for record in records for key, value in record.payload.items()
        if value not in (None, "")
    } | ({"record_type"} if records else set())
    chosen = [field for field in preferred if field in available]
    if dataset_type == "client_maturity_calendar" and not include_client_details:
        chosen = [field for field in chosen if field not in {"client_name", "manager"}]
    # Brokerage is deliberately a strict allowlist: falling back to every
    # source key could reintroduce a client/account field when a later source
    # workbook adds or renames columns.
    if dataset_type == "brokerage_trade_ledger":
        return chosen
    return chosen or sorted(available)


def _field_value(record: Any, field: str) -> Any:
    # record_type lives on the DatasetRecord row itself, not inside payload
    # (see _fields) - resolve it from there instead of always getting None.
    raw = record.record_type if field == "record_type" else record.payload.get(field)
    return _typed(raw, field)


def _typed(value: Any, field: str) -> Any:
    if value is None or value == "": return "Недоступно"
    if isinstance(value, bool): return "Да" if value else "Нет"
    if isinstance(value, list): return ", ".join(str(item) for item in value)
    if field in {"date", "trade_date", "maturity_date", "coupon_payment_date", "opening_date", "settlement_date"} and isinstance(value, str):
        try: return date.fromisoformat(value)
        except ValueError: return value
    if field.endswith("_kzt") or field.endswith("_usd") or field.endswith("_pct") or field in {"quantity", "amount", "units", "unit_value_kzt", "commission_rate", "coupon_percent", "days_to_maturity", "cash_share", "income", "nominal", "market_price", "duration_limit", "modified_duration", "duration_headroom", "amount_native", "fx_rate", "carrying_value_native", "coupon_rate", "purchase_price", "utilization", "near_breach_threshold"}:
        try: return Decimal(str(value))
        except InvalidOperation: return str(value)
    if isinstance(value, (dict, tuple)): return str(value)
    return value


def _sheet_name(dataset_type: str) -> str:
    for prefix, portfolio in (("risk_limits_sobstv_", "SOBSTV"), ("risk_limits_tabys_", "TABYS")):
        if dataset_type.startswith(prefix):
            dimension = dataset_type.removeprefix(prefix)
            label = _RISK_DIMENSION_SHEET_LABELS.get(dimension, dimension)
            # Excel sheet titles reject ":" (and several other characters) -
            # "·" matches the separator already used elsewhere in this export
            # (e.g. the concentration chart titles) for a portfolio/dimension
            # pairing.
            return f"{portfolio} · {label}"[:31]
    names = {
        "fund_valuation": "Оценка фонда", "fund_holdings": "Портфель фонда", "fund_unit_series": "История пая",
        "brokerage_trade_ledger": "Сделки", "derivatives_register": "Производные",
        "client_account_snapshot_client": "Клиенты", "client_account_snapshot_position": "Позиции клиентов",
        "client_open_dates": "Даты открытия", "client_maturity_calendar": "Календарь погашения", "client_dashboard_snapshot": "Сводка клиентов", "corporate_finance_register": "Корпоративные финансы", "accounting_landing": "Бухгалтерские источники",
        "accounting_balance_sheet": "Баланс", "accounting_income_statement": "Прибыли и убытки",
        "accounting_budget": "Бюджет 2023-2025", "accounting_portfolio_detail": "Детализация портфеля",
    }
    return names.get(dataset_type, dataset_type)[:31]
