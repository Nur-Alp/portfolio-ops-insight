"""Controlled Russian XLSX exports for published multi-source datasets."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal
import io
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from osip_dashboard.persistence.models import DatasetVersion
from osip_dashboard.services.brokerage import is_repo_trade
from osip_dashboard.services.excel_safety import neutralize_formulas
# Re-exported (not used directly in this module) so that
# tests/test_multi_source_export_summary.py's
# `monkeypatch.setattr(multi_source_export, "resolve_export_fx_rate", ...)`
# has something to patch - brokerage.py reads this name off this package at
# call time (see the comment in brokerage.py) rather than binding its own
# private copy, so patching it here is actually observed there.
from osip_dashboard.services.fx_rates import resolve_export_fx_rate  # noqa: F401

from .accounting import _write_accounting_summary
from .brokerage import _brokerage_calendar_records, _calendar_as_of, _write_brokerage_summary, _write_brokerage_trade_tables
from .clients import _write_clients_summary
from .corporate_finance import _write_corporate_finance_summary
from .fund import _write_asset_management_summary, _write_unique_etf_positions
from .risk import _write_risk_summary
from .shared import (
    _BORDER,
    _FAILED_TRADE_FILL,
    _LABELS,
    _POINTS_PERCENT_FORMAT,
    _RISK_COLUMN_WIDTHS,
    _RISK_DIMENSION_SHEET_ORDER,
    _fields,
    _field_value,
    _matches_term,
    _sheet_name,
    _summary_decimal,
    _title,
    _write_full_mix_table,
)

# Names that must remain importable exactly as they are today:
#   - routes/multi_source.py: create_module_xlsx
#   - tests/test_multi_source_export_summary.py: _fields, _write_full_mix_table, _write_module_summary
__all__ = ["create_module_xlsx", "_fields", "_write_full_mix_table", "_write_module_summary"]


def create_module_xlsx(module: str, datasets: list[DatasetVersion], term: str = "", include_repo: bool = True) -> bytes:
    normalized_term = term.strip().casefold()
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Portfolio Operations Insight"
    workbook.properties.subject = "Операционные данные из опубликованных наборов; не официальный NAV или бухгалтерская отчётность"
    workbook.properties.created = datetime.now()
    dataset_by_type = {dataset.dataset_type: dataset for dataset in datasets}
    # A derivatives register labels its rows with the reporting-period end,
    # while the workbook itself has the report date carried by another child
    # dataset.  Keep filename warnings tied to that workbook date rather than
    # the period-end label.
    workbook_dates_by_filename = {
        dataset.source_upload.original_filename: dataset.source_report_date
        for dataset in datasets
        if dataset.dataset_type != "derivatives_register" and dataset.source_report_date is not None
    }
    _write_module_summary(workbook, module, dataset_by_type, term=term, include_repo=include_repo)
    sheet_datasets: list[Any] = []
    for dataset in datasets:
        if dataset.dataset_type == "client_account_snapshot":
            # Account-level rows (one per client) and position-level rows
            # (one per holding) are two different shapes of data - stacking
            # them in one table meant every account row showed "Недоступно"
            # for every position-only column and vice versa. Split into two
            # sheets instead; the shared original DatasetVersion (summary,
            # business date, etc.) still backs both.
            for suffix, record_type in (
                ("client", "client"),
                ("position", "client_position"),
            ):
                sheet_datasets.append(SimpleNamespace(
                    dataset_type=f"client_account_snapshot_{suffix}",
                    records=[record for record in dataset.records if record.record_type == record_type],
                    source_upload=dataset.source_upload,
                    business_date=dataset.business_date,
                    source_report_date=dataset.source_report_date,
                    version=dataset.version,
                    scope_code=dataset.scope_code,
                ))
        elif dataset.dataset_type in ("risk_limits_sobstv", "risk_limits_tabys"):
            # One dimension's rows share almost nothing with another's beyond
            # portfolio_code/label/signal - SOBSTV alone spans 8 structurally
            # different source sheets (limit %/KZT/USD controls, duration,
            # two informational detail dimensions with no limit at all).
            # Stacking all of them in one table meant most cells were
            # "Недоступно" for whichever fields belonged to a different
            # dimension. Split into one sheet per dimension instead, same
            # reasoning as the client_account_snapshot split above.
            by_dimension: dict[str, list[Any]] = {}
            for record in dataset.records:
                by_dimension.setdefault(str(record.payload.get("dimension") or ""), []).append(record)
            for dimension in _RISK_DIMENSION_SHEET_ORDER:
                dimension_records = by_dimension.get(dimension)
                if not dimension_records:
                    continue
                sheet_datasets.append(SimpleNamespace(
                    dataset_type=f"{dataset.dataset_type}_{dimension}",
                    records=dimension_records,
                    source_upload=dataset.source_upload,
                    business_date=dataset.business_date,
                    source_report_date=dataset.source_report_date,
                    version=dataset.version,
                    scope_code=dataset.scope_code,
                ))
        else:
            sheet_datasets.append(dataset)
    for dataset in sheet_datasets:
        if dataset.dataset_type == "brokerage_trade_ledger":
            # A full ledger dump can run to thousands of rows spanning years -
            # verbatim, that's a table nobody reads end to end. Two small,
            # curated tables (what happened lately, what's worth a second
            # look) instead; complete row-by-row history remains queryable in
            # the source workbook itself (see the "Исходный файл" metadata
            # line plus each row's "Лист/Строка").
            _write_brokerage_trade_tables(
                workbook, dataset, term=term, include_repo=include_repo,
                workbook_report_date=workbook_dates_by_filename.get(dataset.source_upload.original_filename),
            )
            continue
        sheet = workbook.create_sheet(_sheet_name(dataset.dataset_type))
        _title(sheet, dataset, workbook_report_date=workbook_dates_by_filename.get(dataset.source_upload.original_filename))
        records = [record for record in dataset.records if _matches_term(record, normalized_term) and (include_repo or not is_repo_trade(record.payload))]
        if dataset.dataset_type == "accounting_portfolio_detail":
            # Cash balances (currency, custodian, amount_kzt) share this
            # dataset_type but not this table's columns (category, isin,
            # carrying_value_kzt, ...) - left mixed in here they rendered as
            # rows of near-total "Недоступно" cells. _write_accounting_summary
            # (accounting.py) gives them their own small table instead,
            # mirroring the web UI's separate "Portfolio cash balances" panel.
            records = [record for record in records if record.record_type != "cash_balance"]
        if module == "brokerage" and dataset.dataset_type == "client_maturity_calendar":
            # Brokerage gets an operational maturity view, not the full
            # client-detail calendar.  Keep the source dataset itself
            # untouched for the Clients export.
            records = _brokerage_calendar_records(records, _calendar_as_of(dataset_by_type))
        fields = _fields(dataset.dataset_type, records, include_client_details=module != "brokerage")
        header_row = 8
        holding_weights: list[Decimal | str] = []
        if dataset.dataset_type == "fund_holdings":
            holding_total = sum((_summary_decimal(record.payload.get("purchase_value_kzt")) for record in records), Decimal("0"))
            holding_weights = [
                (_summary_decimal(record.payload.get("purchase_value_kzt")) / holding_total if holding_total else "Недоступно")
                for record in records
            ]
        extra_headers = ["Вес в портфеле, %"] if dataset.dataset_type == "fund_holdings" else []
        # The source filename is constant for every row on this sheet (one
        # sheet = one dataset = one source_upload) - it's already shown once
        # in the "Исходный файл" metadata line above, so repeating it on
        # every row just wastes column width without adding information.
        # Sheet and row do vary per record and stay here.
        headers = [_LABELS.get(field, field.replace("_", " ")) for field in fields] + extra_headers + ["Лист", "Строка"]
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, label)
            cell.font = Font(bold=True, color="000000")
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for record_index, (row_number, record) in enumerate(zip(range(header_row + 1, header_row + 1 + len(records)), records)):
            source = record.source_ref or {}
            values = [_field_value(record, field) for field in fields]
            if dataset.dataset_type == "brokerage_trade_ledger":
                values.append(trade_weights[record_index])
            elif dataset.dataset_type == "fund_holdings":
                values.append(holding_weights[record_index])
            values += [source.get("sheet_name") or "Недоступно", source.get("row_number") or "Недоступно"]
            failed_trade = dataset.dataset_type == "brokerage_trade_ledger" and bool(str(record.payload.get("failure_reason") or "").strip())
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_number, column, value)
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=column <= len(fields))
                if failed_trade:
                    cell.fill = _FAILED_TRADE_FILL
                field = fields[column - 1] if column <= len(fields) else "trade_weight_percent" if dataset.dataset_type == "brokerage_trade_ledger" and column == len(fields) + 1 else "holding_weight_percent" if dataset.dataset_type == "fund_holdings" and column == len(fields) + 1 else ""
                if isinstance(value, Decimal):
                    cell.number_format = (
                        "0.0%;[Red](0.0%);-" if field in {"utilization", "near_breach_threshold"}
                        else _POINTS_PERCENT_FORMAT if field in {"coupon_percent", "limit_pct", "actual_pct"}
                        else "0.0%;[Red](0.0%);-" if field in {"trade_weight_percent", "holding_weight_percent"}
                        else "#,##0;[Red](#,##0);-" if field in {"quantity", "days_to_maturity"}
                        else "#,##0.0000;[Red](#,##0.0000);-" if field in {"unit_value_kzt", "units"}
                        else "#,##0.00;[Red](#,##0.00);-"
                    )
                elif isinstance(value, (date, datetime)):
                    cell.number_format = "dd.mm.yyyy"
        last_row = max(header_row, header_row + len(records))
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
        # Keep the table header/context rows visible and pin column A during
        # horizontal review of wide source exports.
        sheet.freeze_panes = f"B{header_row + 1}"
        for column, label in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(column)].width = min(48, max(14, len(label) + 4))
        if dataset.dataset_type.startswith(("risk_limits_sobstv_", "risk_limits_tabys_")):
            for field, width in _RISK_COLUMN_WIDTHS.items():
                if field in fields:
                    column = fields.index(field) + 1
                    sheet.column_dimensions[get_column_letter(column)].width = max(
                        sheet.column_dimensions[get_column_letter(column)].width or 0,
                        width,
                    )
        if dataset.dataset_type == "brokerage_trade_ledger":
            sheet.cell(7, 1, "Вес каждой сделки рассчитан внутри пары «валюта + сторона» по сумме сделки; валюты между собой не складываются.")
            sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=max(1, len(headers)))
            sheet.cell(7, 1).font = Font(italic=True, color="666666", size=9)
            sheet.cell(7, 1).alignment = Alignment(wrap_text=True)
        elif dataset.dataset_type == "fund_holdings":
            sheet.cell(7, 1, "Вес в портфеле рассчитан как стоимость приобретения строки / сумма стоимости приобретения выбранных строк.")
            sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=max(1, len(headers)))
            sheet.cell(7, 1).font = Font(italic=True, color="666666", size=9)
            sheet.cell(7, 1).alignment = Alignment(wrap_text=True)
        if module == "brokerage" and dataset.dataset_type == "client_maturity_calendar":
            as_of = _calendar_as_of(dataset_by_type)
            end_date = as_of + timedelta(days=180) if as_of is not None else None
            note = (
                f"Период: с {as_of.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')} включительно; "
                # Only this sheet drops client_name/manager (see _fields'
                # include_client_details) - the client-account/position
                # sheets in this same workbook still carry full client
                # identity, IIN, and account numbers. The note must not
                # imply the whole export is identity-free.
                "на этом листе идентификаторы клиентов и менеджеров не показаны "
                "(на листах с клиентскими счетами и позициями эти данные присутствуют)."
                if as_of is not None else
                "Календарь не построен: в источнике отсутствует бизнес-дата для определения шестимесячного периода."
            )
            sheet.cell(7, 1, note)
            sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=max(1, len(headers)))
            sheet.cell(7, 1).font = Font(italic=True, color="666666", size=9)
            sheet.cell(7, 1).alignment = Alignment(wrap_text=True)
    # Chart data is an editable implementation detail, so keep it after the
    # domain tables and source extracts. Charts themselves remain on the
    # module summary sheet and reference this final tab.
    if "Данные графиков" in workbook.sheetnames:
        chart_data_sheet = workbook["Данные графиков"]
        workbook._sheets.remove(chart_data_sheet)
        workbook._sheets.append(chart_data_sheet)
    if not workbook.worksheets:
        workbook.create_sheet("Нет данных")["A1"] = "Нет опубликованных строк"
    neutralize_formulas(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_module_summary(workbook: Workbook, module: str, dataset_by_type: dict[str, DatasetVersion], *, term: str = "", include_repo: bool = True) -> None:
    """One "Сводка" sheet per module, appended before the raw dataset sheets.

    Built entirely from summary aggregates already computed at ingestion
    time (``dataset.summary``) or safe row-level fields - never from a
    currency/unit conversion this app hasn't decided how to do (see the
    corporate-finance placement/demand note below).
    """
    if module == "asset-management":
        _write_asset_management_summary(workbook, dataset_by_type, term=term)
        _write_unique_etf_positions(workbook, dataset_by_type.get("fund_holdings"), term=term)
    elif module == "brokerage":
        _write_brokerage_summary(workbook, dataset_by_type, term=term, include_repo=include_repo)
        _write_clients_summary(workbook, dataset_by_type, term=term)
    elif module == "clients":
        _write_clients_summary(workbook, dataset_by_type, term=term)
    elif module == "corporate-finance":
        _write_corporate_finance_summary(workbook, dataset_by_type, term=term)
    elif module == "risk":
        _write_risk_summary(workbook, dataset_by_type, term=term)
    elif module == "accounting":
        _write_accounting_summary(workbook, dataset_by_type, term=term)
