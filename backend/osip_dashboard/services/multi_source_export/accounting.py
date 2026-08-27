"""Accounting (balance sheet / income statement / budget / portfolio) summary writer."""

from __future__ import annotations

from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from osip_dashboard.persistence.models import DatasetVersion

from .shared import _matches_term, _summary_decimal, _write_multi_series_summary_chart, _write_summary_table_with_chart


def _write_accounting_summary(workbook: Workbook, dataset_by_type: dict[str, DatasetVersion], *, term: str = "") -> None:
    """One "Сводка ФО" sheet: balance sheet and income statement totals plus a reconciliation check.

    Totals are recomputed from the term-filtered records the same way
    corporate finance and risk do it, so the summary always agrees with the
    detail sheets rather than showing an unfiltered ingestion-time figure.
    """
    balance_sheet = dataset_by_type.get("accounting_balance_sheet")
    income_statement = dataset_by_type.get("accounting_income_statement")
    portfolio = dataset_by_type.get("accounting_portfolio_detail")
    has_cash_records = portfolio is not None and any(record.record_type == "cash_balance" for record in portfolio.records)
    if balance_sheet is None and income_statement is None and not has_cash_records:
        # Cash balances (below) don't depend on balance_sheet/income_statement
        # existing - only bail out here when there's truly nothing this sheet
        # would show at all.
        return
    normalized_term = term.strip().casefold()

    def total(dataset: DatasetVersion | None, label: str, field: str) -> Decimal | None:
        if dataset is None:
            return None
        for record in dataset.records:
            if not _matches_term(record, normalized_term):
                continue
            if str(record.payload.get("line_label") or "").strip().casefold() == label:
                value = record.payload.get(field)
                if value not in (None, ""):
                    return Decimal(str(value))
        return None

    assets = total(balance_sheet, "итого активы", "current_period_kzt")
    liabilities = total(balance_sheet, "итого обязательства", "current_period_kzt")
    equity = total(balance_sheet, "итого капитал", "current_period_kzt")
    income = total(income_statement, "итого доходов", "quarter_kzt")
    expenses = total(income_statement, "итого расходов", "quarter_kzt")
    net_profit = total(income_statement, "чистая прибыль (убыток) до уплаты корпоративного подоходного налога", "quarter_kzt")

    # The balance sheet / income statement / budget workbooks state every
    # *_kzt figure in thousands of tenge (source header: "в тысячах тенге" /
    # "в тыс тг") - unlike accounting_portfolio_detail's carrying_value_kzt,
    # which the source states directly "в тенге" (full units). Scale up for
    # the KPI/chart figures below so every KZT number on this sheet is the
    # same real-world unit; the raw per-dataset sheets further down this
    # workbook still show the untouched thousands values, since those mirror
    # the source cell exactly. The reconciliation checks just below compare
    # the unscaled totals to each other, so scaling doesn't affect them.
    def thousand_to_kzt(value: Decimal | None) -> Decimal | None:
        return value * 1000 if value is not None else None

    assets_kzt = thousand_to_kzt(assets)
    liabilities_kzt = thousand_to_kzt(liabilities)
    equity_kzt = thousand_to_kzt(equity)
    income_kzt = thousand_to_kzt(income)
    expenses_kzt = thousand_to_kzt(expenses)
    net_profit_kzt = thousand_to_kzt(net_profit)

    sheet = workbook.create_sheet("Сводка ФО")
    sheet.sheet_view.showGridLines = False
    # No frozen panes: this is a dashboard of stacked tables/charts, read
    # chart-first rather than scrolled as a long table - a frozen column A
    # has nothing to do here. Deliberate, documented exception to the
    # workbook navigation standard (docs/export-column-audit.md), enforced
    # by tests/export_compliance.py's _UNFROZEN_DASHBOARD_SHEETS.
    row = 2
    sheet.cell(row, 1, "Сводка финансовой отчётности").font = Font(bold=True, size=14)
    row += 2

    def kpi(row_index: int, label: str, value: Decimal | None) -> None:
        sheet.cell(row_index, 1, label).font = Font(bold=True)
        cell = sheet.cell(row_index, 2, value if value is not None else "Недоступно")
        if value is not None:
            cell.number_format = "#,##0;[Red](#,##0);-"

    for offset, (label, value) in enumerate((
        ("Активы всего, KZT", assets_kzt), ("Обязательства всего, KZT", liabilities_kzt),
        ("Капитал всего, KZT", equity_kzt), ("Доходы за квартал, KZT", income_kzt),
        ("Расходы за квартал, KZT", expenses_kzt), ("Чистая прибыль за квартал, KZT", net_profit_kzt),
    )):
        kpi(row + offset, label, value)
    row += 8

    balance_check = "Недоступно"
    if assets is not None and liabilities is not None and equity is not None:
        balance_check = "Сходится" if abs(assets - (liabilities + equity)) <= Decimal("1") else "Расхождение"
    income_check = "Недоступно"
    if income is not None and expenses is not None and net_profit is not None:
        income_check = "Сходится" if abs(income - expenses - net_profit) <= Decimal("1") else "Расхождение"
    sheet.cell(row, 1, "Активы = Обязательства + Капитал").font = Font(bold=True)
    sheet.cell(row, 2, balance_check)
    sheet.cell(row + 1, 1, "Доходы − Расходы = Чистая прибыль").font = Font(bold=True)
    sheet.cell(row + 1, 2, income_check)
    row += 3

    if assets_kzt is not None and liabilities_kzt is not None and equity_kzt is not None:
        row = _write_summary_table_with_chart(
            sheet, row, "Состав баланса", ["Статья", "Сумма, KZT"],
            [["Обязательства", liabilities_kzt], ["Капитал", equity_kzt]],
            numeric_formats={2: "#,##0;[Red](#,##0);-"}, chart_column=2,
            chart_title="Обязательства и капитал", chart_kind="pie", widths=[28, 22], add_weight=True,
        )
    if income_kzt is not None and expenses_kzt is not None and net_profit_kzt is not None:
        row = _write_summary_table_with_chart(
            sheet, row, "Доходы, расходы и прибыль", ["Показатель", "Сумма, KZT"],
            [["Доходы", income_kzt], ["Расходы", expenses_kzt], ["Чистая прибыль", net_profit_kzt]],
            numeric_formats={2: "#,##0;[Red](#,##0);-"}, chart_column=2,
            chart_title="Доходы, расходы и чистая прибыль", chart_kind="bar", widths=[28, 22],
        )

    budget = dataset_by_type.get("accounting_budget")
    if budget is not None:
        all_budget_rows = [record for record in budget.records if _matches_term(record, normalized_term)]
        budget_rows = [record for record in all_budget_rows if record.payload.get("section") == "income_statement"]

        # Budget workbook figures are also in thousands of tenge (see the
        # thousand_to_kzt comment above) - scale for these display tables.
        def field_kzt(record: Any, field: str) -> Decimal:
            return _summary_decimal(record.payload.get(field)) * 1000

        def field_kzt_or_none(record: Any, field: str) -> Decimal | None:
            # Unlike field_kzt above (used where every field is expected to
            # be populated), the monthly columns can be genuinely absent for
            # a given month - leave the cell blank rather than writing a
            # fabricated 0, which would be indistinguishable from a real
            # reported zero.
            raw = record.payload.get(field)
            return None if raw in (None, "") else field_kzt(record, field)

        def budget_vs_actual_rows(records_subset: list[Any]) -> list[list[Any]]:
            ranked = sorted(
                (record for record in records_subset if record.payload.get("deviation_kzt") not in (None, "")),
                key=lambda record: abs(_summary_decimal(record.payload.get("deviation_kzt"))), reverse=True,
            )[:12]
            return [
                [
                    record.payload.get("line_label") or "Не указано",
                    field_kzt(record, "budget_2025_kzt"),
                    field_kzt(record, "forecast_2025_kzt"),
                ]
                for record in ranked
            ]

        deviation_rows = budget_vs_actual_rows(budget_rows)
        if deviation_rows:
            row = _write_multi_series_summary_chart(
                sheet, row, "Исполнение бюджета 2025: план и факт (доходы/расходы)", ["Статья", "Бюджет 2025, KZT", "Прогнозный факт 2025, KZT"], deviation_rows,
                numeric_formats={2: "#,##0;[Red](#,##0);-", 3: "#,##0;[Red](#,##0);-"},
                value_col_first=2, value_col_last=3, chart_title="Крупнейшие по отклонению строки: бюджет и факт 2025", widths=[36, 22, 22], horizontal=True,
            )

        cash_flow_rows = [record for record in all_budget_rows if record.payload.get("section") == "cash_flow"]
        cash_flow_deviation_rows = budget_vs_actual_rows(cash_flow_rows)
        if cash_flow_deviation_rows:
            row = _write_multi_series_summary_chart(
                sheet, row, "Исполнение бюджета 2025: план и факт (движение денежных средств)", ["Статья", "Бюджет 2025, KZT", "Прогнозный факт 2025, KZT"], cash_flow_deviation_rows,
                numeric_formats={2: "#,##0;[Red](#,##0);-", 3: "#,##0;[Red](#,##0);-"},
                value_col_first=2, value_col_last=3, chart_title="Крупнейшие по отклонению строки ДДС: бюджет и факт 2025", widths=[36, 22, 22], horizontal=True,
            )

        # Same "largest deviation" ranking as the chart above, just read the
        # monthly columns instead of the annual deviation - shows *when*
        # within Q4 those top lines moved, not just the annual total.
        # Gross figures excluded in favor of their net counterpart so the two
        # don't both take a slot among the top 5 for one underlying concept:
        # "Комиссионные доходы" duplicates "Чистый комиссионный доход" (gross
        # vs net commission), "Процентные доходы" duplicates "Чистый
        # процентный доход" (gross vs net of interest expense).
        monthly_trend_excluded_lines = {"Комиссионные доходы", "Процентные доходы"}
        monthly_candidates = [
            record for record in budget_rows
            if record.payload.get("line_label") not in monthly_trend_excluded_lines
            and (
                record.payload.get("oct_2025_kzt") not in (None, "")
                or record.payload.get("nov_2025_kzt") not in (None, "")
                or record.payload.get("dec_2025_kzt") not in (None, "")
            )
        ]
        monthly_candidates.sort(key=lambda record: abs(_summary_decimal(record.payload.get("deviation_kzt"))), reverse=True)
        # Some lines (e.g. pre-tax vs. net profit, when this quarter's tax is
        # zero) carry the exact same Oct/Nov/Dec figures - two series plotted
        # on identical points draw as one, so the second, lower one is
        # invisible on the chart despite having its own legend entry. Skip
        # any candidate whose full monthly series exactly matches one
        # already picked.
        seen_monthly_series: set[tuple[str, str, str]] = set()
        monthly_top: list[Any] = []
        for candidate in monthly_candidates:
            fingerprint = (
                str(candidate.payload.get("oct_2025_kzt")),
                str(candidate.payload.get("nov_2025_kzt")),
                str(candidate.payload.get("dec_2025_kzt")),
            )
            if fingerprint in seen_monthly_series:
                continue
            seen_monthly_series.add(fingerprint)
            monthly_top.append(candidate)
            if len(monthly_top) == 5:
                break
        if len(monthly_top) >= 2:
            month_labels = ["Октябрь", "Ноябрь", "Декабрь"]
            month_fields = ["oct_2025_kzt", "nov_2025_kzt", "dec_2025_kzt"]
            monthly_headers = ["Месяц"] + [str(record.payload.get("line_label") or "Не указано") for record in monthly_top]
            monthly_rows = [
                [month_label] + [field_kzt_or_none(record, field) for record in monthly_top]
                for month_label, field in zip(month_labels, month_fields)
            ]
            row = _write_multi_series_summary_chart(
                sheet, row, "Динамика по месяцам (Окт-Дек 2025)", monthly_headers, monthly_rows,
                numeric_formats={index: "#,##0;[Red](#,##0);-" for index in range(2, len(monthly_headers) + 1)},
                value_col_first=2, value_col_last=len(monthly_headers),
                chart_title="Крупнейшие по отклонению строки, Окт-Дек 2025", widths=[14] + [24] * len(monthly_top),
            )

        # 2025 dropped from this comparison: the source only gives full-year
        # totals for 2023/2024, while 2025's only available figure
        # (forecast_2025_kzt) mixes actual months with projected ones -
        # keeping it here would put a partially-projected column next to two
        # purely actual ones with no way to flag which was which. Year-over-
        # year stays actual-only; the "Исполнение бюджета" chart above is the
        # one that's explicit about being a plan vs. forecast comparison.
        yoy_candidates = [
            record for record in budget_rows
            if record.payload.get("year_2023_kzt") not in (None, "")
            and record.payload.get("year_2024_kzt") not in (None, "")
        ]
        yoy_candidates.sort(
            key=lambda record: max(
                abs(_summary_decimal(record.payload.get("year_2023_kzt"))),
                abs(_summary_decimal(record.payload.get("year_2024_kzt"))),
            ),
            reverse=True,
        )
        yoy_rows = [
            [
                record.payload.get("line_label") or "Не указано",
                field_kzt(record, "year_2023_kzt"),
                field_kzt(record, "year_2024_kzt"),
            ]
            for record in yoy_candidates[:6]
        ]
        if len(yoy_rows) >= 2:
            row = _write_multi_series_summary_chart(
                sheet, row, "Динамика по годам, 2023-2024", ["Статья", "2023", "2024"], yoy_rows,
                numeric_formats={2: "#,##0;[Red](#,##0);-", 3: "#,##0;[Red](#,##0);-"},
                value_col_first=2, value_col_last=3, chart_title="Крупнейшие строки бюджета по годам", widths=[36, 20, 20], horizontal=True,
            )

    if portfolio is not None:
        portfolio_records = [record for record in portfolio.records if _matches_term(record, normalized_term)]
        # A "portfolio by category" pie used to live here, mirroring the
        # equivalent chart on the web page - that one was deliberately
        # hidden on web ("hidden by request", AccountingCharts.tsx), and
        # this export-only copy was removed to match rather than leave the
        # two channels disagreeing on what data to show.
        # Cash balances (currency, custodian, amount_kzt) share this
        # dataset_type but not the position table's columns - excluded from
        # it in __init__.py to avoid near-blank rows there, given their own
        # small table here instead (mirrors the web UI's separate "Portfolio
        # cash balances" panel). Confirmed this session: the position
        # table's own carrying-value sum falls short of the workbook's
        # stated portfolio total by exactly this cash total.
        cash_rows = [
            [f"{record.payload.get('currency') or '—'} · {record.payload.get('custodian') or '—'}", Decimal(str(record.payload.get("amount_kzt")))]
            for record in portfolio_records
            if record.record_type == "cash_balance" and record.payload.get("amount_kzt") not in (None, "")
        ]
        cash_rows = [row_values for row_values in cash_rows if row_values[1] != 0]
        if cash_rows:
            _write_summary_table_with_chart(
                sheet, row, "Денежные средства портфеля", ["Валюта · Банк/депозитарий", "Сумма, KZT"], cash_rows,
                numeric_formats={2: "#,##0;[Red](#,##0);-"}, chart_column=2,
                chart_title="Денежные средства портфеля", chart_kind="pie", widths=[36, 24], render_chart=len(cash_rows) > 1,
            )
    sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, 32)
    sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 22)
