"""Corporate finance deal register summary writer."""

from __future__ import annotations

from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font

from osip_dashboard.persistence.models import DatasetVersion

from .shared import _matches_term, _summary_decimal, _write_summary_table_with_chart


def _write_corporate_finance_summary(workbook: Workbook, dataset_by_type: dict[str, DatasetVersion], *, term: str = "") -> None:
    register = dataset_by_type.get("corporate_finance_register")
    if register is None:
        return
    # The precomputed dataset.summary (deal_count/active_count) is an
    # ingestion-time aggregate over ALL records - using it here regardless
    # of the search term made every count/chart on this sheet silently
    # disagree with the detail sheet whenever a term filtered it down.
    # Recompute both from the same filtered record set the detail sheet
    # uses instead.
    normalized_term = term.strip().casefold()
    records = [record for record in register.records if _matches_term(record, normalized_term)]
    deal_count = len(records)
    active_count = sum(1 for record in records if record.payload.get("active"))
    sheet = workbook.create_sheet("Сводка сделок")
    sheet.sheet_view.showGridLines = False
    # No frozen panes: this is a dashboard of stacked tables/charts, read
    # chart-first rather than scrolled as a long table - a frozen column A
    # has nothing to do here. Deliberate, documented exception to the
    # workbook navigation standard (docs/export-column-audit.md), enforced
    # by tests/export_compliance.py's _UNFROZEN_DASHBOARD_SHEETS.
    row = 2
    sheet.cell(row, 1, "Сводка сделок").font = Font(bold=True, size=14)
    row += 2
    sheet.cell(row, 1, "Сделок/мандатов").font = Font(bold=True)
    sheet.cell(row, 2, deal_count)
    sheet.cell(row + 1, 1, "Действующие").font = Font(bold=True)
    sheet.cell(row + 1, 2, active_count)
    row += 3
    fee_rows = sorted(
        (
            [record.payload.get("issuer") or "Не указано", _summary_decimal(record.payload.get("fee_received_kzt"))]
            for record in records if record.payload.get("fee_received_kzt")
        ),
        key=lambda values: values[1],
        reverse=True,
    )
    if fee_rows:
        row = _write_summary_table_with_chart(
            sheet, row, "Вклад вознаграждения по сделкам", ["Эмитент", "Вознаграждение, KZT"], fee_rows,
            numeric_formats={2: "#,##0.00;[Red](#,##0.00);-"}, chart_column=2,
            chart_title="Вклад вознаграждения по эмитентам", chart_kind="bar", widths=[36, 24], log_scale=True,
        )
    # Placement/demand remain available as raw source fields on the dataset
    # sheet. They are intentionally not charted until their units/currency
    # are structured; do not add a warning block to the summary layout.
    row += 3

    active_rows = [["Действующие", Decimal(active_count)], ["Не действующие", Decimal(max(deal_count - active_count, 0))]]
    row = _write_summary_table_with_chart(
        sheet, row, "Статус сделок", ["Статус", "Количество"], active_rows,
        numeric_formats={2: "#,##0"}, chart_column=2, chart_title="Действующие и не действующие сделки",
        chart_kind="pie", widths=[28, 18], add_weight=True,
    )

    issuer_counts: dict[str, Decimal] = {}
    for record in records:
        issuer = str(record.payload.get("issuer") or "Не указано")
        issuer_counts[issuer] = issuer_counts.get(issuer, Decimal("0")) + Decimal("1")
    issuer_rows = [[issuer, count] for issuer, count in sorted(issuer_counts.items(), key=lambda pair: (-pair[1], pair[0].casefold()))]
    row = _write_summary_table_with_chart(
        sheet, row, "Количество сделок по эмитентам", ["Эмитент", "Количество сделок"], issuer_rows,
        numeric_formats={2: "#,##0"}, chart_column=2, chart_title="Количество сделок по эмитентам",
        chart_kind="bar", widths=[36, 20], add_weight=True,
    )

    isin_rows = []
    for record in records:
        issuer = str(record.payload.get("issuer") or "Не указано")
        isins = record.payload.get("isins") or []
        isin_rows.append([issuer, Decimal(str(len(isins)))])
    isin_rows.sort(key=lambda values: (-values[1], str(values[0]).casefold()))
    row = _write_summary_table_with_chart(
        sheet, row, "ISIN по сделкам", ["Эмитент", "Количество ISIN"], isin_rows,
        numeric_formats={2: "#,##0"}, chart_column=2, chart_title="Покрытие ISIN по сделкам",
        chart_kind="bar", widths=[36, 20], add_weight=True,
    )
