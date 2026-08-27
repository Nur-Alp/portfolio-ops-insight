"""Brokerage trade-ledger detail tables and summary writer."""

from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from osip_dashboard.persistence.models import DatasetVersion
from osip_dashboard.services.brokerage import is_repo_trade
from osip_dashboard.services.excel_charts import write_bar_chart

# Imported as the parent package (not `from ...fx_rates import
# resolve_export_fx_rate` directly) so that
# tests/test_multi_source_export_summary.py's
# `monkeypatch.setattr(multi_source_export, "resolve_export_fx_rate", ...)`
# - which patches the public package's re-exported name - is actually
# observed here. A direct function import would bind our own private copy
# of the name, invisible to that monkeypatch. Safe under the package's own
# circular import (this module is only imported from within __init__.py,
# by which point `osip_dashboard.services.multi_source_export` is already
# registered in sys.modules with `resolve_export_fx_rate` bound on it).
from osip_dashboard.services import multi_source_export as _multi_source_export_pkg

from .shared import (
    _BORDER,
    _FAILED_TRADE_FILL,
    _LABELS,
    _fields,
    _field_value,
    _fx_method_note,
    _matches_term,
    _open_charts_sheet,
    _payload_date,
    _security_class_bucket,
    _summary_decimal,
    _title,
    _write_chart_source_pie,
    _write_full_mix_table,
    _write_summary_table_with_chart,
)


def _calendar_as_of(dataset_by_type: dict[str, DatasetVersion]) -> date | None:
    """Choose the source date used for the brokerage maturity window.

    The maturity calendar's own business/source date is authoritative.  The
    trade ledger is only a fallback for older uploads that predate that field;
    no wall-clock ``today()`` value is used, so an export is reproducible.
    """
    calendar = dataset_by_type.get("client_maturity_calendar")
    if calendar is not None:
        return calendar.business_date or calendar.source_report_date
    for dataset_type in ("brokerage_trade_ledger", "derivatives_register"):
        dataset = dataset_by_type.get(dataset_type)
        if dataset is not None and (dataset.business_date or dataset.source_report_date):
            return dataset.business_date or dataset.source_report_date
    return None


def _brokerage_calendar_records(records: list[Any], as_of: date | None) -> list[Any]:
    """Keep the next 180 calendar days, inclusive, in chronological order."""
    if as_of is None:
        return []
    end_date = as_of + timedelta(days=180)
    window = [
        record for record in records
        if (maturity := _payload_date(record, "maturity_date")) is not None
        and as_of <= maturity <= end_date
    ]
    return sorted(
        window,
        key=lambda record: (
            _payload_date(record, "maturity_date") or date.max,
            str(record.payload.get("instrument") or "").casefold(),
            str(record.payload.get("isin") or "").casefold(),
        ),
    )


def _write_brokerage_trade_tables(
    workbook: Workbook,
    ledger: DatasetVersion,
    *,
    term: str = "",
    include_repo: bool = True,
    workbook_report_date: date | None = None,
) -> None:
    """Two small, curated trade tables instead of the full ledger dump.

    Recent activity (what to act on) and the largest trades (what's worth a
    second look) are what an operator actually reaches for; the aggregates
    (turnover, venues, instrument mix) already live on the summary sheet.
    Full row-by-row history stays queryable in the source workbook itself -
    every row here still carries its own Лист/Строка (source filename is
    constant for the sheet and shown once in the metadata block instead).
    """
    normalized_term = term.strip().casefold()
    records = [
        record for record in ledger.records
        if _matches_term(record, normalized_term) and (include_repo or not is_repo_trade(record.payload))
    ]
    sheet = workbook.create_sheet("Сделки")
    _title(sheet, ledger, workbook_report_date=workbook_report_date)
    fields = _fields("brokerage_trade_ledger", records, include_client_details=False)
    # The source filename is already shown once in the "Исходный файл"
    # metadata line above (_title) and is constant for the whole sheet -
    # repeating it on every row just wastes column width.
    headers = [_LABELS.get(field, field.replace("_", " ")) for field in fields] + ["Вес оборота в валюте, %", "Лист", "Строка"]

    def _trade_weights(rows: list[Any]) -> list[Decimal | str]:
        # Weighted within the exact rows shown on THIS table, per
        # currency+side pair - the same "no honest cross-currency weight
        # without inventing an FX conversion" reasoning the old ledger sheet
        # used, just recomputed per curated subset instead of once globally.
        totals: dict[tuple[str, str], Decimal] = {}
        for record in rows:
            currency = str(record.payload.get("currency") or "Не указано")
            side = str(record.payload.get("side") or "Не указано").strip() or "Не указано"
            totals[(currency, side)] = totals.get((currency, side), Decimal("0")) + abs(_summary_decimal(record.payload.get("amount")))
        weights = []
        for record in rows:
            currency = str(record.payload.get("currency") or "Не указано")
            side = str(record.payload.get("side") or "Не указано").strip() or "Не указано"
            amount = abs(_summary_decimal(record.payload.get("amount")))
            denominator = totals.get((currency, side), Decimal("0"))
            weights.append(amount / denominator if denominator else "Недоступно")
        return weights

    first_header_row: int | None = None

    def _write_one_table(row: int, title: str, note: str, rows: list[Any]) -> int:
        nonlocal first_header_row
        sheet.cell(row, 1, title).font = Font(bold=True, size=12)
        row += 1
        sheet.cell(row, 1, note).font = Font(italic=True, color="666666", size=9)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1
        header_row = row
        if first_header_row is None:
            first_header_row = header_row
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(header_row, column, label)
            cell.font = Font(bold=True); cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        weights = _trade_weights(rows)
        for offset, (record, weight) in enumerate(zip(rows, weights), header_row + 1):
            source = record.source_ref or {}
            values = [_field_value(record, field) for field in fields] + [weight, source.get("sheet_name") or "Недоступно", source.get("row_number") or "Недоступно"]
            failed_trade = bool(str(record.payload.get("failure_reason") or "").strip())
            for column, value in enumerate(values, 1):
                cell = sheet.cell(offset, column, value)
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=column <= len(fields))
                if failed_trade:
                    cell.fill = _FAILED_TRADE_FILL
                if isinstance(value, Decimal):
                    field = fields[column - 1] if column <= len(fields) else None
                    cell.number_format = (
                        "0.0%;[Red](0.0%);-" if column == len(headers) - 3
                        else "#,##0;[Red](#,##0);-" if field == "quantity"
                        else "#,##0.00;[Red](#,##0.00);-"
                    )
                elif isinstance(value, (date, datetime)):
                    cell.number_format = "dd.mm.yyyy"
        last_row = header_row + len(rows)
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, last_row)}"
        for column, label in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, min(48, max(14, len(label) + 4)))
        return last_row + 3

    row = 8
    window_days = 90
    business_date = ledger.business_date
    if business_date is not None:
        window_start = business_date - timedelta(days=window_days)
        recent = sorted(
            (record for record in records if (_payload_date(record, "trade_date") or date.min) >= window_start and (_payload_date(record, "trade_date") or date.min) <= business_date),
            key=lambda record: (record.payload.get("trade_date") or "", record.payload.get("trade_number") or ""),
            reverse=True,
        )
        row = _write_one_table(
            row, f"Последние сделки ({window_days} дней)",
            f"Сделки с {window_start.strftime('%d.%m.%Y')} по {business_date.strftime('%d.%m.%Y')} включительно, отсортированы по дате (новые сверху). "
            "Полная история сделок остаётся в исходной рабочей книге.",
            recent,
        )
    else:
        sheet.cell(row, 1, "Последние сделки: не построено — в источнике нет бизнес-даты.").font = Font(italic=True, color="666666", size=9)
        row += 2

    top_n = 20
    largest = sorted(records, key=lambda record: abs(_summary_decimal(record.payload.get("amount"))), reverse=True)[:top_n]
    _write_one_table(
        row, f"Крупнейшие сделки (топ {min(top_n, len(largest))} по сумме)",
        "Отсортированы по абсолютной величине суммы сделки в её собственной валюте (валюты между собой не складываются). "
        "Полная история сделок остаётся в исходной рабочей книге.",
        largest,
    )
    if first_header_row is not None:
        # The summary contains stacked tables, so freeze at the first table's
        # header while also keeping the label column visible.
        sheet.freeze_panes = f"B{first_header_row + 1}"


def _write_brokerage_summary(workbook: Workbook, dataset_by_type: dict[str, DatasetVersion], *, term: str = "", include_repo: bool = True) -> None:
    ledger = dataset_by_type.get("brokerage_trade_ledger")
    if ledger is None:
        return
    summary = ledger.summary or {}
    normalized_term = term.strip().casefold()
    records = [
        record for record in ledger.records
        if _matches_term(record, normalized_term) and (include_repo or not is_repo_trade(record.payload))
    ]
    sheet = workbook.create_sheet("Сводка брокерской деятельности")
    sheet.sheet_view.showGridLines = False
    # No frozen panes: this is a dashboard of stacked tables/charts, read
    # chart-first rather than scrolled as a long table - a frozen column A
    # has nothing to do here. Deliberate, documented exception to the
    # workbook navigation standard (docs/export-column-audit.md), enforced
    # by tests/export_compliance.py's _UNFROZEN_DASHBOARD_SHEETS.
    row = 2
    sheet.cell(row, 1, "Сводка брокерской деятельности").font = Font(bold=True, size=14)
    row += 2
    sheet.cell(row, 1, "РЕПО").font = Font(bold=True)
    sheet.cell(row, 2, "Включены" if include_repo else "Исключены")
    row += 2
    buy: dict[str, Decimal] = {}
    sell: dict[str, Decimal] = {}
    venue_mix: dict[str, int] = {}
    instrument_mix: dict[str, int] = {}
    status_mix: dict[str, int] = {}
    charts_sheet, charts_row = _open_charts_sheet(
        workbook,
        "Источник данных для диаграмм листа «Сводка брокерской деятельности»",
        "Редактируемые таблицы-источники диаграмм. Полные категории остаются на листе «Сводка брокерской деятельности»; "
        "мелкие категории объединяются в «Прочее» только в данных соответствующей диаграммы.",
    )
    for record in records:
        payload = record.payload
        currency = str(payload.get("currency") or "Не указано")
        amount = abs(_summary_decimal(payload.get("amount")))
        side = str(payload.get("side") or "").casefold()
        target = buy if side.startswith(("куп", "покуп", "buy")) else sell if side.startswith(("прод", "sell")) else None
        if target is not None:
            target[currency] = target.get(currency, Decimal("0")) + amount
        venue = str(payload.get("venue") or "Не указано")
        venue_mix[venue] = venue_mix.get(venue, 0) + 1
        instrument = str(payload.get("security_type") or "Не указано")
        instrument_mix[instrument] = instrument_mix.get(instrument, 0) + 1
        status = str(payload.get("execution_status") or "Не указано")
        status_mix[status] = status_mix.get(status, 0) + 1
    if buy or sell:
        currencies = sorted(set(buy) | set(sell))
        rate_date = ledger.business_date
        # One NBK lookup per currency in play (KZT is trivially rate 1, no
        # network call) - resolve_export_fx_rate already caches per calendar
        # date, so re-exporting the same period doesn't refetch.
        fx_by_currency = {currency: _multi_source_export_pkg.resolve_export_fx_rate(currency, rate_date) if rate_date is not None else None for currency in currencies}

        def _kzt_equivalent(amount: Decimal, currency: str) -> Decimal | str:
            fx = fx_by_currency.get(currency)
            return amount * fx.rate if fx is not None else "Недоступно"

        turnover_base_rows = [
            [
                currency, _summary_decimal(buy.get(currency, "0")), _summary_decimal(sell.get(currency, "0")),
                _kzt_equivalent(_summary_decimal(buy.get(currency, "0")), currency),
                _kzt_equivalent(_summary_decimal(sell.get(currency, "0")), currency),
            ]
            for currency in currencies
        ]
        buy_kzt_total = sum((values[3] for values in turnover_base_rows if isinstance(values[3], Decimal)), Decimal("0"))
        sell_kzt_total = sum((values[4] for values in turnover_base_rows if isinstance(values[4], Decimal)), Decimal("0"))
        turnover_rows = []
        for values in turnover_base_rows:
            buy_weight = values[3] / buy_kzt_total if isinstance(values[3], Decimal) and buy_kzt_total else "Недоступно"
            sell_weight = values[4] / sell_kzt_total if isinstance(values[4], Decimal) and sell_kzt_total else "Недоступно"
            turnover_rows.append(values + [buy_weight, sell_weight])
        sheet.cell(row, 1, "Оборот по валютам").font = Font(bold=True)
        row += 1
        # Amounts are kept in each trade's own currency, never converted to a
        # KZT equivalent (see "Валюта" column) - a "Покупка, KZT" header
        # previously mislabelled every non-KZT row with the wrong currency.
        sheet.cell(row, 1, "«Покупка»/«Продажа» — суммы в валюте каждой сделки; отдельные столбцы ниже показывают приблизительный эквивалент в KZT.").font = Font(italic=True, color="666666", size=9)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        row += 1
        sheet.cell(row, 1, _fx_method_note(rate_date, currencies, fx_by_currency)).font = Font(italic=True, color="666666", size=9)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        header_row = row
        for column, label in enumerate(["Валюта", "Покупка", "Продажа", "Покупка, ≈KZT", "Продажа, ≈KZT", "Вес покупки, %", "Вес продажи, %"], 1):
            cell = sheet.cell(header_row, column, label)
            cell.font = Font(bold=True); cell.border = _BORDER
        for offset, values in enumerate(turnover_rows, header_row + 1):
            for column, value in enumerate(values, 1):
                cell = sheet.cell(offset, column, value)
                cell.border = _BORDER
                if isinstance(value, Decimal):
                    cell.number_format = "0.0%;[Red](0.0%);-" if column >= 6 else "#,##0.00;[Red](#,##0.00);-"
        last_row = header_row + len(turnover_rows)
        if turnover_rows:
            total_values = [
                "Итого (доступные KZT-эквиваленты)", "Недоступно", "Недоступно",
                buy_kzt_total if buy_kzt_total else "Недоступно",
                sell_kzt_total if sell_kzt_total else "Недоступно",
                Decimal("1") if buy_kzt_total else "Недоступно",
                Decimal("1") if sell_kzt_total else "Недоступно",
            ]
            total_row = last_row + 1
            for column, value in enumerate(total_values, 1):
                cell = sheet.cell(total_row, column, value)
                cell.border = _BORDER
                cell.font = Font(bold=True)
                if isinstance(value, Decimal):
                    cell.number_format = "0.0%;[Red](0.0%);-" if column >= 6 else "#,##0.00;[Red](#,##0.00);-"
            last_row = total_row
        for column, width in enumerate([34, 22, 22, 20, 20, 18, 18], 1):
            sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, width)
        if turnover_rows:
            # Keep the editable chart source separate from the presentation
            # table, while anchoring the native Excel chart on this summary
            # sheet.  The summary table ends at column G ("Вес покупки, %"),
            # so I leaves a deliberate visual gap and avoids covering it.
            charts_sheet.cell(charts_row, 1, "Оборот по валютам (покупка/продажа)").font = Font(bold=True)
            charts_header_row = charts_row + 1
            for column, label in enumerate(["Валюта", "Покупка", "Продажа"], 1):
                charts_sheet.cell(charts_header_row, column, label)
            for offset, values in enumerate(turnover_rows, charts_header_row + 1):
                charts_sheet.cell(offset, 1, values[0])
                charts_sheet.cell(offset, 2, values[1]).number_format = "#,##0.00;[Red](#,##0.00);-"
                charts_sheet.cell(offset, 3, values[2]).number_format = "#,##0.00;[Red](#,##0.00);-"
            charts_last_row = charts_header_row + len(turnover_rows)
            table = Table(displayName="ChartDataTurnover", ref=f"A{charts_header_row}:C{charts_last_row}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            charts_sheet.add_table(table)
            for column, width in (("A", 16), ("B", 22), ("C", 22)):
                charts_sheet.column_dimensions[column].width = max(charts_sheet.column_dimensions[column].width or 0, width)
            # Never combine currencies on one axis - each currency is its
            # own labelled category with buy/sell as the two series.
            write_bar_chart(
                charts_sheet, label_col=1, value_col_first=2, value_col_last=3, header_row=charts_header_row,
                first_row=charts_header_row + 1, last_row=charts_last_row, title="Оборот по валютам (покупка/продажа)",
                anchor=f"I{header_row}", anchor_worksheet=sheet,
                y_axis_title="Сумма (в валюте сделки)", log_scale=True,
            )
            charts_row = max(charts_last_row + 3, charts_row + 18)
        # Charts are 9cm tall; the same "leave enough rows for the chart
        # itself" gap used by _write_summary_table_with_chart applies here
        # too - confirmed against real Excel, this table's own hand-rolled
        # layout had the same overlap bug. The floor must be measured from
        # the chart's own anchor row (header_row) rather than the section's
        # title row above it - the two disclosure lines shift header_row
        # down from the title, so anchoring the gap to the title under-counts
        # the real distance and reproduces the overlap.
        row = max(last_row + 3, header_row + 18)
    mixes = {"venue_mix": venue_mix, "instrument_mix": instrument_mix, "status_mix": status_mix}
    if any(mixes.values()):
        # Full category tables live here; only the <2%-grouped chart sources
        # live on the final chart-data sheet. The native charts stay beside
        # these tables on the summary sheet. A trade ledger can span a dozen+
        # venues/instrument types, and folding small ones into "Прочее" is a
        # materiality decision for the chart alone, not something that should
        # hide real rows from this table (confirmed against a real export
        # where the grouped table hid 8 of 16 actual instrument types).
        for key, sheet_title, chart_title, table_name in (
            ("venue_mix", "Площадки исполнения", "Площадки исполнения", "ChartDataVenue"),
            ("instrument_mix", "Состав по типам инструментов", "Состав по типам инструментов", "ChartDataInstrument"),
            ("status_mix", "Статус исполнения сделок", "Статус исполнения сделок", "ChartDataStatus"),
        ):
            mix = mixes[key]
            if not mix:
                continue
            row, charts_row = _write_full_mix_table(
                sheet, row, sheet_title, mix, charts_sheet=charts_sheet, charts_row=charts_row,
                table_name=table_name, chart_title=chart_title, log_scale=True,
                chart_kind="pie" if key == "venue_mix" else "bar",
            )
    # Stocks vs bonds by KZT turnover - same classification and NBK-rate
    # conversion the web brokerage chart already uses (DomainCharts.tsx),
    # so the two agree. A trade only counts once its currency has a
    # resolvable rate; currencies that don't are named and excluded rather
    # than silently dropped.
    stock_bond_currencies = sorted({
        str(record.payload.get("currency") or "").strip()
        for record in records
        if str(record.payload.get("currency") or "").strip() and str(record.payload.get("security_type") or record.payload.get("instrument_type") or "").strip()
    })
    if stock_bond_currencies:
        rate_date = ledger.business_date
        stock_bond_fx = {currency: _multi_source_export_pkg.resolve_export_fx_rate(currency, rate_date) if rate_date is not None else None for currency in stock_bond_currencies}
        missing_rate_currencies = sorted(currency for currency, fx in stock_bond_fx.items() if fx is None)
        stock_bond_totals: dict[str, Decimal] = {}
        for record in records:
            payload = record.payload
            security_type = str(payload.get("security_type") or payload.get("instrument_type") or "").strip()
            currency = str(payload.get("currency") or "").strip()
            amount = payload.get("amount")
            fx = stock_bond_fx.get(currency)
            if not security_type or not currency or fx is None or amount is None:
                continue
            bucket = _security_class_bucket(security_type)
            stock_bond_totals[bucket] = stock_bond_totals.get(bucket, Decimal("0")) + abs(_summary_decimal(amount)) * fx.rate
        if stock_bond_totals:
            stock_bond_rows = [[label, value] for label, value in sorted(stock_bond_totals.items(), key=lambda pair: pair[1], reverse=True)]
            note = f"Оборот в KZT по виду ценной бумаги на дату {rate_date:%d.%m.%Y}." if rate_date else "Оборот в KZT по виду ценной бумаги."
            note += " Фонды и индексные инструменты не отнесены ни к акциям, ни к облигациям."
            if missing_rate_currencies:
                note += f" Курс недоступен для: {', '.join(missing_rate_currencies)} - эти сделки исключены из суммы."
            sheet.cell(row, 1, note).font = Font(italic=True, color="666666", size=9)
            sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
            chart_start_row = row
            row = _write_summary_table_with_chart(
                sheet, row, "Акции против облигаций", ["Вид ценной бумаги", "Оборот, KZT"], stock_bond_rows,
                numeric_formats={2: "#,##0.00;[Red](#,##0.00);-"}, chart_column=2,
                chart_title="Акции против облигаций", chart_kind="pie", widths=[20, 24], add_weight=True,
                render_chart=False,
            )
            charts_row = _write_chart_source_pie(
                charts_sheet,
                charts_row,
                title="Акции против облигаций",
                table_name="ChartDataStocksBonds",
                headers=("Вид ценной бумаги", "Оборот, KZT"),
                rows=stock_bond_rows,
                anchor=f"E{chart_start_row}",
                anchor_worksheet=sheet,
            )
