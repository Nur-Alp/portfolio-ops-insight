"""Asset management / TABYS fund summary and detail-sheet writers."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from osip_dashboard.persistence.models import DatasetVersion
from osip_dashboard.services.excel_charts import write_line_chart

from .shared import _BORDER, _matches_term, _open_charts_sheet, _summary_decimal, _write_summary_table_with_chart


def _write_asset_management_summary(workbook: Workbook, dataset_by_type: dict[str, DatasetVersion], *, term: str = "") -> None:
    valuation = dataset_by_type.get("fund_valuation")
    holdings = dataset_by_type.get("fund_holdings")
    unit_series = dataset_by_type.get("fund_unit_series")
    if not any((valuation, holdings, unit_series)):
        return
    # NAV composition and the unit-value/NAV time series are fund-level
    # aggregates a search term doesn't meaningfully narrow (they aren't
    # per-instrument). Only the currency allocation below is built from
    # per-instrument holdings rows, so only it needs to match the "Портфель
    # фонда" detail sheet's own term filter - otherwise this summary
    # silently disagreed with the filtered detail sheet in the same file.
    normalized_term = term.strip().casefold()
    sheet = workbook.create_sheet("Сводка фонда")
    sheet.sheet_view.showGridLines = False
    # No frozen panes: this is a dashboard of stacked tables/charts, read
    # chart-first rather than scrolled as a long table - a frozen column A
    # has nothing to do here. Deliberate, documented exception to the
    # workbook navigation standard (docs/export-column-audit.md), enforced
    # by tests/export_compliance.py's _UNFROZEN_DASHBOARD_SHEETS.
    row = 2
    sheet.cell(row, 1, "Сводка фонда TABYS").font = Font(bold=True, size=14)
    row += 2
    if unit_series is not None and unit_series.records:
        row = _write_fund_time_series_chart(
            workbook, sheet, row, unit_series, value_field="unit_value_kzt", value_label="Стоимость пая, KZT",
            chart_title="Стоимость пая во времени", number_format="#,##0.0000;[Red](#,##0.0000);-", line_color="9226A8",
        )
        row = _write_fund_time_series_chart(
            workbook, sheet, row, unit_series, value_field="nav_kzt", value_label="СЧА, KZT",
            chart_title="СЧА фонда во времени", number_format="#,##0.00;[Red](#,##0.00);-", line_color="1F6FB2",
        )
    if valuation is not None and valuation.summary:
        composition_rows = [
            [label, _summary_decimal(valuation.summary.get(field))]
            for label, field in (("Ценные бумаги", "securities_value_kzt"), ("Деньги", "cash_kzt"), ("Обязательства", "liabilities_kzt"))
            if valuation.summary.get(field) is not None
        ]
        if len(composition_rows) >= 2:
            # A pie is the wrong form here: liabilities is a deduction from
            # NAV, not a positive share of it (NAV = securities + cash -
            # liabilities), so plotting all three as pie slices would imply
            # they sum to the whole. A bar just compares their magnitudes -
            # and securities routinely dominates cash/liabilities by 2+
            # orders of magnitude, so a linear axis flattens the smaller two
            # to invisible slivers (confirmed against a real export). Log
            # scale keeps all three readable, same fix already applied to
            # the brokerage turnover and corporate-finance fee charts.
            row = _write_summary_table_with_chart(
                sheet, row, "Состав СЧА: бумаги, деньги, обязательства", ["Статья", "Сумма, KZT"], composition_rows,
                numeric_formats={2: "#,##0.00;[Red](#,##0.00);-"}, chart_column=2,
                chart_title="Состав СЧА", chart_kind="bar", widths=[24, 24], log_scale=True,
            )
    if holdings is not None and holdings.records:
        by_currency: dict[str, Decimal] = {}
        for record in holdings.records:
            if not _matches_term(record, normalized_term):
                continue
            currency = str(record.payload.get("currency") or "Не указано")
            by_currency[currency] = by_currency.get(currency, Decimal("0")) + _summary_decimal(record.payload.get("purchase_value_kzt"))
        allocation_rows = [[currency, value] for currency, value in sorted(by_currency.items(), key=lambda pair: pair[1], reverse=True)]
        row = _write_summary_table_with_chart(
            sheet, row, "Распределение фонда по валютам", ["Валюта", "Стоимость приобретения, KZT"], allocation_rows,
            numeric_formats={2: "#,##0.00;[Red](#,##0.00);-"}, chart_column=2,
            chart_title="Распределение фонда по валютам", chart_kind="pie", widths=[16, 28], add_weight=True,
        )


def _write_unique_etf_positions(workbook: Workbook, holdings: DatasetVersion | None, *, term: str = "") -> None:
    """One row per distinct holding (by ISIN), quantity and purchase value summed across lots.

    Every row in "часть 1 (портфель)" is currently an ETF position - there
    is no instrument-type field in the source to separate ETFs from any
    other instrument that might appear here later, so every fund_holdings
    row is treated as one today rather than guessing at a classification
    rule. If a non-ETF instrument is ever uploaded into this same sheet,
    it will show up here too - an honest reflection of what the source
    can distinguish, not a silent misclassification.
    """
    if holdings is None or not holdings.records:
        return
    normalized_term = term.strip().casefold()
    positions: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for record in holdings.records:
        if not _matches_term(record, normalized_term):
            continue
        isin = str(record.payload.get("isin") or "Не указано")
        quantity = _summary_decimal(record.payload.get("quantity"))
        purchase_kzt = _summary_decimal(record.payload.get("purchase_value_kzt"))
        purchase_native = _summary_decimal(record.payload.get("purchase_value_native"))
        if isin not in positions:
            order.append(isin)
            positions[isin] = {
                "instrument": record.payload.get("instrument") or "Не указано",
                "currency": record.payload.get("currency") or "Не указано",
                "quantity": Decimal("0"), "purchase_value_kzt": Decimal("0"), "purchase_value_native": Decimal("0"),
                "lot_count": 0, "purchase_dates": [],
            }
        position = positions[isin]
        position["quantity"] += quantity
        position["purchase_value_kzt"] += purchase_kzt
        position["purchase_value_native"] += purchase_native
        position["lot_count"] += 1
        purchase_date = record.payload.get("purchase_date")
        if purchase_date:
            position["purchase_dates"].append(purchase_date)
    if not positions:
        return
    sheet = workbook.create_sheet("Уникальные позиции ETF")
    sheet.sheet_view.showGridLines = False
    row = 2
    sheet.cell(row, 1, "Уникальные позиции ETF (TABYS)").font = Font(bold=True, size=14)
    row += 1
    sheet.cell(row, 1, "Один инструмент = одна строка; количество и стоимость приобретения суммированы по всем лотам покупки.").alignment = Alignment(wrap_text=True)
    row += 2
    headers = ["Инструмент", "ISIN", "Валюта", "Количество", "Стоимость приобретения, KZT", "Стоимость приобретения (валюта)", "Лотов", "Первая покупка"]
    header_row = row
    for column, label in enumerate(headers, 1):
        cell = sheet.cell(header_row, column, label)
        cell.font = Font(bold=True, color="000000")
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # A single continuous table - pin the title/header rows and column A
    # (docs/export-column-audit.md navigation standard).
    sheet.freeze_panes = f"B{header_row + 1}"
    isins_by_value = sorted(order, key=lambda isin: positions[isin]["purchase_value_kzt"], reverse=True)
    for offset, isin in enumerate(isins_by_value, header_row + 1):
        position = positions[isin]
        earliest_purchase_raw = min(position["purchase_dates"]) if position["purchase_dates"] else None
        try:
            earliest_purchase = date.fromisoformat(earliest_purchase_raw) if earliest_purchase_raw else "Недоступно"
        except ValueError:
            earliest_purchase = earliest_purchase_raw
        values = [
            position["instrument"], isin, position["currency"], position["quantity"],
            position["purchase_value_kzt"], position["purchase_value_native"], position["lot_count"], earliest_purchase,
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(offset, column, value)
            cell.border = _BORDER
            if column in (5, 6) and isinstance(value, Decimal):
                cell.number_format = "#,##0.00;[Red](#,##0.00);-"
            elif column == 4 and isinstance(value, Decimal):
                # Share counts are always whole numbers - the same integer
                # format "quantity" gets everywhere else in this export
                # (see _typed()). A "#,##0.####" mask was used here at
                # first, which rendered a dangling trailing decimal point
                # in Excel (e.g. "300.") even though every value is a
                # clean integer.
                cell.number_format = "#,##0;[Red](#,##0);-"
            elif column == 8 and isinstance(value, date):
                cell.number_format = "dd.mm.yyyy"
    widths = [24, 18, 10, 14, 24, 24, 8, 16]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(column)].width = max(sheet.column_dimensions[get_column_letter(column)].width or 0, width)


def _write_fund_time_series_chart(
    workbook: Workbook, sheet, row: int, unit_series: DatasetVersion, *,
    value_field: str, value_label: str, chart_title: str, number_format: str, line_color: str = "1F6FB2",
) -> int:
    """One full-history line chart, sourced from a visible "Данные графиков"
    row block instead of a capped inline table - a daily series can run to
    thousands of rows, and writing even a windowed copy inline still pushed
    every section after it down by that many rows (confirmed against a real
    export). The complete series stays queryable on "История пая"; this only
    adds a chart-ready copy, same convention as the brokerage/clients charts.
    """
    dated_rows = sorted(
        ((record.payload.get("date"), _summary_decimal(record.payload.get(value_field))) for record in unit_series.records if record.payload.get("date") and record.payload.get(value_field) is not None),
        key=lambda pair: pair[0],
    )
    if not dated_rows:
        return row
    charts_sheet, charts_row = _open_charts_sheet(
        workbook,
        f"Источник данных для диаграммы «{chart_title}»",
        f"Редактируемая таблица-источник диаграммы; полная история ({len(dated_rows)} наблюдений). Построчные данные с указанием листа/строки источника — на листе «История пая».",
    )
    header_row = charts_row
    charts_sheet.cell(header_row, 1, "Дата").font = Font(bold=True)
    charts_sheet.cell(header_row, 2, value_label).font = Font(bold=True)
    for offset, (when, value) in enumerate(dated_rows, header_row + 1):
        charts_sheet.cell(offset, 1, when)
        charts_sheet.cell(offset, 2, value).number_format = number_format
    last_charts_row = header_row + len(dated_rows)
    for column, width in (("A", 16), ("B", 22)):
        charts_sheet.column_dimensions[column].width = max(charts_sheet.column_dimensions[column].width or 0, width)
    sheet.cell(row, 1, chart_title).font = Font(bold=True)
    row += 1
    sheet.cell(row, 1, f"Полная история ({len(dated_rows)} наблюдений) — таблица на листе «Данные графиков»; построчные источники — на листе «История пая».").font = Font(italic=True, color="666666", size=9)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2
    # No separate y-axis title: openpyxl/Excel renders it rotated right
    # beside the tick labels with no reserved space of its own, so it
    # overlapped the numbers on a chart this narrow (confirmed against a
    # real export). Folding the unit into the chart's own title instead
    # states it just as clearly without fighting that layout.
    unit_suffix = value_label.rsplit(", ", 1)[-1]
    write_line_chart(
        charts_sheet, label_col=1, value_col=2, header_row=header_row,
        first_row=header_row + 1, last_row=last_charts_row, title=f"{chart_title}, {unit_suffix}",
        anchor=f"A{row}", anchor_worksheet=sheet, line_color=line_color,
    )
    return row + 18
