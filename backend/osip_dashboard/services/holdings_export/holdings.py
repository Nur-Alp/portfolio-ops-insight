"""Holdings, position-lot, expected-cash-flow and source-lot exports."""

from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from osip_dashboard.persistence.models import PortfolioSnapshotRecord
from osip_dashboard.services.dividends import dividend_data_status, is_us_listed, load_dividend_history, lot_dividend_contribution
from osip_dashboard.services.fx_rates import resolve_export_usd_kzt_rate
from osip_dashboard.services.hpr import hpr_amount, hpr_percent
from osip_dashboard.services.instrument_dictionary import instrument_class, true_asset_class
from osip_dashboard.services.holdings_export.coupons import (
    _accrued_or_dividend_kzt,
    _coupon_or_dividend_native,
    _dividend_freshness_note,
    estimated_coupon_income_kzt,
    estimated_paid_coupon_income_native,
    expected_coupon_native,
    lot_maturity_amount_native,
)
from osip_dashboard.services.holdings_export.distribution import (
    _append_distribution_sheets,
    _append_weight_validation,
    _hpr_usd_percent_for_item,
    _lot_usd_equivalent,
    _snapshot_usd_rates,
    _weight_validation,
)
from osip_dashboard.services.holdings_export.shared import (
    ExportBasis,
    _BASIS_LABELS,
    _GOOD_FILL,
    _HEADER_FONT,
    _MONEY_FORMAT,
    _NUMBER_FORMAT,
    _POINTS_PERCENT_FORMAT,
    _TABLE_BORDER,
    _asset_class_label,
    _decimal_value,
    _optional_decimal,
    _save,
    _snapshot_title,
    _source_columns,
    _sources_columns,
    _table,
    _total_row,
    _workbook,
)


_VALUE_LABELS = {
    "derived_carrying": "Расчётная балансовая стоимость, KZT",
    "purchase": "Сумма покупки, KZT",
}


def create_holdings_xlsx(snapshot: PortfolioSnapshotRecord, items: list[dict[str, Any]], *, basis: ExportBasis, term: str, asset_class: str | None) -> bytes:
    """Create a typed, filterable XLSX workbook for the exact holdings view."""
    workbook = _workbook("Инструменты портфеля OSIP")
    worksheet = workbook.active
    worksheet.title = "Инструменты"
    value_key = "purchase_amount_kzt" if basis == "purchase" else "derived_carrying_value_kzt"
    weight_key = "purchase_weight_percent" if basis == "purchase" else "derived_weight_percent"
    # An instrument whose derived carrying value is unavailable (e.g. a
    # short-dated deposit the source never marks a current balance for) is
    # excluded from the derived-basis total/weight below, not folded in as
    # zero - see _aggregated_holdings' derived_carrying_incomplete. Purchase
    # basis never has this gap: purchase_amount_kzt is populated for every
    # lot.
    incomplete_items = [item for item in items if item.get("derived_carrying_incomplete")] if basis != "purchase" else []
    total_weight = sum((_decimal_value(item[weight_key]) for item in items if item[weight_key] is not None), Decimal("0")) / Decimal("100")
    weight_validation = _weight_validation(total_weight)
    if incomplete_items:
        excluded_purchase_kzt = sum((_decimal_value(item["purchase_amount_kzt"]) for item in incomplete_items), Decimal("0"))
        weight_validation += (
            f" · {len(incomplete_items)} инстр. без расчётной стоимости исключены из суммы и веса "
            f"(сумма покупки: {excluded_purchase_kzt:,.2f} KZT) - см. столбец «{_VALUE_LABELS['purchase']}» или лист «Позиции по лотам»."
        )
    _snapshot_title(worksheet, snapshot, "Инструменты портфеля OSIP", {
        "Основа оценки": _BASIS_LABELS[basis], "Поиск": term or "Не применён", "Класс актива": _asset_class_label(asset_class) if asset_class else "Все классы активов",
    })
    _append_weight_validation(worksheet, weight_validation)
    # Resolve the same NBK-dated rate (with its offline-fallback and
    # disclosure) that "Распределения" and "Позиции по лотам" already use for
    # their own USD equivalents in this same workbook - this sheet used to
    # call _snapshot_usd_rates directly instead, so "HPR (расч.), FX, %"
    # could silently be computed against a different implied rate than every
    # other USD figure in the download, with no note here explaining why.
    workbook_rate = _snapshot_usd_rates(snapshot)
    resolved_fx = resolve_export_usd_kzt_rate(snapshot.report_date, workbook_rate)
    usd_rate = resolved_fx.rate if resolved_fx is not None else None
    headers = ["Код инструмента", "ISIN", "Эмитент", "Истинный класс актива", "Валюта", "Лоты", "Количество", "HPR (расч.), KZT, %", "HPR (расч.), FX, %", "Текущая YTM, %", _VALUE_LABELS[basis], "Вес, %"]
    rows = [[item["security_code"], item["isin"], item["issuer"], _asset_class_label(item["true_asset_class"]), item["instrument_currency"], item["lot_count"], _decimal_value(item["quantity"]), _optional_decimal(item["hpr_percent"]), _optional_decimal(_hpr_usd_percent_for_item(item, usd_rate=usd_rate)), _optional_decimal(item["current_ytm"]), _optional_decimal(item[value_key]), (_decimal_value(item[weight_key]) / Decimal("100")) if item[weight_key] is not None else "Недоступно"] for item in items]
    numeric_formats = {6: "#,##0", 7: "#,##0;[Red](#,##0);-", 8: _POINTS_PERCENT_FORMAT, 9: _POINTS_PERCENT_FORMAT, 10: _POINTS_PERCENT_FORMAT, 11: _MONEY_FORMAT, 12: "0.0%;[Red](0.0%);-"}
    _table(worksheet, headers, rows, widths=[22, 18, 42, 28, 12, 10, 18, 20, 20, 18, 31, 12], numeric_formats=numeric_formats, wrap_columns={1, 3, 4})
    fx_note_row = getattr(worksheet, "_osip_header_row", 6) - 1
    if resolved_fx is None:
        fx_note = "HPR (расч.), FX, %: USD-эквивалент не рассчитан - официальный курс НБК и согласованный курс рабочей книги недоступны."
    else:
        fallback_note = " (оффлайн-резерв)" if resolved_fx.fallback else ""
        fx_note = (
            f"HPR (расч.), FX, %: 1 USD = {resolved_fx.rate:,.4f} KZT по {resolved_fx.source}{fallback_note}; "
            f"дата курса {resolved_fx.effective_date.isoformat()}; отчётная дата {snapshot.report_date.isoformat()}. "
            "Тот же курс используют USD-эквиваленты на листах «Распределения» и «Позиции по лотам»."
        )
    fx_note_cell = worksheet.cell(fx_note_row, 1, fx_note)
    fx_note_cell.font = Font(italic=True, color="666666", size=9)
    fx_note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.merge_cells(start_row=fx_note_row, start_column=1, end_row=fx_note_row, end_column=12)
    worksheet.row_dimensions[fx_note_row].height = 30
    # Keep the total tied to the exact filtered view. It is a value rather
    # than an Excel formula so data-only readers see the result immediately
    # without requiring recalculation in Excel or LibreOffice.
    total_value = sum((_decimal_value(item[value_key]) for item in items if item[value_key] is not None), Decimal("0"))
    total_row = getattr(worksheet, "_osip_header_row", 6) + len(rows) + 1
    _total_row(worksheet, headers, row_number=total_row, label="Итого", value_column=11, value=total_value, numeric_formats=numeric_formats)
    weight_cell = worksheet.cell(total_row, 12, total_weight)
    weight_cell.number_format = numeric_formats[12]
    weight_cell.border = _TABLE_BORDER
    weight_cell.font = _HEADER_FONT
    weight_cell.alignment = Alignment(horizontal="right", vertical="center")
    _append_distribution_sheets(workbook, snapshot, items, basis=basis, term=term, asset_class=asset_class)
    _append_position_lots_sheet(workbook, snapshot, items, basis=basis, term=term, asset_class=asset_class, total_weight=total_weight, weight_validation=weight_validation)
    _append_expected_cash_flows_sheet(workbook, snapshot, items, term=term, asset_class=asset_class)
    _append_control_sheet(workbook, snapshot, items, total_value=total_value, total_weight=total_weight, weight_validation=weight_validation)
    # Keep the original sheets intact and only place the two detailed views
    # between the allocation and chart/control sheets.
    desired_order = [
        "Инструменты", "Распределения", "Позиции по лотам",
        "Ожидаемые денежные потоки", "Контроль и происхождение", "Данные графиков",
    ]
    workbook._sheets = [workbook[name] for name in desired_order if name in workbook.sheetnames]
    return _save(workbook)


def _append_position_lots_sheet(
    workbook: Workbook,
    snapshot: PortfolioSnapshotRecord,
    items: list[dict[str, Any]],
    *,
    basis: ExportBasis,
    term: str,
    asset_class: str | None,
    total_weight: Decimal,
    weight_validation: str,
) -> None:
    """Export the unaggregated source lots behind the filtered instrument view."""
    sheet = workbook.create_sheet("Позиции по лотам")
    value_key = "purchase_amount_kzt" if basis == "purchase" else "derived_carrying_value_kzt"
    selected_isins = {str(item.get("isin")) for item in items}
    selected_lots = [lot for lot in snapshot.position_lots if lot.isin in selected_isins]
    selected_lots.sort(key=lambda lot: (lot.security_code or "", lot.isin or "", lot.source_row.row_number if lot.source_row else 0))
    total_purchase_value = sum(
        (_decimal_value(lot.purchase_amount_kzt) for lot in selected_lots if lot.purchase_amount_kzt is not None),
        Decimal("0"),
    )
    workbook_rate = _snapshot_usd_rates(snapshot)
    resolved_fx = resolve_export_usd_kzt_rate(snapshot.report_date, workbook_rate)
    usd_rate = resolved_fx.rate if resolved_fx is not None else None
    dividend_history = load_dividend_history()
    dividend_as_of = date.today()
    dividend_status = dividend_data_status(as_of=dividend_as_of)
    # Weight each lot against the whole snapshot's total (every lot, not just
    # the filtered subset) - the same denominator the Instruments sheet's
    # weight column uses. Weighting against only the filtered lots' own sum
    # made every filtered lots-sheet export silently read "100% / OK" even
    # when it was a small slice of the portfolio, contradicting the
    # Instruments sheet's correct "filtered view - N%" disclosure for the
    # exact same download.
    portfolio_total_value = sum(
        (_decimal_value(getattr(lot, value_key)) for lot in snapshot.position_lots if getattr(lot, value_key) is not None),
        Decimal("0"),
    )
    _snapshot_title(sheet, snapshot, "Позиции портфеля по лотам", {
        "Основа оценки": _BASIS_LABELS[basis],
        "Поиск": term or "Не применён",
        "Класс актива": _asset_class_label(asset_class) if asset_class else "Все классы активов",
    })
    _append_weight_validation(sheet, weight_validation)
    note_row = getattr(sheet, "_osip_header_row", 6) - 1
    sheet.cell(
        note_row,
        1,
        "Цена покупки за единицу — исходное значение OSIP; масштаб котировки (например, за 100 "
        "номинала или за 1 единицу) OSIP не указывает. Номинал — отдельная величина, не цена. "
        "Дата погашения облигации берётся из исходного поля OSIP; для строк РЕПО это поле отражает дату закрытия, если она указана. "
        "Текущая балансовая цена за единицу берётся из исходной колонки OSIP «Балансовая цена»; "
        "если Excel сохранил формулу без кэшированного результата, приложение пересчитывает ту же формулу OSIP: "
        "для BF=4 или BE=3 — AA/Q, иначе AA/Q/L×100. "
        "для облигаций это чистая цена (%), для акций и РЕПО — цена в валюте учёта. "
        "Сравнивайте цены только при совпадающей базе котировки; это не официальная рыночная стоимость. "
        "HPR включает дивиденды, прошедшие строгую проверку ex-date/pay-date; для тикеров с отдельным "
        "токеном US удерживается 15%. "
        "Накопленные купоны или дивиденды, KZT, объединяют исходное накопление купона OSIP с подтверждёнными дивидендами Bloomberg после налога. "
        "Оценка выплаченных купонов или дивидендов в валюте цены объединяет приближённую оценку купона "
        "(номинал × количество × ставка купона × дни владения / 360 минус текущее накопление, уже включённое в балансовую стоимость) "
        "с подтверждёнными дивидендами; это не подтверждение фактических купонных выплат. "
        + _dividend_freshness_note(dividend_status)
        + "\n\n"
        + (
            f"USD-эквиваленты: 1 USD = {usd_rate:,.4f} KZT по {resolved_fx.source}; "
            f"дата курса {resolved_fx.effective_date.isoformat()}. Если курс или исходная сумма недоступны, "
            "USD-значение помечено как «Недоступно».\n\n"
            if resolved_fx is not None
            else "USD-эквиваленты не рассчитаны: официальный и исходный курс USD/KZT недоступны.\n\n"
        )
        + "Эквиваленты в USD используют раскрытый курс USD/KZT на отчётную дату и остаются «Недоступно» "
        "при отсутствии исходных данных.",
    )
    sheet.cell(note_row, 1).font = Font(italic=True, color="666666", size=9)
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[note_row].height = 72
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=24)
    true_class_by_isin = {str(item.get("isin")): _asset_class_label(item.get("true_asset_class")) for item in items}
    rows: list[list[Any]] = []
    for lot_number, lot in enumerate(selected_lots, 1):
        lot_value = getattr(lot, value_key)
        lot_value_decimal = _decimal_value(lot_value) if lot_value is not None else None
        quantity = _decimal_value(lot.quantity)
        hpr_kzt = None
        hpr_kzt_percent = None
        purchase_value = _decimal_value(lot.purchase_amount_kzt) if lot.purchase_amount_kzt is not None else None
        derived_value = _decimal_value(lot.derived_carrying_value_kzt) if lot.derived_carrying_value_kzt is not None else None
        native_value_key = "purchase_amount_native" if basis == "purchase" else "carrying_amount_native"
        mv_usd = _lot_usd_equivalent(lot, kzt_value=lot_value_decimal, native_value_key=native_value_key, usd_rate=usd_rate)
        purchase_usd = _lot_usd_equivalent(lot, kzt_value=purchase_value, native_value_key="purchase_amount_native", usd_rate=usd_rate)
        derived_usd = _lot_usd_equivalent(lot, kzt_value=derived_value, native_value_key="carrying_amount_native", usd_rate=usd_rate)
        dividend = lot_dividend_contribution(
            lot, history=dividend_history, current_date=dividend_as_of
        )
        dividend_kzt = dividend.kzt_amount
        dividend_usd = Decimal("0")
        if dividend.matched_count:
            if str(lot.instrument_currency or "").upper() == "USD":
                dividend_usd = dividend.native_amount
            elif str(lot.instrument_currency or "").upper() == "KZT" and usd_rate:
                dividend_usd = dividend.native_amount / usd_rate
            else:
                dividend_usd = None
        coupon_native = estimated_paid_coupon_income_native(lot, snapshot.report_date)
        coupon_kzt = estimated_coupon_income_kzt(lot, coupon_native)
        coupon_usd = None
        if coupon_native is not None:
            currency = str(lot.instrument_currency or "").upper()
            if currency == "USD":
                coupon_usd = coupon_native
            elif currency == "KZT" and usd_rate:
                coupon_usd = coupon_native / usd_rate
        estimated_kzt_income = (dividend_kzt or Decimal("0")) + (coupon_kzt or Decimal("0"))
        estimated_usd_income = (dividend_usd or Decimal("0")) + (coupon_usd or Decimal("0"))
        hpr_kzt = hpr_amount(purchase_value, derived_value, estimated_kzt_income)
        # OSIP's HPR fields are percentage points (for example 0.3, not
        # the Excel fraction 0.003), matching the instrument table. The
        # received dividend and estimated coupon contributions are included
        # in the numerator. The coupon component is a transparent estimate,
        # not a payment ledger.
        hpr_kzt_percent = hpr_percent(purchase_value, derived_value, estimated_kzt_income)
        hpr_usd = hpr_amount(purchase_usd, derived_usd, estimated_usd_income)
        hpr_usd_percent = (
            hpr_kzt_percent
            if str(lot.instrument_currency or "").upper() == "KZT"
            else hpr_percent(purchase_usd, derived_usd, estimated_usd_income)
        )
        weight = lot_value_decimal / portfolio_total_value if lot_value_decimal is not None and portfolio_total_value else None
        accrued_or_dividend_kzt = _accrued_or_dividend_kzt(lot, dividend)
        coupon_or_dividend_native = _coupon_or_dividend_native(coupon_native, dividend)
        purchase_cost_kzt = purchase_value
        rows.append([
            lot.security_code,
            lot_number,
            lot.isin,
            lot.issuer,
            lot.close_date or "Недоступно",
            true_class_by_isin.get(lot.isin, _asset_class_label(getattr(lot.instrument, "normalized_asset_class", None))),
            lot.instrument_currency,
            quantity,
            _optional_decimal(lot.nominal_value),
            purchase_cost_kzt if purchase_cost_kzt is not None else "Недоступно",
            mv_usd if mv_usd is not None else "Недоступно",
            # "Дата открытия" means when this portfolio opened/bought the
            # lot - that's purchase_date. open_date is the underlying
            # instrument's own issue date (e.g. a bond's issuance, which can
            # predate the fund itself by over a decade) and was being shown
            # here instead whenever both were present, since both are
            # populated for nearly every bond - confirmed against the source
            # workbook by cross-checking every lot: purchase_date is always
            # populated (even for a few equities where open_date is blank)
            # and always falls within a plausible ownership window, while
            # open_date does not.
            lot.purchase_date or lot.open_date or "Недоступно",
            _optional_decimal(lot.purchase_price),
            _optional_decimal(lot.purchase_amount_native),
            accrued_or_dividend_kzt,
            coupon_or_dividend_native,
            _optional_decimal(lot.carrying_price_native),
            lot_value_decimal if lot_value_decimal is not None else "Недоступно",
            hpr_kzt if hpr_kzt is not None else "Недоступно",
            hpr_usd if hpr_usd is not None else "Недоступно",
            hpr_kzt_percent if hpr_kzt_percent is not None else "Недоступно",
            hpr_usd_percent if hpr_usd_percent is not None else "Недоступно",
            _optional_decimal(lot.current_ytm),
            weight if weight is not None else "Недоступно",
        ])
    headers = [
        "Код инструмента", "Лот №", "ISIN", "Эмитент", "Дата погашения облигации", "Истинный класс актива", "Валюта цены",
        "Количество", "Номинал (источник)", "Стоимость покупки, ₸", "MV, USD", "Дата открытия", "Цена покупки за единицу (источник)", "Стоимость открытия",
        "Накопленные купоны или дивиденды, KZT", "Оценка выплаченных купонов или дивидендов (валюта цены)", "Текущая балансовая цена за единицу (источник OSIP)", "Текущая стоимость, KZT",
        "HPR (расч.), KZT", "HPR (расч.), FX", "HPR (расч.), KZT, %", "HPR (расч.), FX, %", "Текущая YTM, %", "Вес, %",
    ]
    numeric_formats = {
        2: "#,##0", 5: "dd.mm.yyyy", 8: "#,##0;[Red](#,##0);-", 9: _NUMBER_FORMAT, 10: _MONEY_FORMAT, 11: _MONEY_FORMAT,
        12: "dd.mm.yyyy", 13: _NUMBER_FORMAT, 14: _NUMBER_FORMAT, 15: _MONEY_FORMAT, 16: _MONEY_FORMAT,
        17: _NUMBER_FORMAT, 18: _MONEY_FORMAT, 19: _MONEY_FORMAT, 20: _MONEY_FORMAT, 21: _POINTS_PERCENT_FORMAT, 22: _POINTS_PERCENT_FORMAT, 23: _POINTS_PERCENT_FORMAT, 24: "0.0%;[Red](0.0%);-",
    }
    _table(sheet, headers, rows, widths=[22, 10, 18, 42, 22, 28, 14, 16, 16, 22, 20, 16, 28, 22, 32, 36, 28, 22, 20, 20, 16, 20, 18, 12], numeric_formats=numeric_formats, wrap_columns={4, 6, 16})
    total_row = getattr(sheet, "_osip_header_row", 6) + len(rows) + 1
    _total_row(worksheet=sheet, headers=headers, row_number=total_row, label="Итого", value_column=10, value=total_purchase_value, numeric_formats=numeric_formats)
    weight_cell = sheet.cell(total_row, 24, total_weight)
    weight_cell.number_format = numeric_formats[24]
    weight_cell.border = _TABLE_BORDER
    weight_cell.font = _HEADER_FONT
    weight_cell.alignment = Alignment(horizontal="right", vertical="center")


def _append_expected_cash_flows_sheet(
    workbook: Workbook,
    snapshot: PortfolioSnapshotRecord,
    items: list[dict[str, Any]],
    *,
    term: str,
    asset_class: str | None,
) -> None:
    """Export dated OSIP cash-flow events without inventing unavailable amounts."""
    sheet = workbook.create_sheet("Ожидаемые денежные потоки")
    selected_isins = {str(item.get("isin")) for item in items}
    _snapshot_title(sheet, snapshot, "Ожидаемые денежные потоки", {
        "Поиск": term or "Не применён",
        "Класс актива": _asset_class_label(asset_class) if asset_class else "Все классы активов",
        "Суммы": "Недоступно, если не указаны в источнике",
    })
    window_end = snapshot.report_date + timedelta(days=180)
    note_row = getattr(sheet, "_osip_header_row", 6) - 1
    sheet.cell(note_row, 1, (
        "Предстоящие расчёты OSIP исключены. Показаны только потоки от отчётной даты "
        f"{snapshot.report_date:%d.%m.%Y} до {window_end:%d.%m.%Y} включительно "
        "(180 календарных дней), отсортированные по дате по возрастанию. "
        "Сумма купона берётся из поля «Сумма ожидаемого купона», если заполнена; "
        "иначе рассчитывается как номинал × ставка купона / периодичность выплат × "
        "количество (периодичность определяется по длительности купонного периода) "
        "- поле AN является формулой OSIP, чей сохранённый результат не всегда "
        "доступен вне Excel. Ожидаемые дивиденды Bloomberg добавлены по pay_date "
        "для лотов, которыми владели на ex_date; для тикеров с отдельным токеном US "
        "применено удержание 15%. Если словарь Bloomberg устарел, будущие выплаты "
        "могут отсутствовать. "
        + _dividend_freshness_note(dividend_data_status())
    ))
    sheet.cell(note_row, 1).font = Font(italic=True, color="666666", size=9)
    sheet.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[note_row].height = 64
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    lots = [lot for lot in snapshot.position_lots if lot.isin in selected_isins]
    rows = _expected_cash_flow_rows(snapshot, lots)
    _table(sheet, ["Код инструмента", "Тип потока", "Дата", "Сумма", "Валюта"], rows, widths=[24, 20, 16, 24, 12], numeric_formats={3: "dd.mm.yyyy", 4: _MONEY_FORMAT})


def _lot_raw_true_class(lot: Any) -> str:
    normalized_asset_class = getattr(getattr(lot, "instrument", None), "normalized_asset_class", None) or ""
    return true_asset_class(
        normalized_asset_class,
        getattr(lot, "raw_security_type", "") or "",
        getattr(lot, "raw_sector", "") or "",
        getattr(lot, "security_code", "") or "",
        instrument_class(getattr(lot, "isin", None)),
    )


def _lot_is_repo(lot: Any) -> bool:
    return _lot_raw_true_class(lot) == "Repo"


def _expected_cash_flow_rows(snapshot: PortfolioSnapshotRecord, lots: Iterable[Any], *, window_days: int = 180, dividend_history: Any | None = None) -> list[list[Any]]:
    """Return dated flow rows inside the inclusive report-date window.

    ``snapshot.report_date`` is the workbook's generated/as-of date. Using it
    instead of ``today()`` or the filename keeps exports reproducible for a
    published version downloaded at a later time. Undated events are omitted
    because their inclusion in a date-bounded view cannot be established.

    Rows are grouped by instrument (ISIN), not by lot: an instrument split
    across several lots that all mature/coupon/dividend-pay on the same date
    is one economic cash flow, not several - the portfolio doesn't receive
    (or pay) it once per lot. If any lot contributing to a group has an
    unavailable amount, the whole group's amount is "Недоступно" rather than
    a partial sum silently missing part of the flow. Bloomberg dividends use
    pay_date as the cash date and require ex_date > purchase_date, matching
    the HPR entitlement rule.
    """
    start_date = snapshot.report_date
    end_date = start_date + timedelta(days=window_days)
    groups: dict[tuple[str, str, date], dict[str, Any]] = {}
    dividend_history = dividend_history or load_dividend_history()

    def _add(isin: str, security_code: str, flow_type: str, flow_date: date, currency: str, amount: Decimal | None) -> None:
        key = (isin, flow_type, flow_date)
        group = groups.setdefault(key, {"security_code": security_code, "currency": currency, "amount": Decimal("0"), "available": True})
        if amount is None:
            group["available"] = False
        else:
            group["amount"] += amount

    for lot in lots:
        if lot.close_date and start_date <= lot.close_date <= end_date:
            flow_type = "Закрытие репо" if _lot_is_repo(lot) else "Погашение"
            _add(lot.isin, lot.security_code, flow_type, lot.close_date, lot.instrument_currency, lot_maturity_amount_native(lot))
        if lot.next_coupon_date and start_date <= lot.next_coupon_date <= end_date:
            _add(lot.isin, lot.security_code, "Купон", lot.next_coupon_date, lot.instrument_currency, expected_coupon_native(lot))
        purchase_date = getattr(lot, "purchase_date", None)
        quantity = getattr(lot, "quantity", None)
        if purchase_date is None or quantity is None:
            continue
        tax_factor = Decimal("0.85") if is_us_listed(lot.security_code) else Decimal("1")
        for dividend in dividend_history.for_ticker(lot.security_code):
            if (
                dividend.ex_date is None
                or dividend.pay_date is None
                or dividend.ex_date <= purchase_date
                or not start_date <= dividend.pay_date <= end_date
            ):
                continue
            amount = dividend.dividend * quantity * tax_factor
            _add(lot.isin, lot.security_code, "Дивиденд (Bloomberg)", dividend.pay_date, lot.instrument_currency, amount)

    rows = [
        [group["security_code"], flow_type, flow_date, group["amount"] if group["available"] else "Недоступно", group["currency"]]
        for (_isin, flow_type, flow_date), group in groups.items()
    ]
    rows.sort(key=lambda row: (row[2], str(row[0] or ""), str(row[1] or "")))
    return rows


def _append_control_sheet(
    workbook: Workbook, snapshot: PortfolioSnapshotRecord, items: list[dict[str, Any]], *,
    total_value: Decimal, total_weight: Decimal, weight_validation: str,
) -> None:
    """Add a control-total and source-provenance sheet.

    A reviewer can check this sheet alone against the main table without
    re-deriving anything: the same total/weight figures the main sheet
    already shows, plus exactly which source workbook/sheet/row backs every
    instrument on it.
    """
    sheet = workbook.create_sheet("Контроль и происхождение")
    _snapshot_title(sheet, snapshot, "Контроль и происхождение", {})
    expected_weight = Decimal("1")
    control_rows = [
        ["Итоговая стоимость, KZT", total_value, ""],
        ["Итоговый вес", total_weight, ""],
        ["Ожидаемый вес", expected_weight, ""],
        ["Разница", total_weight - expected_weight, ""],
        ["Статус проверки", weight_validation, ""],
        ["Количество ISIN", len(items), ""],
    ]
    control_row = getattr(sheet, "_osip_header_row", 6)
    sheet.cell(control_row, 1, "Контрольные показатели").font = _HEADER_FONT
    sheet.merge_cells(start_row=control_row, start_column=1, end_row=control_row, end_column=3)
    for offset, (label, value, _blank) in enumerate(control_rows, control_row + 1):
        sheet.cell(offset, 1, label).font = Font(bold=True)
        cell = sheet.cell(offset, 2, value)
        if isinstance(value, Decimal):
            cell.number_format = "0.0%;[Red](0.0%);-" if label != "Итоговая стоимость, KZT" else _MONEY_FORMAT
    provenance_row = control_row + len(control_rows) + 2
    provenance_headers = ["ISIN", "Код инструмента", "Эмитент", "Рабочая книга", "Листы", "Строки"]
    provenance_rows = [
        [item["isin"], item["security_code"], item["issuer"], *_sources_columns(item.get("source_refs", []))]
        for item in sorted(items, key=lambda item: item["isin"])
    ]
    for column, label in enumerate(provenance_headers, 1):
        cell = sheet.cell(provenance_row, column, label)
        cell.font = _HEADER_FONT
        cell.border = _TABLE_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number, values in enumerate(provenance_rows, provenance_row + 1):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row_number, column, value)
            cell.border = _TABLE_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column in {3, 5})
    sheet.auto_filter.ref = f"A{provenance_row}:{get_column_letter(len(provenance_headers))}{max(provenance_row, provenance_row + len(provenance_rows))}"
    # Keep the provenance table's header/context rows visible and pin column A
    # while the reviewer scrolls horizontally through the source references.
    sheet.freeze_panes = f"B{provenance_row + 1}"
    for column, width in enumerate([18, 22, 42, 42, 24, 18], 1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _true_class_label(item: dict[str, Any]) -> str:
    return _asset_class_label(item.get("true_asset_class") or item.get("normalized_asset_class"))


def _lot_true_asset_class(lot: Any) -> str:
    """Reclassify a raw lot exactly like the aggregated holdings view does.

    ``InstrumentRecord.normalized_asset_class`` describes the workbook section
    and may say ``ETF`` even when the true economic exposure is a government
    bond (for example SGOV or TIP) or a security-code/sector match. Calling
    the same shared `true_asset_class` the aggregated view uses (rather than
    a reduced, dictionary-only copy of its logic) means the lot export can
    never silently disagree with the main holdings and distribution sheets
    for the same ISIN. Raw source classification remains available through
    the source workbook and is never overwritten in persistence.
    """
    return _asset_class_label(_lot_raw_true_class(lot))


def create_lots_xlsx(snapshot: PortfolioSnapshotRecord) -> bytes:
    workbook = _workbook("Лоты источника OSIP")
    worksheet = workbook.active
    worksheet.title = "Лоты источника"
    _snapshot_title(worksheet, snapshot, "Лоты источника OSIP", {"Детализация": "Неизменяемые строки рабочей книги"})
    total_quantity = sum((_decimal_value(lot.quantity) for lot in snapshot.position_lots if lot.quantity is not None), Decimal("0"))
    total_purchase_kzt = sum((_decimal_value(lot.purchase_amount_kzt) for lot in snapshot.position_lots if lot.purchase_amount_kzt is not None), Decimal("0"))
    total_derived_kzt = sum((_decimal_value(lot.derived_carrying_value_kzt) for lot in snapshot.position_lots if lot.derived_carrying_value_kzt is not None), Decimal("0"))
    worksheet.cell(5, 1, (
        f"Контроль детализации: {len(snapshot.position_lots)} лотов; количество: {total_quantity:,.0f}; "
        f"сумма покупки: {total_purchase_kzt:,.2f} KZT; расчётная стоимость: {total_derived_kzt:,.2f} KZT. "
        "Строки и значения сохранены как в источнике."
    ))
    worksheet.cell(5, 1).font = Font(italic=True, color="666666", size=9)
    worksheet.cell(5, 1).alignment = Alignment(wrap_text=True)
    worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=25)
    rows = []
    for lot in snapshot.position_lots:
        # coupon_or_repo_rate is ingested as a raw fraction (0.19 for a 19%
        # bond, see ingestion/osip_workbook.py) - scale to the same
        # percentage-point convention as every other "%" column on this row
        # (purchase yield, current YTM), or it reads two orders of magnitude
        # too small next to them.
        coupon_rate_points = lot.coupon_or_repo_rate * Decimal("100") if lot.coupon_or_repo_rate is not None else None
        rows.append([lot.isin, lot.security_code, lot.issuer, _lot_true_asset_class(lot), lot.instrument_currency, lot.source_section, lot.quantity, lot.purchase_date or "Недоступно", _optional_decimal(lot.purchase_price), _optional_decimal(lot.purchase_yield), _optional_decimal(lot.current_ytm), _optional_decimal(lot.purchase_amount_native), _optional_decimal(lot.purchase_amount_kzt), _optional_decimal(lot.derived_carrying_value_kzt), _optional_decimal(lot.reserve_kzt), _optional_decimal(lot.organizer_fee_kzt), _optional_decimal(lot.broker_fee_kzt), _optional_decimal(coupon_rate_points), _optional_decimal(lot.nominal_value), lot.previous_coupon_date or "Недоступно", lot.next_coupon_date or "Недоступно", lot.listing_rating or "Недоступно", *_source_columns({"workbook_name": lot.source_row.workbook_name, "sheet_name": lot.source_row.sheet_name, "row_number": lot.source_row.row_number})])
    lot_headers = ["ISIN", "Код инструмента", "Эмитент", "Истинный класс актива", "Валюта", "Раздел источника", "Количество", "Дата покупки", "Цена покупки", "Доходность покупки, %", "Текущая YTM, %", "Сумма покупки в исходной валюте", "Сумма покупки, KZT", "Расчётная стоимость, KZT", "Резерв, KZT", "Комиссия организатора, KZT", "Брокерская комиссия, KZT", "Купон / ставка РЕПО, %", "Номинал", "Предыдущий купон", "Следующий купон", "Листинг / рейтинг", "Рабочая книга", "Лист", "Строка"]
    _table(worksheet, lot_headers, rows, widths=[18, 24, 42, 28, 12, 28, 18, 16, 18, 20, 18, 30, 24, 28, 18, 28, 28, 22, 18, 18, 18, 25, 42, 22, 12], numeric_formats={7: "#,##0;[Red](#,##0);-", 8: "dd.mm.yyyy", 9: _NUMBER_FORMAT, 10: _POINTS_PERCENT_FORMAT, 11: _POINTS_PERCENT_FORMAT, 12: _NUMBER_FORMAT, 13: _MONEY_FORMAT, 14: _MONEY_FORMAT, 15: _MONEY_FORMAT, 16: _MONEY_FORMAT, 17: _MONEY_FORMAT, 18: _POINTS_PERCENT_FORMAT, 19: _NUMBER_FORMAT, 20: "dd.mm.yyyy", 21: "dd.mm.yyyy", 25: "#,##0"}, wrap_columns={2, 3, 4, 6, 22, 23, 24})
    if total_derived_kzt:
        worksheet["A5"].fill = _GOOD_FILL
    return _save(workbook)
