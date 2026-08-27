"""Currency / class / risk / factor distribution sheets for the holdings export."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from osip_dashboard.persistence.models import PortfolioSnapshotRecord
from osip_dashboard.services.excel_charts import chart_series, write_pie_chart
from osip_dashboard.services.fx_rates import resolve_export_usd_kzt_rate
from osip_dashboard.services.hpr import hpr_percent
from osip_dashboard.services.instrument_dictionary import instrument_focus, instrument_rating_group
from osip_dashboard.services.holdings_export.shared import (
    ExportBasis,
    _BASIS_LABELS,
    _HEADER_FONT,
    _MONEY_FORMAT,
    _TABLE_BORDER,
    _TITLE_FONT,
    _asset_class_label,
    _decimal_value,
    _snapshot_title,
)


def _weight_validation(weight: Decimal) -> str:
    if abs(weight - Decimal("1")) <= Decimal("0.0001"):
        return "ОК — 100.0%"
    return f"Фильтрованный вид — {weight * Decimal('100'):.4f}%"


def _append_weight_validation(worksheet, validation: str) -> None:
    disclosure_row = getattr(worksheet, "_osip_header_row", 6) - 2
    worksheet.cell(disclosure_row, 1).value = (
        f"Операционные / расчётные данные из опубликованной версии OSIP; "
        f"не являются официальным NAV или доходностью. Проверка веса: {validation}."
    )


def _distribution_total_row(worksheet, headers: list[str], *, row_number: int, values: dict[int, Decimal | str], label: str = "Итого") -> None:
    for column in range(1, len(headers) + 1):
        cell = worksheet.cell(row_number, column)
        cell.border = _TABLE_BORDER
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="right" if column in values else "left", vertical="center")
    worksheet.cell(row_number, 1, label)
    for column, value in values.items():
        worksheet.cell(row_number, column, value)
        worksheet.cell(row_number, column).font = _HEADER_FONT


def _append_distribution_sheets(workbook: Workbook, snapshot: PortfolioSnapshotRecord, items: list[dict[str, Any]], *, basis: ExportBasis, term: str, asset_class: str | None) -> None:
    """Add source-backed currency, true-class and risk distributions.

    ISIN is the aggregation key used by ``items``.  Every distribution is
    derived from the same filtered instrument view and therefore reconciles to
    the main sheet rather than re-counting individual lots.
    """
    value_key = "purchase_amount_kzt" if basis == "purchase" else "derived_carrying_value_kzt"
    weight_key = "purchase_weight_percent" if basis == "purchase" else "derived_weight_percent"
    native_value_key = "purchase_amount_native" if basis == "purchase" else "carrying_amount_native"
    # Same exclusion the Инструменты sheet applies: an instrument with no
    # derived carrying value (derived_carrying_incomplete) is left out of
    # every one of these distributions instead of folding in as zero, which
    # would silently understate every currency/class/risk/factor bucket it
    # would otherwise belong to.
    priced_items = [item for item in items if item[value_key] is not None]
    total_value = sum((_decimal_value(item[value_key]) for item in priced_items), Decimal("0"))
    total_weight = sum((_decimal_value(item[weight_key]) for item in priced_items), Decimal("0")) / Decimal("100")
    metadata = {"Основа оценки": _BASIS_LABELS[basis], "Поиск": term or "Не применён", "Класс актива": _asset_class_label(asset_class) if asset_class else "Все классы активов"}

    # Keep the four reconciled views together.  This makes the downloaded
    # workbook easier to review and avoids forcing a portfolio manager to
    # switch between tabs while comparing allocations.  Each block remains
    # an independent table and chart; all values come from the same ISIN-keyed
    # filtered holdings list.
    distribution_sheet = workbook.create_sheet("Распределения")
    _snapshot_title(distribution_sheet, snapshot, "Распределение инструментов OSIP", metadata)
    distribution_weight_note = _weight_validation(total_weight)
    excluded_items = [item for item in items if item[value_key] is None]
    if excluded_items:
        excluded_purchase_kzt = sum((_decimal_value(item["purchase_amount_kzt"]) for item in excluded_items), Decimal("0"))
        distribution_weight_note += (
            f" · {len(excluded_items)} инстр. без расчётной стоимости исключены из всех распределений "
            f"(сумма покупки: {excluded_purchase_kzt:,.2f} KZT)."
        )
    _append_weight_validation(distribution_sheet, distribution_weight_note)
    # This is four independent stacked sections (currency/class/risk/factor),
    # each with its own local title, so there is no single meaningful row to
    # freeze.  Column A is still pinned so labels remain visible during
    # horizontal scrolling.
    distribution_sheet.freeze_panes = "B1"
    distribution_sheet.sheet_view.showGridLines = False

    # Each chart reads from its own visible, real Excel Table here rather
    # than a hidden range in "Распределения" - editing a value on this sheet
    # moves the corresponding chart. A slice grouped into "Прочее" reflects a
    # materiality decision (see chart_series) made once at export time, not
    # something re-derived live from the main table if it's edited afterward.
    charts_sheet = workbook.create_sheet("Данные графиков")
    charts_sheet.sheet_view.showGridLines = False
    # Chart source tables are wide enough to require horizontal scrolling;
    # keep the category/label column visible without pinning an unrelated
    # section header.
    charts_sheet.freeze_panes = "B1"
    charts_sheet.cell(1, 1, "Источник данных для диаграмм листа «Распределения»").font = _TITLE_FONT
    charts_sheet.cell(2, 1, (
        "Категории с долей менее 2% объединяются в «Прочее», только если таких категорий "
        "две и более. Изменение значения здесь переместит соответствующую диаграмму; "
        "таблицы на листе «Распределения» не пересчитываются от этого листа."
    )).alignment = Alignment(wrap_text=True)
    charts_sheet.merge_cells("A2:D2")
    charts_row = 4

    workbook_rate = _snapshot_usd_rates(snapshot)
    fx_rate = resolve_export_usd_kzt_rate(snapshot.report_date, workbook_rate)
    rates = fx_rate.rate if fx_rate is not None else None
    currencies: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"kzt": Decimal("0"), "usd": Decimal("0"), "weight": Decimal("0")})
    for item in priced_items:
        currency = item.get("instrument_currency") or "Не указано"
        currencies[currency]["kzt"] += _decimal_value(item[value_key])
        currencies[currency]["weight"] += _decimal_value(item[weight_key]) / Decimal("100")
        usd_value = _usd_equivalent(item, value_key=value_key, native_value_key=native_value_key, usd_rate=rates)
        if usd_value is not None:
            currencies[currency]["usd"] += usd_value
    currency_rows = [[currency, values["kzt"], values["usd"] if values["usd"] else "Недоступно", values["weight"]] for currency, values in sorted(currencies.items(), key=lambda pair: pair[1]["weight"], reverse=True)]
    currency_headers = ["Валюта", "MV, ₸", "MV, $", "Вес, %"]
    # The source-date warning can add a metadata row to the title block. Start
    # the first distribution section after that block so its merge never
    # overlaps the explanatory note row.
    currency_row = getattr(distribution_sheet, "_osip_header_row", 6)
    currency_total, charts_row = _write_distribution_block(distribution_sheet, currency_row, "Распределение по валютам", currency_headers, currency_rows, widths=[18, 24, 24, 14], numeric_formats={2: _MONEY_FORMAT, 3: _MONEY_FORMAT, 4: "0.0%;[Red](0.0%);-"}, chart_column=4, chart_title="Вес по валютам", charts_sheet=charts_sheet, charts_row=charts_row, table_name="ChartDataCurrency", chart_anchor_column="G")

    classes: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"value": Decimal("0"), "weight": Decimal("0")})
    for item in priced_items:
        bucket = classes[item.get("true_asset_class") or "Unrated"]
        bucket["value"] += _decimal_value(item[value_key])
        bucket["weight"] += _decimal_value(item[weight_key]) / Decimal("100")
    class_rows = [[_asset_class_label(key), value["value"], value["weight"]] for key, value in sorted(classes.items(), key=lambda pair: pair[1]["weight"], reverse=True)]
    class_headers = ["Истинный класс актива", "MV, ₸", "Вес, %"]
    # Leave enough vertical room for the chart anchored beside each block;
    # otherwise a short table could cause the next block's chart to overlap it.
    class_row = max(currency_total + 3, currency_row + 16)
    class_total, charts_row = _write_distribution_block(distribution_sheet, class_row, "Распределение по истинному классу актива", class_headers, class_rows, widths=[32, 28, 14], numeric_formats={2: _MONEY_FORMAT, 3: "0.0%;[Red](0.0%);-"}, wrap_columns={1}, chart_column=3, chart_title="Вес по истинному классу", charts_sheet=charts_sheet, charts_row=charts_row, table_name="ChartDataClass")

    risk: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"value": Decimal("0"), "weight": Decimal("0")})
    for item in priced_items:
        bucket = risk[_risk_bucket(item)]
        bucket["value"] += _decimal_value(item[value_key])
        bucket["weight"] += _decimal_value(item[weight_key]) / Decimal("100")
    risk_order = ["[A-]+", "[BBB-]+", "[BB-]+", "[B-]+", "[CCC+]-", "Товары / сырьё", "Акции", "Рейтинг не указан"]
    risk_rows = [[bucket, risk[bucket]["value"], risk[bucket]["weight"]] for bucket in risk_order if risk[bucket]["value"]]
    risk_headers = ["Рейтинг / тип риска", "MV, ₸", "Вес, %"]
    risk_row = max(class_total + 3, class_row + 16)
    risk_total, charts_row = _write_distribution_block(distribution_sheet, risk_row, "Распределение по рейтингу и типу риска", risk_headers, risk_rows, widths=[28, 28, 14], numeric_formats={2: _MONEY_FORMAT, 3: "0.0%;[Red](0.0%);-"}, chart_column=3, chart_title="Вес по риску", charts_sheet=charts_sheet, charts_row=charts_row, table_name="ChartDataRisk")

    factors: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"value": Decimal("0"), "weight": Decimal("0")})
    for item in priced_items:
        factor = _instrument_focus(item)
        factors[factor]["value"] += _decimal_value(item[value_key])
        factors[factor]["weight"] += _decimal_value(item[weight_key]) / Decimal("100")
    factor_rows = [[factor, values["value"], values["weight"]] for factor, values in sorted(factors.items(), key=lambda pair: pair[1]["weight"], reverse=True)]
    factor_headers = ["Focus/sector/factor", "MV, ₸", "Вес, %"]
    factor_row = max(risk_total + 3, risk_row + 16)
    _write_distribution_block(distribution_sheet, factor_row, "Распределение по фокусу / сектору / фактору", factor_headers, factor_rows, widths=[34, 28, 14], numeric_formats={2: _MONEY_FORMAT, 3: "0.0%;[Red](0.0%);-"}, wrap_columns={1}, chart_column=3, chart_title="Вес по фокусу / сектору / фактору", charts_sheet=charts_sheet, charts_row=charts_row, table_name="ChartDataFactor")

    if fx_rate is None:
        fx_method = "USD-эквивалент не рассчитан: официальный курс НБК и согласованный курс рабочей книги недоступны."
    else:
        rate_text = f"{fx_rate.rate:,.4f} KZT"
        effective = fx_rate.effective_date.isoformat()
        url_text = f" URL: {fx_rate.source_url}." if fx_rate.source_url else ""
        if fx_rate.fallback:
            fx_method = f"USD-эквивалент: 1 USD = {rate_text}; источник: {fx_rate.source}; дата отчёта {snapshot.report_date.isoformat()}.{url_text}"
        else:
            fx_method = f"USD-эквивалент: 1 USD = {rate_text}; источник: {fx_rate.source}; дата курса {effective}; отчётная дата {snapshot.report_date.isoformat()}.{url_text}"
    note_row = currency_row - 1
    distribution_sheet.cell(note_row, 1, fx_method)
    distribution_sheet.cell(note_row, 1).font = Font(italic=True, color="666666", size=9)
    distribution_sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    distribution_sheet.column_dimensions["A"].width = 34
    distribution_sheet.column_dimensions["B"].width = 28
    distribution_sheet.column_dimensions["C"].width = 24
    distribution_sheet.column_dimensions["D"].width = 14


def _snapshot_usd_rates(snapshot: PortfolioSnapshotRecord) -> Decimal | None:
    """Return the report-date KZT-per-USD rate carried by the workbook.

    OSIP stores a native-currency amount and a report FX rate on each lot.  A
    KZT lot has a neutral rate of 1, while USD lots carry the useful
    KZT-per-USD rate.  We only infer a single rate when the USD source rows
    agree (within 1%); otherwise the USD equivalent is deliberately shown as
    unavailable rather than silently mixing rates.
    """
    rates = sorted(
        lot.report_fx_rate
        for lot in snapshot.position_lots
        if lot.instrument_currency == "USD"
        and lot.report_fx_rate is not None
        and lot.report_fx_rate > 0
    )
    if not rates:
        return None
    if rates[-1] / rates[0] - Decimal("1") > Decimal("0.01"):
        return None
    return rates[len(rates) // 2]


def _lot_usd_equivalent(
    lot: Any,
    *,
    kzt_value: Decimal | None,
    native_value_key: str,
    usd_rate: Decimal | None,
) -> Decimal | None:
    """Return one lot's selected value in USD without hiding missing inputs."""
    if kzt_value is None:
        return None
    currency = str(getattr(lot, "instrument_currency", "") or "").upper()
    if currency == "USD":
        native_value = getattr(lot, native_value_key, None)
        if native_value is not None:
            return _decimal_value(native_value)
        return kzt_value / usd_rate if usd_rate else None
    if currency == "KZT" and usd_rate:
        return kzt_value / usd_rate
    return None


def _usd_equivalent(item: dict[str, Any], *, value_key: str, native_value_key: str, usd_rate: Decimal | None) -> Decimal | None:
    currency = item.get("instrument_currency")
    native = item.get(native_value_key)
    if currency == "USD" and native is not None:
        return _decimal_value(native)
    if currency == "KZT" and usd_rate:
        return _decimal_value(item[value_key]) / usd_rate
    return None


def _hpr_usd_percent_for_item(item: dict[str, Any], *, usd_rate: Decimal | None) -> Decimal | None:
    """Return an instrument's HPR in USD percentage points when convertible.

    The KZT and USD percentages use the same carrying/purchase/dividend
    numerator. USD conversion is only allowed for USD-native and KZT lots;
    other currencies remain explicitly unavailable rather than being treated
    as if they were KZT or USD.
    """
    def decimal_or_none(value: Any) -> Decimal | None:
        return None if value in (None, "Недоступно") else Decimal(value)

    currency = str(item.get("instrument_currency") or "").upper()
    purchase_kzt = decimal_or_none(item.get("purchase_amount_kzt"))
    carrying_kzt = decimal_or_none(item.get("derived_carrying_value_kzt"))
    coupon_kzt = decimal_or_none(item.get("coupon_income_kzt_estimated")) or Decimal("0")
    coupon_native = decimal_or_none(item.get("coupon_income_native_estimated")) or Decimal("0")
    if currency == "USD":
        purchase_usd = decimal_or_none(item.get("purchase_amount_native"))
        carrying_usd = decimal_or_none(item.get("carrying_amount_native"))
        if purchase_usd is None and purchase_kzt is not None and usd_rate:
            purchase_usd = purchase_kzt / usd_rate
        if carrying_usd is None and carrying_kzt is not None and usd_rate:
            carrying_usd = carrying_kzt / usd_rate
        # A USD-native dividend remains usable in USD even when its KZT
        # conversion is unavailable; the two percentage columns have distinct
        # availability rules.
        dividend_usd = (decimal_or_none(item.get("dividend_income_native")) or Decimal("0")) + coupon_native
    elif currency == "KZT":
        # A common KZT/USD conversion factor cancels in a percentage return.
        # Keep the USD percentage available even for a KZT-only portfolio
        # whose export cannot establish a USD amount/rate.
        if item.get("dividend_unavailable"):
            return None
        return hpr_percent(
            purchase_kzt,
            carrying_kzt,
            (decimal_or_none(item.get("dividend_income_kzt")) or Decimal("0")) + coupon_kzt,
        )
    else:
        return None
    return hpr_percent(purchase_usd, carrying_usd, dividend_usd)


def _write_distribution_block(
    worksheet,
    section_row: int,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    widths: list[int],
    numeric_formats: dict[int, str],
    chart_column: int,
    chart_title: str,
    charts_sheet,
    charts_row: int,
    table_name: str,
    chart_anchor_column: str = "I",
    wrap_columns: set[int] | None = None,
) -> tuple[int, int]:
    """Write one bordered table and its pie chart.

    Returns ``(total_row, next_charts_row)``: ``total_row`` is this block's
    own total row in ``worksheet``; ``next_charts_row`` is the first free row
    in ``charts_sheet`` after this block's chart-source table, for the
    caller to thread into the next block.
    """
    wrap_columns = wrap_columns or set()
    worksheet.cell(section_row, 1, title)
    worksheet.cell(section_row, 1).font = _HEADER_FONT
    worksheet.cell(section_row, 1).alignment = Alignment(vertical="center")
    worksheet.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=len(headers))
    header_row = section_row + 1
    for column, value in enumerate(headers, 1):
        cell = worksheet.cell(header_row, column, value)
        cell.font = _HEADER_FONT
        cell.border = _TABLE_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 30
    for row_number, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row_number, column, value)
            cell.border = _TABLE_BORDER
            cell.alignment = Alignment(horizontal="right" if column in numeric_formats else "left", vertical="top", wrap_text=column in wrap_columns)
            if column in numeric_formats and value != "Недоступно":
                cell.number_format = numeric_formats[column]
    total_row = header_row + len(rows) + 1
    totals: dict[int, Decimal | str] = {}
    for column in range(2, len(headers) + 1):
        column_values = [value[column - 1] for value in rows]
        # A "Недоступно" cell means that row's contribution to this column
        # genuinely could not be computed (e.g. no resolvable FX rate for
        # that currency) - summing only the rows that DID resolve would
        # print a confidently-formatted total that's silently short by
        # whatever was skipped. Disclose the same way the row itself does
        # rather than hiding the gap in the total.
        if any(value == "Недоступно" for value in column_values):
            totals[column] = "Недоступно"
        else:
            numeric_values = [value for value in column_values if isinstance(value, Decimal)]
            if numeric_values:
                totals[column] = sum(numeric_values, Decimal("0"))
    _distribution_total_row(worksheet, headers, row_number=total_row, values=totals)
    for column, number_format in numeric_formats.items():
        worksheet.cell(total_row, column).number_format = number_format
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(worksheet.column_dimensions[get_column_letter(index)].width or 0, width)

    next_charts_row = charts_row
    if rows:
        # The pie reads from its own visible table on "Данные графиков"
        # rather than from this sheet's columns directly, so slices under 2%
        # can be grouped into one "Прочее" wedge for the chart alone - the
        # table above keeps every row exactly as parsed. Keeping this source
        # on a real, visible Excel Table (not a hidden range) means editing
        # a value there and watching the chart move is possible, unlike the
        # earlier hidden-helper-column approach.
        chart_labels, chart_values = chart_series(rows, chart_column)
        charts_title_row = charts_row
        charts_sheet.cell(charts_title_row, 1, chart_title).font = _HEADER_FONT
        charts_header_row = charts_title_row + 1
        charts_sheet.cell(charts_header_row, 1, "Категория")
        charts_sheet.cell(charts_header_row, 2, "Доля")
        for offset, (label, value) in enumerate(zip(chart_labels, chart_values)):
            charts_sheet.cell(charts_header_row + 1 + offset, 1, label)
            charts_sheet.cell(charts_header_row + 1 + offset, 2, value).number_format = "0.0%;[Red](0.0%);-"
        charts_last_row = charts_header_row + len(chart_labels)
        table = Table(displayName=table_name, ref=f"A{charts_header_row}:B{charts_last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        charts_sheet.add_table(table)
        for column, width in (("A", 34), ("B", 14)):
            charts_sheet.column_dimensions[column].width = max(charts_sheet.column_dimensions[column].width or 0, width)
        next_charts_row = charts_last_row + 3

        # Anchor with a visible gap after the table (which ends around
        # column D) rather than right against it - column I leaves a few
        # blank columns of margin between the table and the chart.
        write_pie_chart(
            charts_sheet,
            label_col=1, value_col=2,
            first_row=charts_header_row + 1, last_row=charts_last_row,
            values=chart_values, title=chart_title, anchor=f"{chart_anchor_column}{section_row + 2}",
            anchor_worksheet=worksheet,
        )
    return total_row, next_charts_row


_RATING_ORDER = {
    # S&P/Fitch style values.  Lower numbers are stronger ratings.
    "AAA": 0, "AA+": 1, "AA": 2, "AA-": 3, "A+": 4, "A": 5, "A-": 6,
    "BBB+": 7, "BBB": 8, "BBB-": 9, "BB+": 10, "BB": 11, "BB-": 12,
    "B+": 13, "B": 14, "B-": 15, "CCC+": 16, "CCC": 17, "CCC-": 18,
    "CC": 19, "C": 20, "D": 21,
    # Moody's equivalent bands used by the supplied OSIP workbook.
    "AAA_M": 0, "AA1": 1, "AA2": 2, "AA3": 3, "A1": 4, "A2": 5, "A3": 6,
    "BAA1": 9, "BAA2": 9, "BAA3": 9, "BA1": 12, "BA2": 12, "BA3": 12,
    "B1": 15, "B2": 15, "B3": 15, "CAA1": 16, "CAA2": 16, "CAA3": 16,
}


_CYRILLIC_LATIN_HOMOGLYPHS = str.maketrans(
    {"А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T", "У": "Y", "Х": "X"}
)


def _normalise_rating(value: str | None) -> str:
    # The portfolio team maintains the ratings dictionary by hand in Excel,
    # where autocorrect/keyboard-layout slips substitute visually identical
    # Cyrillic letters for Latin ones (e.g. Cyrillic "С" in "ССС+"). Without
    # this, such a row silently fails to match any known rating bucket.
    text = (value or "").strip().upper().replace(" ", "").replace("—", "-")
    return text.translate(_CYRILLIC_LATIN_HOMOGLYPHS)


def _rating_score(item: dict[str, Any]) -> int | None:
    scores: list[int] = []
    for value in (item.get("rating_sp"), item.get("rating_moodys"), item.get("rating_fitch"), item.get("listing_rating")):
        normalized = _normalise_rating(value)
        if not normalized:
            continue
        # Listing values such as «ГЦБ» are classifications, not ratings.
        if normalized in {"ГЦБ", "ETF", "ЦЕННЫЕБУМАГИИНВЕСТИЦИОННЫХФОНДОВ"}:
            continue
        score = _RATING_ORDER.get(normalized)
        if score is None and normalized.startswith("A") and normalized.endswith("+"):
            score = 4
        if score is not None:
            scores.append(score)
    return max(scores) if scores else None


def _instrument_focus(item: dict[str, Any]) -> str:
    true_class = item.get("true_asset_class") or item.get("normalized_asset_class")
    if true_class == "Repo":
        # Same reasoning as the repo rating rule below: auto-repo ISINs roll
        # over every period, so a per-ISIN dictionary lookup can't be relied
        # on to classify them. Per the portfolio team, авторепо's true focus
        # is always Money market, regardless of which ISIN it landed under.
        return "Money market"
    return instrument_focus(item.get("isin")) or "Не указано"


def _risk_bucket(item: dict[str, Any]) -> str:
    true_class = item.get("true_asset_class") or item.get("normalized_asset_class")
    if true_class == "Commodity":
        return "Товары / сырьё"
    if true_class == "Equity":
        return "Акции"
    if true_class == "Repo":
        # Auto-repo agreements roll over each period under fresh ISINs, so a
        # static per-ISIN dictionary entry can never keep up with them - and
        # they never carry a public S&P/Moody's/Fitch rating of their own.
        # Per the portfolio team, авторепо is always rated [BBB-].
        return "[BBB-]+"
    dictionary_rating = _normalise_rating(instrument_rating_group(item.get("isin")))
    dictionary_bucket = {
        "A-": "[A-]+", "BBB-": "[BBB-]+", "BB-": "[BB-]+", "B-": "[B-]+", "CCC+": "[CCC+]-",
    }.get(dictionary_rating)
    if dictionary_bucket:
        return dictionary_bucket
    if dictionary_rating == "EQUITIES":
        return "Акции"
    score = _rating_score(item)
    if score is None:
        return "Рейтинг не указан"
    if score <= 6:
        return "[A-]+"
    if score <= 9:
        return "[BBB-]+"
    if score <= 12:
        return "[BB-]+"
    if score <= 15:
        return "[B-]+"
    return "[CCC+]-"
