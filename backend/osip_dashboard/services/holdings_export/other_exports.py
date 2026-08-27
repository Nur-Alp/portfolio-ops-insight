"""Cash calendar, data-quality issue and import-registry exports."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal
import json
from typing import Any, Iterable

from openpyxl.styles import Alignment, Font

from osip_dashboard.persistence.models import ImportBatch, PortfolioSnapshotRecord
from osip_dashboard.services.excel_charts import write_pie_chart
from osip_dashboard.services.holdings_export.distribution import _distribution_total_row
from osip_dashboard.services.holdings_export.shared import (
    _GOOD_FILL,
    _HEADER_FONT,
    _MONEY_FORMAT,
    _NUMBER_FORMAT,
    _WARN_FILL,
    _BAD_FILL,
    _amount_basis,
    _date_value,
    _decimal_value,
    _global_title,
    _optional_decimal,
    _save,
    _snapshot_title,
    _source_columns,
    _sources_columns,
    _table,
    _workbook,
)


_EVENT_TYPES = {
    "settlement": "Расчёт по сделке",
    "repo_open": "Открытие РЕПО",
    "repo_close": "Закрытие РЕПО",
    "instrument_open": "Открытие инструмента",
    "maturity": "Погашение",
    "previous_coupon": "Предыдущий купон",
    "next_coupon": "Следующий купон",
}


_EVENT_STATUS = {"upcoming": "Предстоящее", "overdue": "Просроченное", "historical": "Прошедшее"}


_SEVERITY = {"blocker": "Блокирующая", "high": "Высокая", "medium": "Средняя", "low": "Низкая"}


_IMPORT_STATUS = {
    "draft": "Черновик", "validating": "Проверяется", "validated": "Проверено", "approved": "Утверждено",
    "published": "Опубликовано", "failed": "Ошибка", "rejected": "Отклонено", "superseded": "Заменено", "withdrawn": "Снято с публикации",
}


def create_cash_calendar_xlsx(snapshot: PortfolioSnapshotRecord, cash_items: list[dict[str, Any]], calendar_items: list[dict[str, Any]], *, include_inactive: bool) -> bytes:
    """Export cash, its currency summary and operational calendar in one workbook."""
    workbook = _workbook("Деньги и календарь OSIP")
    cash_sheet = workbook.active
    cash_sheet.title = "Денежные средства"
    visible_cash = [item for item in cash_items if include_inactive or item["active"]]
    _snapshot_title(cash_sheet, snapshot, "Денежные средства OSIP", {"Нулевые шаблоны": "Включены" if include_inactive else "Исключены"})
    visible_cash_kzt = sum((_decimal_value(item.get("kzt_amount")) for item in visible_cash), Decimal("0"))
    cash_sheet.cell(5, 1, (
        f"Контроль вида: {len(visible_cash)} строк; общий эквивалент: "
        f"{visible_cash_kzt:,.2f} KZT. Каждая строка сохраняет рабочую книгу, лист и строку источника."
    ))
    cash_sheet.cell(5, 1).font = Font(italic=True, color="666666", size=9)
    cash_sheet.cell(5, 1).alignment = Alignment(wrap_text=True)
    cash_sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=9)
    cash_rows = [[item.get("custodian") or "Не указано", item["raw_label"], item["currency"], _decimal_value(item["native_amount"]), _decimal_value(item["kzt_amount"]), "Да" if item["active"] else "Нет", *_source_columns(item["source"])] for item in visible_cash]
    _table(cash_sheet, ["Кастодиан", "Исходное наименование", "Валюта", "Остаток в исходной валюте", "Эквивалент, KZT", "Активна", "Рабочая книга", "Лист", "Строка"], cash_rows, widths=[28, 40, 12, 25, 22, 12, 42, 22, 12], numeric_formats={4: _NUMBER_FORMAT, 5: _MONEY_FORMAT, 9: "#,##0"}, wrap_columns={1, 2, 7})
    cash_header = getattr(cash_sheet, "_osip_header_row", 6)
    for row_number, item in enumerate(visible_cash, cash_header + 1):
        if item.get("active"):
            cash_sheet.cell(row_number, 6).fill = _GOOD_FILL
        else:
            cash_sheet.cell(row_number, 6).fill = _WARN_FILL
    # Keep the source rows untouched in the filter range; the control value is
    # deliberately above the header so filtering still shows exactly the
    # rows delivered by the API.
    if visible_cash_kzt:
        cash_sheet["A5"].fill = _GOOD_FILL

    summary_sheet = workbook.create_sheet("Сводка валют")
    _snapshot_title(summary_sheet, snapshot, "Сводка денежных средств по валютам", {"Нулевые шаблоны": "Включены" if include_inactive else "Исключены"})
    summary: dict[str, list[Decimal]] = defaultdict(lambda: [Decimal("0"), Decimal("0")])
    for item in visible_cash:
        summary[item["currency"]][0] += _decimal_value(item["native_amount"])
        summary[item["currency"]][1] += _decimal_value(item["kzt_amount"])
    summary_rows = [[currency, values[0], values[1], sum(1 for item in visible_cash if item["currency"] == currency)] for currency, values in sorted(summary.items())]
    _table(summary_sheet, ["Валюта", "Остаток в исходной валюте", "Эквивалент, KZT", "Строк источника"], summary_rows, widths=[16, 28, 24, 20], numeric_formats={2: _NUMBER_FORMAT, 3: _MONEY_FORMAT, 4: "#,##0"})
    summary_header = getattr(summary_sheet, "_osip_header_row", 6)
    summary_total_row = summary_header + len(summary_rows) + 1
    _distribution_total_row(
        summary_sheet,
        ["Валюта", "Остаток в исходной валюте", "Эквивалент, KZT", "Строк источника"],
        row_number=summary_total_row,
        values={3: sum((row[2] for row in summary_rows), Decimal("0")), 4: Decimal(len(visible_cash))},
    )
    summary_sheet.cell(summary_total_row, 3).number_format = _MONEY_FORMAT
    summary_sheet.cell(summary_total_row, 4).number_format = "#,##0"
    if summary_rows:
        # A native chart is intentionally tied to the visible summary table,
        # not a hidden helper range: changing the KZT totals in Excel updates
        # the chart without rebuilding the workbook.
        chart_values = [row[2] for row in summary_rows]
        write_pie_chart(
            summary_sheet,
            label_col=1,
            value_col=3,
            first_row=summary_header + 1,
            last_row=summary_header + len(summary_rows),
            values=chart_values,
            title="Денежные средства по валютам",
            anchor="F6",
        )

    calendar_sheet = workbook.create_sheet("Календарь событий")
    _snapshot_title(calendar_sheet, snapshot, "Календарь событий OSIP", {"Предстоящие расчёты": "Исключены"})
    # The section header above promises upcoming settlements are excluded -
    # matching how OSIP's own "предстоящие расчёты" section is already
    # treated as out-of-scope everywhere else this app discloses the same
    # wording (e.g. the Expected Cash Flows sheet). snapshot_calendar()
    # itself still includes them (the live Cash & Calendar page legitimately
    # shows pending settlements), so the export is the one place that must
    # filter them out to keep its own disclosure true.
    calendar_items = [item for item in calendar_items if item.get("event_type") != "settlement"]
    known_calendar_kzt = sum((_decimal_value(item.get("amount_kzt")) for item in calendar_items if item.get("amount_kzt") is not None), Decimal("0"))
    calendar_sheet.cell(5, 1, (
        f"Контроль календаря: {len(calendar_items)} событий; доступная сумма в KZT: "
        f"{known_calendar_kzt:,.2f}. Недоступные суммы не заменяются нулём."
    ))
    calendar_sheet.cell(5, 1).font = Font(italic=True, color="666666", size=9)
    calendar_sheet.cell(5, 1).alignment = Alignment(wrap_text=True)
    calendar_sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=12)
    calendar_rows = [[_date_value(item["event_date"]), _EVENT_TYPES.get(item["event_type"], item["event_type"]), item.get("security_code") or "Недоступно", item.get("isin") or "Недоступно", _EVENT_STATUS.get(item["status"], item["status"]), _optional_decimal(item.get("amount_native")), _optional_decimal(item.get("amount_kzt")), item.get("currency") or "Недоступно", _amount_basis(item.get("amount_basis")), *_sources_columns(item.get("source_refs", []))] for item in calendar_items]
    calendar_headers = ["Дата", "Тип", "Инструмент", "ISIN", "Статус", "Сумма в исходной валюте", "Сумма, KZT", "Валюта", "Основа суммы", "Рабочая книга", "Листы", "Строки"]
    _table(calendar_sheet, calendar_headers, calendar_rows, widths=[16, 24, 24, 18, 18, 26, 20, 12, 42, 42, 24, 18], numeric_formats={1: "dd.mm.yyyy", 6: _NUMBER_FORMAT, 7: _MONEY_FORMAT}, wrap_columns={2, 3, 9, 10, 11, 12})
    calendar_header = getattr(calendar_sheet, "_osip_header_row", 6)
    calendar_total_row = calendar_header + len(calendar_rows) + 1
    _distribution_total_row(
        calendar_sheet,
        calendar_headers,
        row_number=calendar_total_row,
        values={7: known_calendar_kzt},
    )
    calendar_sheet.cell(calendar_total_row, 7).number_format = _MONEY_FORMAT
    return _save(workbook)


def create_dq_issues_xlsx(snapshot: PortfolioSnapshotRecord, issues: list[dict[str, Any]], *, term: str, severity: str) -> bytes:
    workbook = _workbook("Проверки качества данных OSIP")
    worksheet = workbook.active
    worksheet.title = "Замечания DQ"
    _snapshot_title(worksheet, snapshot, "Проверки качества данных OSIP", {"Поиск": term or "Не применён", "Серьёзность": "Все уровни" if severity == "all" else _SEVERITY.get(severity, severity)})
    blocker_high = sum(1 for issue in issues if issue.get("severity") in {"blocker", "high"})
    acknowledged = sum(1 for issue in issues if issue.get("acknowledgement"))
    worksheet.cell(5, 1, (
        f"Контроль вида: {len(issues)} замечаний; блокирующих/высоких: {blocker_high}; "
        f"подтверждено: {acknowledged}. Фильтры применены к текущему опубликованному снимку."
    ))
    worksheet.cell(5, 1).font = Font(italic=True, color="666666", size=9)
    worksheet.cell(5, 1).alignment = Alignment(wrap_text=True)
    worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=13)
    if blocker_high:
        worksheet["A5"].fill = _WARN_FILL
    rows = []
    for issue in issues:
        acknowledgement = issue.get("acknowledgement") or {}
        rows.append([issue["code"], _SEVERITY.get(issue["severity"], issue["severity"]), issue["message"], ", ".join(issue.get("affected_fields", [])) or "Портфель", *_sources_columns(issue.get("source_refs", [])), acknowledgement.get("actor_id") or "Не подтверждено", acknowledgement.get("comment") or "Недоступно", _date_value(acknowledgement.get("acknowledged_at")), issue.get("owner_id") or "Не назначен", _date_value(issue.get("due_date")), "Да" if issue.get("is_overdue") else "Нет"])
    _table(worksheet, ["Правило", "Серьёзность", "Замечание", "Затронутые поля", "Рабочая книга", "Листы", "Строки", "Подтвердил", "Комментарий подтверждения", "Дата подтверждения (UTC)", "Ответственный", "Срок", "Просрочено"], rows, widths=[14, 18, 55, 30, 42, 22, 18, 22, 48, 20, 24, 16, 16], numeric_formats={7: "#,##0", 10: "dd.mm.yyyy hh:mm", 12: "dd.mm.yyyy"}, wrap_columns={3, 4, 5, 6, 7, 9, 11})
    header_row = getattr(worksheet, "_osip_header_row", 6)
    for offset, issue in enumerate(issues, header_row + 1):
        severity_fill = _BAD_FILL if issue.get("severity") in {"blocker", "high"} else _WARN_FILL if issue.get("severity") == "medium" else None
        if severity_fill:
            worksheet.cell(offset, 2).fill = severity_fill
        if issue.get("is_overdue"):
            worksheet.cell(offset, 13).fill = _BAD_FILL
        elif issue.get("acknowledgement"):
            worksheet.cell(offset, 8).fill = _GOOD_FILL

    # A compact severity chart lives beside the filtered table. Its source
    # values are kept in visible cells so a reviewer can edit or audit the
    # chart without relying on hidden formulas. For a single filtered issue,
    # the table remains the clearer representation and no redundant chart is
    # drawn.
    severity_counts: dict[str, Decimal] = defaultdict(Decimal)
    for issue in issues:
        severity_counts[_SEVERITY.get(issue.get("severity"), issue.get("severity") or "Не указано")] += Decimal("1")
    if len(severity_counts) >= 2:
        worksheet.cell(2, 15, "Сводка серьёзности для диаграммы").font = _HEADER_FONT
        worksheet.cell(3, 15, "Серьёзность").font = _HEADER_FONT
        worksheet.cell(3, 16, "Количество").font = _HEADER_FONT
        severity_rows = sorted(severity_counts.items(), key=lambda pair: pair[1], reverse=True)
        for row_number, (label, count) in enumerate(severity_rows, 4):
            worksheet.cell(row_number, 15, label)
            worksheet.cell(row_number, 16, count).number_format = "#,##0"
        write_pie_chart(
            worksheet,
            label_col=15,
            value_col=16,
            first_row=4,
            last_row=3 + len(severity_rows),
            values=[count for _, count in severity_rows],
            title="Замечания по серьёзности",
            anchor="O8",
        )
        worksheet.column_dimensions["O"].width = 22
        worksheet.column_dimensions["P"].width = 14
    return _save(workbook)


def create_import_registry_xlsx(batches: Iterable[ImportBatch]) -> bytes:
    materialized = list(batches)
    workbook = _workbook("Реестр загрузок OSIP")
    registry = workbook.active
    registry.title = "Реестр загрузок"
    _global_title(registry, "Реестр загрузок OSIP", {"Версий в выгрузке": str(len(materialized)), "Данные": "Операционные / расчётные"})
    status_counts: dict[str, int] = defaultdict(int)
    portfolio_codes: set[str] = set()
    total_critical = 0
    for batch in materialized:
        status_counts[_IMPORT_STATUS.get(batch.status.value, batch.status.value)] += 1
        if batch.portfolio_code:
            portfolio_codes.add(batch.portfolio_code)
        if batch.snapshot:
            total_critical += sum(1 for issue in batch.snapshot.issues if issue.severity in {"blocker", "high"})
    status_note = "; ".join(f"{label}: {count}" for label, count in sorted(status_counts.items())) or "нет строк"
    registry.cell(4, 1, (
        f"Контроль реестра: портфели {', '.join(sorted(portfolio_codes)) or 'не определены'}; "
        f"статусы — {status_note}; блокирующих/высоких DQ: {total_critical}. "
        "Исходные файлы и события аудита сохранены без удаления доказательств."
    ))
    registry.cell(4, 1).font = Font(italic=True, color="666666", size=9)
    registry.cell(4, 1).alignment = Alignment(wrap_text=True)
    registry.merge_cells(start_row=4, start_column=1, end_row=4, end_column=21)
    registry_rows = []
    audit_rows = []
    for batch in materialized:
        snapshot = batch.snapshot
        counts = defaultdict(int)
        if snapshot:
            for issue in snapshot.issues:
                counts[issue.severity] += 1
        registry_rows.append([batch.original_filename, batch.source_sha256, batch.portfolio_code or "Не определён", batch.report_date or "Недоступно", batch.version if batch.version is not None else "Недоступно", _IMPORT_STATUS.get(batch.status.value, batch.status.value), snapshot.position_count if snapshot else "Недоступно", snapshot.unique_isin_count if snapshot else "Недоступно", _optional_decimal(snapshot.purchase_amount_kzt if snapshot else None), _optional_decimal(snapshot.derived_carrying_value_kzt if snapshot else None), _optional_decimal(snapshot.cash_kzt if snapshot else None), _optional_decimal(snapshot.derived_operational_total_kzt if snapshot else None), counts["blocker"] + counts["high"], batch.uploader_id, batch.reviewer_id or "Недоступно", batch.publisher_id or "Недоступно", batch.created_at, batch.validated_at or "Недоступно", batch.approved_at or "Недоступно", batch.published_at or "Недоступно", batch.rejection_reason or batch.error_message or "Недоступно"])
        for event in sorted(batch.audit_events, key=lambda value: value.created_at):
            audit_rows.append([batch.portfolio_code or "Не определён", batch.report_date or "Недоступно", batch.version if batch.version is not None else "Недоступно", batch.original_filename, event.created_at, event.action, event.actor_id, json.dumps(event.detail, ensure_ascii=False, sort_keys=True)])
    registry_headers = ["Исходный файл", "SHA-256", "Портфель", "Отчётная дата", "Версия", "Статус", "Лоты", "Инструменты", "Сумма покупки, KZT", "Расчётная стоимость, KZT", "Денежные средства, KZT", "Операционный итог, KZT", "DQ: блок./высокие", "Загрузил", "Проверил", "Опубликовал", "Создано (UTC)", "Проверено (UTC)", "Утверждено (UTC)", "Опубликовано (UTC)", "Причина отклонения / снятия"]
    _table(registry, registry_headers, registry_rows, widths=[55, 32, 18, 18, 12, 20, 12, 16, 26, 28, 26, 26, 22, 22, 22, 22, 22, 22, 22, 22, 45], numeric_formats={4: "dd.mm.yyyy", 5: "#,##0", 7: "#,##0", 8: "#,##0", 9: _MONEY_FORMAT, 10: _MONEY_FORMAT, 11: _MONEY_FORMAT, 12: _MONEY_FORMAT, 13: "#,##0", 17: "dd.mm.yyyy hh:mm", 18: "dd.mm.yyyy hh:mm", 19: "dd.mm.yyyy hh:mm", 20: "dd.mm.yyyy hh:mm"}, wrap_columns={1, 2, 21})
    registry_header = getattr(registry, "_osip_header_row", 5)
    for row_number, batch in enumerate(materialized, registry_header + 1):
        status = batch.status.value
        if status == "published":
            registry.cell(row_number, 6).fill = _GOOD_FILL
        elif status in {"failed", "rejected", "withdrawn"}:
            registry.cell(row_number, 6).fill = _BAD_FILL
        elif status in {"draft", "validating", "validated", "approved"}:
            registry.cell(row_number, 6).fill = _WARN_FILL
    audit = workbook.create_sheet("Аудит")
    _global_title(audit, "Аудит экспортируемых загрузок OSIP", {"Событий": str(len(audit_rows)), "Временная зона": "UTC"})
    _table(audit, ["Портфель", "Отчётная дата", "Версия", "Исходный файл", "Дата и время (UTC)", "Действие", "Исполнитель", "Параметры"], audit_rows, widths=[18, 18, 12, 55, 24, 30, 24, 64], numeric_formats={2: "dd.mm.yyyy", 3: "#,##0", 5: "dd.mm.yyyy hh:mm"}, wrap_columns={4, 6, 8})
    return _save(workbook)
