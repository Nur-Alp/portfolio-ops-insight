"""Generic Excel-writing helpers shared by every OSIP export type."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
import io
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from osip_dashboard.persistence.models import PortfolioSnapshotRecord
from osip_dashboard.services.date_provenance import filename_date_warning
from osip_dashboard.services.excel_safety import neutralize_formulas


ExportBasis = Literal["derived_carrying", "purchase"]


_ASSET_CLASS_LABELS = {
    "Corporate bond": "Корпоративные облигации",
    "Government bond": "Государственные облигации",
    "Government bonds": "Государственные облигации",
    "Equity": "Акции",
    "ETF": "ETF",
    "Repo": "РЕПО",
    "Development Institutions": "Институты развития",
    "Not supplied": "Не указано",
    "Commodity": "Товары / сырьё",
    "Unrated": "Рейтинг не указан",
}


_BASIS_LABELS = {
    "derived_carrying": "Расчётная балансовая стоимость",
    "purchase": "Сумма покупки",
}


_TITLE_FONT = Font(color="000000", bold=True, size=14)
_HEADER_FONT = Font(color="000000", bold=True)
_THIN_LINE = Side(style="thin", color="000000")
_TABLE_BORDER = Border(left=_THIN_LINE, right=_THIN_LINE, top=_THIN_LINE, bottom=_THIN_LINE)
_NUMBER_FORMAT = "#,##0.0000;[Red](#,##0.0000);-"
_MONEY_FORMAT = "#,##0.00;[Red](#,##0.00);-"
# OSIP stores percentage fields as percentage points (for example, 16.70
# means 16.70%). Quoting the percent sign prevents Excel multiplying it by 100.
_POINTS_PERCENT_FORMAT = '0.00"%";[Red](0.00"%");-'
_GOOD_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_BAD_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


def _workbook(title: str) -> Workbook:
    workbook = Workbook()
    workbook.properties.title = title
    workbook.properties.subject = "Операционные / расчётные данные; не официальный NAV"
    workbook.properties.creator = "OSIP Insight"
    workbook.properties.created = datetime.now()
    return workbook


def _snapshot_title(worksheet, snapshot: PortfolioSnapshotRecord, title: str, metadata: dict[str, str]) -> None:
    metadata = {"Портфель": snapshot.portfolio_code, "Отчётная дата": snapshot.report_date, "Версия": snapshot.version, **metadata}
    batch = getattr(snapshot, "import_batch", None)
    warning = filename_date_warning(
        getattr(batch, "original_filename", None),
        getattr(batch, "source_report_date", None) or snapshot.report_date,
    )
    if warning:
        metadata["Предупреждение"] = warning
    _title(worksheet, title, metadata)


def _global_title(worksheet, title: str, metadata: dict[str, str]) -> None:
    _title(worksheet, title, metadata)


def _title(worksheet, title: str, metadata: dict[str, Any]) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:C1")
    cell = worksheet["A1"]
    cell.value = title
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 26
    for index, (label, value) in enumerate(metadata.items()):
        row = 2 + index // 3
        column = 1 + (index % 3) * 3
        label_cell = worksheet.cell(row, column, label)
        label_cell.font = Font(bold=True, color="000000")
        value_cell = worksheet.cell(row, column + 1, value)
        if isinstance(value, (date, datetime)):
            value_cell.number_format = "dd.mm.yyyy"
    disclosure_row = 2 + (len(metadata) + 2) // 3
    worksheet.merge_cells(start_row=disclosure_row, start_column=1, end_row=disclosure_row, end_column=3)
    disclosure = worksheet.cell(disclosure_row, 1, "Операционные / расчётные данные из опубликованной версии OSIP; не являются официальным NAV или доходностью.")
    disclosure.font = Font(italic=True, color="000000")
    disclosure.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.row_dimensions[disclosure_row].height = 30
    worksheet._osip_header_row = disclosure_row + 2  # type: ignore[attr-defined]


def _table(worksheet, headers: list[str], rows: list[list[Any]], *, widths: list[int], numeric_formats: dict[int, str] | None = None, wrap_columns: set[int] | None = None) -> None:
    header_row = getattr(worksheet, "_osip_header_row", 6)
    numeric_formats = numeric_formats or {}
    wrap_columns = wrap_columns or set()
    for index, value in enumerate(headers, 1):
        cell = worksheet.cell(header_row, index, value)
        cell.font = _HEADER_FONT
        cell.border = _TABLE_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 34
    for row_number, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = worksheet.cell(row_number, column, value)
            cell.border = _TABLE_BORDER
            cell.alignment = Alignment(horizontal="right" if column in numeric_formats else "left", vertical="top", wrap_text=column in wrap_columns)
            if column in numeric_formats and value != "Недоступно":
                cell.number_format = numeric_formats[column]
    worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(header_row, header_row + len(rows))}"
    # Freeze the title/metadata rows through the table header and pin column A.
    # For the usual six-row title block this is B7; the dynamic row keeps the
    # same rule correct for exports with a different metadata layout.
    worksheet.freeze_panes = f"B{header_row + 1}"
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _total_row(worksheet, headers: list[str], *, row_number: int, label: str, value_column: int, value: Decimal, numeric_formats: dict[int, str]) -> None:
    """Append a bold total below a table without expanding its filter range."""
    for column in range(1, len(headers) + 1):
        cell = worksheet.cell(row_number, column)
        cell.border = _TABLE_BORDER
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="right" if column in numeric_formats else "left", vertical="center")
    worksheet.cell(row_number, 1, label)
    value_cell = worksheet.cell(row_number, value_column, value)
    value_cell.number_format = numeric_formats[value_column]
    worksheet.row_dimensions[row_number].height = 22


def _save(workbook: Workbook) -> bytes:
    neutralize_formulas(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _source_columns(source: dict[str, Any]) -> list[Any]:
    return [source.get("workbook_name") or "Недоступно", source.get("sheet_name") or "Недоступно", source.get("row_number") if source.get("row_number") is not None else "Недоступно"]


def _sources_columns(sources: list[dict[str, Any]]) -> list[str]:
    return ["; ".join(dict.fromkeys(str(source.get("workbook_name") or "Недоступно") for source in sources)) or "Недоступно", "; ".join(dict.fromkeys(str(source.get("sheet_name") or "Недоступно") for source in sources)) or "Недоступно", "; ".join(str(source.get("row_number")) for source in sources if source.get("row_number") is not None) or "Недоступно"]


def _asset_class_label(value: str | None) -> str:
    return _ASSET_CLASS_LABELS.get(value or "Not supplied", value or "Не указано")


def _decimal_value(value: str | Decimal) -> Decimal:
    return Decimal(value)


def _optional_decimal(value: str | Decimal | None) -> Decimal | str:
    return Decimal(value) if value is not None else "Недоступно"


def _date_value(value: str | date | datetime | None) -> date | datetime | str:
    if value is None:
        return "Недоступно"
    if isinstance(value, (date, datetime)):
        return value
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return "Недоступно"


def _amount_basis(value: str | None) -> str:
    if value == "source_purchase_amount":
        return "Источник: сумма покупки"
    if value == "derived_maturity_amount":
        return "Расчёт: номинал × количество"
    if value == "derived_expected_coupon":
        return "Расчёт: ожидаемый купон"
    if value == "unavailable" or not value:
        return "Недоступно"
    if value.startswith("source;"):
        return "Источник: семантика валюты расчёта не подтверждена"
    return value
