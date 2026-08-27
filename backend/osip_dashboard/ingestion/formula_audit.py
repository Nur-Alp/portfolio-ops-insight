"""Workbook formula evidence without pretending to be an Excel engine.

The application publishes cached source values.  This module records whether
formula cells have a cached result, contain an error, or reference an external
workbook.  It deliberately does not evaluate arbitrary Excel formulas.
"""

from __future__ import annotations

from collections import OrderedDict, defaultdict
from contextlib import closing, contextmanager
from pathlib import Path
import struct
from typing import Any, Iterator

from olefile import DEFECT_FATAL, OleFileIO
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.workbook.workbook import Workbook


OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
XLSX_SIGNATURE = b"PK\x03\x04"
BIFF_BOUNDSHEET = 0x0085
BIFF_EOF = 0x000A
BIFF_FORMULA = 0x0006
_ERROR_TOKENS = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!")


_AUDIT_CACHE_MAXSIZE = 256
_audit_cache: "OrderedDict[str, dict[str, Any]]" = OrderedDict()


@contextmanager
def open_xlsx_pair(path: str | Path) -> Iterator[tuple[Workbook, Workbook]]:
    """Open the (formulas, cached-values) workbook pair once, for reuse.

    ``parse_detected_dataset`` needs both a formula-text view and a
    cached-value view of the same ``.xlsx`` for its own contract parser
    (when that parser does the same formula-error/external-link check
    ``_parse_unit_history`` and ``_parse_tabys_valuation`` do) and for the
    audit steps below - each of those was independently calling
    ``load_workbook`` on the exact same file, and that call alone measured
    ~5s on a real ~1900-row workbook (dominated by openpyxl's own
    stylesheet parsing, not row count - see ``audit_workbook``'s docstring).
    Opening the pair once here and threading it through every consumer that
    can accept a pre-opened workbook turns what used to be up to six
    ``load_workbook`` calls per parse into two. ``keep_links=True`` on both
    (rather than only the formula view) is safe: verified it returns
    byte-identical cell values to a plain ``data_only=True`` open on a real
    workbook - it only affects external-reference resolution, never a plain
    cell's cached value.
    """
    with closing(load_workbook(path, read_only=True, data_only=False, keep_links=True)) as formulas, closing(
        load_workbook(path, read_only=True, data_only=True, keep_links=True)
    ) as cached:
        yield formulas, cached


def audit_workbook(path: str | Path, *, workbooks: tuple[Workbook, Workbook] | None = None) -> dict[str, Any]:
    """Return a stable, JSON-safe formula audit for an ``.xlsx`` or ``.xls``.

    ``.xlsx`` files expose both formula text and cached values through
    openpyxl.  Legacy ``.xls`` files are read by Calamine elsewhere in the
    ingestion pipeline; BIFF lets us count formula records, but not reliably
    expose their cached result or formula text, so that limitation is explicit
    in the returned status.

    Memoized by path: every ``parse_detected_dataset`` call runs this once,
    and a workbook offering several datasets (e.g. TABYS unit history's SAQ
    and TABYS keys) gets parsed once per selected key against the exact same
    file - each call was independently re-opening and re-scanning it. Safe to
    cache indefinitely: blob storage is content-addressed by SHA-256, so the
    same path is guaranteed to mean the same immutable bytes for as long as
    the blob exists - never a staleness risk the way caching a mutable file
    would be. The returned dict is only ever read by callers (spread into a
    new dict, never mutated in place), so sharing the same cached object
    across calls is safe.

    ``workbooks``, if given, is only used on a cache miss - lets a caller
    that already has the pair open (see ``open_xlsx_pair``) avoid yet
    another redundant open instead of this function silently opening its
    own.
    """
    key = str(Path(path))
    cached_value = _audit_cache.get(key)
    if cached_value is not None:
        _audit_cache.move_to_end(key)
        return cached_value
    source = Path(path)
    suffix = source.suffix.casefold()
    with source.open("rb") as stream:
        signature = stream.read(8)
    if suffix == ".xlsx" and signature[:4] == XLSX_SIGNATURE:
        result = _audit_xlsx(source, workbooks=workbooks)
    elif suffix == ".xls" and signature == OLE_SIGNATURE:
        result = _audit_xls(source)
    else:
        result = _empty_audit(suffix.lstrip(".") or "unknown", "unsupported_format")
    _audit_cache[key] = result
    if len(_audit_cache) > _AUDIT_CACHE_MAXSIZE:
        _audit_cache.popitem(last=False)
    return result


def audit_consumed_formula_results(
    path: str | Path, records: list[dict[str, Any]], *, workbooks: tuple[Workbook, Workbook] | None = None
) -> dict[str, Any]:
    """Audit only formula cells that back values the parser will publish.

    Workbook-wide formula errors are useful evidence, but many operational
    workbooks contain unused helper areas and template rows.  A publication
    gate must distinguish those from a missing/error result in a field that is
    actually persisted to the dashboard.  Legacy ``.xls`` cannot expose the
    formula/cached-value pair reliably, so it remains explicitly observable
    rather than being falsely blocked.

    ``workbooks``, if given, is an already-open (formulas, cached) pair (see
    ``open_xlsx_pair``) - reused instead of opening a third copy of the same
    file this call would otherwise need.
    """
    source = Path(path)
    if source.suffix.casefold() != ".xlsx":
        return {
            "status": "not_inspectable",
            "checked_formula_cells": 0,
            "invalid_cells": [],
            "message": "Проверка использованных результатов формул доступна только для .xlsx; для .xls сохраняются вычисленные значения источника.",
        }

    requested: dict[str, dict[str, list[str]]] = defaultdict(dict)
    for record in records:
        source_ref = record.get("source_ref") or {}
        sheet_name = source_ref.get("sheet_name")
        for field, cell_ref in (source_ref.get("field_columns") or {}).items():
            cell = cell_ref.get("source_cell") if isinstance(cell_ref, dict) else None
            if isinstance(sheet_name, str) and isinstance(cell, str):
                requested[sheet_name].setdefault(cell, []).append(str(field))
    if not requested:
        return {"status": "not_applicable", "checked_formula_cells": 0, "invalid_cells": []}

    if workbooks is not None:
        formulas, cached = workbooks
        return _scan_consumed_formula_results(formulas, cached, requested)
    with closing(load_workbook(source, read_only=True, data_only=False, keep_links=True)) as formulas, closing(
        load_workbook(source, read_only=True, data_only=True, keep_links=True)
    ) as cached:
        return _scan_consumed_formula_results(formulas, cached, requested)


def _scan_consumed_formula_results(
    formulas: Workbook, cached: Workbook, requested: dict[str, dict[str, list[str]]]
) -> dict[str, Any]:
    invalid: list[dict[str, Any]] = []
    checked = 0
    for sheet_name, cells in requested.items():
        if sheet_name not in formulas.sheetnames:
            continue
        # A read-only worksheet re-scans from row 1 on every random
        # ws[cell_ref] access (no random-access index), so looking up
        # each requested cell individually is O(rows) per lookup - O(rows
        # * lookups) overall, which took 9+ minutes on a real ~1900-row
        # workbook with ~1900 requested cells. One sequential pass over
        # both sheets, collecting only the requested coordinates, is
        # O(rows) total - matches the pattern _audit_xlsx already uses.
        wanted_coords = {coordinate_to_tuple(ref): ref for ref in cells}
        max_row = max(coord[0] for coord in wanted_coords)
        formula_cells: dict[str, Any] = {}
        for row_number, row in enumerate(formulas[sheet_name].iter_rows(max_row=max_row), 1):
            for col_number, cell in enumerate(row, 1):
                ref = wanted_coords.get((row_number, col_number))
                if ref is not None:
                    formula_cells[ref] = cell
        cached_cells: dict[str, Any] = {}
        if sheet_name in cached.sheetnames:
            for row_number, row in enumerate(cached[sheet_name].iter_rows(max_row=max_row), 1):
                for col_number, cell in enumerate(row, 1):
                    ref = wanted_coords.get((row_number, col_number))
                    if ref is not None:
                        cached_cells[ref] = cell
        for cell_ref, fields in cells.items():
            formula_cell = formula_cells.get(cell_ref)
            if formula_cell is None:
                continue
            is_formula = formula_cell.data_type == "f" or (
                isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
            )
            if not is_formula:
                continue
            checked += 1
            cached_cell = cached_cells.get(cell_ref)
            value = cached_cell.value if cached_cell is not None else None
            data_type = cached_cell.data_type if cached_cell is not None else None
            if value in (None, "") or data_type == "e" or (
                isinstance(value, str) and value in _ERROR_TOKENS
            ):
                invalid.append(
                    {
                        "sheet_name": sheet_name,
                        "source_cell": cell_ref,
                        "fields": sorted(set(fields)),
                        "reason": "blank_cached_result" if value in (None, "") else "cached_formula_error",
                    }
                )
    return {
        "status": "blocked" if invalid else "passed",
        "checked_formula_cells": checked,
        "invalid_cells": invalid,
    }


def _audit_xlsx(path: Path, *, workbooks: tuple[Workbook, Workbook] | None = None) -> dict[str, Any]:
    if workbooks is not None:
        formulas, cached = workbooks
        return _scan_xlsx(formulas, cached)
    with closing(load_workbook(path, read_only=True, data_only=False, keep_links=True)) as formulas, closing(
        load_workbook(path, read_only=True, data_only=True, keep_links=True)
    ) as cached:
        return _scan_xlsx(formulas, cached)


def _scan_xlsx(formulas: Workbook, cached: Workbook) -> dict[str, Any]:
    by_sheet: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "formula_count": 0,
            "blank_cached_formula_count": 0,
            "formula_error_count": 0,
            "error_value_count": 0,
            "external_formula_count": 0,
        }
    )
    formula_count = blank_cached = formula_errors = error_values = external = 0
    for sheet_name in formulas.sheetnames:
        formula_sheet = formulas[sheet_name]
        cached_sheet = cached[sheet_name]
        formula_rows = formula_sheet.iter_rows()
        cached_rows = cached_sheet.iter_rows()
        sheet_stats = by_sheet[sheet_name]
        for formula_row, cached_row in zip(formula_rows, cached_rows):
            for formula_cell, cached_cell in zip(formula_row, cached_row):
                formula = formula_cell.value
                is_formula = formula_cell.data_type == "f" or (
                    isinstance(formula, str) and formula.startswith("=")
                )
                cached_text = str(cached_cell.value) if cached_cell.value is not None else ""
                if not is_formula and (cached_cell.data_type == "e" or any(token == cached_text for token in _ERROR_TOKENS)):
                    error_values += 1
                    sheet_stats["error_value_count"] += 1
                if not is_formula:
                    continue
                formula_count += 1
                sheet_stats["formula_count"] += 1
                if cached_cell.value is None or cached_cell.value == "":
                    blank_cached += 1
                    sheet_stats["blank_cached_formula_count"] += 1
                formula_text = str(formula)
                if any(token in formula_text or token in cached_text for token in _ERROR_TOKENS):
                    formula_errors += 1
                    sheet_stats["formula_error_count"] += 1
                if "[" in formula_text and "]" in formula_text:
                    external += 1
                    sheet_stats["external_formula_count"] += 1

    status = "no_formulas" if formula_count == 0 else "ok"
    if formula_errors:
        status = "formula_errors"
    elif error_values:
        status = "source_errors"
    elif blank_cached:
        status = "blank_cached_results"
    return {
        "format": "xlsx",
        "formula_count": formula_count,
        "blank_cached_formula_count": blank_cached,
        "formula_error_count": formula_errors,
        "error_value_count": error_values,
        "external_formula_count": external,
        "formula_status": status,
        "cached_result_status": "inspected",
        "recalculation_status": "not_performed",
        "by_sheet": dict(by_sheet),
    }


def _audit_xls(path: Path) -> dict[str, Any]:
    by_sheet: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "formula_count": 0,
            "blank_cached_formula_count": None,
            "formula_error_count": None,
            "error_value_count": None,
            "external_formula_count": None,
        }
    )
    formula_count = 0
    try:
        with OleFileIO(path, raise_defects=DEFECT_FATAL) as container:
            stream_name = "Workbook" if container.exists("Workbook") else "Book"
            if not container.exists(stream_name):
                return _empty_audit("xls", "formula_stream_unavailable")
            stream = container.openstream(stream_name).read()
        offsets = _biff_sheet_offsets(stream)
        ordered = sorted(offsets.items(), key=lambda item: item[1])
        for index, (sheet_name, start) in enumerate(ordered):
            end = ordered[index + 1][1] if index + 1 < len(ordered) else len(stream)
            sheet_stats = by_sheet[sheet_name]
            position = start
            while position + 4 <= end:
                record_type, length = struct.unpack_from("<HH", stream, position)
                position += 4
                if position + length > end:
                    break
                payload = stream[position : position + length]
                position += length
                if record_type == BIFF_EOF:
                    break
                if record_type == BIFF_FORMULA and length >= 6:
                    formula_count += 1
                    sheet_stats["formula_count"] += 1
    except (OSError, ValueError, struct.error):
        return _empty_audit("xls", "formula_stream_unavailable")

    status = "no_formulas" if formula_count == 0 else "formula_records_detected"
    return {
        "format": "xls",
        "formula_count": formula_count,
        "blank_cached_formula_count": None,
        "formula_error_count": None,
        "error_value_count": None,
        "external_formula_count": None,
        "formula_status": status,
        "cached_result_status": "not_exposed_by_reader",
        "recalculation_status": "not_available",
        "by_sheet": dict(by_sheet),
    }


def _biff_sheet_offsets(stream: bytes) -> dict[str, int]:
    offsets: dict[str, int] = {}
    position = 0
    while position + 4 <= len(stream):
        record_type, length = struct.unpack_from("<HH", stream, position)
        position += 4
        if position + length > len(stream):
            break
        payload = stream[position : position + length]
        position += length
        if record_type != BIFF_BOUNDSHEET or length < 8:
            continue
        offset = struct.unpack_from("<I", payload)[0]
        character_count = payload[6]
        flags = payload[7]
        encoded_name = payload[8:]
        if flags & 0x01:
            name = encoded_name[: character_count * 2].decode("utf-16le", errors="replace")
        else:
            name = encoded_name[:character_count].decode("latin-1", errors="replace")
        offsets[name] = offset
    return offsets


def _empty_audit(file_format: str, status: str) -> dict[str, Any]:
    return {
        "format": file_format,
        "formula_count": 0,
        "blank_cached_formula_count": None,
        "formula_error_count": None,
        "error_value_count": None,
        "external_formula_count": None,
        "formula_status": status,
        "cached_result_status": "not_inspected",
        "recalculation_status": "not_performed",
        "by_sheet": {},
    }
