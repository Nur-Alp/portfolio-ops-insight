"""Shared dataclasses, constants, and low-level parsing helpers for multi-source ingestion."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from contextlib import closing, contextmanager
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
from typing import Any, Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


OLE_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
XLSX_SIGNATURE = b"PK\x03\x04"
PARSER_VERSION = "forte-multisource-v6"


class SourceDetectionError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedIssue:
    code: str
    severity: str
    message: str
    affected_fields: tuple[str, ...] = ()
    source_refs: tuple[dict[str, Any], ...] = ()


@dataclass
class ParsedDataset:
    dataset_type: str
    detected_key: str
    scope_type: str
    scope_code: str
    source_report_date: date | None
    business_date: date | None
    summary: dict[str, Any] = field(default_factory=dict)
    records: list[dict[str, Any]] = field(default_factory=list)
    issues: list[ParsedIssue] = field(default_factory=list)


def validate_source_filename(filename: str, content: bytes, max_bytes: int) -> tuple[str, str]:
    safe = Path(filename.replace("\\", "/")).name
    lowered = safe.casefold()
    if not safe or len(safe) > 500:
        raise SourceDetectionError("Имя файла отсутствует или слишком длинное")
    if lowered.startswith("~$") or lowered == "desktop.ini":
        raise SourceDetectionError("Временные и системные файлы не загружаются")
    if not content:
        raise SourceDetectionError("Файл пуст")
    if len(content) > max_bytes:
        raise SourceDetectionError(f"Размер файла превышает лимит {max_bytes} байт")
    suffix = Path(safe).suffix.lower()
    if suffix == ".xls" and content.startswith(OLE_SIGNATURE):
        return safe, "xls"
    if suffix == ".xlsx" and content.startswith(XLSX_SIGNATURE):
        return safe, "xlsx"
    raise SourceDetectionError("Поддерживаются только корректные рабочие книги .xls и .xlsx")


def _proposal(key: str, dataset_type: str, scope_type: str, scope_code: str) -> dict[str, str]:
    return {"key": key, "dataset_type": dataset_type, "scope_type": scope_type, "scope_code": scope_code}


def _sheet_names(path: Path, file_format: str) -> list[str]:
    if file_format == "xlsx":
        with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
            return list(workbook.sheetnames)
    from osip_dashboard.ingestion import multi_source as _multi_source

    workbook = _multi_source.CalamineWorkbook.from_path(path)
    return list(workbook.sheet_names)


def _xlsx_contains(path: Path, sheet: str, needle: str) -> bool:
    with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
        for row in workbook[sheet].iter_rows(min_row=1, max_row=50, values_only=True):
            if any(needle.casefold() in str(value).casefold() for value in row if value is not None):
                return True
    return False


def _contains_any(path: Path, file_format: str, sheet: str, needles: Iterable[str]) -> bool:
    wanted = tuple(value.casefold() for value in needles)
    if file_format == "xlsx":
        with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
            rows = list(workbook[sheet].iter_rows(min_row=1, max_row=12, values_only=True))
    else:
        from osip_dashboard.ingestion import multi_source as _multi_source

        rows = _multi_source.CalamineWorkbook.from_path(path).get_sheet_by_name(sheet).to_python()[:12]
    text = " ".join(str(value).casefold() for row in rows for value in row if value is not None)
    return all(value in text for value in wanted)


@contextmanager
def _workbook_or(existing: Workbook | None, path: Path, *, data_only: bool, keep_links: bool = False) -> Iterator[Workbook]:
    """Use an already-open workbook if given, otherwise open (and close) one.

    Lets a parser accept an optional pre-opened workbook from a caller that
    already paid the cost of opening it (see ``multi_source/__init__.py``'s
    ``parse_detected_dataset``, and ``open_xlsx_pair`` in
    ``ingestion/formula_audit.py`` for why that cost is worth avoiding a
    second/third time) while still working standalone - callers/tests that
    invoke a parser function directly with just a path see no behavior
    change; ``data_only``/``keep_links`` are only used for that standalone
    open, since a passed-in workbook was already opened with its own settings.
    """
    if existing is not None:
        yield existing
        return
    with closing(load_workbook(path, read_only=True, data_only=data_only, keep_links=keep_links)) as workbook:
        yield workbook


def _record(
    record_type: str, key: str, payload: dict[str, Any], sheet: str, row: int, *,
    column: int | None = None, field_columns: dict[str, int | None] | None = None,
    field_cells: dict[str, tuple[int, int] | None] | None = None,
) -> dict[str, Any]:
    # Parsed/cached values are intentionally separated from immutable raw evidence
    # and formulas so later parser revisions never overwrite source provenance.
    source_ref: dict[str, Any] = {"sheet_name": sheet, "row_number": row}
    if column is not None:
        # column is zero-based, matching to_python()'s row tuples; Excel
        # addressing (both the column number shown to a reviewer and the
        # A1-style cell reference) is one-based.
        letter = _excel_column_letter(column + 1)
        source_ref.update(source_column=column + 1, source_column_letter=letter, source_cell=f"{letter}{row}")
    # A single column/cell pointer per record only ever points at its
    # classification label. Recording which cell backs each payload field too
    # lets a reviewer jump straight to (say) the actual_kzt cell that caused
    # a breach, instead of always landing on the label. Most fields share the
    # record's own row (field_columns, column varies); a few sources (e.g.
    # TABYS's transposed NAV-history sheet) store one field per row instead,
    # sharing a single column - field_cells covers that case explicitly.
    combined: dict[str, tuple[int, int]] = {}
    if field_columns:
        combined.update({field: (row, field_column) for field, field_column in field_columns.items() if field_column is not None})
    if field_cells:
        combined.update({field: cell for field, cell in field_cells.items() if cell is not None})
    if combined:
        source_ref["field_columns"] = {field: _cell_ref(field_column, field_row) for field, (field_row, field_column) in combined.items()}
    return {"record_type": record_type, "record_key": str(key), "payload": payload, "source_ref": source_ref, "raw_values": dict(payload), "formulas": {}, "cached_values": dict(payload)}


def _cell_ref(column: int, row: int) -> dict[str, Any]:
    letter = _excel_column_letter(column + 1)
    return {"source_column": column + 1, "source_column_letter": letter, "source_cell": f"{letter}{row}"}


def _excel_column_letter(column: int) -> str:
    letters = ""
    while column > 0:
        column, remainder = divmod(column - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _cell(row: tuple[Any, ...], column: int | None) -> Any:
    """Read a zero-based cell without allowing a changed layout to shift fields."""
    if column is None or column < 0 or column >= len(row):
        return None
    return row[column]


def _client_trade_columns(sheet: Any) -> tuple[dict[str, int], int, list[str]]:
    """Locate the client-trade contract by header text, not by blank spacer columns.

    The supplied workbook has intentionally spaced columns (for example, amount is
    in column 33 while currency is in column 36). The previous positional parser
    treated those spacers as fields and consequently exposed prices as execution
    statuses. Header matching keeps the adapter tolerant of inserted rows and
    harmless column spacing changes while returning a DQ finding when a required
    business field disappears.
    """
    aliases: dict[str, tuple[str, ...]] = {
        "trade_number": ("№ п/п",),
        "order_number": ("номер клиентского заказа",),
        "trade_date": ("дата заключения сделки",),
        "venue": ("место заключения сделки",),
        "client_name": ("наименование/ ф.и.о. клиента", "наименование/фио клиента"),
        "account": ("№ лицевого счета", "номер лицевого счета"),
        "counterparty": ("сведения о контрагенте",),
        "side": ("вид сделки",),
        "issuer": ("эмитент",),
        "security_type": ("вид цб",),
        "isin": ("нин / isin", "нин/isin"),
        "quantity": ("количество цб",),
        "amount": ("сумма сделки",),
        "currency": ("валюта сделки",),
        "clean_price": ("чистая цена одной цб",),
        "dirty_price": ("грязная цена одной цб",),
        "yield_percent": ("доходность",),
        "settlement_date": ("дата исполнения обязательств по сделке",),
        "execution_status": ("отметка об исполнении",),
        "failure_reason": ("причина неисполнения",),
    }
    found: dict[str, int] = {}
    header_row = 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        for column, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            for field, candidates in aliases.items():
                if field in found:
                    continue
                if any(candidate in normalized for candidate in candidates):
                    found[field] = column
                    header_row = max(header_row, row_number)
    required = (
        "trade_number", "trade_date", "venue", "client_name", "account", "side",
        "issuer", "security_type", "isin", "quantity", "amount", "currency",
        "execution_status",
    )
    missing = [field for field in required if field not in found]
    return found, header_row or 8, missing


def _normalize_header(value: Any) -> str:
    text = _text(value).casefold().replace("ё", "е")
    text = re.sub(r"[\n\r\t]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _header_columns(
    sheet: Any, aliases: dict[str, tuple[str, ...]], *, max_row: int = 40
) -> tuple[dict[str, int], int]:
    """Shared header-scan core for ``_client_ledger_columns``/``_corporate_finance_columns``.

    Scans the first ``max_row`` rows for label text (not a single fixed
    header row) so a multi-row header, where different fields' labels live
    on different physical rows (e.g. a currency sub-header under a spanning
    "cash" label), still resolves every field in one pass. First match per
    field wins, mirroring ``_client_trade_columns`` above.
    """
    found: dict[str, int] = {}
    header_row = 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        for column, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            for field, candidates in aliases.items():
                if field in found:
                    continue
                if any(candidate in normalized for candidate in candidates):
                    found[field] = column
                    header_row = max(header_row, row_number)
    return found, header_row


def _client_ledger_columns(sheet: Any) -> tuple[dict[str, int], list[str]]:
    """Locate the Лист4 client/position register contract by header text.

    Header labels here span two physical rows (the cash total's own label,
    "Итого по всем валютам в тенге", sits one row below the sheet's main
    header row, under a spanning "Деньги на счетах" label) - a plain
    single-row header check would miss it, so this scans a small window
    like ``_client_trade_columns`` does for Лист8.
    """
    aliases: dict[str, tuple[str, ...]] = {
        "account": ("л/счет",),
        "client_name": ("клиент",),
        "iin": ("иин",),
        "citizenship": ("гражданство",),
        "resident": ("резиденство",),
        "economic_sector": ("сектор экономики",),
        "document_type": ("тип документа",),
        "branch": ("филиал",),
        "category": ("категория",),
        "agent": ("агент",),
        "manager": ("менеджер",),
        "issuer": ("эмитент",),
        "security_type": ("вид и категория ценных бумаг",),
        "security_code": ("код цб",),
        "isin": ("нин",),
        "nominal": ("номинал",),
        "nominal_currency": ("валюта номинала",),
        "quantity": ("цб на основном разделе",),
        "market_price": ("рыночная цена цб",),
        "price_currency": ("валюта цены",),
        "market_value_kzt": ("рыночная стоимость цб в тенге",),
        "cash_kzt": ("итого по всем валютам в тенге",),
        "total_assets_kzt": ("итого стоимость активов в тенге",),
    }
    found, _header_row = _header_columns(sheet, aliases, max_row=10)
    required = ("account", "client_name", "isin", "security_code")
    missing = [field for field in required if field not in found]
    return found, missing


def _tabys_holdings_columns(sheet: Any) -> tuple[dict[str, int], list[str]]:
    """Locate TABYS's "часть 1 (портфель)" holdings columns by header text.

    The header spans several rows (a "Покупная стоимость" label at row 1
    spans two sub-columns whose own currency-vs-KZT labels sit at row 3),
    like Лист4's cash header - scanning a small window instead of one row
    resolves both.
    """
    aliases: dict[str, tuple[str, ...]] = {
        "instrument": ("наименование эмитента",),
        "isin": ("нин (isin)", "нин(isin)"),
        "quantity": ("количество (штук ценных бумаг)", "количество"),
        "purchase_date": ("дата постановки на учет",),
        "maturity_date": ("дата погашения",),
        "currency": ("валюта номинальной стоимости",),
        "coupon_rate": ("ставка купона",),
        "purchase_value_kzt": ("чистая стоимость при покупке (в тг.)",),
        "purchase_value_native": ("чистая стоимость при покупке (в валюте)",),
        "clean_price_native": ("чистая цена в валюте",),
        "purchase_value_payment_currency": ("в валюте платежа",),
    }
    found, _header_row = _header_columns(sheet, aliases, max_row=6)
    required = ("instrument", "isin", "quantity")
    missing = [field for field in required if field not in found]
    return found, missing


def _corporate_finance_columns(sheet: Any) -> tuple[dict[str, int], list[str]]:
    """Locate the corporate-finance deal register contract by header text."""
    aliases: dict[str, tuple[str, ...]] = {
        "issuer": ("эмитент",),
        "subject": ("предмет договора",),
        "placement_amount": ("объем размещения",),
        "demand": ("фактически удовлетворенный спрос", "удовлетворенный спрос"),
        "investors": ("инвесторы",),
        "commission_rate": ("ставка комиссионного вознаграждения", "ставка комиссии"),
        "fee_received_kzt": ("размер полученного вознаграждения",),
        "duration_raw": ("длительность проекта",),
    }
    found, _header_row = _header_columns(sheet, aliases, max_row=10)
    required = ("issuer", "subject", "placement_amount", "demand")
    missing = [field for field in required if field not in found]
    return found, missing


# Aliases shared with backend/osip_dashboard/ingestion/osip_workbook.py's
# _FIELD_LABELS: the accounting workbook's own "ОСИП_ПОРТФЕЛЬ" sheet is the
# same generator/template family as the standalone OSIP workbook that broke
# this session (five rating/classification columns inserted, shifting every
# later field), so it carries the same risk and the same known label text.
_ACCOUNTING_PORTFOLIO_ALIASES: dict[str, tuple[str, ...]] = {
    "security_code": ("код ценной бумаги",),
    "isin": ("нин",),
    "security_type": ("тип ценной бумаги",),
    "issuer": ("эмитент",),
    "coupon_rate": ("ставка купона/ репо",),
    "nominal": ("номинальная стоимость",),
    "currency": ("валюта инструмента", "валюта инструмета"),
    "quantity": ("количество (шт.)", "количество"),
    "purchase_price": ("цена покупки",),
    "purchase_value_kzt": ("объем покупки в тенге",),
    "carrying_value_kzt": ("балансовая стоимость, в тенге",),
    "market_value_kzt": ("рыночная стоимость в тенге на отчетную дату",),
    "reserve_kzt": ("сумма резерва, в тенге",),
    "accrued_income_kzt": ("накопленный купон в тенге",),
}


def _rows_header_columns(
    rows: list[list[Any]], aliases: dict[str, tuple[str, ...]], *, max_row: int = 8
) -> dict[str, int]:
    """Locate columns by header text in already-materialized rows.

    Same scanning approach as ``_header_columns`` (a small window, not one
    fixed row - a multi-row/spanning header still resolves), but for
    python_calamine's ``to_python()`` output (a plain list of lists), not
    an openpyxl sheet with ``iter_rows`` - the .xls reader used for several
    of these sheets has no such object to scan.

    Per cell, only the LONGEST matching candidate across all not-yet-found
    fields is assigned, not just the first one checked. Some real header
    sets genuinely contain one label that is a plain substring of another
    (e.g. risk.py's TABYS sheet has "Установленный лимит" right next to
    "Установленный лимит (%)") - taking the first match in field-iteration
    order let the shorter field's candidate also match the longer cell
    that actually belongs to the other field, before the shorter field's
    own true (later) column was ever reached. Longest-match-wins is
    strictly more specific and never changes the result for a cell only
    one candidate matches.
    """
    found: dict[str, int] = {}
    for row in rows[:max_row]:
        for column, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            best_field: str | None = None
            best_length = -1
            for field, candidates in aliases.items():
                if field in found:
                    continue
                for candidate in candidates:
                    if candidate in normalized and len(candidate) > best_length:
                        best_field, best_length = field, len(candidate)
            if best_field is not None:
                found[best_field] = column
    return found


def _accounting_portfolio_columns(rows: list[list[Any]]) -> tuple[dict[str, int], list[str]]:
    """Locate the accounting workbook's ОСИП_ПОРТФЕЛЬ detail columns by header text."""
    found = _rows_header_columns(rows, _ACCOUNTING_PORTFOLIO_ALIASES, max_row=8)
    required = ("security_code", "isin", "security_type", "issuer")
    missing = [field for field in required if field not in found]
    return found, missing


def _accounting_portfolio_report_date(rows: list[list[Any]]) -> date | None:
    """Find the report date above the header, wherever it landed.

    Mirrors ``osip_workbook._find_report_date``'s fix this session: the
    same generator that moved this date from column J to column O in the
    standalone OSIP workbook moves it in this sheet too, so a fixed column
    index is not safe here either.
    """
    for row in rows[:3]:
        for value in row:
            candidate = _as_date(value)
            if candidate is not None:
                return candidate
    return None


def _issue(code: str, severity: str, message: str, fields: tuple[str, ...], sheet: str, row: int) -> ParsedIssue:
    return ParsedIssue(code, severity, message, fields, ({"sheet_name": sheet, "row_number": row},))


def _metric(metrics: dict[str, Any], label: str) -> Any:
    for key, value in metrics.items():
        if label.casefold() in key.casefold(): return value
    return None


def _metric_cell(metric_rows: dict[str, int], label: str, column: int) -> tuple[int, int] | None:
    for key, row in metric_rows.items():
        if label.casefold() in key.casefold(): return (row, column - 1)
    return None


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal_text(value: Any) -> str | None:
    if value in (None, "", "-"): return None
    try:
        return format(Decimal(str(value).replace(" ", "").replace(",", ".")), "f")
    except (InvalidOperation, ValueError):
        return None


def _d(value: Any) -> Decimal:
    return Decimal(str(value or "0"))


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value
    if not isinstance(value, str): return None
    for pattern in (r"(\d{4})-(\d{2})-(\d{2})", r"(\d{2})\.(\d{2})\.(\d{4})"):
        match = re.search(pattern, value)
        if match:
            parts = [int(item) for item in match.groups()]
            try: return date(*parts) if pattern.startswith("(\\d{4}") else date(parts[2], parts[1], parts[0])
            except ValueError: return None
    return None


def _extract_date(value: str) -> date | None:
    return _as_date(value)


_REPORT_DATE_LABELS = (
    "отчетная дата",
    "дата отчета",
    "по состоянию на",
    "report date",
    "reporting date",
    "as of",
)


def _explicit_report_dates(rows: Iterable[Iterable[Any]]) -> set[date]:
    """Return dates explicitly attached to a report-date label.

    Workbooks contain many legitimate dates (maturities, coupon dates,
    trade dates, opening dates, etc.).  Treating the maximum date in a sheet
    range as the report date makes those values look authoritative.  Only a
    labelled metadata cell, plus its immediately following value cells, can
    establish a report date; all other dates remain evidence in the sheet
    record.
    """
    explicit: set[date] = set()
    for row in rows:
        values = list(row)
        for index, value in enumerate(values):
            label = _text(value).casefold().replace("ё", "е")
            if not any(marker in label for marker in _REPORT_DATE_LABELS):
                continue
            # A common layout is ``Отчётная дата: 20.07.2026``.  Also accept
            # the next few cells when the label and value are separated.
            for candidate_value in values[index : index + 4]:
                candidate = _extract_date(_text(candidate_value))
                if candidate:
                    explicit.add(candidate)
    return explicit


def _accounting_explicit_dates(rows: Iterable[Iterable[Any]]) -> set[date]:
    """Backward-compatible name for accounting parser contract tests."""
    return _explicit_report_dates(rows)


def _explicit_report_date(sheet: Any, first: int, last: int) -> date | None:
    """Same signature as the old blanket date scan, but label-gated.

    Sheets like brokerage's ``Лист8`` have no report-date label at all, so
    this deliberately returns ``None`` rather than falling back to any
    date-shaped cell in the range - an absent report date is safer than a
    fabricated one taken from unrelated trade/maturity data.
    """
    dates = _explicit_report_dates(sheet.iter_rows(min_row=first, max_row=last, values_only=True))
    return max(dates) if dates else None


def _date_text(value: Any) -> str | None:
    parsed = _as_date(value)
    return parsed.isoformat() if parsed else None


def _looks_isin(value: Any) -> bool:
    return bool(re.fullmatch(r"[A-Z]{2}[A-Z0-9]{10}", _text(value).upper()))


def _client_position_lookup(sheet: Any) -> dict[tuple[str | None, str | None], list[tuple[int, str, Any, str | None]]]:
    """Index Лист4's client/position hierarchy by (client_name, issuer).

    Mirrors the row classification the "clients" key already uses above
    (an account-number cell marks a new client header row; a
    ISIN-shaped value in column P marks a position row belonging to the
    most recently seen client) so this reads the same authoritative
    register the same way, just keyed for lookup instead of emitted as
    its own records. Each entry also carries the instrument description
    (e.g. "ORCL 4.2 09/27/29"), the one place a coupon rate is available
    to confirm a match - see _extract_coupon_rate.
    """
    columns, _missing = _client_ledger_columns(sheet)
    index: dict[tuple[str | None, str | None], list[tuple[int, str, Any, str | None]]] = defaultdict(list)
    current_client: str | None = None
    for row_number, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), 7):
        account = _text(_cell(row, columns.get("account")))
        account_marker = account.casefold().replace(" ", "") if account else ""
        if account and account_marker.startswith(("итого", "всегопосчету", "всего")):
            continue
        if account:
            current_client = _text(_cell(row, columns.get("client_name")))
            continue
        isin = _text(_cell(row, columns.get("isin")))
        if current_client and _looks_isin(isin):
            issuer = _text(_cell(row, columns.get("issuer")))
            index[(current_client, issuer)].append((row_number, isin, _cell(row, columns.get("market_value_kzt")), _text(_cell(row, columns.get("security_code")))))
    return index


def _extract_coupon_rate(description: str | None) -> Decimal | None:
    """Pull a coupon rate out of a Лист4 instrument description.

    These follow a "TICKER RATE MM/DD/YY[ extra]" shape (e.g.
    "ORCL 4.2 09/27/29", "BAC 1.734 07/22/27") often enough to use as a
    confirmation signal, but not always (e.g. "FRTBS.0230", plain ISIN
    placeholders) - returns None rather than guessing when it doesn't.
    """
    if not description:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)\s+\d{2}/\d{2}/\d{2}", description)
    if not match:
        return None
    try:
        return Decimal(match.group(1))
    except InvalidOperation:
        return None


_RATE_TOLERANCE = Decimal("0.05")


def _resolve_calendar_isin(
    candidates: list[tuple[int, str, Any, str | None]], calendar_coupon_percent: Decimal | None, calendar_value_kzt: Decimal
) -> str | None:
    """Pick the Лист4 position a calendar row actually refers to.

    A client holding only one line from an issuer is not, by itself,
    reliable enough to accept blindly - the same issuer can appear on the
    calendar for a position that isn't in Лист4 at all (see the Oracle
    Corporation case this was built for: one client has both a real
    4.2% 2029 bond in Лист4 and an unrelated calendar row at 2.65%/2026
    naming the same issuer). A coupon-rate match against the Lист4
    description is trusted when available; a same-issuer position whose
    rate is confirmed *not* to match is never picked, even if it is the
    only candidate. Only once no rate signal exists at all does this fall
    back to accepting a single unambiguous candidate, or the closest
    current value among several - a weaker, but not fabricated, signal.
    """
    if not candidates:
        return None
    rated = [(candidate, _extract_coupon_rate(candidate[3])) for candidate in candidates]
    if calendar_coupon_percent is not None:
        confirmed = [
            candidate
            for candidate, rate in rated
            if rate is not None
            and abs(rate - calendar_coupon_percent) <= _RATE_TOLERANCE
            # A rate of exactly zero is true of every zero-coupon
            # instrument at once (T-bills, discount notes, ...) - "0
            # matches 0" carries no discriminating information the way "4.2
            # matches 4.2" does (confirmed a real bug: it made every
            # zero-coupon candidate in a multi-line group win regardless of
            # which one the row's own value actually pointed at).
            and not (rate == 0 and calendar_coupon_percent == 0)
        ]
        if confirmed:
            return min(confirmed, key=lambda item: abs(_d(item[2]) - calendar_value_kzt))[1]
        contradicted = any(rate is not None and abs(rate - calendar_coupon_percent) > _RATE_TOLERANCE for _candidate, rate in rated)
        if contradicted:
            return None
    if len(candidates) == 1:
        return candidates[0][1]
    return min(candidates, key=lambda item: abs(_d(item[2]) - calendar_value_kzt))[1]


def _extract_isins(value: Any) -> list[str]:
    """Extract ISINs while tolerating labels such as ``ISINXS...``.

    Corporate-finance subjects are free text and often concatenate the
    literal label ``ISIN`` directly to the two-letter country prefix.  Strip
    that label before matching so the normalized value is the actual 12
    character ISIN; the original subject remains in the payload unchanged.
    """
    text = _text(value).upper()
    text = re.sub(r"ISIN(?=[A-Z]{2}[A-Z0-9]{10})", " ", text)
    return sorted(set(re.findall(r"(?<![A-Z0-9])[A-Z]{2}[A-Z0-9]{10}(?![A-Z0-9])", text)))


def _normalize_name(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value).upper()).strip()


def _period_end(value: str) -> date | None:
    match = re.search(r"(\d)H(\d{4})", value.upper())
    if match:
        half, year = int(match.group(1)), int(match.group(2))
        return date(year, 6 if half == 1 else 12, 30 if half == 1 else 31)
    return _extract_date(value)


def _period_end_ddmmyy(value: str) -> date | None:
    match = re.search(r"(\d{2})\.(\d{2})\.(\d{2})", value)
    if not match:
        return None
    day, month, year = (int(part) for part in match.groups())
    try:
        return date(2000 + year, month, day)
    except ValueError:
        return None


def _parse_amount(value: Any) -> tuple[str | None, str | None, bool]:
    raw = _text(value).upper().replace("$", "").replace(" ", "").replace(",", ".")
    if not raw or raw == "-": return None, None, False
    currency = "USD" if "USD" in raw else "KZT" if "KZT" in raw else None
    match = re.search(r"([-+]?\d+(?:\.\d+)?)", raw)
    if not match or currency is None: return None, currency, False
    amount = Decimal(match.group(1))
    # "МЛРД" (Cyrillic) must be checked before the million branch below - it
    # starts with Cyrillic "М", the same character the million branch matches
    # on its own, so an unordered check would silently read a billion as a
    # million (1000x understatement) while still reporting the amount as
    # unambiguously resolved, suppressing the CORPFIN-01 DQ flag that would
    # otherwise catch it.
    if "BLN" in raw or "МЛРД" in raw or re.search(r"\dB(?:USD|KZT|$)", raw): amount *= Decimal("1000000000")
    elif "M" in raw or "М" in raw: amount *= Decimal("1000000")
    else: return format(amount, "f"), currency, False
    return format(amount, "f"), currency, True


def _russian_period_end(value: str) -> date | None:
    months = {"январ": 1, "феврал": 2, "март": 3, "апрел": 4, "ма": 5, "июн": 6, "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12}
    year_match = re.search(r"(20\d{2})", value)
    folded = value.casefold()
    month = next((number for token, number in months.items() if token in folded), None)
    if not year_match or month is None: return _extract_date(value)
    import calendar
    year = int(year_match.group(1))
    return date(year, month, calendar.monthrange(year, month)[1])
