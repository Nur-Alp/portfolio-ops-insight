"""Parser for the corporate finance workbook contract."""

from __future__ import annotations

from contextlib import closing
import hashlib
from pathlib import Path

from openpyxl import load_workbook

from .shared import (
    ParsedDataset,
    ParsedIssue,
    _cell,
    _corporate_finance_columns,
    _decimal_text,
    _extract_isins,
    _issue,
    _normalize_name,
    _parse_amount,
    _period_end,
    _record,
    _text,
)


def _parse_corporate_finance(path: Path, scope: str) -> ParsedDataset:
    with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
        sheet = workbook[workbook.sheetnames[0]]
        period = _text(sheet["B2"].value)
        business_date = _period_end(period)
        columns, missing_headers = _corporate_finance_columns(sheet)
        records = []
        issues: list[ParsedIssue] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=5, values_only=True), 5):
            issuer = _text(_cell(row, columns.get("issuer")))
            if not issuer: continue
            subject = _text(_cell(row, columns.get("subject"))); isins = _extract_isins(subject)
            placement_raw = _cell(row, columns.get("placement_amount"))
            demand_raw = _cell(row, columns.get("demand"))
            duration_raw = _cell(row, columns.get("duration_raw"))
            placement, placement_currency, placement_ok = _parse_amount(placement_raw)
            demand, demand_currency, demand_ok = _parse_amount(demand_raw)
            rate = _decimal_text(_cell(row, columns.get("commission_rate")))
            deal_key = hashlib.sha256(f"{_normalize_name(issuer)}|{_normalize_name(subject)}|{'|'.join(isins)}".encode()).hexdigest()[:24]
            payload = {
                "deal_key": deal_key, "issuer": issuer, "subject": subject, "isins": isins,
                "placement_amount": placement, "placement_currency": placement_currency, "placement_raw": _text(placement_raw),
                "satisfied_demand": demand, "demand_currency": demand_currency, "demand_raw": _text(demand_raw),
                "investors": _text(_cell(row, columns.get("investors"))), "commission_rate": rate,
                "fee_received_kzt": _decimal_text(_cell(row, columns.get("fee_received_kzt"))),
                "duration_raw": _text(duration_raw), "active": "действующ" in _text(duration_raw).casefold(), "period": period,
            }
            records.append(_record("deal", deal_key, payload, sheet.title, row_number, column=columns.get("issuer"), field_columns={
                "issuer": columns.get("issuer"), "subject": columns.get("subject"), "isins": columns.get("subject"),
                "placement_amount": columns.get("placement_amount"), "placement_currency": columns.get("placement_amount"), "placement_raw": columns.get("placement_amount"),
                "satisfied_demand": columns.get("demand"), "demand_currency": columns.get("demand"), "demand_raw": columns.get("demand"),
                "investors": columns.get("investors"), "commission_rate": columns.get("commission_rate"), "fee_received_kzt": columns.get("fee_received_kzt"),
                "duration_raw": columns.get("duration_raw"), "active": columns.get("duration_raw"),
            }))
            if not placement_ok or not demand_ok:
                issues.append(_issue("CORPFIN-01", "medium", "Сумма или единица измерения неоднозначна; исходный текст сохранён", ("placement_amount", "satisfied_demand"), sheet.title, row_number))
            if not isins:
                issues.append(_issue("CORPFIN-02", "medium", "Для сделки не найден ISIN", ("isins",), sheet.title, row_number))
            if not rate:
                issues.append(_issue("CORPFIN-03", "medium", "Ставка комиссии отсутствует", ("commission_rate",), sheet.title, row_number))
        if missing_headers:
            issues.append(_issue(
                "CORPFIN-MAP-01",
                "high",
                "Не удалось однозначно сопоставить обязательные столбцы реестра сделок: " + ", ".join(missing_headers),
                tuple(missing_headers),
                sheet.title,
                3,
            ))
        result = ParsedDataset("corporate_finance_register", "deals", "business_domain", scope, business_date, business_date, records=records, issues=issues)
        result.summary = {"deal_count": len(records), "period": period, "active_count": sum(bool(item["payload"]["active"]) for item in records)}
        return result
