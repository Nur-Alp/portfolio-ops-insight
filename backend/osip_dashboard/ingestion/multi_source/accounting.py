"""Parsers for the accounting workbook contracts (balance sheet, income statement, budget, portfolio, landing)."""

from __future__ import annotations

from contextlib import closing
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from osip_dashboard.ingestion.osip_workbook import CASH_PATTERN, CASH_PREFIX

from .shared import (
    ParsedDataset,
    ParsedIssue,
    _accounting_portfolio_columns,
    _accounting_portfolio_report_date,
    _as_date,
    _cell,
    _d,
    _decimal_text,
    _explicit_report_dates,
    _extract_date,
    _issue,
    _record,
    _text,
    _workbook_or,
)


_BALANCE_SHEET_SECTIONS = {"активы", "обязательства", "собственный капитал"}


def _parse_accounting_balance_sheet(path: Path, scope: str) -> ParsedDataset:
    with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
        sheet = workbook["f1_uip"]
        report_date = _as_date(sheet["A3"].value)
        records: list[dict[str, Any]] = []
        issues: list[ParsedIssue] = []
        totals: dict[str, Decimal] = {}
        total_refs: dict[str, tuple[str, int]] = {}
        section = ""
        # Starts at row 11, not 12: row 11 is the "Активы" section header
        # itself (confirmed against the real workbook) - starting one row
        # later skipped it, so every asset-section line ended up with
        # section="" while "Обязательства" (row 68+) and "Собственный
        # капитал" (row 103+) were correctly captured, since those header
        # rows do fall inside the old scan window.
        for row_number, row in enumerate(sheet.iter_rows(min_row=11, values_only=True), 11):
            label = _text(_cell(row, 0))
            code = _text(_cell(row, 1))
            if not label:
                continue
            if label.casefold().startswith("примечан"):
                break
            if not code:
                if label.casefold() in _BALANCE_SHEET_SECTIONS:
                    section = label
                continue
            current = _decimal_text(_cell(row, 2))
            prior = _decimal_text(_cell(row, 3))
            payload = {
                "line_code": code, "line_label": label, "section": section,
                "current_period_kzt": current, "prior_period_kzt": prior,
            }
            records.append(_record("balance_sheet_line", code, payload, sheet.title, row_number, column=0, field_columns={
                "line_code": 1, "current_period_kzt": 2, "prior_period_kzt": 3,
            }))
            normalized = label.casefold()
            if normalized in ("итого активы", "итого обязательства", "итого капитал"):
                key = {"итого активы": "assets", "итого обязательства": "liabilities", "итого капитал": "equity"}[normalized]
                if current is not None:
                    totals[key] = Decimal(current)
                    total_refs[key] = (sheet.title, row_number)
        if {"assets", "liabilities", "equity"}.issubset(totals):
            diff = totals["assets"] - (totals["liabilities"] + totals["equity"])
            if abs(diff) > Decimal("1"):
                ref_sheet, ref_row = total_refs["assets"]
                issues.append(_issue("ACCOUNTING-BS-01", "blocker", "Активы не равны сумме обязательств и капитала", ("current_period_kzt",), ref_sheet, ref_row))
        result = ParsedDataset("accounting_balance_sheet", "balance_sheet", "business_domain", scope, report_date, report_date, records=records, issues=issues)
        result.summary = {
            "line_count": len(records),
            "total_assets_kzt": _decimal_text(totals.get("assets")),
            "total_liabilities_kzt": _decimal_text(totals.get("liabilities")),
            "total_equity_kzt": _decimal_text(totals.get("equity")),
        }
        return result


def _parse_accounting_income_statement(path: Path, scope: str) -> ParsedDataset:
    with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
        sheet = workbook["f2_uip"]
        report_date = _as_date(sheet["A4"].value)
        records: list[dict[str, Any]] = []
        issues: list[ParsedIssue] = []
        totals: dict[str, dict[str, Decimal]] = {"quarter": {}, "ytd": {}}
        total_ref: tuple[str, int] | None = None
        for row_number, row in enumerate(sheet.iter_rows(min_row=12, values_only=True), 12):
            label = _text(_cell(row, 0))
            code = _text(_cell(row, 1))
            if not label:
                continue
            if not code:
                continue
            quarter = _decimal_text(_cell(row, 2))
            ytd = _decimal_text(_cell(row, 3))
            prior_quarter = _decimal_text(_cell(row, 4))
            prior_ytd = _decimal_text(_cell(row, 5))
            payload = {
                "line_code": code, "line_label": label,
                "quarter_kzt": quarter, "ytd_kzt": ytd,
                "prior_quarter_kzt": prior_quarter, "prior_ytd_kzt": prior_ytd,
            }
            records.append(_record("income_statement_line", code, payload, sheet.title, row_number, column=0, field_columns={
                "line_code": 1, "quarter_kzt": 2, "ytd_kzt": 3, "prior_quarter_kzt": 4, "prior_ytd_kzt": 5,
            }))
            normalized = label.casefold()
            if normalized == "итого доходов":
                if quarter is not None: totals["quarter"]["income"] = Decimal(quarter)
                if ytd is not None: totals["ytd"]["income"] = Decimal(ytd)
            elif normalized == "итого расходов":
                if quarter is not None: totals["quarter"]["expenses"] = Decimal(quarter)
                if ytd is not None: totals["ytd"]["expenses"] = Decimal(ytd)
            elif normalized == "чистая прибыль (убыток) до уплаты корпоративного подоходного налога":
                total_ref = (sheet.title, row_number)
                if quarter is not None: totals["quarter"]["net_profit"] = Decimal(quarter)
                if ytd is not None: totals["ytd"]["net_profit"] = Decimal(ytd)
        for period in ("quarter", "ytd"):
            values = totals[period]
            if {"income", "expenses", "net_profit"}.issubset(values):
                diff = values["income"] - values["expenses"] - values["net_profit"]
                if abs(diff) > Decimal("1"):
                    ref_sheet, ref_row = total_ref or (sheet.title, 12)
                    field = "quarter_kzt" if period == "quarter" else "ytd_kzt"
                    issues.append(_issue("ACCOUNTING-IS-01", "blocker", "Доходы за вычетом расходов не равны чистой прибыли до налога", (field,), ref_sheet, ref_row))
        result = ParsedDataset("accounting_income_statement", "income_statement", "business_domain", scope, report_date, report_date, records=records, issues=issues)
        result.summary = {
            "line_count": len(records),
            "total_income_kzt": _decimal_text(totals["quarter"].get("income")),
            "total_expenses_kzt": _decimal_text(totals["quarter"].get("expenses")),
            "net_profit_kzt": _decimal_text(totals["quarter"].get("net_profit")),
        }
        return result


_BUDGET_PERIOD_COLUMNS = {
    "year_2023_kzt": 1, "year_2024_kzt": 2, "budget_9m_2025_kzt": 3,
    "actual_9m_2025_kzt": 5, "budget_2025_kzt": 7,
    "oct_2025_kzt": 8, "nov_2025_kzt": 9, "dec_2025_kzt": 10,
    "forecast_2025_kzt": 11, "execution_pct": 12, "deviation_kzt": 13,
    # A separate, full Jan-Dec monthly breakdown sits to the right of the
    # summary block above (columns O-Z) - independent of it, not a repeat:
    # a spot-check against the real file found its own Oct/Nov/Dec figures
    # disagree with the oct/nov/dec_2025_kzt columns above, which look like
    # they may be reading stale cached formula results. Column month_01_kzt
    # is real, populated January data even for cash-flow lines, whose
    # summary columns (year_2023_kzt etc.) are genuinely blank in the
    # source - so this is the one reliable month-by-month series here.
    "month_01_kzt": 14, "month_02_kzt": 15, "month_03_kzt": 16, "month_04_kzt": 17,
    "month_05_kzt": 18, "month_06_kzt": 19, "month_07_kzt": 20, "month_08_kzt": 21,
    "month_09_kzt": 22, "month_10_kzt": 23, "month_11_kzt": 24, "month_12_kzt": 25,
}
_BUDGET_SECTION_MARKERS = {
    "отчет о прибылях и убытках": "income_statement",
    "оддс": "cash_flow",
    "баланс": "balance",
}
_BUDGET_RATIO_ONLY_LABELS = {"cost to income", "roa", "roe"}


def _parse_accounting_budget(path: Path, scope: str) -> ParsedDataset:
    """Real budget/actual figures from the Бюджет sheet's three sections.

    Column positions (2023 / 2024 / budget-9m-2025 / actual-9m-2025 /
    budget-2025 / Oct-Nov-Dec 2025 / forecast-2025 / % execution /
    deviation) are identical across all three sections and match the real
    file exactly (verified against sources/Бухгалтерия_Бюджет 2026.xlsx
    this session). Despite the filename there is no 2026 column anywhere
    in this sheet - the newest period covered is a 2025 forecast built
    from actual data through Q3 plus an Oct-Dec budget. The one explicit,
    literal date in the sheet ("Факт" as of 30.09.2025, next to the
    balance section headers) is used as the business date rather than
    inventing one.

    The cash-flow section (ОДДС) has every row label but its summary block
    (year_2023_kzt through deviation_kzt) is genuinely blank in the source
    file - month_01_kzt..month_12_kzt is where its real values live. Every
    section is scanned identically and rows with every period column blank
    are dropped rather than being special-cased, so any section starts
    reporting a given figure automatically the moment the source workbook
    has it - no code change needed on that day.
    """
    with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
        sheet = workbook["Бюджет"]
        business_date = _as_date(sheet.cell(66, 6).value)
        records: list[dict[str, Any]] = []
        current_section: str | None = None
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), 4):
            label = _text(_cell(row, 0))
            if not label:
                continue
            folded = label.casefold()
            matched_section = next((section for marker, section in _BUDGET_SECTION_MARKERS.items() if folded.startswith(marker)), None)
            if matched_section is not None:
                current_section = matched_section
                continue
            if current_section is None or folded in _BUDGET_RATIO_ONLY_LABELS:
                continue
            values = {field: _decimal_text(_cell(row, column)) for field, column in _BUDGET_PERIOD_COLUMNS.items()}
            if all(value is None for value in values.values()):
                continue
            payload = {"section": current_section, "line_label": label, **values}
            records.append(_record(
                "budget_line", f"{current_section}:{label}", payload, sheet.title, row_number, column=0,
                field_columns=dict(_BUDGET_PERIOD_COLUMNS),
            ))
        result = ParsedDataset("accounting_budget", "budget_detail", "business_domain", scope, business_date, business_date, records=records, issues=[])
        result.summary = {
            "line_count": len(records),
            "income_statement_line_count": sum(1 for record in records if record["payload"]["section"] == "income_statement"),
            "cash_flow_line_count": sum(1 for record in records if record["payload"]["section"] == "cash_flow"),
            "balance_line_count": sum(1 for record in records if record["payload"]["section"] == "balance"),
        }
        return result


def _parse_accounting_portfolio(path: Path, scope: str) -> ParsedDataset:
    """Position-level portfolio detail from the ОСИП_ПОРТФЕЛЬ sheet (.xls).

    Position rows are grouped under plain-text category headers (e.g.
    "Корпоративные облигации", "ETF", "ГЦБ") - each category can repeat
    (verified against sources/Бухгалтерия_Портфель.xls this session, e.g.
    "ГЦБ" appears at both row 14 and row 23 for genuinely different
    lots). After the last category's positions, the sheet switches to a
    "ПО ..." (all-caps, "по" as a case-insensitive prefix) cash-by-bank/
    currency breakdown block with an unrelated column layout ("ПО
    ПОРТФЕЛЮ" being its grand-total row) - detecting that all-caps prefix
    stops position parsing before that block is reached, rather than
    misreading its columns as position fields.

    The cash rows within that block (label prefix CASH_PREFIX, same
    generator/format as the standalone OSIP workbook's own cash rows - see
    osip_workbook.CASH_PATTERN, reused here rather than duplicated) are
    extracted as their own "cash_balance" records. Confirmed this session:
    portfolio_total's carrying_value_kzt legitimately equals the sum of the
    position rows' carrying_value_kzt PLUS these cash rows' - previously
    the cash rows were silently dropped, so the position table's own sum
    fell tens of millions of KZT short of the displayed total with no
    explanation anywhere on the page. The duration and raw-holdings sheets
    are still deferred to a later pass.

    Columns are resolved by header label (see _accounting_portfolio_columns):
    this sheet is the same generator/template as the standalone OSIP
    workbook whose column layout changed this session (five rating columns
    inserted, shifting everything after them), so a fixed column map would
    have broken here too.

    record_key includes the row number: the same ISIN legitimately repeats
    across separate lots (e.g. two SGOV US purchases at different prices),
    so ISIN alone is not a stable identity.
    """
    from osip_dashboard.ingestion import multi_source as _multi_source

    workbook = _multi_source.CalamineWorkbook.from_path(path)
    rows = workbook.get_sheet_by_name("ОСИП_ПОРТФЕЛЬ").to_python(skip_empty_area=False)
    business_date = _accounting_portfolio_report_date(rows)
    columns, missing_headers = _accounting_portfolio_columns(rows)
    records: list[dict[str, Any]] = []
    cash_records: list[dict[str, Any]] = []
    issues: list[ParsedIssue] = []
    portfolio_total: dict[str, Any] = {}
    current_category: str | None = None
    carrying_value_column = columns.get("carrying_value_kzt")
    for row_number, row in enumerate(rows, 1):
        if row_number <= 3:
            continue
        label = _text(_cell(row, 0))
        isin = _text(_cell(row, columns.get("isin")))
        if not label and not isin:
            continue
        if not isin:
            if label.strip().casefold() == "по портфелю":
                portfolio_total = {field: _decimal_text(_cell(row, column)) for field, column in columns.items() if field.endswith("_kzt")}
            elif label.startswith(CASH_PREFIX):
                match = CASH_PATTERN.search(label)
                currency, custodian = match.groups() if match else (None, None)
                cash_records.append(_record(
                    "cash_balance", f"{row_number}", {
                        "currency": currency,
                        "custodian": custodian.strip() if custodian and custodian.strip() else None,
                        "amount_kzt": _decimal_text(_cell(row, carrying_value_column)),
                        "raw_label": label,
                    }, "ОСИП_ПОРТФЕЛЬ", row_number, column=carrying_value_column,
                ))
            elif not label.casefold().startswith("по "):
                # The cash-by-bank/currency breakdown block that follows the
                # last real category uses this same "по ..." label shape but
                # an unrelated column layout - once one is seen, later plain
                # labels are its rows, not a new position category, so
                # `current_category` is deliberately left untouched here.
                current_category = label
            continue
        if current_category is None:
            continue
        values = {field: (_decimal_text(_cell(row, column)) if field not in {"security_code", "isin", "security_type", "issuer", "currency"} else _text(_cell(row, column))) for field, column in columns.items()}
        payload = {"category": current_category, **values}
        records.append(_record(
            "portfolio_position", f"{isin or values.get('security_code')}:{row_number}", payload, "ОСИП_ПОРТФЕЛЬ", row_number, column=columns.get("isin"),
            field_columns=dict(columns),
        ))
    position_count = len(records)
    cash_kzt = sum((_d(record["payload"]["amount_kzt"]) for record in cash_records), Decimal("0")) if cash_records else None
    records.extend(cash_records)
    if missing_headers:
        issues.append(_issue(
            "ACCOUNTING-PORT-01",
            "high",
            "Не удалось однозначно сопоставить обязательные столбцы портфеля (ОСИП_ПОРТФЕЛЬ): " + ", ".join(missing_headers),
            tuple(missing_headers),
            "ОСИП_ПОРТФЕЛЬ",
            3,
        ))
    result = ParsedDataset("accounting_portfolio_detail", "portfolio_detail", "business_domain", scope, business_date, business_date, records=records, issues=issues)
    result.summary = {
        "position_count": position_count,
        "cash_kzt": format(cash_kzt, "f") if cash_kzt is not None else None,
        "total_carrying_value_kzt": portfolio_total.get("carrying_value_kzt"),
        "total_market_value_kzt": portfolio_total.get("market_value_kzt"),
        "total_reserve_kzt": portfolio_total.get("reserve_kzt"),
    }
    return result


def _parse_accounting_landing(
    path: Path, detection: dict[str, Any], key: str, scope: str, *, formula_workbook: Workbook | None = None
) -> ParsedDataset:
    file_format = "xlsx" if path.suffix.lower() == ".xlsx" else "xls"
    sheets = detection["sheets"]
    records = []
    issues: list[ParsedIssue] = []
    report_date: date | None = None
    detected_dates: set[date] = set()
    explicit_dates: set[date] = set()
    if file_format == "xlsx":
        with _workbook_or(formula_workbook, path, data_only=False) as workbook:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                formula_errors = 0
                external_links = 0
                formula_refs: dict[str, str] = {}
                sheet_dates: set[date] = set()
                rows = list(sheet.iter_rows())
                metadata_rows = [[cell.value for cell in row] for row in rows[:24]]
                explicit_dates.update(_explicit_report_dates(metadata_rows))
                for row in rows:
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and ("#REF!" in value or "#N/A" in value): formula_errors += 1
                        if isinstance(value, str) and value.startswith("="):
                            formula_refs[cell.coordinate] = value
                            if "[" in value and "]" in value:
                                external_links += 1
                        candidate = _extract_date(_text(value))
                        if candidate: sheet_dates.add(candidate); detected_dates.add(candidate)
                record = _record("sheet_evidence", sheet_name, {"sheet": sheet_name, "rows": sheet.max_row, "columns": sheet.max_column, "formula_error_count": formula_errors, "external_link_count": external_links, "dates": sorted(value.isoformat() for value in sheet_dates)}, sheet_name, 1)
                record["formulas"] = formula_refs
                records.append(record)
                if formula_errors:
                    issues.append(_issue("ACCOUNTING-01", "high", f"Лист содержит ошибок формул/значений: {formula_errors}", ("formula",), sheet_name, 1))
                if external_links:
                    issues.append(_issue("ACCOUNTING-03", "medium", f"Лист содержит внешних ссылок: {external_links}", ("formula",), sheet_name, 1))
    else:
        from osip_dashboard.ingestion import multi_source as _multi_source

        workbook = _multi_source.CalamineWorkbook.from_path(path)
        for sheet_name in sheets:
            rows = workbook.get_sheet_by_name(sheet_name).to_python()
            width = max((len(row) for row in rows), default=0)
            explicit_dates.update(_explicit_report_dates(rows[:24]))
            sheet_dates: set[date] = set()
            for row in rows[:12]:
                for value in row:
                    candidate = _extract_date(_text(value))
                    if candidate:
                        sheet_dates.add(candidate)
                        detected_dates.add(candidate)
            records.append(_record("sheet_evidence", sheet_name, {"sheet": sheet_name, "rows": len(rows), "columns": width, "formula_error_count": None, "dates": sorted(value.isoformat() for value in sheet_dates)}, sheet_name, 1))
        # python_calamine (the .xls reader) exposes only computed cell
        # values, not formula source text, so ACCOUNTING-01/03's formula-error
        # and external-link scans (which inspect formula text on the .xlsx
        # branch above) cannot run here at all. Silently skipping them would
        # let a reviewer mistake "no findings" for "no formula errors" on a
        # legacy .xls source - make the gap an explicit, visible finding
        # instead of an absent one.
        issues.append(ParsedIssue("ACCOUNTING-04", "medium", "Проверка ошибок формул и внешних ссылок недоступна для источников формата .xls; она выполняется только для .xlsx", ("formula",), ()))
    # Accounting landing files do not have a stable date contract. A date-shaped
    # value in a data row (for example, a bond maturity date) is evidence only,
    # never the dataset date. Use a date only when it is explicitly labelled in
    # the workbook's metadata area; otherwise leave both dates unavailable.
    if len(explicit_dates) == 1:
        report_date = next(iter(explicit_dates))
    if len(explicit_dates) > 1:
        issues.append(ParsedIssue("ACCOUNTING-02", "medium", "В метаданных явно указаны конфликтующие отчётные даты", ("source_report_date",), ()))
    issues.append(ParsedIssue("ACCOUNTING-00", "high", "Источник сохранён только для предварительного контроля; финансовые показатели не публикуются", ("publication",), ()))
    result = ParsedDataset("accounting_landing", key, "business_domain", scope, report_date, report_date, records=records, issues=issues)
    result.summary = {"sheet_count": len(records), "publication_basis": "landing_only", "source_type": detection["source_type"]}
    return result
