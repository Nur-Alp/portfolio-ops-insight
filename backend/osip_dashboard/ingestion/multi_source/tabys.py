"""Parsers for the TABYS fund workbook contracts (valuation and unit-value history)."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl.workbook.workbook import Workbook

from .shared import (
    ParsedDataset,
    _as_date,
    _cell,
    _d,
    _date_text,
    _decimal_text,
    _issue,
    _looks_isin,
    _metric,
    _metric_cell,
    _record,
    _tabys_holdings_columns,
    _text,
    _workbook_or,
)


def _parse_tabys_valuation(
    path: Path, key: str, scope: str, *, data_workbook: Workbook | None = None, formula_workbook: Workbook | None = None
) -> ParsedDataset:
    with _workbook_or(data_workbook, path, data_only=True) as workbook:
        current = workbook["дата"]
        report_date = _as_date(current["B2"].value)
        metrics = {str(current.cell(row, 1).value).strip(): current.cell(row, 4).value for row in range(1, 42) if current.cell(row, 1).value}
        metric_rows = {str(current.cell(row, 1).value).strip(): row for row in range(1, 42) if current.cell(row, 1).value}
        common = ParsedDataset(
            dataset_type={
                "valuation": "fund_valuation", "holdings": "fund_holdings",
                "cash_liabilities": "fund_cash_liabilities", "nav_history": "fund_nav_history",
                "prices": "fund_prices", "inactive_evidence": "fund_inactive_evidence",
            }[key],
            detected_key=key, scope_type="fund", scope_code=scope,
            source_report_date=report_date, business_date=report_date,
        )
        if key in {"valuation", "cash_liabilities"}:
            values = {
                "securities_value_kzt": _decimal_text(_metric(metrics, "Текущая стоимость портфеля ЦБ")),
                "cash_kzt": _decimal_text((_metric(metrics, "Деньги на инвестиционном счете (в тенге)") or 0) + (_metric(metrics, "Деньги на инвестиционном счете (в USD)") or 0) + (_metric(metrics, "Деньги на инвестиционном счете (в EUR)") or 0)),
                "liabilities_kzt": _decimal_text(_metric(metrics, "Обязательства")),
                "nav_kzt": _decimal_text(_metric(metrics, 'Текущая "чистая" стоимость активов')),
                "units": _decimal_text(_metric(metrics, "Количество паев в обращении")),
                "unit_value_kzt": _decimal_text(_metric(metrics, "Расчетная стоимость пая")),
                "unit_value_usd": _decimal_text(_metric(metrics, "Расчетная стоимость в пая USD")),
            }
            value_field_cells = {
                "securities_value_kzt": _metric_cell(metric_rows, "Текущая стоимость портфеля ЦБ", 4),
                # cash_kzt sums three separate metric rows, so it has no single cell to point at.
                "liabilities_kzt": _metric_cell(metric_rows, "Обязательства", 4),
                "nav_kzt": _metric_cell(metric_rows, 'Текущая "чистая" стоимость активов', 4),
                "units": _metric_cell(metric_rows, "Количество паев в обращении", 4),
                "unit_value_kzt": _metric_cell(metric_rows, "Расчетная стоимость пая", 4),
                "unit_value_usd": _metric_cell(metric_rows, "Расчетная стоимость в пая USD", 4),
            }
            common.records.append(_record("valuation", scope, values, "дата", 19, field_cells=value_field_cells))
            common.summary = values
            nav_expected = _d(values["securities_value_kzt"]) + _d(values["cash_kzt"]) - _d(values["liabilities_kzt"])
            nav_diff = abs(nav_expected - _d(values["nav_kzt"]))
            if nav_diff > Decimal("1"):
                common.issues.append(_issue("TABYS-NAV-01", "blocker", "СЧА не равна стоимости портфеля плюс деньги минус обязательства", ("nav_kzt",), "дата", 38))
            if _d(values["units"]) and abs(_d(values["nav_kzt"]) / _d(values["units"]) - _d(values["unit_value_kzt"])) > Decimal("0.0001"):
                common.issues.append(_issue("TABYS-NAV-02", "high", "Стоимость пая не согласуется с СЧА и количеством паёв", ("unit_value_kzt",), "дата", 40))
            return common
        if key == "holdings":
            sheet = workbook["часть 1 (портфель)"]
            columns, missing_headers = _tabys_holdings_columns(sheet)
            for row_number, row in enumerate(sheet.iter_rows(min_row=6, values_only=True), 6):
                if not _looks_isin(_cell(row, columns.get("isin"))):
                    continue
                payload = {
                    "instrument": _text(_cell(row, columns.get("instrument"))), "isin": _text(_cell(row, columns.get("isin"))),
                    "quantity": _decimal_text(_cell(row, columns.get("quantity"))),
                    "purchase_date": _date_text(_cell(row, columns.get("purchase_date"))), "maturity_date": _date_text(_cell(row, columns.get("maturity_date"))),
                    "currency": _text(_cell(row, columns.get("currency"))), "coupon_rate": _decimal_text(_cell(row, columns.get("coupon_rate"))),
                    "purchase_value_kzt": _decimal_text(_cell(row, columns.get("purchase_value_kzt"))),
                    "purchase_value_native": _decimal_text(_cell(row, columns.get("purchase_value_native"))),
                    "clean_price_native": _decimal_text(_cell(row, columns.get("clean_price_native"))),
                    "purchase_value_payment_currency": _decimal_text(_cell(row, columns.get("purchase_value_payment_currency"))),
                }
                common.records.append(_record("holding", f"{payload['isin']}:{row_number}", payload, sheet.title, row_number, column=columns.get("isin"), field_columns=dict(columns)))
            common.summary = {"holding_count": len(common.records)}
            if missing_headers:
                common.issues.append(_issue(
                    "TABYS-HOLD-MAP-01",
                    "high",
                    "Не удалось однозначно сопоставить обязательные столбцы портфеля: " + ", ".join(missing_headers),
                    tuple(missing_headers),
                    sheet.title,
                    1,
                ))
            return common
        if key == "prices":
            sheet = workbook["Цены"]
            missing = 0
            latest: date | None = None
            for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
                if not row[0]:
                    continue
                price = row[2]
                price_date = _as_date(row[4])
                if price_date and (latest is None or price_date > latest): latest = price_date
                unavailable = isinstance(price, str) and "#N/A" in price
                missing += int(unavailable)
                common.records.append(_record("price", _text(row[0]), {
                    "ticker": _text(row[0]), "name": _text(row[1]), "price": None if unavailable else _decimal_text(price),
                    "currency": _text(row[3]), "price_date": _date_text(row[4]),
                }, sheet.title, row_number, column=0, field_columns={"ticker": 0, "name": 1, "price": 2, "currency": 3, "price_date": 4}))
            common.summary = {"price_count": len(common.records), "missing_price_count": missing, "latest_price_date": _date_text(latest)}
            if missing:
                common.issues.append(_issue("TABYS-PRICE-01", "high", f"В справочнике отсутствуют {missing} цен", ("price",), "Цены", 2))
            if latest and report_date and (report_date - latest).days > 1:
                common.issues.append(_issue("TABYS-PRICE-02", "medium", "Дата цен отстаёт от даты оценки более чем на один день", ("price_date",), "Цены", 2))
            return common
        if key == "nav_history":
            sheet = workbook["справка о ст-ти ЧА"]
            rows = list(sheet.iter_rows(values_only=True))
            dates = [(_as_date(value), column) for column, value in enumerate(rows[2]) if _as_date(value)]
            metric_rows = []
            for row_number, row in enumerate(rows, 1):
                label = _text(row[1] if len(row) > 1 else None)
                folded = label.casefold()
                if any(token in folded for token in ('"чист', "количество паев", "стоимость пая")):
                    metric_rows.append((row_number, label))
            for item_date, column in dates:
                payload = {"date": item_date.isoformat()}
                # Each metric lives on its own row but shares this date's
                # column, the mirror image of a normal row-shaped record -
                # field_cells lets the reviewer land on the metric's actual
                # (row, column) instead of the arbitrary anchor row below.
                field_cells: dict[str, tuple[int, int] | None] = {}
                for row_number, label in metric_rows:
                    row = rows[row_number - 1]
                    payload[label] = _decimal_text(row[column] if column < len(row) else None)
                    field_cells[label] = (row_number, column)
                common.records.append(_record("nav_history", item_date.isoformat(), payload, sheet.title, 3, field_cells=field_cells))
            common.summary = {"observation_count": len(common.records), "metric_count": len(metric_rows)}
            return common
        # Evidence-only partitions are deliberately not interpreted as active assets.
        evidence_sheets = ["часть 2 (депозиты)", "об РЕПО", "часть 4 (иное имущество)", "ОПиУ", "ББ (2)"]
        with _workbook_or(formula_workbook, path, data_only=False, keep_links=True) as formula_book:
            for sheet_name in evidence_sheets:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    formulas: dict[str, str] = {}
                    broken = 0
                    external = 0
                    for row in formula_book[sheet_name].iter_rows():
                        for cell in row:
                            value = cell.value
                            if isinstance(value, str) and value.startswith("="):
                                formulas[cell.coordinate] = value
                                broken += int("#REF!" in value or "#N/A" in value)
                                external += int("[" in value and "]" in value)
                    record = _record("sheet_evidence", sheet_name, {"sheet": sheet_name, "rows": sheet.max_row, "columns": sheet.max_column, "active": False, "formula_error_count": broken, "external_formula_count": external}, sheet_name, 1)
                    record["formulas"] = formulas
                    common.records.append(record)
                    if broken or external:
                        common.issues.append(_issue("TABYS-EVIDENCE-02", "high", f"Раздел содержит ошибок формул/внешних ссылок: {broken + external}", ("formula",), sheet_name, 1))
        common.summary = {"sheet_count": len(common.records), "publication_basis": "evidence_only"}
        common.issues.append(_issue("TABYS-EVIDENCE-01", "medium", "Исторические/шаблонные разделы сохранены как доказательство и исключены из текущей оценки", ("active",), evidence_sheets[0], 1))
        return common


def _parse_unit_history(
    path: Path, key: str, scope: str, *, data_workbook: Workbook | None = None, formula_workbook: Workbook | None = None
) -> ParsedDataset:
    with _workbook_or(data_workbook, path, data_only=True) as workbook:
        sheet = workbook["Форма"]
        columns = (2, 3, 4, 5, 6, 7, 8) if key == "SAQ" else (10, 13, 14, 15, 16, 17, 18)
        records: list[dict[str, Any]] = []
        seen: set[date] = set()
        duplicates = 0
        previous: date | None = None
        non_monotonic = 0
        missing_values = 0
        discontinuities = 0
        previous_unit_value: Decimal | None = None
        for row_number, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), 4):
            item_date = _as_date(row[columns[0] - 1] if len(row) >= columns[0] else None)
            if item_date is None:
                continue
            duplicates += int(item_date in seen)
            non_monotonic += int(previous is not None and item_date < previous)
            seen.add(item_date); previous = item_date
            values = [row[column - 1] if len(row) >= column else None for column in columns[1:]]
            payload = {
                "date": item_date.isoformat(), "units": _decimal_text(values[0]), "investors": _decimal_text(values[1]),
                "nav_usd": _decimal_text(values[2]), "unit_value_usd": _decimal_text(values[3]),
                "nav_kzt": _decimal_text(values[4]), "unit_value_kzt": _decimal_text(values[5]),
            }
            missing_values += int(payload["nav_kzt"] is None or payload["unit_value_kzt"] is None)
            current_unit_value = _d(payload["unit_value_kzt"]) if payload["unit_value_kzt"] is not None else None
            if previous_unit_value and current_unit_value and abs(current_unit_value / previous_unit_value - Decimal("1")) > Decimal("0.25"):
                discontinuities += 1
            if current_unit_value: previous_unit_value = current_unit_value
            records.append(_record("unit_observation", f"{item_date.isoformat()}:{row_number}", payload, sheet.title, row_number, column=columns[0] - 1, field_columns={
                "units": columns[1] - 1, "investors": columns[2] - 1, "nav_usd": columns[3] - 1,
                "unit_value_usd": columns[4] - 1, "nav_kzt": columns[5] - 1, "unit_value_kzt": columns[6] - 1,
            }))
        latest = max(seen) if seen else None
        result = ParsedDataset("fund_unit_series", key, "fund", scope, latest, latest, records=records)
        result.summary = {"observation_count": len(records), "latest_date": _date_text(latest), "active": key != "SAQ", "missing_value_count": missing_values, "discontinuity_count": discontinuities}
        if duplicates:
            result.issues.append(_issue("UNIT-01", "medium", f"Обнаружено повторных дат: {duplicates}", ("date",), sheet.title, 4))
        if non_monotonic:
            result.issues.append(_issue("UNIT-02", "high", "Даты временного ряда расположены не по порядку", ("date",), sheet.title, 4))
        if missing_values:
            result.issues.append(_issue("UNIT-04", "high", f"В ряду отсутствуют СЧА или стоимость пая: {missing_values}", ("nav_kzt", "unit_value_kzt"), sheet.title, 4))
        if discontinuities:
            result.issues.append(_issue("UNIT-05", "medium", f"Изменения стоимости пая свыше 25% требуют объяснения: {discontinuities}", ("unit_value_kzt",), sheet.title, 4))
        with _workbook_or(formula_workbook, path, data_only=False, keep_links=True) as formula_book:
            external_formulas = 0
            broken_formulas = 0
            for row in formula_book["Форма"].iter_rows():
                for cell in row:
                    formula = cell.value
                    if isinstance(formula, str) and formula.startswith("="):
                        external_formulas += int("[" in formula and "]" in formula)
                        broken_formulas += int("#REF!" in formula or "#N/A" in formula)
            if external_formulas:
                result.issues.append(_issue("UNIT-06", "high", f"Формулы содержат внешние ссылки: {external_formulas}", ("formula",), sheet.title, 1))
            if broken_formulas:
                result.issues.append(_issue("UNIT-07", "high", f"Формулы содержат ошибки ссылок: {broken_formulas}", ("formula",), sheet.title, 1))
        if key == "SAQ":
            result.issues.append(_issue("UNIT-03", "high", "Ряд SAQ устарел и требует явного подтверждения перед публикацией", ("latest_date",), sheet.title, 2))
        return result
