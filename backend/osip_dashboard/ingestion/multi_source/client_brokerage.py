"""Parser for the client/brokerage workbook contract."""

from __future__ import annotations

from contextlib import closing
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from osip_dashboard.services.brokerage import is_repo_trade

from .shared import (
    ParsedDataset,
    _as_date,
    _cell,
    _client_ledger_columns,
    _client_position_lookup,
    _client_trade_columns,
    _d,
    _date_text,
    _decimal_text,
    _explicit_report_date,
    _issue,
    _looks_isin,
    _normalize_name,
    _record,
    _resolve_calendar_isin,
    _russian_period_end,
    _text,
)


def _parse_client_brokerage(path: Path, key: str, scope: str) -> ParsedDataset:
    with closing(load_workbook(path, read_only=True, data_only=True)) as workbook:
        if key == "clients":
            sheet = workbook["Лист4"]
            report_date = _explicit_report_date(sheet, 1, 6)
            columns, missing_headers = _client_ledger_columns(sheet)
            records: list[dict[str, Any]] = []
            clients: dict[str, dict[str, Any]] = {}
            # Лист4 is the account/position register and its branch/manager
            # columns are blank in the supplied client workbook. The companion
            # Клиенты sheet is the authoritative manager directory; use an
            # exact normalized-name join so the client table does not silently
            # discard assignments that are already present in the workbook.
            manager_by_client: dict[str, str] = {}
            if "Клиенты" in workbook.sheetnames:
                for dashboard_row in workbook["Клиенты"].iter_rows(min_row=3, values_only=True):
                    dashboard_name = _normalize_name(dashboard_row[1] if len(dashboard_row) > 1 else None)
                    dashboard_manager = _text(dashboard_row[2] if len(dashboard_row) > 2 else None)
                    if dashboard_name and dashboard_manager:
                        manager_by_client[dashboard_name] = dashboard_manager
            current_account = ""
            current_client = ""
            position_count = 0
            for row_number, row in enumerate(sheet.iter_rows(min_row=7, values_only=True), 7):
                account = _text(_cell(row, columns.get("account")))
                # The workbook contains account subtotals and a grand total in
                # the same rectangular range as client rows.  They carry
                # aggregate cash/assets in columns 29/30, so letting them pass
                # through would overwrite the last real account and double
                # the domain totals.  Keep them as source evidence only.
                account_marker = account.casefold().replace(" ", "")
                if account and account_marker.startswith(("итого", "всегопосчету", "всего")):
                    continue
                if account:
                    current_account, current_client = account, _text(_cell(row, columns.get("client_name")))
                    manager = _text(_cell(row, columns.get("manager"))) or manager_by_client.get(_normalize_name(current_client), "")
                    clients[account] = _record("client", account, {
                        "account": account, "client_name": current_client, "iin": _text(_cell(row, columns.get("iin"))),
                        "citizenship": _text(_cell(row, columns.get("citizenship"))), "resident": _text(_cell(row, columns.get("resident"))),
                        "economic_sector": _text(_cell(row, columns.get("economic_sector"))),
                        "document_type": _text(_cell(row, columns.get("document_type"))), "branch": _text(_cell(row, columns.get("branch"))),
                        "category": _text(_cell(row, columns.get("category"))),
                        "agent": _text(_cell(row, columns.get("agent"))), "manager": manager, "cash_kzt": None, "total_assets_kzt": None,
                    }, sheet.title, row_number, column=columns.get("client_name"), field_columns={
                        field: columns.get(field)
                        for field in (
                            "account", "client_name", "iin", "citizenship", "resident",
                            "economic_sector", "document_type", "branch", "category", "agent", "manager",
                        )
                        # cash_kzt/total_assets_kzt are filled in from a later row in this
                        # account's block (see below), so no single cell backs them here.
                    })
                isin = _text(_cell(row, columns.get("isin")))
                if current_account and _looks_isin(isin):
                    position_count += 1
                    records.append(_record("client_position", f"{current_account}:{isin}:{row_number}", {
                        "account": current_account, "client_name": current_client, "issuer": _text(_cell(row, columns.get("issuer"))),
                        "security_type": _text(_cell(row, columns.get("security_type"))), "security_code": _text(_cell(row, columns.get("security_code"))), "isin": isin,
                        "nominal": _decimal_text(_cell(row, columns.get("nominal"))), "nominal_currency": _text(_cell(row, columns.get("nominal_currency"))),
                        "quantity": _decimal_text(_cell(row, columns.get("quantity"))), "market_price": _decimal_text(_cell(row, columns.get("market_price"))),
                        "price_currency": _text(_cell(row, columns.get("price_currency"))), "market_value_kzt": _decimal_text(_cell(row, columns.get("market_value_kzt"))),
                    }, sheet.title, row_number, column=columns.get("isin"), field_columns={
                        field: columns.get(field)
                        for field in (
                            "issuer", "security_type", "security_code", "isin", "nominal",
                            "nominal_currency", "quantity", "market_price", "price_currency", "market_value_kzt",
                        )
                    }))
                cash_cell = _cell(row, columns.get("cash_kzt"))
                if current_account and cash_cell not in (None, ""):
                    clients[current_account]["payload"]["cash_kzt"] = _decimal_text(cash_cell)
                total_assets_cell = _cell(row, columns.get("total_assets_kzt"))
                if current_account and total_assets_cell not in (None, ""):
                    clients[current_account]["payload"]["total_assets_kzt"] = _decimal_text(total_assets_cell)
            records = list(clients.values()) + records
            result = ParsedDataset("client_account_snapshot", key, "business_domain", scope, report_date, report_date, records=records)
            if missing_headers:
                result.issues.append(_issue(
                    "CLIENT-MAP-01",
                    "high",
                    "Не удалось однозначно сопоставить обязательные столбцы реестра клиентов/позиций (Лист4): " + ", ".join(missing_headers),
                    tuple(missing_headers),
                    sheet.title,
                    7,
                ))
            result.summary = {
                "client_count": len(clients), "position_count": position_count,
                "cash_kzt": _decimal_text(sum((_d(item["payload"].get("cash_kzt")) for item in clients.values()), Decimal("0"))),
                "total_assets_kzt": _decimal_text(sum((_d(item["payload"].get("total_assets_kzt")) for item in clients.values()), Decimal("0"))),
            }
            manager_mix: dict[str, dict[str, Any]] = {}
            for item in clients.values():
                manager = _text(item["payload"].get("manager")) or "Не указано"
                bucket = manager_mix.setdefault(manager, {"client_count": 0, "cash_kzt": Decimal("0"), "total_assets_kzt": Decimal("0")})
                bucket["client_count"] += 1
                bucket["cash_kzt"] += _d(item["payload"].get("cash_kzt"))
                bucket["total_assets_kzt"] += _d(item["payload"].get("total_assets_kzt"))
            result.summary["manager_mix"] = {
                manager: {"client_count": values["client_count"], "cash_kzt": _decimal_text(values["cash_kzt"]), "total_assets_kzt": _decimal_text(values["total_assets_kzt"])}
                for manager, values in sorted(manager_mix.items(), key=lambda pair: pair[1]["total_assets_kzt"], reverse=True)
            }
            if any(_text(item["payload"].get("cash_kzt")) == "" for item in clients.values()):
                result.issues.append(_issue("CLIENT-01", "medium", "Для части клиентов не найден итог денежных средств", ("cash_kzt",), sheet.title, 3))
            return result
        if key == "trades":
            sheet = workbook["Лист8"]
            header_columns, header_row, missing_headers = _client_trade_columns(sheet)
            # Лист8 (trade ledger) carries no report-date label of its own;
            # Лист4 (client register) does, so reuse it rather than treating
            # any date-shaped cell near the trade rows as a report date.
            report_date = _explicit_report_date(workbook["Лист4"], 1, 6)
            records = []
            first_data_row = max(header_row + 2, 10)
            for row_number, row in enumerate(sheet.iter_rows(min_row=first_data_row, values_only=True), first_data_row):
                trade_number = _cell(row, header_columns.get("trade_number"))
                if trade_number in (None, ""):
                    continue
                payload = {
                    "trade_number": _text(trade_number),
                    "order_number": _text(_cell(row, header_columns.get("order_number"))),
                    "trade_date": _date_text(_cell(row, header_columns.get("trade_date"))),
                    "venue": _text(_cell(row, header_columns.get("venue"))),
                    "client_name": _text(_cell(row, header_columns.get("client_name"))),
                    "account": _text(_cell(row, header_columns.get("account"))),
                    "counterparty": _text(_cell(row, header_columns.get("counterparty"))),
                    "side": _text(_cell(row, header_columns.get("side"))),
                    "issuer": _text(_cell(row, header_columns.get("issuer"))),
                    "security_type": _text(_cell(row, header_columns.get("security_type"))),
                    "isin": _text(_cell(row, header_columns.get("isin"))),
                    "quantity": _decimal_text(_cell(row, header_columns.get("quantity"))),
                    "amount": _decimal_text(_cell(row, header_columns.get("amount"))),
                    "currency": _text(_cell(row, header_columns.get("currency"))),
                    "clean_price": _decimal_text(_cell(row, header_columns.get("clean_price"))),
                    "dirty_price": _decimal_text(_cell(row, header_columns.get("dirty_price"))),
                    "yield_percent": _decimal_text(_cell(row, header_columns.get("yield_percent"))),
                    "settlement_date": _date_text(_cell(row, header_columns.get("settlement_date"))),
                    "execution_status": _text(_cell(row, header_columns.get("execution_status"))),
                    "failure_reason": _text(_cell(row, header_columns.get("failure_reason"))),
                }
                payload["is_repo"] = is_repo_trade(payload)
                # header_columns already maps every payload field to its own
                # column (it's built from the sheet's own header row), so it
                # doubles directly as the field->column provenance map -
                # excluding is_repo, which is derived, not a source cell.
                records.append(_record("trade", f"{payload['trade_number']}:{row_number}", payload, sheet.title, row_number, column=header_columns.get("trade_number"), field_columns=dict(header_columns)))
            latest = max((_as_date(item["payload"]["trade_date"]) for item in records if item["payload"]["trade_date"]), default=report_date)
            result = ParsedDataset("brokerage_trade_ledger", key, "business_domain", scope, report_date, latest, records=records)
            header_values = next(
                (tuple(row) for row_number, row in enumerate(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True), header_row)),
                tuple(),
            )
            mapping_fields = []
            for field, column in sorted(header_columns.items()):
                samples = []
                for sample_row in sheet.iter_rows(min_row=first_data_row, max_row=first_data_row + 2, values_only=True):
                    value = _cell(sample_row, column)
                    if value not in (None, ""):
                        samples.append(_text(value))
                mapping_fields.append({
                    "normalized_field": field,
                    "source_header": _text(_cell(header_values, column)) or None,
                    "source_sheet": sheet.title,
                    "source_row": header_row,
                    "source_column": column + 1,
                    "sample_values": samples,
                })
            matched_headers = sorted(header_columns)
            if missing_headers:
                result.issues.append(_issue(
                    "BROKERAGE-MAP-01",
                    "high",
                    "Не удалось однозначно сопоставить обязательные столбцы журнала сделок: " + ", ".join(missing_headers),
                    tuple(missing_headers),
                    sheet.title,
                    header_row,
                ))
            turnover: dict[str, Decimal] = {}
            buys: dict[str, Decimal] = {}
            sells: dict[str, Decimal] = {}
            venues: dict[str, int] = {}
            instruments: dict[str, int] = {}
            statuses: dict[str, int] = {}
            repo_trade_count = 0
            repo_turnover: dict[str, Decimal] = {}
            for item in records:
                payload = item["payload"]
                currency = payload.get("currency") or "Не указано"
                amount = abs(_d(payload.get("amount")))
                if payload.get("is_repo"):
                    repo_trade_count += 1
                    repo_turnover[currency] = repo_turnover.get(currency, Decimal("0")) + amount
                turnover[currency] = turnover.get(currency, Decimal("0")) + amount
                side = _text(payload.get("side")).casefold()
                target = buys if side.startswith(("куп", "покуп", "buy")) else sells if side.startswith(("прод", "sell")) else None
                if target is not None: target[currency] = target.get(currency, Decimal("0")) + amount
                for bucket, value in ((venues, payload.get("venue")), (instruments, payload.get("security_type")), (statuses, payload.get("execution_status"))):
                    label = _text(value) or "Не указано"; bucket[label] = bucket.get(label, 0) + 1
            result.summary = {
                "trade_count": len(records), "latest_trade_date": _date_text(latest),
                "date_basis": "latest_transaction",
                "gross_turnover_by_currency": {key: _decimal_text(value) for key, value in sorted(turnover.items())},
                "buy_turnover_by_currency": {key: _decimal_text(value) for key, value in sorted(buys.items())},
                "sell_turnover_by_currency": {key: _decimal_text(value) for key, value in sorted(sells.items())},
                "venue_mix": venues, "instrument_mix": instruments, "execution_status_mix": statuses,
                "repo_trade_count": repo_trade_count,
                "repo_turnover_by_currency": {key: _decimal_text(value) for key, value in sorted(repo_turnover.items())},
                "mapping": {
                    "header_row": header_row,
                    "confidence": "high" if not missing_headers else "low",
                    "matched_fields": matched_headers,
                    "missing_fields": missing_headers,
                    "fields": mapping_fields,
                    "mapping_confirmed": not missing_headers,
                },
            }
            return result
        if key == "derivatives":
            sheet = workbook["Лист7"]
            report_date = _russian_period_end(_text(sheet["A9"].value))
            records = []
            excluded_non_derivative = 0
            for row_number, row in enumerate(sheet.iter_rows(min_row=15, values_only=True), 15):
                if not isinstance(row[0], (int, float)): continue
                payload = {"number": _text(row[0]), "trade_datetime": _date_text(row[1]), "settlement_date": _date_text(row[3]), "instrument_type": _text(row[4]), "identifier": _text(row[5]), "market": _text(row[6]), "underlying": _text(row[7]), "counterparty": _text(row[9]), "deal_type": _text(row[10]), "quantity": _decimal_text(row[11]), "price": _decimal_text(row[12]), "amount": _decimal_text(row[13]), "currency": _text(row[14]), "obligation_status": _text(row[22])}
                # "Лист7" is meant to be the derivatives register, but in
                # practice most of its rows are plain ETF purchases (no
                # strike, no expiry - just quantity x price) that duplicate
                # trades already correctly recorded on the trade ledger
                # (Лист8) with security_type=ETF. An ETF unit is direct fund
                # ownership, not a derivative contract, so it doesn't belong
                # in this register regardless of which sheet the source
                # workbook happened to put it on.
                if payload["instrument_type"].casefold() == "etf":
                    excluded_non_derivative += 1
                    continue
                records.append(_record("derivative", f"{payload['number']}:{row_number}", payload, sheet.title, row_number, column=0, field_columns={
                    "number": 0, "trade_datetime": 1, "settlement_date": 3, "instrument_type": 4, "identifier": 5,
                    "market": 6, "underlying": 7, "counterparty": 9, "deal_type": 10, "quantity": 11, "price": 12,
                    "amount": 13, "currency": 14, "obligation_status": 22,
                }))
            result = ParsedDataset("derivatives_register", key, "business_domain", scope, report_date, report_date, records=records)
            result.summary = {"derivative_count": len(records), "date_basis": "reporting_period_end"}
            if excluded_non_derivative:
                result.summary["excluded_etf_rows"] = excluded_non_derivative
                result.issues.append(_issue(
                    "DERIV-01", "medium",
                    f"Исключено строк ETF (не являются производными инструментами): {excluded_non_derivative}. "
                    "Эти сделки уже учтены в реестре сделок (Лист8) как ETF.",
                    ("instrument_type",), sheet.title, 15,
                ))
            return result
        if key == "client_dashboard":
            sheet = workbook["Клиенты"]
            report_date = _explicit_report_date(workbook["Лист4"], 1, 6)
            records = []
            manager_mix: dict[str, dict[str, Any]] = {}
            for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
                client = _text(row[1])
                if not client:
                    continue
                manager = _text(row[2]) or "Не указано"
                payload = {
                    "client_name": client,
                    "manager": manager,
                    "client_type": _text(row[3]),
                    "opening_date": _date_text(row[4]),
                    "cash_kzt": _decimal_text(row[5]),
                    "securities_value_kzt": _decimal_text(row[6]),
                    "total_assets_kzt": _decimal_text(row[7]),
                    "cash_share": _decimal_text(row[8]),
                    "income": _decimal_text(row[9]),
                    "status": _text(row[10]),
                    "source_key": _text(row[11]),
                }
                record = _record("client_summary", f"{client}:{row_number}", payload, sheet.title, row_number, column=1, field_columns={
                    "client_name": 1, "manager": 2, "client_type": 3, "opening_date": 4, "cash_kzt": 5,
                    "securities_value_kzt": 6, "total_assets_kzt": 7, "cash_share": 8, "income": 9,
                    "status": 10, "source_key": 11,
                })
                records.append(record)
                bucket = manager_mix.setdefault(manager, {"client_count": 0, "cash_kzt": Decimal("0"), "securities_value_kzt": Decimal("0"), "total_assets_kzt": Decimal("0")})
                bucket["client_count"] += 1
                bucket["cash_kzt"] += _d(payload["cash_kzt"])
                bucket["securities_value_kzt"] += _d(payload["securities_value_kzt"])
                bucket["total_assets_kzt"] += _d(payload["total_assets_kzt"])
            result = ParsedDataset("client_dashboard_snapshot", key, "business_domain", scope, report_date, report_date, records=records)
            result.summary = {
                "client_count": len(records),
                "cash_kzt": _decimal_text(sum((_d(item["payload"].get("cash_kzt")) for item in records), Decimal("0"))),
                "securities_value_kzt": _decimal_text(sum((_d(item["payload"].get("securities_value_kzt")) for item in records), Decimal("0"))),
                "total_assets_kzt": _decimal_text(sum((_d(item["payload"].get("total_assets_kzt")) for item in records), Decimal("0"))),
                "manager_mix": {manager: {key: _decimal_text(value) if isinstance(value, Decimal) else value for key, value in bucket.items()} for manager, bucket in sorted(manager_mix.items(), key=lambda pair: pair[1]["total_assets_kzt"], reverse=True)},
            }
            ledger_totals = next((row for row in workbook["Лист4"].iter_rows(min_row=7, values_only=True) if _text(row[1]).casefold().replace(" ", "") == "всегопосчету"), None)
            if ledger_totals is not None:
                ledger_cash = _decimal_text(ledger_totals[28] if len(ledger_totals) > 28 else None)
                ledger_assets = _decimal_text(ledger_totals[29] if len(ledger_totals) > 29 else None)
                result.summary["ledger_cash_kzt"] = ledger_cash
                result.summary["ledger_total_assets_kzt"] = ledger_assets
                if abs(_d(ledger_cash) - _d(result.summary["cash_kzt"])) > Decimal("1") or abs(_d(ledger_assets) - _d(result.summary["total_assets_kzt"])) > Decimal("1"):
                    result.issues.append(_issue("CLIENT-DASH-01", "medium", "Сводный лист клиентов не совпадает с итогом реестра Лист4; оба значения сохранены как источник", ("total_assets_kzt", "cash_kzt"), sheet.title, 2))
            return result
        if key == "maturity_calendar":
            sheet = workbook["календарь погашения"]
            report_date = _explicit_report_date(workbook["Лист4"], 1, 6)
            # The calendar sheet's own ISIN column is an array formula that
            # joins to Лист4 by (issuer, current market value) - fragile by
            # construction, since a position's market value moves daily and
            # the two sheets are not guaranteed to be valued on the same
            # day. Лист4 is the actual position register (the same one
            # "clients" parses above); it is authoritative, so look the
            # ISIN up there directly instead of trusting the calendar's own
            # (frequently unresolved) formula result. The calendar's own
            # value is still used, but only to disambiguate on the rare
            # occasion a client holds more than one line from the same
            # issuer - a confirmation signal, not the join key.
            position_lookup = _client_position_lookup(workbook["Лист4"])
            records = []
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                client = _text(row[0]); instrument = _text(row[2]); maturity = _date_text(row[4])
                if not client and not instrument and not maturity:
                    continue
                if not maturity:
                    continue
                candidates = position_lookup.get((client, instrument), [])
                calendar_coupon = _d(_decimal_text(row[6])) if row[6] not in (None, "") else None
                isin = _resolve_calendar_isin(candidates, calendar_coupon, _d(_decimal_text(row[8])))
                payload = {
                    "client_name": client,
                    "manager": _text(row[1]),
                    "instrument": instrument,
                    "isin": isin,
                    "maturity_date": maturity,
                    "coupon_payment_date": _date_text(row[5]),
                    "coupon_percent": _decimal_text(row[6]),
                    "days_to_maturity": _decimal_text(row[7]),
                    "value_kzt": _decimal_text(row[8]),
                }
                # isin/coupon_payment_date are excluded from field_columns
                # deliberately: isin is now sourced from Лист4 (a different
                # sheet than this record's own "календарь погашения"), and
                # this preview mechanism only ever opens one sheet per
                # record, so pointing at a Лист4 cell here would show the
                # wrong sheet's content - worse than no pointer at all.
                # coupon_payment_date depends on a column not sourced here
                # at all and, like isin's own broken formula, this workbook
                # saved it with no cached result for most rows - not a
                # parsing bug, a stale/unrecalculated source file. Claiming
                # a "consumed formula" source cell for a value that never
                # resolves would either block every other (genuinely fine)
                # field in this row too, or - if the audit
                # gate is ever loosened - silently publish a blank as if it
                # were evidence-backed. The values still publish honestly as
                # whatever the cache holds (usually empty) via the payload
                # above; they just aren't claimed as sourced-from-a-cell.
                records.append(_record("maturity_event", f"{client}:{instrument}:{maturity}:{row_number}", payload, sheet.title, row_number, column=0, field_columns={
                    "client_name": 0, "manager": 1, "instrument": 2, "maturity_date": 4,
                    "coupon_percent": 6, "days_to_maturity": 7, "value_kzt": 8,
                }))
            dates = sorted(item["payload"]["maturity_date"] for item in records if item["payload"].get("maturity_date"))
            total = sum((_d(item["payload"].get("value_kzt")) for item in records), Decimal("0"))
            result = ParsedDataset("client_maturity_calendar", key, "business_domain", scope, report_date, report_date, records=records)
            result.summary = {
                "event_count": len(records),
                "total_value_kzt": _decimal_text(total),
                "nearest_maturity_date": dates[0] if dates else None,
                "latest_maturity_date": dates[-1] if dates else None,
            }
            return result
        sheet = workbook["Лист6"]
        records = []
        names: set[str] = set()
        duplicates = 0
        master_counts: dict[str, int] = {}
        for master_row in workbook["Лист4"].iter_rows(min_row=7, values_only=True):
            marker = _text(master_row[1]).casefold().replace(" ", "")
            if _text(master_row[1]) and not marker.startswith(("итого", "всегопосчету", "всего")):
                normalized = _normalize_name(master_row[2])
                if normalized: master_counts[normalized] = master_counts.get(normalized, 0) + 1
        unmatched = 0
        ambiguous = 0
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            name = _normalize_name(row[0])
            if not name: continue
            duplicates += int(name in names); names.add(name)
            count = master_counts.get(name, 0)
            match_status = "exact" if count == 1 else "ambiguous" if count > 1 else "unmatched"
            unmatched += int(match_status == "unmatched"); ambiguous += int(match_status == "ambiguous")
            records.append(_record("client_open_date", f"{name}:{row_number}", {"normalized_name": name, "source_name": _text(row[0]), "open_date": _date_text(row[1]), "match_status": match_status}, sheet.title, row_number, column=0, field_columns={"source_name": 0, "normalized_name": 0, "open_date": 1}))
        result = ParsedDataset("client_open_dates", key, "business_domain", scope, None, None, records=records)
        result.summary = {"reference_count": len(records), "exact_matches": len(records) - unmatched - ambiguous, "unmatched": unmatched, "ambiguous": ambiguous}
        if duplicates:
            result.issues.append(_issue("CLIENT-02", "medium", "Справочник дат открытия содержит повторные нормализованные имена", ("client_name",), sheet.title, 1))
        if unmatched:
            result.issues.append(_issue("CLIENT-03", "medium", f"Имена без точного соответствия в клиентском источнике: {unmatched}", ("client_name",), sheet.title, 1))
        if ambiguous:
            result.issues.append(_issue("CLIENT-04", "high", f"Имена с неоднозначным точным соответствием: {ambiguous}", ("client_name",), sheet.title, 1))
        return result
