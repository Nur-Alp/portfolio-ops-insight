from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import UUID

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session as SqlaSession

from osip_dashboard.persistence import Base
from osip_dashboard.persistence.database import create_database_engine
from osip_dashboard.persistence.models import DatasetRecord, DatasetVersion, ImportStatus, SourceUpload, utcnow
from osip_dashboard.ingestion.formula_audit import audit_consumed_formula_results, audit_workbook
from osip_dashboard.ingestion.multi_source import RISK_NEAR_BREACH_POLICY_VERSION, RISK_NEAR_BREACH_THRESHOLD, _explicit_report_dates, _extract_coupon_rate, _extract_isins, _parse_accounting_landing, _parse_amount, _parse_client_brokerage, _parse_corporate_finance, _parse_tabys_valuation, _parse_unit_history, _resolve_calendar_isin, _risk_near_breach, _risk_utilization, detect_source, parse_detected_dataset
from osip_dashboard.services.multi_source import _all_dataset_type_scopes, _freshness, _readiness_status, approve_dataset, create_source_upload, latest_published, list_datasets, module_payload
from osip_dashboard.storage import LocalBlobStore

from export_compliance import assert_workbook_is_compliant
from test_snapshot_api import approve_and_publish as osip_approve_and_publish, upload as osip_upload


def test_extract_coupon_rate_reads_the_rate_before_a_date_and_ignores_unparseable_text():
    assert _extract_coupon_rate("ORCL 4.2 09/27/29") == Decimal("4.2")
    assert _extract_coupon_rate("BAC 1.734 07/22/27") == Decimal("1.734")
    assert _extract_coupon_rate("JWN 4 03/15/27") == Decimal("4")
    assert _extract_coupon_rate("FRTBS.0230") is None
    assert _extract_coupon_rate("B 08/27/2026") is None
    assert _extract_coupon_rate(None) is None


def test_resolve_calendar_isin_rejects_a_same_issuer_candidate_whose_rate_contradicts():
    # Regression for a real case: a client's only "Oracle Corporation" line
    # in the register is a 4.2% bond, but a calendar row also names Oracle
    # Corporation at a 2.65% rate - a different, unregistered position, not
    # that bond. Accepting the sole same-issuer candidate anyway would have
    # fabricated an ISIN for the wrong instrument.
    candidates = [(631, "US68389XCS27", 942009150, "ORCL 4.2 09/27/29")]
    assert _resolve_calendar_isin(candidates, Decimal("4.2"), Decimal("931743550")) == "US68389XCS27"
    assert _resolve_calendar_isin(candidates, Decimal("2.65"), Decimal("462386450")) is None


def test_resolve_calendar_isin_does_not_let_a_zero_coupon_match_short_circuit_the_value_tiebreak():
    # Regression: two zero-coupon T-bills for the same client/issuer both
    # "match" a calendar row's 0% coupon equally (every zero-coupon
    # instrument does), so that must not be treated as a discriminating
    # confirmation the way a real rate like 4.2% is - it has to fall back to
    # picking by current value like an unconfirmed multi-candidate group.
    candidates = [
        (611, "US912797VG91", 76582290, "B 0 12/10/26"),
        (612, "US912797TY36", 106181580, "B 08/27/2026"),
    ]
    assert _resolve_calendar_isin(candidates, Decimal("0"), Decimal("105024460")) == "US912797TY36"
    assert _resolve_calendar_isin(candidates, Decimal("0"), Decimal("75747730")) == "US912797VG91"


def test_resolve_calendar_isin_accepts_a_single_unambiguous_candidate_with_no_rate_signal():
    candidates = [(1, "KZX000005186", 1879320000, "FRTB 09/27")]
    assert _resolve_calendar_isin(candidates, Decimal("0"), Decimal("1879320000")) == "KZX000005186"


def test_resolve_calendar_isin_returns_none_with_no_candidates():
    assert _resolve_calendar_isin([], Decimal("4.2"), Decimal("100")) is None


def test_formula_audit_reports_cached_status_errors_and_external_links(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence"
    sheet["A1"] = 10
    sheet["A2"] = "=A1+1"  # openpyxl has no recalculated cache for new formulas
    sheet["A3"] = "=A1+#REF!"
    sheet["A4"] = "=[other.xlsx]Sheet1!A1"
    sheet["A5"] = "#N/A"  # literal error value, not a formula
    path = tmp_path / "formula-audit.xlsx"
    workbook.save(path)

    audit = audit_workbook(path)

    assert audit["format"] == "xlsx"
    assert audit["formula_count"] == 3
    assert audit["blank_cached_formula_count"] == 3
    assert audit["formula_error_count"] == 1
    assert audit["error_value_count"] == 1
    assert audit["external_formula_count"] == 1
    assert audit["formula_status"] == "formula_errors"
    assert audit["cached_result_status"] == "inspected"
    assert audit["by_sheet"]["Evidence"]["formula_count"] == 3


def test_formula_audit_is_attached_to_every_parsed_xlsx_dataset(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр"
    sheet["A1"] = "Направление корпоративного финансирования"
    sheet["A2"] = "=1+1"
    path = tmp_path / "corpfin-formula.xlsx"
    workbook.save(path)

    detection = detect_source(path, "xlsx")
    parsed = parse_detected_dataset(path, detection, "deals", "CORPFIN")

    assert parsed.summary["formula_audit"]["formula_count"] == 1
    assert parsed.summary["formula_audit"]["format"] == "xlsx"


def test_consumed_formula_audit_blocks_only_published_field_formula_errors(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence"
    sheet["A1"] = "input"
    sheet["B2"] = "=A1+#REF!"
    sheet["C2"] = "=A1+#REF!"  # Helper error, not a parser-backed field.
    path = tmp_path / "consumed-formula.xlsx"
    workbook.save(path)

    audit = audit_consumed_formula_results(
        path,
        [{"source_ref": {"sheet_name": "Evidence", "field_columns": {"published_metric": {"source_cell": "B2"}}}}],
    )

    assert audit["status"] == "blocked"
    assert audit["checked_formula_cells"] == 1
    assert audit["invalid_cells"] == [{
        "sheet_name": "Evidence", "source_cell": "B2", "fields": ["published_metric"], "reason": "blank_cached_result",
    }]


def test_legacy_consumed_formula_audit_is_explicitly_not_inspectable(tmp_path):
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"not-a-real-xls")
    audit = audit_consumed_formula_results(path, [])
    assert audit["status"] == "not_inspectable"


def test_formula_audit_counts_legacy_xls_formula_records():
    path = Path("Portfolio operations/ОСИП Портфель о состоянии текущего портфеля и  предстоящих расчетах по сделкам  СОБСТВ 15.07.2026.xls")
    if not path.exists():
        pytest.skip("legacy workbook fixture is not checked out")

    audit = audit_workbook(path)

    assert audit["format"] == "xls"
    assert audit["formula_count"] > 0
    assert audit["formula_status"] == "formula_records_detected"
    assert audit["cached_result_status"] == "not_exposed_by_reader"
    assert audit["recalculation_status"] == "not_available"


# A solo domain operator holds every role - visibility is uploader-scoped
# (see docs/domain-upload-instructions.md "Visibility rule"), so a
# genuinely separate reviewer/publisher/reader identity couldn't see this
# actor's own datasets anyway. REVIEWER/PUBLISHER/READER below remain for
# tests that specifically exercise a distinct actor identity.
UPLOADER = {"X-Actor-Id": "uploader", "X-Actor-Roles": "uploader,reviewer,publisher,reader"}
REVIEWER = {"X-Actor-Id": "reviewer", "X-Actor-Roles": "reviewer,reader"}
PUBLISHER = {"X-Actor-Id": "publisher", "X-Actor-Roles": "publisher,reader"}
READER = {"X-Actor-Id": "reader", "X-Actor-Roles": "reader"}


def test_domain_scope_is_independent_from_portfolio_scope(api):
    context = api.get(
        "/api/v1/session/context",
        headers={**READER, "X-Actor-Domains": "client_ops", "X-Actor-Portfolios": "*"},
    )
    assert context.status_code == 200
    assert context.json()["domains"] == ["client_ops"]

    denied = api.get(
        "/api/v1/corporate-finance/overview",
        headers={**READER, "X-Actor-Domains": "client_ops"},
    )
    assert denied.status_code == 403

    allowed = api.get(
        "/api/v1/corporate-finance/overview",
        headers={**READER, "X-Actor-Domains": "corpfin"},
    )
    assert allowed.status_code == 200

    legacy_denied = api.get(
        "/api/v1/portfolios",
        headers={**READER, "X-Actor-Domains": "client_ops", "X-Actor-Portfolios": "*"},
    )
    assert legacy_denied.status_code == 403


def test_accounting_disclosures_follow_accept_language(api):
    accounting = api.get(
        "/api/v1/accounting/source-readiness",
        headers={**READER, "X-Actor-Domains": "accounting", "Accept-Language": "en"},
    )
    assert accounting.status_code == 200
    assert accounting.json()["records"]["readiness"][0]["requirement"] == "Authoritative accounting package"


def test_source_dataset_materialization_rejects_cross_domain_assignment(api):
    content = _corporate_finance_workbook()
    uploaded = api.post(
        "/api/v1/source-uploads",
        headers={**UPLOADER, "X-Actor-Domains": "corpfin"},
        files={"file": ("corp.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert uploaded.status_code == 201, uploaded.text
    upload_id = uploaded.json()["id"]
    denied = api.post(
        f"/api/v1/source-uploads/{upload_id}/datasets",
        headers={**UPLOADER, "X-Actor-Domains": "client_ops"},
        json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]},
    )
    assert denied.status_code == 403


def test_client_identity_exception_queue_keeps_source_record_immutable(api):
    session = api.app.state.session_factory()
    try:
        upload = SourceUpload(
            source_sha256="d" * 64,
            original_filename="client.xlsx",
            storage_key="client.xlsx",
            file_format="xlsx",
            detected_source_type="client_brokerage",
            detection={},
            uploader_id="uploader",
        )
        snapshot = DatasetVersion(
            source_upload=upload, dataset_type="client_account_snapshot", detected_key="clients",
            scope_type="business_domain", scope_code="BROKERAGE", source_report_date=date(2026, 7, 20),
            business_date=date(2026, 7, 20), parser_version="test", version=1,
            status=ImportStatus.PUBLISHED, summary={}, uploader_id="uploader",
        )
        open_dates = DatasetVersion(
            source_upload=upload, dataset_type="client_open_dates", detected_key="open_dates",
            scope_type="business_domain", scope_code="BROKERAGE", source_report_date=date(2026, 7, 20),
            business_date=date(2026, 7, 20), parser_version="test", version=1,
            status=ImportStatus.VALIDATED, summary={}, uploader_id="uploader",
        )
        session.add_all([upload, snapshot, open_dates]); session.flush()
        client_record = DatasetRecord(dataset_id=snapshot.id, record_type="client", record_key="ACC-1", payload={"account": "ACC-1", "client_name": "Иванов Иван", "iin": "900101123456"}, source_ref={"sheet_name": "Лист4", "row_number": 7}, raw_values={}, formulas={}, cached_values={})
        session.add(client_record)
        source_record = DatasetRecord(dataset_id=open_dates.id, record_type="client_open_date", record_key="IVANOV", payload={"normalized_name": "ИВАНОВ ИВАН", "source_name": "Иванов Иван", "open_date": "2020-01-01", "match_status": "unmatched"}, source_ref={"sheet_name": "Лист6", "row_number": 12}, raw_values={}, formulas={}, cached_values={})
        session.add(source_record); session.commit(); record_id = source_record.id
    finally:
        session.close()

    headers = {**UPLOADER, "X-Actor-Domains": "client_ops"}
    pending = api.get("/api/v1/client-exceptions", headers=headers)
    assert pending.status_code == 200, pending.text
    item = pending.json()["items"][0]
    assert item["original_match_status"] == "unmatched"
    assert item["candidate_accounts"] == ["ACC-1"]

    detail = api.get(f"/api/v1/clients/records/{client_record.id}", headers=headers)
    assert detail.status_code == 200, detail.text
    assert any(row.get("iin") == "900101123456" for row in detail.json()["records"]["client_account_snapshot"])

    # A different operator (even with the right role/domain) must not see or
    # resolve another uploader's exception - the same per-uploader
    # visibility rule that applies to every other dataset read/write.
    other_operator = {"X-Actor-Id": "other-operator", "X-Actor-Roles": "reviewer,reader", "X-Actor-Domains": "client_ops"}
    assert api.get("/api/v1/client-exceptions", headers=other_operator).json()["items"] == []
    denied = api.post(
        f"/api/v1/client-exceptions/{record_id}/resolve",
        headers=other_operator,
        json={"disposition": "confirmed", "account": "ACC-1", "comment": "Проверено вручную"},
    )
    assert denied.status_code == 403, denied.text

    resolved = api.post(
        f"/api/v1/client-exceptions/{record_id}/resolve",
        headers={**UPLOADER, "X-Actor-Domains": "client_ops"},
        json={"disposition": "confirmed", "account": "ACC-1", "comment": "Проверено вручную"},
    )
    assert resolved.status_code == 200, resolved.text
    session = api.app.state.session_factory()
    try:
        source = session.get(DatasetRecord, record_id)
        assert source is not None
        assert source.payload["match_status"] == "unmatched"
    finally:
        session.close()
    assert api.get("/api/v1/client-exceptions", headers=headers).json()["items"] == []


def _tabys_valuation_workbook(*, nav_kzt: float, unit_value_kzt: float) -> bytes:
    """A minimal "дата" sheet covering the fields NAV/unit-value DQ checks read."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "дата"
    sheet["B2"] = date(2026, 7, 1)
    rows = [
        ("Текущая стоимость портфеля ЦБ", 1_000_000),
        ("Деньги на инвестиционном счете (в тенге)", 500_000),
        ("Деньги на инвестиционном счете (в USD)", 0),
        ("Деньги на инвестиционном счете (в EUR)", 0),
        ("Обязательства", 100_000),
        ('Текущая "чистая" стоимость активов', nav_kzt),
        ("Количество паев в обращении", 1_000),
        ("Расчетная стоимость пая", unit_value_kzt),
        ("Расчетная стоимость в пая USD", 1),
    ]
    for offset, (label, value) in enumerate(rows, start=1):
        sheet.cell(offset, 1, label)
        sheet.cell(offset, 4, value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_tabys_nav_reconciliation_dq_codes_fire_on_a_genuine_mismatch(tmp_path):
    # Correct NAV would be 1,000,000 + 500,000 - 100,000 = 1,400,000, and a
    # correct unit value would be 1,400,000 / 1,000 = 1,400 - both are wrong
    # here on purpose.
    path = tmp_path / "tabys-valuation.xlsx"
    path.write_bytes(_tabys_valuation_workbook(nav_kzt=999_999, unit_value_kzt=1))
    parsed = _parse_tabys_valuation(path, "valuation", "TABYS")
    codes = {issue.code for issue in parsed.issues}
    assert "TABYS-NAV-01" in codes
    assert "TABYS-NAV-02" in codes


def test_tabys_nav_reconciliation_dq_codes_do_not_fire_on_consistent_figures(tmp_path):
    path = tmp_path / "tabys-valuation.xlsx"
    path.write_bytes(_tabys_valuation_workbook(nav_kzt=1_400_000, unit_value_kzt=1_400))
    parsed = _parse_tabys_valuation(path, "valuation", "TABYS")
    codes = {issue.code for issue in parsed.issues}
    assert "TABYS-NAV-01" not in codes
    assert "TABYS-NAV-02" not in codes


def test_tabys_prices_flags_missing_prices_and_a_stale_price_date(tmp_path):
    workbook = Workbook()
    dата = workbook.active
    dата.title = "дата"
    dата["B2"] = date(2026, 7, 10)
    prices = workbook.create_sheet("Цены")
    prices.append([])
    prices.append([])
    # Missing price (#N/A) and a latest price date more than a day before the
    # valuation date (2026-07-05, five days before 2026-07-10).
    prices.append(["KZ0000000001", "МинФин РК", 100.5, "KZT", date(2026, 7, 5)])
    prices.append(["KZ0000000002", "МинФин РК-2", "#N/A", "KZT", date(2026, 7, 5)])
    buffer = BytesIO()
    workbook.save(buffer)
    path = tmp_path / "tabys-valuation.xlsx"
    path.write_bytes(buffer.getvalue())
    parsed = _parse_tabys_valuation(path, "prices", "TABYS")
    codes = {issue.code for issue in parsed.issues}
    assert "TABYS-PRICE-01" in codes
    assert "TABYS-PRICE-02" in codes
    assert parsed.summary["missing_price_count"] == 1


def test_tabys_prices_do_not_fire_when_prices_are_complete_and_current(tmp_path):
    workbook = Workbook()
    dата = workbook.active
    dата.title = "дата"
    dата["B2"] = date(2026, 7, 10)
    prices = workbook.create_sheet("Цены")
    prices.append([])
    prices.append([])
    prices.append(["KZ0000000001", "МинФин РК", 100.5, "KZT", date(2026, 7, 10)])
    buffer = BytesIO()
    workbook.save(buffer)
    path = tmp_path / "tabys-valuation.xlsx"
    path.write_bytes(buffer.getvalue())
    parsed = _parse_tabys_valuation(path, "prices", "TABYS")
    codes = {issue.code for issue in parsed.issues}
    assert "TABYS-PRICE-01" not in codes
    assert "TABYS-PRICE-02" not in codes


def test_tabys_evidence_sheets_are_flagged_inactive_and_broken_formulas_are_reported(tmp_path):
    workbook = Workbook()
    dата = workbook.active
    dата.title = "дата"
    dата["B2"] = date(2026, 7, 10)
    evidence = workbook.create_sheet("часть 2 (депозиты)")
    evidence["A1"] = "=A2+#REF!"
    buffer = BytesIO()
    workbook.save(buffer)
    path = tmp_path / "tabys-valuation.xlsx"
    path.write_bytes(buffer.getvalue())
    parsed = _parse_tabys_valuation(path, "inactive_evidence", "TABYS")
    codes = {issue.code for issue in parsed.issues}
    assert "TABYS-EVIDENCE-01" in codes
    assert "TABYS-EVIDENCE-02" in codes
    evidence_record = next(record for record in parsed.records if record["record_key"] == "часть 2 (депозиты)")
    assert evidence_record["payload"]["active"] is False


def _unit_history_workbook() -> bytes:
    """TABYS columns are J/M/N/O/P/Q/R (date/units/investors/nav_usd/
    unit_value_usd/nav_kzt/unit_value_kzt) starting at row 4 - see
    _parse_unit_history's non-SAQ `columns` tuple."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Форма"

    def row(row_number: int, item_date, units, investors, nav_usd, unit_value_usd, nav_kzt, unit_value_kzt) -> None:
        sheet.cell(row_number, 10, item_date)
        sheet.cell(row_number, 13, units)
        sheet.cell(row_number, 14, investors)
        sheet.cell(row_number, 15, nav_usd)
        sheet.cell(row_number, 16, unit_value_usd)
        sheet.cell(row_number, 17, nav_kzt)
        sheet.cell(row_number, 18, unit_value_kzt)

    row(4, date(2026, 6, 1), 1000, 5, 100, 1.5, 47_000, 470)
    row(5, date(2026, 6, 1), 1000, 5, 100, 1.5, 47_000, 470)  # duplicate date -> UNIT-01
    row(6, date(2026, 5, 1), 1000, 5, 100, 1.5, 47_000, 470)  # earlier than the prior row -> UNIT-02
    row(7, date(2026, 7, 1), 1000, 5, 100, 1.5, None, 470)  # nav_kzt missing -> UNIT-04
    row(8, date(2026, 8, 1), 1000, 5, 100, 1.5, 94_000, 940)  # unit value doubled -> UNIT-05
    # External and broken-reference formulas, anywhere on the sheet.
    sheet.cell(20, 1, "=[other.xlsx]Sheet1!A1")
    sheet.cell(21, 1, "=A1+#REF!")
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_unit_history_flags_duplicate_non_monotonic_missing_and_discontinuous_rows(tmp_path):
    path = tmp_path / "unit-history.xlsx"
    path.write_bytes(_unit_history_workbook())
    parsed = _parse_unit_history(path, "TABYS", "TABYS")
    codes = {issue.code for issue in parsed.issues}
    assert "UNIT-01" in codes
    assert "UNIT-02" in codes
    assert "UNIT-04" in codes
    assert "UNIT-05" in codes
    assert "UNIT-06" in codes
    assert "UNIT-07" in codes
    assert "UNIT-03" not in codes  # only fires for key == "SAQ"


def test_unit_history_saq_series_is_flagged_stale_regardless_of_data_quality(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Форма"
    sheet.cell(4, 2, date(2026, 6, 1)); sheet.cell(4, 3, 1000); sheet.cell(4, 4, 5)
    sheet.cell(4, 5, 100); sheet.cell(4, 6, 1.5); sheet.cell(4, 7, 47_000); sheet.cell(4, 8, 470)
    buffer = BytesIO()
    workbook.save(buffer)
    path = tmp_path / "unit-history-saq.xlsx"
    path.write_bytes(buffer.getvalue())
    parsed = _parse_unit_history(path, "SAQ", "SAQ")
    codes = {issue.code for issue in parsed.issues}
    assert "UNIT-03" in codes
    assert parsed.summary["active"] is False


def test_accounting_landing_flags_that_formula_scanning_is_unavailable_for_xls(tmp_path, monkeypatch):
    """.xls accounting sources are read with python_calamine, which exposes
    only computed cell values, not formula source text - so the formula-error
    and external-link scans that run for .xlsx (ACCOUNTING-01/03) cannot run
    at all here. ACCOUNTING-04 makes that gap an explicit, visible finding
    instead of a silently absent one."""
    class _FakeSheet:
        def to_python(self):
            return [["Бюджет 2026"], ["Отчётная дата", "20.07.2026"]]

    class _FakeWorkbook:
        def __init__(self, *_args, **_kwargs):
            pass

        def get_sheet_by_name(self, _name):
            return _FakeSheet()

        @classmethod
        def from_path(cls, _path):
            return cls()

    monkeypatch.setattr("osip_dashboard.ingestion.multi_source.CalamineWorkbook", _FakeWorkbook)
    detection = {"sheets": ["Бюджет"], "source_type": "accounting_budget_landing"}
    parsed = _parse_accounting_landing(Path("fake.xls"), detection, "budget", "ACCOUNTING")
    codes = {issue.code for issue in parsed.issues}
    assert "ACCOUNTING-04" in codes
    assert "ACCOUNTING-00" in codes


def test_accounting_landing_flags_formula_errors_and_external_links_for_xlsx(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Бюджет"
    sheet["A1"] = "бюджет 2026"
    sheet["A2"] = "=A1+#REF!"
    sheet["A3"] = "=[external.xlsx]Sheet1!A1"
    path = tmp_path / "budget.xlsx"
    workbook.save(path)
    detection = {"sheets": ["Бюджет"], "source_type": "accounting_budget_landing"}
    parsed = _parse_accounting_landing(path, detection, "budget", "ACCOUNTING")
    codes = {issue.code for issue in parsed.issues}
    assert "ACCOUNTING-01" in codes
    assert "ACCOUNTING-03" in codes
    assert "ACCOUNTING-04" not in codes  # only applies to the .xls branch



def test_client_open_dates_flags_duplicates_unmatched_and_ambiguous_names(tmp_path):
    workbook = Workbook()
    лист4 = workbook.active
    лист4.title = "Лист4"
    лист4["B7"] = "ACC-1"; лист4["C7"] = "Иванов Иван"
    лист4["B8"] = "ACC-2"; лист4["C8"] = "Иванов Иван"  # same name on a second account -> ambiguous match
    лист4["B9"] = "ACC-3"; лист4["C9"] = "Петров Пётр"
    лист6 = workbook.create_sheet("Лист6")
    лист6["A2"] = "Иванов Иван"; лист6["B2"] = date(2020, 1, 1)
    лист6["A3"] = "Иванов Иван"; лист6["B3"] = date(2020, 1, 1)  # duplicate normalized name within Лист6
    лист6["A4"] = "Сидоров Сидор"; лист6["B4"] = date(2021, 1, 1)  # no match at all in Лист4
    path = tmp_path / "client-open-dates.xlsx"; workbook.save(path)
    parsed = _parse_client_brokerage(path, "open_dates", "BROKERAGE")
    codes = {issue.code for issue in parsed.issues}
    assert "CLIENT-02" in codes
    assert "CLIENT-03" in codes
    assert "CLIENT-04" in codes
    assert parsed.summary["unmatched"] == 1
    assert parsed.summary["ambiguous"] == 2


def _corporate_finance_workbook(*, issuer: str = "АО Тест", duplicate: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр"
    sheet["A1"] = "Направление корпоративного финансирования"
    sheet["B2"] = "1H2026"
    # Header labels mirror sources/Направление_Корпфин_01072026.xlsx (verified
    # this session) - the parser resolves columns by this text, not position.
    sheet.append([
        None, "ЭМИТЕНТ", "ПРЕДМЕТ ДОГОВОРА", "ОБЪЕМ РАЗМЕЩЕНИЯ", "ФАКТИЧЕСКИ УДОВЛЕТВОРЕННЫЙ СПРОС",
        "ИНВЕСТОРЫ PORTFOLIO OPS INSIGHT", "СТАВКА КОМИССИОННОГО ВОЗНАГРАЖДЕНИЯ",
        "РАЗМЕР ПОЛУЧЕННОГО ВОЗНАГРАЖДЕНИЯ (KZT)", "ДЛИТЕЛЬНОСТЬ ПРОЕКТА",
    ])
    sheet.append([])
    row = [1, issuer, "Размещение KZ0000000001", "1.5B KZT", "500M KZT", "3 инвестора", "0.25", "1250000", "90 дней, проект действующий"]
    sheet.append(row)
    if duplicate:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _risk_workbook_sobstv() -> bytes:
    """Minimal SOBSTV risk-limits workbook covering all 5 extracted dimensions.

    Column positions mirror the real file exactly (verified against
    ``sources/Риски_Собств_Лимиты на 01.07.26.xls`` this session) - see
    ``_parse_risk_sobstv``'s docstring for the per-sheet shape. One
    deliberately ambiguous "Лимит на Отрасль" row (blank label) checks that
    RISK-02 fires instead of silently dropping it.
    """
    workbook = Workbook()
    limits = workbook.active
    limits.title = "Лимиты"
    limits.cell(1, 3, "СОБСТВЕННЫЙ")
    limits.cell(4, 3, "Перечень инструментов инвестирования")
    limits.cell(4, 6, "Утвержденный лимит в денежном эквиваленте")
    limits.cell(5, 7, "Фактические значения (% от собственных активов под управлением)")
    limits.cell(5, 8, "Фактические значения в денежном эквиваленте")
    limits.cell(5, 9, "Свободный лимит в денежном эквиваленте")
    limits.cell(6, 2, 1)
    limits.cell(6, 3, "Деньги, обратное РЕПО, вклады, в том числе:")
    limits.cell(6, 6, 1000000)
    limits.cell(6, 7, 0.5)
    limits.cell(6, 8, 500000)
    limits.cell(6, 9, 500000)
    # A numeric category marker with no label text - must be skipped with a
    # RISK-02 finding rather than silently classified under an empty name.
    limits.cell(9, 2, 2)

    countries = workbook.create_sheet("Лимит по странам")
    countries.cell(1, 2, "СОБСТВЕННЫЙ")
    countries.cell(2, 2, "на утро")
    countries.cell(2, 3, date(2026, 7, 1))
    countries.cell(3, 3, "Страна")
    countries.cell(3, 5, "Лимит, долл.США")
    countries.cell(3, 6, "Лимит, тенге")
    countries.cell(3, 11, "Фактическое освоение, долл.США")
    countries.cell(3, 12, "Свободный лимит, долл.США")
    countries.cell(5, 2, 1)
    countries.cell(5, 3, "ГЕРМАНИЯ")
    countries.cell(5, 5, 1000.0)
    countries.cell(5, 6, 480000.0)
    countries.cell(5, 11, 0.0)
    countries.cell(5, 12, 1000.0)
    countries.cell(5, 21, 0)

    issuers = workbook.create_sheet("Лимит на Эмитента")
    issuers.cell(1, 2, "портфель")
    issuers.cell(1, 3, "СОБСТВЕННЫЙ")
    issuers.cell(2, 2, "на начало")
    issuers.cell(2, 3, date(2026, 7, 1))
    issuers.cell(5, 2, "Наименование")
    issuers.cell(5, 3, "Инвестировано в тенге")
    issuers.cell(5, 4, "Лимит в тенге")
    issuers.cell(5, 5, "Инвестировано в % от базы лимита")
    issuers.cell(5, 6, "Лимит в % от базы лимита")
    issuers.cell(5, 8, "База лимита")
    issuers.cell(5, 9, "Свободный лимит в тенге")
    issuers.cell(5, 11, "Сигнал")
    issuers.cell(6, 2, 'АО "Тест"')
    issuers.cell(6, 3, 100000)
    issuers.cell(6, 4, 500000)
    issuers.cell(6, 5, 0.2)
    issuers.cell(6, 6, 1.0)
    issuers.cell(6, 7, "не более")
    issuers.cell(6, 8, "Капитал")
    issuers.cell(6, 9, 400000)
    issuers.cell(6, 11, "OK")

    sectors = workbook.create_sheet("Лимит на Отрасль")
    sectors.cell(1, 2, "портфель")
    sectors.cell(1, 3, "СОБСТВЕННЫЙ")
    sectors.cell(2, 2, "на начало")
    sectors.cell(2, 3, date(2026, 7, 1))
    sectors.cell(5, 2, "Наименование")
    sectors.cell(5, 3, "Инвестировано в тенге")
    sectors.cell(5, 4, "Лимит в тенге")
    sectors.cell(5, 5, "Инвестировано в % от базы лимита")
    sectors.cell(5, 6, "Лимит в % от базы лимита")
    sectors.cell(5, 8, "База лимита")
    sectors.cell(5, 9, "Свободный лимит в тенге")
    sectors.cell(5, 11, "Сигнал")
    sectors.cell(6, 2, "Banks")
    sectors.cell(6, 3, 50000)
    sectors.cell(6, 4, 500000)
    sectors.cell(6, 5, 0.1)
    sectors.cell(6, 6, 1.0)
    sectors.cell(6, 7, "не более")
    sectors.cell(6, 8, "СИП")
    sectors.cell(6, 9, 450000)
    sectors.cell(6, 11, "OK")
    sectors.cell(9, 2, "Итого:")
    sectors.cell(9, 3, 500000)
    sectors.cell(12, 2, "Banks")
    sectors.cell(12, 3, 50000)
    sectors.cell(12, 5, 0.1)

    ifrs = workbook.create_sheet("Лимит по МСФО")
    ifrs.cell(1, 2, "портфель")
    ifrs.cell(1, 3, "СОБСТВЕННЫЙ")
    ifrs.cell(2, 2, "на начало")
    ifrs.cell(2, 3, date(2026, 7, 1))
    ifrs.cell(3, 3, "Финансовые инструменты в соответствии с бизнес-моделью по МСФО:")
    ifrs.cell(4, 4, "%")
    ifrs.cell(4, 5, "В денежном эквиваленте")
    ifrs.cell(4, 6, "%")
    ifrs.cell(4, 7, "В денежном эквиваленте")
    ifrs.cell(4, 8, "В денежном эквиваленте")
    ifrs.cell(5, 2, 1)
    ifrs.cell(5, 3, "Учитываемые по справедливой стоимости")
    ifrs.cell(5, 4, 1.0)
    ifrs.cell(5, 5, 1000000)
    ifrs.cell(5, 6, 0.5)
    ifrs.cell(5, 7, 500000)
    ifrs.cell(5, 8, 500000)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _risk_workbook_tabys() -> bytes:
    """Minimal TABYS risk-limits workbook (single sheet, 4 relevant sections)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Пр2-16"
    sheet.cell(1, 2, "Отчет о соблюдении (использовании) установленных лимитов инвестирования ")
    sheet.cell(2, 4, "ИПИФ  «TABYS CAPITAL»")
    sheet.cell(3, 2, 'под управлением АО "PORTFOLIO OPS INSIGHT"')
    sheet.cell(4, 2, "за 30.06.26")
    sheet.cell(8, 2, "Классификация инвестиции")
    sheet.cell(8, 7, "Сигнал")
    sheet.cell(9, 3, "Установленный лимит (%)")
    sheet.cell(9, 4, "Фактический размер инвестирования (в % от активов)")
    sheet.cell(9, 5, "Установленный лимит")
    sheet.cell(9, 6, "Фактический размер инвестирования")
    sheet.cell(11, 2, "По стране")
    sheet.cell(12, 2, "ИРЛАНДИЯ")
    sheet.cell(12, 3, 1.0)
    sheet.cell(12, 4, 0.11)
    sheet.cell(12, 5, 65978492.25)
    sheet.cell(12, 6, 7784282.06)
    sheet.cell(12, 7, "OK")
    sheet.cell(23, 2, "По эмитенту")
    sheet.cell(24, 2, "SPDR")
    sheet.cell(24, 3, 0.19)
    sheet.cell(24, 4, 0.04)
    sheet.cell(24, 5, 12535913.53)
    sheet.cell(24, 6, 2587460.39)
    sheet.cell(24, 7, "OK")
    sheet.cell(38, 2, "По виду финансового инструмента")
    sheet.cell(39, 2, "Паи")
    sheet.cell(39, 3, 1.0)
    sheet.cell(39, 4, 0.82)
    sheet.cell(39, 5, 65978492.25)
    sheet.cell(39, 6, 54015355.20)
    sheet.cell(39, 7, "OK")
    sheet.cell(49, 2, "По GICS отраслям")
    sheet.cell(50, 2, "Diversified Banks")
    sheet.cell(50, 3, 1.0)
    sheet.cell(50, 4, 0.005)
    sheet.cell(50, 5, 65978492.25)
    sheet.cell(50, 6, 325181.46)
    sheet.cell(50, 7, "OK")
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _accounting_budget_workbook() -> bytes:
    """Minimal Бюджет sheet: one real income-statement line, one empty
    cash-flow line (mirrors the real file having labels but no values yet),
    and one balance line - column positions match the real file exactly
    (verified against sources/Бухгалтерия_Бюджет 2026.xlsx this session).
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Бюджет"
    sheet.cell(1, 1, "ПРИЛОЖЕНИЕ 2 - БЮДЖЕТ НА 2021 ГОД")
    sheet.cell(4, 1, "ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ, в тыс тг")
    sheet.cell(4, 20, "2026")  # matches the real file's incidental month/2026 header text used for detection
    sheet.cell(6, 1, "Процентные доходы")
    for column, value in zip((2, 3, 4, 6, 8, 9, 10, 11, 12, 13, 14), (100, 110, 90, 95, 120, 10, 11, 12, 130, 1.05, 10)):
        sheet.cell(6, column, value)
    sheet.cell(9, 1, "ОДДС, в тыс тг")
    sheet.cell(11, 1, "Процентные доходы")  # every value cell left blank on purpose
    sheet.cell(14, 1, "БАЛАНС, в тыс тг")
    sheet.cell(16, 1, "Касса и корр.счета")
    for column, value in zip((2, 3, 4, 6, 8, 9, 10, 11, 12, 13, 14), (500, 520, 480, 510, 600, 50, 51, 52, 610, 1.02, 10)):
        sheet.cell(16, column, value)
    sheet.cell(66, 6, date(2025, 9, 30))  # matches _parse_accounting_budget's fixed business-date cell
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _accounting_portfolio_workbook() -> bytes:
    """Minimal ОСИП_ПОРТФЕЛЬ sheet: two categories, one position each, plus
    the "ПО ПОРТФЕЛЮ" grand-total row. Column positions match the real file
    exactly (verified against sources/Бухгалтерия_Портфель.xls this session).
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ОСИП_ПОРТФЕЛЬ"
    workbook.create_sheet("Лист1")
    sheet.cell(1, 11, date(2026, 7, 17))
    # Header labels mirror sources/Бухгалтерия_Портфель.xls (verified this
    # session) - the parser resolves columns by this text, not position.
    sheet.cell(3, 7, "Код ценной бумаги"); sheet.cell(3, 8, "НИН"); sheet.cell(3, 9, "Тип ценной бумаги")
    sheet.cell(3, 10, "Эмитент"); sheet.cell(3, 12, "Ставка купона/ репо"); sheet.cell(3, 13, "Номинальная стоимость")
    sheet.cell(3, 14, "Валюта инструмета"); sheet.cell(3, 18, "Количество (шт.)"); sheet.cell(3, 20, "Цена покупки")
    sheet.cell(3, 23, "Объем покупки в тенге"); sheet.cell(3, 30, "Балансовая стоимость, в тенге")
    sheet.cell(3, 31, "Рыночная стоимость в ТЕНГЕ на отчетную дату"); sheet.cell(3, 32, "Сумма резерва, в тенге")
    sheet.cell(3, 46, "Накопленный купон в ТЕНГЕ / Начисленное вознаграждение по депозиту")
    sheet.cell(4, 1, "Корпоративные облигации")
    sheet.cell(5, 1, 1)
    sheet.cell(5, 7, "BOND1"); sheet.cell(5, 8, "US0000000001"); sheet.cell(5, 9, "Корпоративные облигации")
    sheet.cell(5, 10, "Test Issuer"); sheet.cell(5, 12, 0.05); sheet.cell(5, 13, 1000); sheet.cell(5, 14, "USD")
    sheet.cell(5, 18, 100); sheet.cell(5, 20, 99.5); sheet.cell(5, 23, 45000000)
    sheet.cell(5, 30, 46000000); sheet.cell(5, 31, 46500000); sheet.cell(5, 32, 50000); sheet.cell(5, 46, 300000)
    sheet.cell(7, 1, "ETF")
    sheet.cell(8, 1, 2)
    sheet.cell(8, 7, "ETF1"); sheet.cell(8, 8, "US9999999999"); sheet.cell(8, 9, "АКЦИИ")
    sheet.cell(8, 10, "Test ETF Trust"); sheet.cell(8, 13, 1); sheet.cell(8, 14, "USD")
    sheet.cell(8, 18, 50); sheet.cell(8, 20, 400.0); sheet.cell(8, 23, 9000000)
    sheet.cell(8, 30, 9200000); sheet.cell(8, 31, 9300000)
    sheet.cell(10, 1, "ПО ПОРТФЕЛЮ")
    sheet.cell(10, 30, 55200000); sheet.cell(10, 31, 55800000); sheet.cell(10, 32, 50000)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _accounting_statements_workbook(*, unbalance: bool = False, period_date: str = "01.07.2026 года", code_31_label: str = "Кредиторская задолженность", add_code: tuple[str, str] | None = None) -> bytes:
    """Minimal f1_uip/f2_uip workbook mirroring the real Бухгалтерия_ФО layout.

    Column positions (label/code/period columns) match the real file exactly
    (verified against ``sources/Бухгалтерия_ФО 2-2026.xlsx`` this session).
    ``unbalance`` breaks the assets total so ACCOUNTING-BS-01 can be tested.
    ``period_date``/``code_31_label``/``add_code`` let a test build a second,
    later period with a relabeled or newly-introduced line code, to exercise
    the account-code registry's drift/new-code detection without needing a
    second real workbook fixture.
    """
    workbook = Workbook()
    bs = workbook.active
    bs.title = "f1_uip"
    bs.cell(2, 1, "Бухгалтерский баланс")
    bs.cell(3, 1, f"по состоянию на {period_date}")
    bs.cell(11, 1, "Активы")
    bs.cell(12, 1, "Денежные средства"); bs.cell(12, 2, "1"); bs.cell(12, 3, 80187); bs.cell(12, 4, 63945)
    bs.cell(13, 1, "в том числе:")
    bs.cell(14, 1, "наличные деньги в кассе"); bs.cell(14, 2, "1.1"); bs.cell(14, 3, 0); bs.cell(14, 4, 0)
    bs.cell(15, 1, "Итого активы"); bs.cell(15, 2, "25"); bs.cell(15, 3, 4200000 if not unbalance else 9999999); bs.cell(15, 4, 4000000)
    bs.cell(16, 1, "Обязательства")
    bs.cell(17, 1, code_31_label); bs.cell(17, 2, "31"); bs.cell(17, 3, 1435); bs.cell(17, 4, 28176)
    bs.cell(18, 1, "Итого обязательства"); bs.cell(18, 2, "42"); bs.cell(18, 3, 200000); bs.cell(18, 4, 171372)
    bs.cell(19, 1, "Собственный капитал")
    bs.cell(20, 1, "Уставный капитал"); bs.cell(20, 2, "43"); bs.cell(20, 3, 3500000); bs.cell(20, 4, 3500000)
    bs.cell(21, 1, "Итого капитал"); bs.cell(21, 2, "52"); bs.cell(21, 3, 4000000); bs.cell(21, 4, 3828628)
    bs.cell(22, 1, "Итого капитал и обязательства"); bs.cell(22, 2, "53"); bs.cell(22, 3, 4200000); bs.cell(22, 4, 4000000)
    note_row = 23
    if add_code is not None:
        new_code, new_label = add_code
        bs.cell(23, 1, new_label); bs.cell(23, 2, new_code); bs.cell(23, 3, 500); bs.cell(23, 4, 0)
        note_row = 24
    bs.cell(note_row, 1, "Примечание")

    is_sheet = workbook.create_sheet("f2_uip")
    is_sheet.cell(3, 1, "Отчет о прибылях и убытках")
    is_sheet.cell(4, 1, "по состоянию на 01.07.2026 года")
    is_sheet.cell(12, 1, "Доходы, связанные с получением вознаграждения"); is_sheet.cell(12, 2, "1")
    is_sheet.cell(12, 3, 160325); is_sheet.cell(12, 4, 318469); is_sheet.cell(12, 5, 168322); is_sheet.cell(12, 6, 323460)
    is_sheet.cell(13, 1, "в том числе:")
    is_sheet.cell(14, 1, "Итого доходов"); is_sheet.cell(14, 2, "13")
    is_sheet.cell(14, 3, 287651); is_sheet.cell(14, 4, 695279); is_sheet.cell(14, 5, 446018); is_sheet.cell(14, 6, 1016651)
    is_sheet.cell(15, 1, "Операционные расходы"); is_sheet.cell(15, 2, "26")
    is_sheet.cell(15, 3, 128326); is_sheet.cell(15, 4, 240156); is_sheet.cell(15, 5, 127845); is_sheet.cell(15, 6, 238028)
    is_sheet.cell(16, 1, "Итого расходов"); is_sheet.cell(16, 2, "28")
    is_sheet.cell(16, 3, 184920); is_sheet.cell(16, 4, 388685); is_sheet.cell(16, 5, 365752); is_sheet.cell(16, 6, 714003)
    is_sheet.cell(17, 1, "Чистая прибыль (убыток) до уплаты корпоративного подоходного налога"); is_sheet.cell(17, 2, "29")
    is_sheet.cell(17, 3, 102731); is_sheet.cell(17, 4, 306594); is_sheet.cell(17, 5, 80266); is_sheet.cell(17, 6, 302648)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_corporate_finance_flags_ambiguous_amount_missing_isin_and_missing_rate(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр"
    sheet["A1"] = "Направление корпоративного финансирования"
    sheet["B2"] = "1H2026"
    sheet.append([
        None, "ЭМИТЕНТ", "ПРЕДМЕТ ДОГОВОРА", "ОБЪЕМ РАЗМЕЩЕНИЯ", "ФАКТИЧЕСКИ УДОВЛЕТВОРЕННЫЙ СПРОС",
        "ИНВЕСТОРЫ PORTFOLIO OPS INSIGHT", "СТАВКА КОМИССИОННОГО ВОЗНАГРАЖДЕНИЯ",
        "РАЗМЕР ПОЛУЧЕННОГО ВОЗНАГРАЖДЕНИЯ (KZT)", "ДЛИТЕЛЬНОСТЬ ПРОЕКТА",
    ])
    sheet.append([])
    # "1000000 KZT" has no B/M magnitude suffix -> ambiguous placement amount.
    # The subject has no ISIN. The commission rate (column 7) is left blank.
    row = [1, "АО Тест", "Размещение без ISIN", "1000000 KZT", "500M KZT", "3 инвестора", None, "1250000", "90 дней, действующий"]
    sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    path = tmp_path / "corpfin.xlsx"
    path.write_bytes(buffer.getvalue())
    parsed = _parse_corporate_finance(path, "CORPFIN")
    codes = {issue.code for issue in parsed.issues}
    assert "CORPFIN-01" in codes
    assert "CORPFIN-02" in codes
    assert "CORPFIN-03" in codes
    assert parsed.records[0]["payload"]["isins"] == []


def test_parse_amount_recognizes_the_cyrillic_billion_abbreviation():
    """Regression: "МЛРД" starts with the same Cyrillic "М" the million
    branch matches on its own, so an unordered check silently read a
    billion as a million (1000x understatement) while still reporting the
    amount as unambiguously resolved - the fixed code checks "МЛРД" before
    falling through to the million branch."""
    amount, currency, ok = _parse_amount("5 млрд KZT")
    assert amount == "5000000000"
    assert currency == "KZT"
    assert ok is True
    amount, currency, ok = _parse_amount("10 МЛРД USD")
    assert amount == "10000000000"
    assert currency == "USD"
    assert ok is True
    # The plain-million case must still resolve to a million, not regress.
    amount, currency, ok = _parse_amount("500М KZT")
    assert amount == "500000000"
    assert ok is True


def test_corporate_finance_placement_in_cyrillic_billions_is_not_flagged_ambiguous(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Реестр"
    sheet["A1"] = "Направление корпоративного финансирования"
    sheet["B2"] = "1H2026"
    sheet.append([
        None, "ЭМИТЕНТ", "ПРЕДМЕТ ДОГОВОРА", "ОБЪЕМ РАЗМЕЩЕНИЯ", "ФАКТИЧЕСКИ УДОВЛЕТВОРЕННЫЙ СПРОС",
        "ИНВЕСТОРЫ PORTFOLIO OPS INSIGHT", "СТАВКА КОМИССИОННОГО ВОЗНАГРАЖДЕНИЯ",
        "РАЗМЕР ПОЛУЧЕННОГО ВОЗНАГРАЖДЕНИЯ (KZT)", "ДЛИТЕЛЬНОСТЬ ПРОЕКТА",
    ])
    sheet.append([])
    row = [1, "АО Тест", "Размещение облигаций", "5 млрд KZT", "2 млрд KZT", "3 инвестора", "US0000000001", "1250000", "90 дней"]
    sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    path = tmp_path / "corpfin.xlsx"
    path.write_bytes(buffer.getvalue())
    parsed = _parse_corporate_finance(path, "CORPFIN")
    codes = {issue.code for issue in parsed.issues}
    assert "CORPFIN-01" not in codes
    assert parsed.records[0]["payload"]["placement_amount"] == "5000000000"
    assert parsed.records[0]["payload"]["satisfied_demand"] == "2000000000"


def test_risk_country_row_with_no_signal_and_no_actual_value_raises_risk_01(tmp_path):
    """RISK-02 (ambiguous label) already has coverage; RISK-01 - a row whose
    OK/breach status can't be derived because there's no source signal text
    AND no actual value to compare against the limit - did not."""
    path = tmp_path / "risk-sobstv.xlsx"
    path.write_bytes(_risk_workbook_sobstv())
    workbook = load_workbook(path)
    countries = workbook["Лимит по странам"]
    countries.cell(9, 2, 2)
    countries.cell(9, 3, "ФРАНЦИЯ")
    countries.cell(9, 6, 480_000.0)  # limit_kzt present, actual_usd/actual_kzt/signal_raw all absent
    workbook.save(path)
    parsed = parse_detected_dataset(path, detect_source(path, "xlsx"), "limits", "SOBSTV")
    france = next(record for record in parsed.records if record["payload"].get("label") == "ФРАНЦИЯ")
    assert france["payload"]["signal"] is None
    assert any(issue.code == "RISK-01" for issue in parsed.issues)


def test_risk_sobstv_parser_tolerates_missing_optional_dimension_sheets(tmp_path):
    # Лимит на Эмитента/Отрасль/МСФО are optional dimensions the same way
    # Лимит по дюрации/Расшифровка/Detail already are (each independently
    # present or absent in a real workbook) - a workbook with only the two
    # sheets detection itself requires must still parse, not raise
    # WorksheetNotFound for one of the three that used to be read
    # unconditionally.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лимиты"
    sheet["B1"] = 1
    workbook.create_sheet("Лимит по странам")
    path = tmp_path / "risk-minimal.xlsx"
    workbook.save(path)

    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "risk_limits_sobstv"
    parsed = parse_detected_dataset(path, detection, "limits", "SOBSTV")
    assert parsed.records == []


def test_detection_ignores_unrelated_extra_sheets(tmp_path):
    # A stray leftover/template/hidden sheet alongside a real contract's
    # sheets must not confuse detection - every check here is "is the
    # required sheet present", never "are these the ONLY sheets", so extra
    # sheets should already be harmless. Locking that in explicitly.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "f1_uip"
    sheet["A1"] = "Бухгалтерский баланс"
    workbook.create_sheet("f2_uip")
    workbook.create_sheet("Черновик")
    workbook.create_sheet("Sheet1 (2)")
    path = tmp_path / "fo-with-extras.xlsx"
    workbook.save(path)

    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "accounting_statements"
    assert {item["key"] for item in detection["datasets"]} == {"balance_sheet", "income_statement"}


def test_лист1_discriminates_accounting_portfolio_landing_from_standalone_osip(tmp_path):
    # "Лист1" alongside "ОСИП_ПОРТФЕЛЬ" is a real discriminator, not dead
    # weight like fund_unit_history's was: it distinguishes a portfolio
    # sheet arriving inside an accounting package (accounting_landing +
    # accounting_portfolio_detail) from a standalone legacy OSIP portfolio
    # file (portfolio_snapshot, imported through a different, dedicated
    # flow entirely - see raise in materialize_datasets). Getting this
    # wrong would silently misclassify one contract as the other.
    with_лист1 = Workbook()
    with_лист1.active.title = "ОСИП_ПОРТФЕЛЬ"
    with_лист1.create_sheet("Лист1")
    path_with = tmp_path / "portfolio-in-accounting-package.xlsx"
    with_лист1.save(path_with)
    detection_with = detect_source(path_with, "xlsx")
    assert detection_with["source_type"] == "accounting_portfolio_landing"
    assert {item["key"] for item in detection_with["datasets"]} == {"portfolio", "portfolio_detail"}

    without_лист1 = Workbook()
    without_лист1.active.title = "ОСИП_ПОРТФЕЛЬ"
    path_without = tmp_path / "standalone-osip-portfolio.xlsx"
    without_лист1.save(path_without)
    detection_without = detect_source(path_without, "xlsx")
    assert detection_without["source_type"] == "osip_portfolio"
    assert {item["key"] for item in detection_without["datasets"]} == {"portfolio"}


def test_risk_parser_covers_duration_exposure_currency_and_fx_sections(tmp_path):
    sobstv_path = tmp_path / "risk-sobstv.xlsx"
    sobstv_path.write_bytes(_risk_workbook_sobstv())
    workbook = load_workbook(sobstv_path)
    duration = workbook.create_sheet("Лимит по дюрации")
    duration.cell(2, 3, date(2026, 7, 1))
    duration.cell(3, 10, "Утвержденный лимит на дюрацию, не более")
    duration.cell(3, 11, "Модифицированная дьюрация")
    duration.cell(3, 13, "Торговый код")
    duration.cell(3, 14, "ISIN")
    duration.cell(3, 15, "Наименование эмитента")
    duration.cell(3, 16, "Балансовая стоимость в валюте ЦБ")
    duration.cell(4, 10, 7.0)
    duration.cell(4, 11, 3.0)
    duration.cell(4, 13, "MUM")
    duration.cell(4, 14, "KZ0000000001")
    duration.cell(4, 15, "МинФин РК")
    duration.cell(4, 16, 100000)
    exposure = workbook.create_sheet("Расшифровка")
    exposure.append(["№", "№ группы", "Тип инструмента", "ФИ", "Валюта", "Сумма в валюте", "Курс", "Сумма в тенге"])
    exposure.append([1, 1, "Деньги", "SOBSTV, USD, Account", "USD", 10, 480, 4800])
    detail = workbook.create_sheet("Detail")
    detail.cell(1, 1, "КАЗАХСТАН")
    detail.cell(1, 4, "KZT")
    detail.cell(2, 1, "Деньги, обратное РЕПО, вклады в финансовые институты")
    detail.cell(2, 4, 12345.67)
    detail.cell(3, 1, 999.0)
    detail.cell(3, 2, "ДС СОБСТВЕННЫЙ, KZT, KZ000, АО Банк")
    workbook.save(sobstv_path)
    parsed = parse_detected_dataset(sobstv_path, detect_source(sobstv_path, "xlsx"), "limits", "SOBSTV")
    dimensions = {record["payload"]["dimension"] for record in parsed.records}
    assert {"duration", "exposure_detail", "country_instrument_detail"}.issubset(dimensions)
    assert parsed.summary["duration_count"] == 1
    assert parsed.summary["exposure_detail_count"] == 1
    assert parsed.summary["country_detail_count"] == 1
    assert parsed.summary["unknown_count"] == 0
    detail_record = next(record for record in parsed.records if record["payload"]["dimension"] == "country_instrument_detail")
    assert detail_record["payload"]["country"] == "КАЗАХСТАН"
    assert detail_record["payload"]["currency"] == "KZT"
    assert detail_record["payload"]["amount_native"] == "12345.67"

    # A country row's breach is driven by USD columns (it carries no KZT
    # actual/limit), so field_columns must point a reviewer at the USD cells
    # specifically rather than only the label cell.
    country_record = next(record for record in parsed.records if record["payload"]["dimension"] == "country")
    country_field_columns = country_record["source_ref"]["field_columns"]
    assert "actual_usd" in country_field_columns
    assert country_field_columns["actual_usd"]["source_cell"] != country_record["source_ref"]["source_cell"]

    duration_record = next(record for record in parsed.records if record["payload"]["dimension"] == "duration")
    assert "modified_duration" in duration_record["source_ref"]["field_columns"]

    tabys_path = tmp_path / "risk-tabys.xlsx"
    tabys_path.write_bytes(_risk_workbook_tabys())
    workbook = load_workbook(tabys_path)
    sheet = workbook["Пр2-16"]
    for row_number, section, label in ((57, "По валюте", "USD"), (60, "По открытой валютной позиции", "Иностранная валюта (USD)"), (63, "По виду финансового инструмента одного эмитента", "Акции SPDR")):
        sheet.cell(row_number, 2, section)
        sheet.cell(row_number + 1, 2, label)
        sheet.cell(row_number + 1, 3, 0.5)
        sheet.cell(row_number + 1, 4, 0.2)
        sheet.cell(row_number + 1, 5, 1000)
        sheet.cell(row_number + 1, 6, 200)
        sheet.cell(row_number + 1, 7, "OK")
    workbook.save(tabys_path)
    parsed = parse_detected_dataset(tabys_path, detect_source(tabys_path, "xlsx"), "limits", "TABYS")
    dimensions = {record["payload"]["dimension"] for record in parsed.records}
    assert {"currency", "fx_position", "instrument_issuer"}.issubset(dimensions)
    assert parsed.summary["unknown_count"] == 0


def _client_trade_workbook() -> bytes:
    """A spaced-header trade sheet guarding against positional column drift."""
    workbook = Workbook()
    client_sheet = workbook.active
    client_sheet.title = "Лист4"
    client_sheet["A2"] = "Реестр в разрезе а/счетов по состоянию на 20.07.2026 г."
    sheet = workbook.create_sheet("Лист8")
    sheet["B8"] = "№ п/п"
    sheet["C8"] = "Номер клиентского заказа"
    sheet["E8"] = "Дата заключения сделки"
    sheet["G8"] = "Место заключения сделки"
    sheet["J8"] = "Наименование/ Ф.И.О. клиента"
    sheet["M8"] = "№ лицевого счета"
    sheet["P8"] = "Сведения о контрагенте"
    sheet["S8"] = "Вид сделки"
    sheet["U8"] = "Эмитент"
    sheet["X8"] = "Вид ЦБ"
    sheet["AA8"] = "НИН / ISIN"
    sheet["AD8"] = "Количество ЦБ"
    sheet["AG8"] = "Сумма сделки"
    sheet["AJ8"] = "Валюта сделки"
    sheet["AL8"] = "Чистая цена одной ЦБ, %"
    sheet["AO8"] = "Грязная цена одной ЦБ, в валюте сделки"
    sheet["AQ8"] = "Доходность, % годовых"
    sheet["AS8"] = "Дата исполнения обязательств по сделке"
    sheet["AU8"] = "Отметка об исполнении/ не исполнении"
    sheet["AV8"] = "Причина неисполнения"
    sheet["B10"] = 7
    sheet["E10"] = "20.07.2026 09:00:00"
    sheet["G10"] = "KASE"
    sheet["J10"] = "Клиент"
    sheet["M10"] = "ACC-1"
    sheet["S10"] = "Покупка"
    sheet["U10"] = "Эмитент"
    sheet["X10"] = "ГЦБ"
    sheet["AA10"] = "KZ0000000001"
    sheet["AD10"] = 12
    sheet["AG10"] = 1_200_000
    sheet["AJ10"] = "KZT"
    sheet["AL10"] = 99.5
    sheet["AO10"] = 100.1
    sheet["AQ10"] = 9.25
    sheet["AS10"] = "21.07.2026 09:00:00"
    sheet["AU10"] = "исполненная"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_trade_parser_uses_headers_instead_of_spacer_positions(tmp_path):
    path = tmp_path / "client.xlsx"
    path.write_bytes(_client_trade_workbook())
    parsed = _parse_client_brokerage(path, "trades", "BROKERAGE")
    assert parsed.summary["mapping"]["confidence"] == "high"
    assert parsed.summary["gross_turnover_by_currency"] == {"KZT": "1200000"}
    assert parsed.summary["instrument_mix"] == {"ГЦБ": 1}
    assert parsed.summary["execution_status_mix"] == {"исполненная": 1}
    assert parsed.records[0]["payload"]["quantity"] == "12"
    assert parsed.records[0]["payload"]["amount"] == "1200000"
    assert parsed.records[0]["payload"]["clean_price"] == "99.5"
    assert parsed.records[0]["payload"]["dirty_price"] == "100.1"
    # Column-level provenance follows the same header-driven mapping used to
    # parse the row, so a reviewer can jump straight to the amount/price cell
    # instead of only the trade-number cell.
    field_columns = parsed.records[0]["source_ref"]["field_columns"]
    assert "amount" in field_columns and "clean_price" in field_columns
    assert "is_repo" not in field_columns


def test_trade_ledger_flags_a_required_header_that_could_not_be_mapped(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Лист4"
    sheet = workbook.create_sheet("Лист8")
    # Every required header except ISIN ("НИН / ISIN") - left out on purpose.
    sheet["B8"] = "№ п/п"
    sheet["E8"] = "Дата заключения сделки"
    sheet["G8"] = "Место заключения сделки"
    sheet["J8"] = "Наименование/ Ф.И.О. клиента"
    sheet["M8"] = "№ лицевого счета"
    sheet["S8"] = "Вид сделки"
    sheet["U8"] = "Эмитент"
    sheet["X8"] = "Вид ЦБ"
    sheet["AD8"] = "Количество ЦБ"
    sheet["AG8"] = "Сумма сделки"
    sheet["AJ8"] = "Валюта сделки"
    sheet["AU8"] = "Отметка об исполнении/ не исполнении"
    path = tmp_path / "client.xlsx"; workbook.save(path)
    parsed = _parse_client_brokerage(path, "trades", "BROKERAGE")
    assert any(issue.code == "BROKERAGE-MAP-01" for issue in parsed.issues)
    assert parsed.summary["mapping"]["missing_fields"] == ["isin"]
    assert parsed.summary["mapping"]["confidence"] == "low"


def test_client_brokerage_detection_needs_only_лист4(tmp_path):
    # A real-world revision of this workbook ("1.1") reorganized/renamed
    # Лист8/Лист7/Лист6 into differently-named sheets while leaving Лист4
    # untouched - detection must still recognize the workbook and offer
    # whatever sheets are actually present, not reject the whole file for
    # missing sheets it never needed to build client_account_snapshot.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "List1"
    list4 = workbook.create_sheet("Лист4")
    list4["B6"] = "Л/счет"
    workbook.create_sheet("календарь погашения")
    workbook.create_sheet("Клиенты")
    del workbook["List1"]
    path = tmp_path / "client-v1.1.xlsx"
    workbook.save(path)

    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "client_brokerage"
    assert {item["key"] for item in detection["datasets"]} == {"clients", "maturity_calendar", "client_dashboard"}


def test_accounting_statements_detection_needs_only_f1_uip(tmp_path):
    # f1_uip (balance sheet) and f2_uip (income statement) are parsed
    # independently of each other - a workbook missing f2_uip should still
    # be recognized and offer the balance sheet alone, not be rejected.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "f1_uip"
    sheet["A1"] = "Бухгалтерский баланс"
    path = tmp_path / "fo-balance-only.xlsx"
    workbook.save(path)

    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "accounting_statements"
    assert {item["key"] for item in detection["datasets"]} == {"balance_sheet"}


def test_tabys_valuation_detection_needs_only_дата(tmp_path):
    # "часть 1 (портфель)"/"Цены"/"справка о ст-ти ЧА" each back exactly one
    # key and nothing else reads them - a workbook missing all three should
    # still be recognized off "дата" alone and offer only the keys that
    # only ever need "дата" itself.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "дата"
    sheet["A1"] = "Текущая стоимость портфеля ЦБ"
    path = tmp_path / "tabys-minimal.xlsx"
    workbook.save(path)

    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "tabys_valuation"
    assert {item["key"] for item in detection["datasets"]} == {"valuation", "cash_liabilities", "inactive_evidence"}


def test_client_parser_ignores_account_subtotals_and_grand_total(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист4"
    # Header labels mirror sources/Клиентский_дашборд 1.1.xlsx (verified this
    # session) - the parser resolves columns by this text, not position.
    sheet["B4"] = "Л/счет"; sheet["C4"] = "Клиент"
    sheet["AC4"] = "Итого по всем валютам в тенге"; sheet["AD4"] = "Итого стоимость активов в тенге"
    sheet["B7"] = "ACC-1"; sheet["C7"] = "Клиент 1"; sheet["AC7"] = 100; sheet["AD7"] = 1000
    sheet["B8"] = "Итого"; sheet["AC8"] = 999; sheet["AD8"] = 9999
    sheet["B9"] = "ВСЕГО по счету"; sheet["AC9"] = 1099; sheet["AD9"] = 10999
    path = tmp_path / "client.xlsx"; workbook.save(path)
    parsed = _parse_client_brokerage(path, "clients", "BROKERAGE")
    assert parsed.summary["client_count"] == 1
    assert parsed.summary["cash_kzt"] == "100"
    assert parsed.summary["total_assets_kzt"] == "1000"
    client_record = next(record for record in parsed.records if record["record_type"] == "client")
    assert client_record["source_ref"]["field_columns"]["client_name"]["source_cell"] == "C7"


def test_client_parser_joins_manager_from_clients_sheet_when_register_column_is_blank(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист4"
    sheet["B4"] = "Л/счет"; sheet["C4"] = "Клиент"
    sheet["AC4"] = "Итого по всем валютам в тенге"; sheet["AD4"] = "Итого стоимость активов в тенге"
    sheet["B7"] = "ACC-1"; sheet["C7"] = "Клиент 1"; sheet["AC7"] = 100; sheet["AD7"] = 1000
    dashboard = workbook.create_sheet("Клиенты")
    dashboard.append(["КЛИЕНТЫ — ДЕНЬГИ И АКТИВЫ"])
    dashboard.append(["№", "Клиент", "Менеджер"])
    dashboard.append([1, "Клиент 1", "Менеджер 1"])
    path = tmp_path / "client.xlsx"; workbook.save(path)

    parsed = _parse_client_brokerage(path, "clients", "BROKERAGE")

    client_record = next(record for record in parsed.records if record["record_type"] == "client")
    assert client_record["payload"]["manager"] == "Менеджер 1"
    assert client_record["payload"]["branch"] == ""
    assert parsed.summary["manager_mix"]["Менеджер 1"]["client_count"] == 1


def test_client_parser_flags_a_client_with_no_cash_total(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист4"
    sheet["B4"] = "Л/счет"; sheet["C4"] = "Клиент"
    sheet["AC4"] = "Итого по всем валютам в тенге"; sheet["AD4"] = "Итого стоимость активов в тенге"
    sheet["B7"] = "ACC-1"; sheet["C7"] = "Клиент 1"
    sheet["AD7"] = None  # widen the row so row[28] (cash, column AC) exists as an explicit blank, not a missing index
    path = tmp_path / "client.xlsx"; workbook.save(path)
    parsed = _parse_client_brokerage(path, "clients", "BROKERAGE")
    assert any(issue.code == "CLIENT-01" for issue in parsed.issues)


def test_corporate_finance_deal_has_field_level_column_provenance(tmp_path):
    path = tmp_path / "corpfin.xlsx"
    path.write_bytes(_corporate_finance_workbook())
    parsed = _parse_corporate_finance(path, "CORPFIN")
    field_columns = parsed.records[0]["source_ref"]["field_columns"]
    assert field_columns["placement_amount"]["source_cell"] == field_columns["placement_raw"]["source_cell"]
    assert field_columns["issuer"]["source_cell"] != field_columns["fee_received_kzt"]["source_cell"]


def test_derivatives_parser_excludes_etf_rows_and_flags_the_exclusion(tmp_path):
    # "Лист7" is nominally the derivatives register, but in real source
    # files most of its rows turn out to be plain ETF purchases (no strike,
    # no expiry) that duplicate trades already correctly recorded on the
    # trade ledger - an ETF unit is direct fund ownership, not a derivative.
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лист7"
    sheet["A9"] = "1H2026"
    for row_number, (number, instrument_type) in enumerate(((1, "Опцион"), (2, "ETF")), 15):
        sheet.cell(row_number, 1, number)
        sheet.cell(row_number, 2, "20.07.2026")
        sheet.cell(row_number, 4, "21.07.2026")
        sheet.cell(row_number, 5, instrument_type)
        sheet.cell(row_number, 6, f"ID{number}")
        sheet.cell(row_number, 7, "неорганизованный рынок")
        sheet.cell(row_number, 8, "Underlying")
        sheet.cell(row_number, 10, "Counterparty")
        sheet.cell(row_number, 11, "Покупка")
        sheet.cell(row_number, 12, 10)
        sheet.cell(row_number, 13, 100)
        sheet.cell(row_number, 14, 1000)
        sheet.cell(row_number, 15, "USD")
        sheet.cell(row_number, 23, "статус")
    path = tmp_path / "client.xlsx"
    workbook.save(path)
    parsed = _parse_client_brokerage(path, "derivatives", "BROKERAGE")
    assert parsed.summary["derivative_count"] == 1
    assert parsed.summary["excluded_etf_rows"] == 1
    assert [record["payload"]["instrument_type"] for record in parsed.records] == ["Опцион"]
    assert any(issue.code == "DERIV-01" for issue in parsed.issues)


def test_corporate_isin_extraction_removes_literal_isin_prefix():
    assert _extract_isins("Series 22: ISINXS3363342927, ISINUS48129VAH96") == ["US48129VAH96", "XS3363342927"]


def test_accounting_date_requires_an_explicit_report_label():
    rows = [
        ["ОСИП Портфель", None, None, None],
        ["Облигация", "KZ0000000001", "Дата погашения", "15.05.2028"],
        ["Отчётная дата", "20.07.2026", None, None],
    ]
    assert _explicit_report_dates(rows) == {date(2026, 7, 20)}


def test_accounting_date_is_unavailable_without_explicit_report_label():
    rows = [["Облигация", "KZ0000000001", "Дата погашения", "15.05.2028"]]
    assert _explicit_report_dates(rows) == set()


def test_future_business_dates_are_not_reported_as_fresh():
    assert _freshness(date(2099, 1, 1)) == "future"


def test_a_period_end_date_within_the_current_month_is_not_flagged_as_future():
    # derivatives_register's report date is computed as the last day of its
    # stated reporting month (_russian_period_end, "Лист7" header) - a report
    # titled for the current month legitimately carries a date later in that
    # same month even before it's over. Only a date outside the current
    # calendar month is a genuinely suspicious future date.
    import calendar

    today = utcnow().date()
    month_end = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    assert _freshness(month_end) == "fresh"
    next_month = today.month % 12 + 1
    next_month_year = today.year + (1 if today.month == 12 else 0)
    assert _freshness(date(next_month_year, next_month, 1)) == "future"


def test_maturity_calendar_is_parsed_as_independent_source_dataset(tmp_path):
    workbook = Workbook()
    master = workbook.active
    master.title = "Лист4"
    master["A1"] = "Дата отчёта"; master["B1"] = date(2026, 7, 20)
    calendar = workbook.create_sheet("календарь погашения")
    calendar.append(["Клиент", "Менеджер", "Бумага", "ISIN", "Дата погашения", "Дата выплаты купона", "Купон %", "Дней до погашения", "Стоимость"])
    calendar.append(["Клиент 1", "Менеджер 1", "Bond", "KZ0000000001", date(2026, 8, 1), date(2026, 8, 1), 5.5, 12, 100000])
    path = tmp_path / "client.xlsx"; workbook.save(path)
    parsed = _parse_client_brokerage(path, "maturity_calendar", "BROKERAGE")
    assert parsed.dataset_type == "client_maturity_calendar"
    assert parsed.summary["event_count"] == 1
    assert parsed.records[0]["payload"]["maturity_date"] == "2026-08-01"
    assert parsed.records[0]["payload"]["value_kzt"] == "100000"


def test_client_dashboard_summary_preserves_manager_source_and_register_difference(tmp_path):
    workbook = Workbook()
    master = workbook.active; master.title = "Лист4"
    master["B7"] = "ВСЕГО по счету"; master["AC7"] = 100; master["AD7"] = 1000
    dashboard = workbook.create_sheet("Клиенты")
    dashboard.append(["КЛИЕНТЫ — ДЕНЬГИ И АКТИВЫ"])
    dashboard.append(["№", "Клиент", "Менеджер", "Тип", "Дата открытия", "Остаток денег", "Стоимость бумаг", "Итого активов", "Доля денег", "Доход", "Статус", "Ключ"])
    dashboard.append([1, "Клиент 1", "Менеджер 1", "Юрлицо", date(2020, 1, 1), 100, 950, 1050, 0.1, 0.02, "Активен", "ACC-1"])
    path = tmp_path / "client.xlsx"; workbook.save(path)
    parsed = _parse_client_brokerage(path, "client_dashboard", "BROKERAGE")
    assert parsed.summary["manager_mix"]["Менеджер 1"]["client_count"] == 1
    assert any(issue.code == "CLIENT-DASH-01" for issue in parsed.issues)


def test_trade_dataset_requires_explicit_confirmation_for_low_confidence_mapping(api):
    session = api.app.state.session_factory()
    try:
        upload = SourceUpload(
            source_sha256="c" * 64,
            original_filename="client.xlsx",
            storage_key="client.xlsx",
            file_format="xlsx",
            detected_source_type="client_brokerage",
            detection={},
            uploader_id="uploader",
        )
        dataset = DatasetVersion(
            source_upload=upload,
            dataset_type="brokerage_trade_ledger",
            detected_key="trades",
            scope_type="business_domain",
            scope_code="BROKERAGE",
            source_report_date=date(2026, 7, 20),
            business_date=date(2026, 7, 20),
            parser_version="test",
            version=1,
            status=ImportStatus.VALIDATED,
            summary={"mapping": {"confidence": "low", "mapping_confirmed": False}},
            uploader_id="uploader",
        )
        session.add(dataset)
        session.flush()
        with pytest.raises(ValueError, match="сопоставление столбцов"):
            approve_dataset(session, dataset, "reviewer", "Проверено", [], mapping_confirmed=False)
        approve_dataset(session, dataset, "reviewer", "Сопоставление проверено вручную", [], mapping_confirmed=True)
        assert dataset.status == ImportStatus.APPROVED
        assert dataset.summary["mapping"]["mapping_confirmed"] is True
    finally:
        session.rollback()
        session.close()


def test_content_detection_independent_child_workflow_and_export(api):
    content = _corporate_finance_workbook()
    uploaded = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("arbitrary-name.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    source = uploaded.json()
    assert source["detected_source_type"] == "corporate_finance"
    assert source["datasets"][0]["dataset_type"] == "corporate_finance_register"

    duplicate = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("renamed.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert duplicate.status_code == 200
    assert duplicate.json()["id"] == source["id"]

    materialized = api.post(f"/api/v1/source-uploads/{source['id']}/datasets", headers=UPLOADER, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    assert materialized.status_code == 200, materialized.text
    dataset = materialized.json()["items"][0]
    assert dataset["status"] == "validated"
    assert dataset["business_date"] == "2026-06-30"
    assert dataset["summary"]["deal_count"] == 1

    mapping = api.get(f"/api/v1/dataset-versions/{dataset['id']}/mapping", headers=UPLOADER)
    assert mapping.status_code == 200, mapping.text
    assert mapping.json()["fields"]
    confirmed = api.post(
        f"/api/v1/dataset-versions/{dataset['id']}/mapping/confirm",
        headers=UPLOADER,
        json={"comment": "Проверены исходные заголовки и значения"},
    )
    assert confirmed.status_code == 200, confirmed.text
    assert confirmed.json()["mapping_confirmed"] is True

    critical = [item["code"] for item in dataset["issues"] if item["severity"] in {"blocker", "high"}]
    approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=UPLOADER, json={"comment": "Source checked", "acknowledged_dq_codes": critical})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=UPLOADER)
    assert published.status_code == 200, published.text

    overview = api.get("/api/v1/corporate-finance/overview", headers=UPLOADER)
    assert overview.status_code == 200
    assert overview.json()["available"] is True
    assert overview.json()["records"]["corporate_finance_register"][0]["issuer"] == "АО Тест"

    exported = api.get("/api/v1/corporate-finance/export", headers=UPLOADER)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert_workbook_is_compliant(exported.content)
    result = load_workbook(BytesIO(exported.content), data_only=True)
    assert "Корпоративные финансы" in result.sheetnames
    assert result["Корпоративные финансы"]["B2"].value == "corporate_finance_register"

    filtered_export = api.get("/api/v1/corporate-finance/export", params={"term": "АО Тест"}, headers=UPLOADER)
    assert filtered_export.status_code == 200
    filtered_result = load_workbook(BytesIO(filtered_export.content), data_only=True)
    filtered_sheet = filtered_result["Корпоративные финансы"]
    assert filtered_sheet["A9"].value == "АО Тест"
    assert filtered_sheet["A10"].value is None


def test_risk_near_breach_policy_uses_percentage_utilization_not_absolute_headroom():
    """Documented policy: a single global 90% utilization threshold, using
    whichever limit/actual pair is available (%% preferred, then KZT, then
    USD) - never an absolute-headroom cutoff, since limit scale varies too
    much across dimensions for a fixed KZT/USD number to be meaningful."""
    assert RISK_NEAR_BREACH_THRESHOLD == Decimal("0.9")
    assert RISK_NEAR_BREACH_POLICY_VERSION == "utilization-ratio-v1"
    # Below threshold: comfortably within limit.
    assert _risk_utilization(Decimal("1.0"), Decimal("0.5"), None, None) == Decimal("0.5")
    assert _risk_near_breach("OK", Decimal("0.5")) is False
    # At/above threshold but not yet breached: this is exactly "near-breach".
    assert _risk_near_breach("OK", Decimal("0.9")) is True
    assert _risk_near_breach("OK", Decimal("0.95")) is True
    # Already breached: never double-counted as near-breach too.
    assert _risk_near_breach("breach", Decimal("1.2")) is False
    # An "OK" row can still compute a ratio over 100% (the source's own flag
    # overrides a derived comparison, e.g. a supranational issuer exemption)
    # - must defer to that "OK" signal rather than contradict it.
    assert _risk_near_breach("OK", Decimal("3.8")) is False
    # No computable ratio: never guessed as near-breach.
    assert _risk_utilization(None, None, None, None) is None
    assert _risk_near_breach("OK", None) is False
    # KZT falls back to when %% pair is unavailable (SOBSTV country dimension).
    assert _risk_utilization(None, None, Decimal("900"), Decimal("1000")) == Decimal("1000") / Decimal("900")
    # USD is the last resort.
    assert _risk_utilization(None, None, None, None, Decimal("100"), Decimal("91")) == Decimal("91") / Decimal("100")


def test_risk_limits_detected_parsed_published_and_exported_for_both_portfolios(api):
    def publish_risk(content: bytes, filename: str, dataset_type: str, scope_code: str) -> None:
        uploaded = api.post("/api/v1/source-uploads", headers={**UPLOADER, "X-Actor-Domains": "risk"}, files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        source = uploaded.json()
        assert source["detected_source_type"] == dataset_type
        assert source["datasets"][0]["dataset_type"] == dataset_type
        materialized = api.post(f"/api/v1/source-uploads/{source['id']}/datasets", headers={**UPLOADER, "X-Actor-Domains": "risk"}, json={"datasets": [{"detected_key": "limits", "scope_code": scope_code}]})
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        assert dataset["status"] == "validated"
        critical = [item["code"] for item in dataset["issues"] if item["severity"] in {"blocker", "high"}]
        approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers={**UPLOADER, "X-Actor-Domains": "risk"}, json={"comment": "Проверено", "acknowledged_dq_codes": critical})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers={**UPLOADER, "X-Actor-Domains": "risk"})
        assert published.status_code == 200, published.text

    publish_risk(_risk_workbook_sobstv(), "risk-sobstv.xlsx", "risk_limits_sobstv", "SOBSTV")
    publish_risk(_risk_workbook_tabys(), "risk-tabys.xlsx", "risk_limits_tabys", "TABYS")

    headers = {**UPLOADER, "X-Actor-Domains": "risk"}
    overview = api.get("/api/v1/risk/overview", headers=headers)
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert body["available"] is True
    sobstv_records = body["records"]["risk_limits_sobstv"]
    tabys_records = body["records"]["risk_limits_tabys"]
    assert {record["dimension"] for record in sobstv_records} == {"instrument_category", "country", "issuer", "sector", "ifrs"}
    assert {record["dimension"] for record in tabys_records} == {"country", "issuer", "instrument_category", "sector"}
    assert all(record["portfolio_code"] == "SOBSTV" for record in sobstv_records)
    assert all(record["portfolio_code"] == "TABYS" for record in tabys_records)
    forte_issuer = next(record for record in sobstv_records if record["dimension"] == "issuer")
    assert forte_issuer["signal"] == "OK"
    assert forte_issuer["limit_kzt"] == "500000.0"
    assert Decimal(forte_issuer["utilization"]) == Decimal("0.2")
    assert forte_issuer["near_breach"] is False
    assert forte_issuer["near_breach_threshold"] == "0.9"
    assert body["summaries"]["risk_limits_sobstv"]["near_breach_threshold"] == "0.9"
    assert body["summaries"]["risk_limits_sobstv"]["near_breach_count"] == 0

    history = {entry["business_date"]: entry for entry in body["history"]}
    assert history["2026-07-01"]["sobstv"]["breach_count"] == body["summaries"]["risk_limits_sobstv"]["breach_count"]
    assert "tabys" not in history["2026-07-01"]
    assert history["2026-06-30"]["tabys"]["breach_count"] == body["summaries"]["risk_limits_tabys"]["breach_count"]
    assert "sobstv" not in history["2026-06-30"]
    assert history["2026-07-01"]["breach_count"] == history["2026-07-01"]["sobstv"]["breach_count"]
    utilization_history = body["risk_utilization_history"]
    assert utilization_history
    forte_utilization = next(item for item in utilization_history if item["portfolio_code"] == "SOBSTV" and item["dimension"] == "issuer")
    assert forte_utilization["label"] == forte_issuer["label"]
    assert Decimal(forte_utilization["utilization"]) == Decimal("0.2")

    denied_history = api.get("/api/v1/risk/overview", headers={**READER, "X-Actor-Domains": "risk"})
    assert denied_history.status_code == 200
    assert denied_history.json()["history"] == []

    sobstv_dataset = api.get("/api/v1/dataset-versions", params={"dataset_type": "risk_limits_sobstv", "scope_code": "SOBSTV"}, headers=headers)
    assert sobstv_dataset.status_code == 200, sobstv_dataset.text
    sobstv_issue_codes = {issue["code"] for item in sobstv_dataset.json()["items"] for issue in item.get("issues", [])}
    assert "RISK-02" in sobstv_issue_codes
    sobstv_version_id = sobstv_dataset.json()["items"][0]["id"]
    pinned = api.get(f"/api/v1/risk/overview?dataset_versions={sobstv_version_id}", headers=headers)
    assert pinned.status_code == 200, pinned.text
    assert pinned.json()["pinned_dataset_types"] == ["risk_limits_sobstv"]
    assert pinned.json()["sources"]

    denied = api.get("/api/v1/risk/overview", headers={**READER, "X-Actor-Domains": "corpfin"})
    assert denied.status_code == 403

    exported = api.get("/api/v1/risk/export", headers=headers)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert_workbook_is_compliant(exported.content)

    readiness = api.get("/api/v1/operations/source-readiness", headers=headers)
    assert readiness.status_code == 200, readiness.text
    register = {(item["dataset_type"], item["scope_code"]): item for item in readiness.json()["readiness"]}
    sobstv_entry = register[("risk_limits_sobstv", "SOBSTV")]
    assert sobstv_entry["business_date"] == "2026-07-01"
    assert sobstv_entry["sla_days"] == 3
    assert sobstv_entry["dq_blocker_count"] == 0
    assert sobstv_entry["status"] == _readiness_status(date(2026, 7, 1), 3, 0)
    # A dataset_type nobody has published yet is "missing", not silently
    # absent from the register.
    assert register[("fund_valuation", "TABYS")]["status"] == "missing"

    # Uploader-scoped like every other read: a reader who never published
    # anything sees the same dataset_types as "missing", not the uploader's data.
    denied_readiness = api.get("/api/v1/operations/source-readiness", headers={**READER, "X-Actor-Domains": "risk"})
    denied_register = {(item["dataset_type"], item["scope_code"]): item for item in denied_readiness.json()["readiness"]}
    assert denied_register[("risk_limits_sobstv", "SOBSTV")]["status"] == "missing"


def test_readiness_status_covers_ready_due_overdue_blocked_and_unavailable():
    today = utcnow().date()
    assert _readiness_status(today, 3, 0) == "ready"
    assert _readiness_status(today - timedelta(days=3), 3, 0) == "ready"
    assert _readiness_status(today - timedelta(days=5), 3, 0) == "due"
    assert _readiness_status(today - timedelta(days=10), 3, 0) == "overdue"
    # A blocker-severity DQ finding always wins, even on a fresh version.
    assert _readiness_status(today, 3, 1) == "blocked"
    assert _readiness_status(None, 3, 0) == "unavailable"
    assert _readiness_status(None, 3, 1) == "blocked"


def test_readiness_and_freshness_use_the_local_date_not_utc(monkeypatch):
    """Regression: this app runs on a single machine configured for the
    business's own timezone (confirmed UTC+5, Kazakhstan), not a UTC-
    everywhere distributed service - services/dividends.py already used
    date.today() for its "as of" calculations for that reason, but
    _readiness_status/_freshness used utcnow().date() instead, which lags
    the real local date for several hours every night (Kazakhstan midnight
    to ~5am, while UTC's date hasn't rolled over yet) - silently
    under-counting every staleness/overdue check by a day during that
    window. Proven here by deliberately breaking utcnow() and checking the
    result still matches the real local date, not the broken one."""
    from datetime import datetime, timezone

    import osip_dashboard.services.multi_source as multi_source_service

    real_today = date.today()

    def wrong_utcnow() -> datetime:
        # One day ahead of the real date - chosen together with the two
        # boundary dates below so that, if either function still called
        # utcnow() at all, the computed age would cross a category
        # threshold (ready->due, fresh->aging) and produce a different
        # string, not just a numerically-off age that happens to land in
        # the same bucket.
        return datetime.combine(real_today + timedelta(days=1), datetime.min.time(), tzinfo=timezone.utc)

    monkeypatch.setattr(multi_source_service, "utcnow", wrong_utcnow)
    # age == sla_days is the last "ready" day; utcnow()+1 would push it to "due".
    assert _readiness_status(real_today - timedelta(days=3), sla_days=3, blocker_count=0) == "ready"
    # age == 1 is the last "fresh" day; utcnow()+1 would push it to "aging".
    assert _freshness(real_today - timedelta(days=1)) == "fresh"


def test_readiness_register_scope_covers_landing_and_saq_unit_history():
    """Regression: an earlier draft of _all_dataset_type_scopes omitted
    accounting_landing (a real, independently published dataset_type - see
    _parse_accounting_landing) and fund_unit_series/SAQ (a distinct scope
    from TABYS - see docs/domain-upload-instructions.md's "TABYS or SAQ
    unit-value history" row), so both silently never appeared in the
    readiness register despite having real published data."""
    all_pairs = {pair for pairs in _all_dataset_type_scopes().values() for pair in pairs}
    assert ("accounting_landing", "ACCOUNTING") in all_pairs
    assert ("fund_unit_series", "SAQ") in all_pairs
    assert ("fund_unit_series", "TABYS") in all_pairs


def test_accounting_balance_sheet_and_income_statement_reconcile_cleanly(tmp_path):
    path = tmp_path / "fo.xlsx"
    path.write_bytes(_accounting_statements_workbook())
    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "accounting_statements"

    bs = parse_detected_dataset(path, detection, "balance_sheet")
    assert bs.business_date == date(2026, 7, 1)
    assert not bs.issues
    codes = {record["payload"]["line_code"] for record in bs.records}
    assert {"1", "1.1", "25", "31", "42", "43", "52", "53"}.issubset(codes)
    total_assets = next(record for record in bs.records if record["payload"]["line_code"] == "25")
    assert total_assets["payload"]["current_period_kzt"] == "4200000"
    assert bs.summary["total_assets_kzt"] == "4200000"

    income_statement = parse_detected_dataset(path, detection, "income_statement")
    assert income_statement.business_date == date(2026, 7, 1)
    assert not income_statement.issues
    assert income_statement.summary["net_profit_kzt"] == "102731"


def test_accounting_unbalanced_balance_sheet_raises_accounting_bs_01(tmp_path):
    path = tmp_path / "fo-unbalanced.xlsx"
    path.write_bytes(_accounting_statements_workbook(unbalance=True))
    detection = detect_source(path, "xlsx")
    bs = parse_detected_dataset(path, detection, "balance_sheet")
    assert any(issue.code == "ACCOUNTING-BS-01" for issue in bs.issues)


def test_accounting_statements_detected_parsed_published_and_exported(api):
    uploaded = api.post("/api/v1/source-uploads", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, files={"file": ("fo.xlsx", _accounting_statements_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    source = uploaded.json()
    assert source["detected_source_type"] == "accounting_statements"
    assert {item["dataset_type"] for item in source["datasets"]} == {"accounting_balance_sheet", "accounting_income_statement"}

    materialized = api.post(
        f"/api/v1/source-uploads/{source['id']}/datasets", headers={**UPLOADER, "X-Actor-Domains": "accounting"},
        json={"datasets": [
            {"detected_key": "balance_sheet", "scope_code": "ACCOUNTING"},
            {"detected_key": "income_statement", "scope_code": "ACCOUNTING"},
        ]},
    )
    assert materialized.status_code == 200, materialized.text
    for dataset in materialized.json()["items"]:
        assert dataset["status"] == "validated"
        approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, json={"comment": "Проверено", "acknowledged_dq_codes": []})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers={**UPLOADER, "X-Actor-Domains": "accounting"})
        assert published.status_code == 200, published.text

    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    overview = api.get("/api/v1/accounting/source-readiness", headers=headers)
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert body["available"] is True
    assert body["summaries"]["accounting_balance_sheet"]["total_assets_kzt"] == "4200000"
    assert body["summaries"]["accounting_income_statement"]["net_profit_kzt"] == "102731"

    denied = api.get("/api/v1/accounting/source-readiness", headers={**READER, "X-Actor-Domains": "corpfin"})
    assert denied.status_code == 403

    exported = api.get("/api/v1/accounting/export", headers=headers)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert_workbook_is_compliant(exported.content)


def test_accounting_budget_parses_real_lines_and_honestly_empty_cash_flow(tmp_path):
    path = tmp_path / "budget.xlsx"
    path.write_bytes(_accounting_budget_workbook())
    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "accounting_budget_landing"
    assert {item["key"] for item in detection["datasets"]} == {"budget", "budget_detail"}

    budget = parse_detected_dataset(path, detection, "budget_detail")
    assert budget.business_date == date(2025, 9, 30)
    assert not budget.issues
    sections = {record["payload"]["section"] for record in budget.records}
    assert sections == {"income_statement", "balance"}
    assert budget.summary["cash_flow_line_count"] == 0
    income_line = next(record for record in budget.records if record["payload"]["section"] == "income_statement")
    assert income_line["payload"]["forecast_2025_kzt"] == "130"
    assert income_line["payload"]["oct_2025_kzt"] == "10"


def test_accounting_portfolio_detail_parses_positions_and_grand_total(tmp_path):
    path = tmp_path / "portfolio.xlsx"
    path.write_bytes(_accounting_portfolio_workbook())
    detection = detect_source(path, "xlsx")
    assert detection["source_type"] == "accounting_portfolio_landing"
    assert {item["key"] for item in detection["datasets"]} == {"portfolio", "portfolio_detail"}

    portfolio = parse_detected_dataset(path, detection, "portfolio_detail")
    assert portfolio.business_date == date(2026, 7, 17)
    assert not portfolio.issues
    assert {record["payload"]["category"] for record in portfolio.records} == {"Корпоративные облигации", "ETF"}
    assert portfolio.summary["total_carrying_value_kzt"] == "55200000.0"
    bond = next(record for record in portfolio.records if record["payload"]["isin"] == "US0000000001")
    assert bond["payload"]["carrying_value_kzt"] == "46000000.0"


def test_accounting_portfolio_cash_rows_are_extracted_and_reconcile_the_grand_total(tmp_path):
    """Regression: the "по..." block below the last position category holds
    cash-by-bank/currency rows (same generator/shape as the standalone OSIP
    workbook's own cash rows - see osip_workbook.CASH_PATTERN). These used
    to be silently dropped: the grand total row legitimately includes them,
    so the position table's own sum fell short of the displayed total with
    no explanation anywhere on the page (found by hand-verifying a real
    workbook's numbers end to end - the gap was exactly the cash sum)."""
    workbook = load_workbook(BytesIO(_accounting_portfolio_workbook()))
    sheet = workbook["ОСИП_ПОРТФЕЛЬ"]
    # Insert two cash rows between the last position (row 8) and the
    # existing "ПО ПОРТФЕЛЮ" total row (row 10) - column 30 is the same
    # "Балансовая стоимость, в тенге" column the position rows use.
    sheet.cell(9, 1, "ОСТАТОК ДЕНЕЖНЫХ СРЕДСТВ В KZT в SAMPLE BANK")
    sheet.cell(9, 30, 500000)
    sheet.cell(10, 1, "ОСТАТОК ДЕНЕЖНЫХ СРЕДСТВ В USD в SAXOBANK")
    sheet.cell(10, 30, 300000)
    sheet.cell(11, 1, "ПО ПОРТФЕЛЮ")
    sheet.cell(11, 30, 56000000)
    content = BytesIO()
    workbook.save(content)
    path = tmp_path / "portfolio-with-cash.xlsx"
    path.write_bytes(content.getvalue())
    detection = detect_source(path, "xlsx")

    portfolio = parse_detected_dataset(path, detection, "portfolio_detail")
    assert not portfolio.issues
    positions = [r for r in portfolio.records if r["record_type"] == "portfolio_position"]
    cash = [r for r in portfolio.records if r["record_type"] == "cash_balance"]
    assert len(positions) == 2
    assert [(c["payload"]["currency"], c["payload"]["custodian"], c["payload"]["amount_kzt"]) for c in cash] == [
        ("KZT", "SAMPLE BANK", "500000.0"),
        ("USD", "SAXOBANK", "300000.0"),
    ]
    assert portfolio.summary["position_count"] == 2
    assert portfolio.summary["cash_kzt"] == "800000.0"
    positions_sum = sum(Decimal(r["payload"]["carrying_value_kzt"]) for r in positions)
    cash_sum = sum(Decimal(c["payload"]["amount_kzt"]) for c in cash)
    assert positions_sum + cash_sum == Decimal(portfolio.summary["total_carrying_value_kzt"])


def test_accounting_export_keeps_cash_balances_out_of_the_position_table(api):
    """Regression: the Excel export's "Детализация портфеля" sheet builds
    its rows from every accounting_portfolio_detail record without
    filtering by record_type. Once cash_balance records existed (previous
    test), that sheet started including them as extra rows where every
    column except "Валюта" read "Недоступно" (cash_balance's payload has
    no category/isin/carrying_value_kzt etc.) - confirmed on the real
    export before this fix. Cash belongs in its own small table on the
    summary sheet instead (mirrors the web UI's separate panel)."""
    workbook = load_workbook(BytesIO(_accounting_portfolio_workbook()))
    sheet = workbook["ОСИП_ПОРТФЕЛЬ"]
    sheet.cell(9, 1, "ОСТАТОК ДЕНЕЖНЫХ СРЕДСТВ В KZT в SAMPLE BANK")
    sheet.cell(9, 30, 500000)
    sheet.cell(10, 1, "ПО ПОРТФЕЛЮ")
    sheet.cell(10, 30, 55700000)
    content = BytesIO()
    workbook.save(content)

    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("portfolio-with-cash.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=headers, json={"datasets": [{"detected_key": "portfolio_detail", "scope_code": "ACCOUNTING"}]})
    assert materialized.status_code == 200, materialized.text
    dataset = materialized.json()["items"][0]
    approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=headers)
    assert published.status_code == 200, published.text

    exported = api.get("/api/v1/accounting/export", headers=headers)
    assert exported.status_code == 200, exported.text
    workbook_out = load_workbook(BytesIO(exported.content))

    detail_rows = list(workbook_out["Детализация портфеля"].iter_rows(min_row=9, values_only=True))
    assert len(detail_rows) == 2, "cash_balance rows must not appear in the position table"
    assert all(row[0] != "Недоступно" for row in detail_rows)

    summary_rows = list(workbook_out["Сводка ФО"].iter_rows(values_only=True))
    cash_header_row = next(i for i, row in enumerate(summary_rows) if row and row[0] == "Денежные средства портфеля")
    assert summary_rows[cash_header_row + 2][:2] == ("KZT · SAMPLE BANK", 500000)


def test_publish_is_blocked_when_a_consumed_formula_has_no_saved_result(api):
    workbook = load_workbook(BytesIO(_accounting_portfolio_workbook()))
    # AD is the parser-backed carrying-value column. openpyxl intentionally
    # cannot create a cached formula result, which models an unsafe upload.
    workbook["ОСИП_ПОРТФЕЛЬ"]["AD5"] = "=1+1"
    content = BytesIO()
    workbook.save(content)
    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    uploaded = api.post(
        "/api/v1/source-uploads", headers=headers,
        files={"file": ("unsafe-portfolio.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert uploaded.status_code == 201, uploaded.text
    materialized = api.post(
        f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=headers,
        json={"datasets": [{"detected_key": "portfolio_detail", "scope_code": "ACCOUNTING"}]},
    )
    assert materialized.status_code == 200, materialized.text
    dataset = materialized.json()["items"][0]
    assert any(issue["code"] == "FORMULA-01" for issue in dataset["issues"])
    approved = api.post(
        f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=headers,
        json={"comment": "Формула проверена", "acknowledged_dq_codes": ["FORMULA-01"]},
    )
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=headers)
    assert published.status_code == 409
    assert "результаты формул" in published.json()["detail"]


def test_accounting_budget_and_portfolio_detected_published_and_exported(api):
    def publish(content: bytes, filename: str, detected_key: str, dataset_type: str) -> None:
        uploaded = api.post("/api/v1/source-uploads", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        source = uploaded.json()
        assert dataset_type in {item["dataset_type"] for item in source["datasets"]}
        materialized = api.post(f"/api/v1/source-uploads/{source['id']}/datasets", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, json={"datasets": [{"detected_key": detected_key, "scope_code": "ACCOUNTING"}]})
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, json={"comment": "Проверено", "acknowledged_dq_codes": []})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers={**UPLOADER, "X-Actor-Domains": "accounting"})
        assert published.status_code == 200, published.text

    publish(_accounting_budget_workbook(), "budget.xlsx", "budget_detail", "accounting_budget")
    publish(_accounting_portfolio_workbook(), "portfolio.xlsx", "portfolio_detail", "accounting_portfolio_detail")

    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    overview = api.get("/api/v1/accounting/source-readiness", headers=headers)
    assert overview.status_code == 200, overview.text
    body = overview.json()
    assert body["available"] is True
    assert body["summaries"]["accounting_budget"]["cash_flow_line_count"] == 0
    assert body["summaries"]["accounting_portfolio_detail"]["total_carrying_value_kzt"] == "55200000.0"
    assert len(body["records"]["accounting_budget"]) == 2
    assert len(body["records"]["accounting_portfolio_detail"]) == 2

    exported = api.get("/api/v1/accounting/export", headers=headers)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert_workbook_is_compliant(exported.content)


def test_accounting_statement_types_can_be_pinned_independently_of_each_other(api):
    """The four accounting dataset types each come from their own physical
    workbook (docs/phase-2-groundwork-risk-accounting.md) - pinning an older
    balance-sheet version via dataset_versions must not disturb the budget
    dataset, which stays on its own latest published version."""
    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}

    def publish_statements(content: bytes, codes: list[str]) -> str:
        uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("fo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(
            f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=headers,
            json={"datasets": [{"detected_key": "balance_sheet", "scope_code": "ACCOUNTING"}, {"detected_key": "income_statement", "scope_code": "ACCOUNTING"}]},
        )
        assert materialized.status_code == 200, materialized.text
        balance_sheet = next(item for item in materialized.json()["items"] if item["dataset_type"] == "accounting_balance_sheet")
        approved = api.post(f"/api/v1/dataset-versions/{balance_sheet['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": codes})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{balance_sheet['id']}/publish", headers=headers)
        assert published.status_code == 200, published.text
        # income_statement is a sibling dataset from the same upload - not
        # this test's concern, but it must still reach a terminal state so
        # it doesn't dangle as "validated" and confuse the next assertion.
        income_statement = next(item for item in materialized.json()["items"] if item["dataset_type"] == "accounting_income_statement")
        api.post(f"/api/v1/dataset-versions/{income_statement['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": codes})
        api.post(f"/api/v1/dataset-versions/{income_statement['id']}/publish", headers=headers)
        return balance_sheet["id"]

    def publish_budget() -> None:
        uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("budget.xlsx", _accounting_budget_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=headers, json={"datasets": [{"detected_key": "budget_detail", "scope_code": "ACCOUNTING"}]})
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": []})
        api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=headers)

    first_balance_sheet_id = publish_statements(_accounting_statements_workbook(), [])
    publish_statements(_accounting_statements_workbook(unbalance=True), ["ACCOUNTING-BS-01"])
    publish_budget()

    latest = api.get("/api/v1/accounting/source-readiness", headers=headers)
    assert latest.status_code == 200, latest.text
    latest_body = latest.json()
    assert latest_body["pinned_dataset_types"] == []
    latest_balance_sheet_source = next(s for s in latest_body["sources"] if s["dataset_type"] == "accounting_balance_sheet")
    assert latest_balance_sheet_source["dataset_id"] != first_balance_sheet_id

    pinned = api.get(f"/api/v1/accounting/source-readiness?dataset_versions={first_balance_sheet_id}", headers=headers)
    assert pinned.status_code == 200, pinned.text
    pinned_body = pinned.json()
    assert pinned_body["pinned_dataset_types"] == ["accounting_balance_sheet"]
    pinned_balance_sheet_source = next(s for s in pinned_body["sources"] if s["dataset_type"] == "accounting_balance_sheet")
    assert pinned_balance_sheet_source["dataset_id"] == first_balance_sheet_id
    # The budget dataset must be completely unaffected by pinning balance_sheet.
    pinned_budget_source = next(s for s in pinned_body["sources"] if s["dataset_type"] == "accounting_budget")
    latest_budget_source = next(s for s in latest_body["sources"] if s["dataset_type"] == "accounting_budget")
    assert pinned_budget_source["dataset_id"] == latest_budget_source["dataset_id"]


def test_accounting_account_mapping_flags_label_drift_and_new_codes_across_periods(api):
    """account_code_registry is derived purely from published workbooks - no
    external chart of accounts exists to import against (Accounting #6).
    A code relabeled between periods must be flagged (label_drift), and a
    code that only appears in the latest period must be flagged (is_new),
    so a period comparison never silently treats a relabel as unrelated or
    a genuinely new line as a continuation of an old one."""
    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}

    def publish_statements(content: bytes) -> None:
        uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("fo.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(
            f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=headers,
            json={"datasets": [{"detected_key": "balance_sheet", "scope_code": "ACCOUNTING"}, {"detected_key": "income_statement", "scope_code": "ACCOUNTING"}]},
        )
        assert materialized.status_code == 200, materialized.text
        for dataset_type in ("accounting_balance_sheet", "accounting_income_statement"):
            item = next(entry for entry in materialized.json()["items"] if entry["dataset_type"] == dataset_type)
            approved = api.post(f"/api/v1/dataset-versions/{item['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": []})
            assert approved.status_code == 200, approved.text
            published = api.post(f"/api/v1/dataset-versions/{item['id']}/publish", headers=headers)
            assert published.status_code == 200, published.text

    publish_statements(_accounting_statements_workbook(period_date="01.07.2026 года"))
    publish_statements(_accounting_statements_workbook(
        period_date="01.10.2026 года",
        code_31_label="Кредиторская задолженность (краткосрочная)",
        add_code=("60", "Резерв по отпускам"),
    ))

    overview = api.get("/api/v1/accounting/source-readiness", headers=headers)
    assert overview.status_code == 200, overview.text
    registry = {entry["line_code"]: entry for entry in overview.json()["account_mapping"]["accounting_balance_sheet"]}

    assert registry["31"]["label_drift"] is True
    assert [item["label"] for item in registry["31"]["label_history"]] == ["Кредиторская задолженность", "Кредиторская задолженность (краткосрочная)"]
    assert registry["31"]["current_label"] == "Кредиторская задолженность (краткосрочная)"
    assert registry["31"]["is_new"] is False

    assert registry["60"]["is_new"] is True
    assert registry["60"]["label_drift"] is False
    assert registry["60"]["first_seen"] == "2026-10-01"

    assert registry["25"]["label_drift"] is False
    assert registry["25"]["is_new"] is False
    assert registry["25"]["first_seen"] == "2026-07-01"
    assert registry["25"]["last_seen"] == "2026-10-01"


def test_accounting_portfolio_reconciles_against_the_matching_osip_snapshot(api, workbook_paths):
    """reconcile_accounting_portfolio must fire on publish and record a
    real ACCOUNTING-PORTFOLIO reconciliation result, using the OSIP
    portfolio the uploader explicitly selected (reconciliation_portfolio_code)
    - never guessed from the accounting workbook's sheet name."""
    osip_imported = osip_upload(api, workbook_paths["SOBSTV"]).json()
    osip_approve_and_publish(api, osip_imported["id"])

    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("portfolio.xlsx", _accounting_portfolio_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    source = uploaded.json()
    materialized = api.post(
        f"/api/v1/source-uploads/{source['id']}/datasets", headers=headers,
        json={"datasets": [{"detected_key": "portfolio_detail", "scope_code": "ACCOUNTING", "reconciliation_portfolio_code": "SOBSTV"}]},
    )
    assert materialized.status_code == 200, materialized.text
    dataset = materialized.json()["items"][0]
    approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=headers)
    assert published.status_code == 200, published.text

    readiness = api.get("/api/v1/operations/source-readiness", headers=headers)
    assert readiness.status_code == 200, readiness.text
    reconciliations = {item["rule_code"]: item for item in readiness.json()["reconciliations"]}
    assert "ACCOUNTING-PORTFOLIO" in reconciliations
    result = reconciliations["ACCOUNTING-PORTFOLIO"]
    assert result["scope_code"] == "SOBSTV"
    assert result["evidence"]["dates_match"] is False
    assert result["status"] == "date_mismatch"
    assert result["actual_values"]["accounting"] == "55200000.0"


def test_accounting_portfolio_does_not_reconcile_when_no_portfolio_was_selected(api, workbook_paths):
    """No reconciliation_portfolio_code means no ACCOUNTING-PORTFOLIO result -
    the OSIP portfolio a workbook represents is never guessed from its
    sheet name or filename."""
    osip_imported = osip_upload(api, workbook_paths["SOBSTV"]).json()
    osip_approve_and_publish(api, osip_imported["id"])

    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("portfolio.xlsx", _accounting_portfolio_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    source = uploaded.json()
    materialized = api.post(
        f"/api/v1/source-uploads/{source['id']}/datasets", headers=headers,
        json={"datasets": [{"detected_key": "portfolio_detail", "scope_code": "ACCOUNTING"}]},
    )
    dataset = materialized.json()["items"][0]
    api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=headers)
    assert published.status_code == 200, published.text

    readiness = api.get("/api/v1/operations/source-readiness", headers=headers)
    rule_codes = {item["rule_code"] for item in readiness.json()["reconciliations"]}
    assert "ACCOUNTING-PORTFOLIO" not in rule_codes


def test_risk_record_source_preview_reopens_the_exact_workbook_cell(api):
    """Risk records aren't backed by a persisted SourceRow table like OSIP -
    the preview endpoint must re-open the original workbook from blob
    storage instead, using the source_cell/row_number the parser recorded."""
    uploaded = api.post("/api/v1/source-uploads", headers={**UPLOADER, "X-Actor-Domains": "risk"}, files={"file": ("risk-sobstv.xlsx", _risk_workbook_sobstv(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    source = uploaded.json()
    materialized = api.post(f"/api/v1/source-uploads/{source['id']}/datasets", headers={**UPLOADER, "X-Actor-Domains": "risk"}, json={"datasets": [{"detected_key": "limits", "scope_code": "SOBSTV"}]})
    assert materialized.status_code == 200, materialized.text
    dataset = materialized.json()["items"][0]
    critical = [item["code"] for item in dataset["issues"] if item["severity"] in {"blocker", "high"}]
    approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers={**UPLOADER, "X-Actor-Domains": "risk"}, json={"comment": "Проверено", "acknowledged_dq_codes": critical})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers={**UPLOADER, "X-Actor-Domains": "risk"})
    assert published.status_code == 200, published.text

    headers = {**UPLOADER, "X-Actor-Domains": "risk"}
    overview = api.get("/api/v1/risk/overview", headers=headers)
    assert overview.status_code == 200, overview.text
    record = overview.json()["records"]["risk_limits_sobstv"][0]
    ref = record["source"]
    assert ref["source_cell"]

    preview = api.get(f"/api/v1/source-rows/{record['id']}/preview", params={"cell": ref["source_cell"]}, headers=headers)
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["sheet_name"] == ref["sheet_name"]
    assert body["target_cell"] == ref["source_cell"]
    assert body["target_row"] == ref["row_number"]
    assert body["target_value"] == record["label"]
    assert body["original_filename"] == ref["filename"]
    assert any(row["row_number"] == ref["row_number"] for row in body["rows"])
    assert body["header_row"] is not None
    assert any(str(value).strip() for value in body["column_labels"]), "the drawer must be able to label columns by real header text, not bare Excel letters"

    # Regression: body["import_id"] here is a DatasetVersion id, not an
    # ImportBatch id - GET /imports/{import_id}/source only knows
    # ImportBatch rows and 404s on it. The "Download original" button must
    # use body["source_upload_id"] against /source-uploads/{id}/source
    # instead, which this proves actually resolves and serves the file.
    assert body["source_upload_id"] == source["id"]
    original = api.get(f"/api/v1/imports/{body['import_id']}/source", headers=headers)
    assert original.status_code == 404
    downloaded = api.get(f"/api/v1/source-uploads/{body['source_upload_id']}/source", headers=headers)
    assert downloaded.status_code == 200, downloaded.text

    mismatch = api.get(f"/api/v1/source-rows/{record['id']}/preview", params={"cell": "ZZ999"}, headers=headers)
    assert mismatch.status_code == 409

    denied = api.get(f"/api/v1/source-rows/{record['id']}/preview", params={"cell": ref["source_cell"]}, headers={**READER, "X-Actor-Domains": "corpfin"})
    assert denied.status_code == 403


def test_risk_record_preview_includes_header_band_even_when_target_row_is_within_ten(api, tmp_path):
    """A target row <= 10 can still have a window that misses the real header.

    The preview window is centered on the target row (roughly row-4..row+4),
    so a record at row 9 produces a window of rows 5-13 - which never
    reaches back to rows 1-2 where this sheet's header actually lives. The
    "always include the header band" fix must not be gated on the target
    row being > 10; it has to run unconditionally.
    """
    path = tmp_path / "risk-sobstv.xlsx"
    path.write_bytes(_risk_workbook_sobstv())
    workbook = load_workbook(path)
    countries = workbook["Лимит по странам"]
    countries.cell(9, 2, 1)
    countries.cell(9, 3, "ФРАНЦИЯ")
    countries.cell(9, 5, 2000.0)
    countries.cell(9, 6, 960000.0)
    countries.cell(9, 11, 0.0)
    countries.cell(9, 12, 2000.0)
    countries.cell(9, 21, 0)
    workbook.save(path)

    uploaded = api.post("/api/v1/source-uploads", headers={**UPLOADER, "X-Actor-Domains": "risk"}, files={"file": ("risk-sobstv.xlsx", path.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers={**UPLOADER, "X-Actor-Domains": "risk"}, json={"datasets": [{"detected_key": "limits", "scope_code": "SOBSTV"}]})
    dataset = materialized.json()["items"][0]
    critical = [item["code"] for item in dataset["issues"] if item["severity"] in {"blocker", "high"}]
    api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers={**UPLOADER, "X-Actor-Domains": "risk"}, json={"comment": "Проверено", "acknowledged_dq_codes": critical})
    api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers={**UPLOADER, "X-Actor-Domains": "risk"})

    headers = {**UPLOADER, "X-Actor-Domains": "risk"}
    overview = api.get("/api/v1/risk/overview", headers=headers)
    record = next(item for item in overview.json()["records"]["risk_limits_sobstv"] if item.get("label") == "ФРАНЦИЯ")
    ref = record["source"]
    assert ref["row_number"] == 9

    preview = api.get(f"/api/v1/source-rows/{record['id']}/preview", params={"cell": ref["source_cell"]}, headers=headers)
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert any(row["row_number"] == 9 for row in body["rows"])
    assert any(row["row_number"] <= 2 for row in body["rows"]), "header rows must be present even though the target row is <= 10"
    assert body["header_row"] is not None and body["header_row"] <= 10


def test_rejects_temporary_and_invalid_workbooks_before_children(api):
    content = _corporate_finance_workbook()
    temporary = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("~$source.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert temporary.status_code == 422
    invalid = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("source.xlsx", b"not-a-workbook", "application/octet-stream")})
    assert invalid.status_code == 422


def test_duplicate_source_keys_are_retained_with_unique_internal_keys(api):
    content = _corporate_finance_workbook(duplicate=True)
    uploaded = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("duplicate-rows.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    materialized = api.post(
        f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets",
        headers=UPLOADER,
        json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]},
    )
    assert materialized.status_code == 200, materialized.text
    dataset_id = UUID(materialized.json()["items"][0]["id"])
    session = api.app.state.session_factory()
    try:
        keys = list(session.scalars(select(DatasetRecord.record_key).where(DatasetRecord.dataset_id == dataset_id)))
    finally:
        session.close()
    assert len(keys) == 2
    assert len(set(keys)) == 2


def test_same_date_dataset_publication_supersedes_prior_version(api):
    def publish(content: bytes) -> str:
        uploaded = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("version.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(
            f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets",
            headers=UPLOADER,
            json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]},
        )
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        approved = api.post(
            f"/api/v1/dataset-versions/{dataset['id']}/approve",
            headers=UPLOADER,
            json={"comment": "Проверено", "acknowledged_dq_codes": []},
        )
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=UPLOADER)
        assert published.status_code == 200, published.text
        return UUID(dataset["id"])

    first_id = publish(_corporate_finance_workbook(issuer="АО Первый"))
    second_id = publish(_corporate_finance_workbook(issuer="АО Второй"))
    comparison = api.get(f"/api/v1/dataset-versions/{second_id}/compare?with_id={first_id}", headers=UPLOADER)
    assert comparison.status_code == 200, comparison.text
    body = comparison.json()
    assert body["added_count"] == 1
    assert body["removed_count"] == 1
    # "Added" must mean present in the subject ({dataset_id}=second_id, the
    # newer publish with "АО Второй") but not the baseline ({with_id}=
    # first_id, "АО Первый") - not the other way around.
    session = api.app.state.session_factory()
    try:
        second_record_key = session.scalar(select(DatasetRecord.record_key).where(DatasetRecord.dataset_id == second_id))
        first_record_key = session.scalar(select(DatasetRecord.record_key).where(DatasetRecord.dataset_id == first_id))
        assert body["added_keys"] == [second_record_key]
        assert body["removed_keys"] == [first_record_key]
        statuses = dict(session.execute(select(DatasetVersion.id, DatasetVersion.status).where(DatasetVersion.id.in_([first_id, second_id]))).all())
    finally:
        session.close()
    assert statuses[first_id] == ImportStatus.SUPERSEDED
    assert statuses[second_id] == ImportStatus.PUBLISHED


def test_publishing_the_same_date_as_a_different_uploader_does_not_supersede_them(api):
    """Two operators can each have their own published version for the same
    dataset_type/scope_code/business_date - per-uploader visibility means
    operator B's publish must never supersede operator A's publish, since A
    cannot even see B's data to know it happened."""
    def publish(headers: dict, issuer: str) -> UUID:
        uploaded = api.post("/api/v1/source-uploads", headers=headers, files={"file": ("cross-uploader.xlsx", _corporate_finance_workbook(issuer=issuer), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=headers, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=headers, json={"comment": "Проверено", "acknowledged_dq_codes": []})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=headers)
        assert published.status_code == 200, published.text
        return UUID(dataset["id"])

    operator_a = {"X-Actor-Id": "operator-a", "X-Actor-Roles": "uploader,reviewer,publisher,reader"}
    operator_b = {"X-Actor-Id": "operator-b", "X-Actor-Roles": "uploader,reviewer,publisher,reader"}
    first_id = publish(operator_a, "АО Первый")
    second_id = publish(operator_b, "АО Второй")

    session = api.app.state.session_factory()
    try:
        statuses = dict(session.execute(select(DatasetVersion.id, DatasetVersion.status).where(DatasetVersion.id.in_([first_id, second_id]))).all())
    finally:
        session.close()
    assert statuses[first_id] == ImportStatus.PUBLISHED
    assert statuses[second_id] == ImportStatus.PUBLISHED


def test_admin_role_bypasses_per_uploader_visibility_but_a_plain_wide_domain_claim_does_not(api):
    """The per-uploader visibility rule (docs/domain-upload-instructions.md)
    has exactly one deliberate bypass: the literal "admin" role. A broad
    domain claim (domains=["*"]) alone must not grant it - that would let
    anyone with domain access read anyone else's uploads just by asking for
    every domain, defeating the whole point of the rule."""
    uploaded = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("corp.xlsx", _corporate_finance_workbook(issuer="АО Первый"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    upload_id = uploaded.json()["id"]
    materialized = api.post(f"/api/v1/source-uploads/{upload_id}/datasets", headers=UPLOADER, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    assert materialized.status_code == 200, materialized.text
    dataset_id = materialized.json()["items"][0]["id"]
    approved = api.post(f"/api/v1/dataset-versions/{dataset_id}/approve", headers=UPLOADER, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset_id}/publish", headers=UPLOADER)
    assert published.status_code == 200, published.text

    wide_domain_reader = {"X-Actor-Id": "some-other-actor", "X-Actor-Roles": "reader", "X-Actor-Domains": "*"}
    overview = api.get("/api/v1/corporate-finance/overview", headers=wide_domain_reader)
    assert overview.status_code == 200, overview.text
    assert overview.json()["available"] is False
    assert "corporate_finance_register" not in overview.json()["records"]

    admin_reader = {"X-Actor-Id": "some-other-actor", "X-Actor-Roles": "admin,reader", "X-Actor-Domains": "*"}
    admin_overview = api.get("/api/v1/corporate-finance/overview", headers=admin_reader)
    assert admin_overview.status_code == 200, admin_overview.text
    assert len(admin_overview.json()["records"]["corporate_finance_register"]) == 1

    # This is the exact scenario that surfaced the bug: "Скачать оригинал"
    # in the source-preview drawer downloads by source_upload_id, a
    # different code path from the dataset-version reads above (and from a
    # different actor_id than the uploader, since the admin persona's own
    # actor_id never matches any real uploader).
    non_admin_download = api.get(f"/api/v1/source-uploads/{upload_id}/source", headers=wide_domain_reader)
    assert non_admin_download.status_code == 403, non_admin_download.text
    admin_download = api.get(f"/api/v1/source-uploads/{upload_id}/source", headers=admin_reader)
    assert admin_download.status_code == 200, admin_download.text


def test_a_solo_operator_can_reject_or_withdraw_their_own_upload(api):
    """A solo domain operator (one actor holding uploader/reviewer/publisher
    roles) must be able to pull back their own mistaken upload - before or
    after publish, and approve their own dataset - without needing a second
    person. Visibility is uploader-scoped, so a genuinely separate reviewer
    couldn't see this dataset to approve it anyway."""
    uploader_as_reviewer = {"X-Actor-Id": "solo-operator", "X-Actor-Roles": "uploader,reviewer,publisher,reader"}

    uploaded = api.post("/api/v1/source-uploads", headers=uploader_as_reviewer, files={"file": ("solo-1.xlsx", _corporate_finance_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=uploader_as_reviewer, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    dataset_id = materialized.json()["items"][0]["id"]

    rejected = api.post(f"/api/v1/dataset-versions/{dataset_id}/reject", headers=uploader_as_reviewer, json={"reason": "Загрузил не тот файл"})
    assert rejected.status_code == 200, rejected.text

    uploaded = api.post("/api/v1/source-uploads", headers=uploader_as_reviewer, files={"file": ("solo-2.xlsx", _corporate_finance_workbook(issuer="АО Второй"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=uploader_as_reviewer, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    dataset_id = materialized.json()["items"][0]["id"]
    approved = api.post(f"/api/v1/dataset-versions/{dataset_id}/approve", headers=uploader_as_reviewer, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset_id}/publish", headers=uploader_as_reviewer)
    assert published.status_code == 200, published.text

    withdrawn = api.post(f"/api/v1/dataset-versions/{dataset_id}/withdraw", headers=uploader_as_reviewer, json={"reason": "Опубликовано по ошибке"})
    assert withdrawn.status_code == 200, withdrawn.text


def test_solo_operator_can_approve_and_publish_their_own_dataset(api):
    """approve_dataset (and mapping confirmation) no longer require a
    separate reviewer identity - the normal model is one responsible
    operator per domain (docs/domain-upload-instructions.md), and
    visibility is uploader-scoped, so a genuinely separate reviewer
    couldn't even see the dataset to approve it."""
    same_actor = {"X-Actor-Id": "solo-operator-2", "X-Actor-Roles": "uploader,reviewer,publisher,reader"}
    uploaded = api.post("/api/v1/source-uploads", headers=same_actor, files={"file": ("self-approve.xlsx", _corporate_finance_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=same_actor, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    dataset_id = materialized.json()["items"][0]["id"]
    approved = api.post(f"/api/v1/dataset-versions/{dataset_id}/approve", headers=same_actor, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    assert approved.status_code == 200, approved.text
    published = api.post(f"/api/v1/dataset-versions/{dataset_id}/publish", headers=same_actor)
    assert published.status_code == 200, published.text


def test_publish_race_returns_409_not_500(api, monkeypatch):
    """publish_dataset relies on uq_published_dataset_scope_date (a partial
    unique DB index) to be the actual guarantee against two concurrent
    publish requests for the same key both succeeding - its own in-process
    "supersede the prior published row" check can't see an overlapping,
    not-yet-committed transaction. Simulates losing that race by making
    publish_dataset raise the IntegrityError the constraint would, and
    checks the route degrades to a clean 409 instead of an unhandled 500 -
    mirrors the identical pattern already used for the OSIP upload race
    in services/imports.py."""
    from sqlalchemy.exc import IntegrityError

    from osip_dashboard.routes import multi_source as multi_source_routes

    same_actor = {"X-Actor-Id": "race-operator", "X-Actor-Roles": "uploader,reviewer,publisher,reader"}
    uploaded = api.post("/api/v1/source-uploads", headers=same_actor, files={"file": ("race.xlsx", _corporate_finance_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=same_actor, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    dataset_id = materialized.json()["items"][0]["id"]
    approved = api.post(f"/api/v1/dataset-versions/{dataset_id}/approve", headers=same_actor, json={"comment": "Проверено", "acknowledged_dq_codes": []})
    assert approved.status_code == 200, approved.text

    def raise_integrity_error(*args, **kwargs):
        raise IntegrityError("INSERT", {}, Exception("uq_published_dataset_scope_date"))

    monkeypatch.setattr(multi_source_routes, "publish_dataset", raise_integrity_error)
    published = api.post(f"/api/v1/dataset-versions/{dataset_id}/publish", headers=same_actor)
    assert published.status_code == 409, published.text


def test_source_upload_dedup_survives_a_concurrent_race(tmp_path, monkeypatch):
    """Same bug class as the publish race above, in a more likely-to-happen
    place: create_source_upload's "does this (sha256, uploader_id) already
    exist" check only sees already-committed state, so two requests
    uploading identical bytes at the same time (a double-click, a browser
    retrying a dropped response) can both pass it before either commits.
    uq_source_upload_hash_uploader is the actual guarantee. Simulates losing
    that race the same way test_publish_race_returns_409_not_500 does -
    SQLite's single-writer-transaction locking makes a genuine two-session
    reproduction here report a lock error rather than the constraint
    violation a real concurrent-writer database (this app's actual
    PostgreSQL target) would - so the flush itself is forced to fail with
    the specific IntegrityError instead."""
    from sqlalchemy.exc import IntegrityError

    database_path = tmp_path / "race.sqlite3"
    engine = create_database_engine(f"sqlite:///{database_path}")
    Base.metadata.create_all(engine)
    blob_store = LocalBlobStore(tmp_path / "blobs")
    content = _corporate_finance_workbook()

    session = SqlaSession(engine)
    try:
        winner, duplicate = create_source_upload(
            session, blob_store, filename="race.xlsx", content=content, uploader_id="racer", max_upload_bytes=10_000_000
        )
        session.commit()
        assert not duplicate

        # The loser's own "does this already exist" check has to miss the
        # winner (as it would if that commit landed in the race window
        # right after this check ran) for its own insert to be the one that
        # then collides - forced here since winner is already committed and
        # would otherwise just be found normally.
        real_scalar = SqlaSession.scalar
        real_flush = SqlaSession.flush
        calls = {"scalar": 0, "flush": 0}

        def scalar_misses_once(self, *args, **kwargs):
            calls["scalar"] += 1
            if calls["scalar"] == 1:
                return None
            return real_scalar(self, *args, **kwargs)

        def flush_fails_once(self, *args, **kwargs):
            calls["flush"] += 1
            if calls["flush"] == 1:
                raise IntegrityError("INSERT", {}, Exception("uq_source_upload_hash_uploader"))
            return real_flush(self, *args, **kwargs)

        monkeypatch.setattr(SqlaSession, "scalar", scalar_misses_once)
        monkeypatch.setattr(SqlaSession, "flush", flush_fails_once)
        loser, duplicate_loser = create_source_upload(
            session, blob_store, filename="race.xlsx", content=content, uploader_id="racer", max_upload_bytes=10_000_000
        )
        assert duplicate_loser
        assert loser.id == winner.id
    finally:
        session.close()
        engine.dispose()


def test_materialize_retries_a_version_number_collision_instead_of_500ing(api, monkeypatch):
    """Third instance of the same bug class: materialize_datasets computes
    its version number the same "read max, add one" way create_source_upload
    and publish_dataset do, and hits it from the same real trigger - a
    double-click or retried request on "Create selected datasets".
    uq_dataset_scope_date_version is the actual guarantee; the fix here is a
    bounded retry (_add_with_version_retry) rather than "return the existing
    row", since the new dataset is still meant to be created, just under a
    version number one higher than its first, collided guess."""
    from sqlalchemy.exc import IntegrityError

    uploaded = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("materialize-race.xlsx", _corporate_finance_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text

    real_flush = SqlaSession.flush
    calls = {"n": 0}

    def flush_fails_once(self, *args, **kwargs):
        calls["n"] += 1
        if calls["n"] == 1:
            raise IntegrityError("INSERT", {}, Exception("uq_dataset_scope_date_version"))
        return real_flush(self, *args, **kwargs)

    monkeypatch.setattr(SqlaSession, "flush", flush_fails_once)
    materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=UPLOADER, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    assert materialized.status_code == 200, materialized.text
    assert calls["n"] >= 2, "expected at least one retry after the simulated collision"


def test_source_first_materialize_publish_race_returns_409_not_500(source_first_api, monkeypatch):
    """In source_first_mode (this app's actual deployed mode - config.py),
    materialize_datasets calls publish_dataset directly for every validated
    dataset - a separate call site from the dedicated /publish endpoint that
    test_publish_race_returns_409_not_500 already covers. Same race, same
    guarantee (uq_published_dataset_scope_date), needed its own except
    clause on this route since exceptions don't propagate between them."""
    from sqlalchemy.exc import IntegrityError

    from osip_dashboard.routes import multi_source as multi_source_routes

    actor = {"X-Actor-Id": "source-first-racer", "X-Actor-Roles": "uploader,reader,publisher", "X-Actor-Domains": "corpfin"}
    uploaded = source_first_api.post("/api/v1/source-uploads", headers=actor, files={"file": ("sf-race.xlsx", _corporate_finance_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert uploaded.status_code == 201, uploaded.text

    def raise_integrity_error(*args, **kwargs):
        raise IntegrityError("INSERT", {}, Exception("uq_published_dataset_scope_date"))

    monkeypatch.setattr(multi_source_routes, "publish_dataset", raise_integrity_error)
    materialized = source_first_api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers=actor, json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]})
    assert materialized.status_code == 409, materialized.text


def test_module_read_route_can_pin_a_superseded_dataset_version(api):
    def publish(content: bytes) -> str:
        uploaded = api.post("/api/v1/source-uploads", headers=UPLOADER, files={"file": ("version.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(
            f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets",
            headers=UPLOADER,
            json={"datasets": [{"detected_key": "deals", "scope_code": "CORPFIN"}]},
        )
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers=UPLOADER, json={"comment": "Проверено", "acknowledged_dq_codes": []})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers=UPLOADER)
        assert published.status_code == 200, published.text
        return dataset["id"]

    first_id = publish(_corporate_finance_workbook(issuer="АО Первый"))
    publish(_corporate_finance_workbook(issuer="АО Второй"))

    latest = api.get("/api/v1/corporate-finance/overview", headers=UPLOADER)
    assert latest.status_code == 200, latest.text
    assert latest.json()["pinned_dataset_types"] == []
    assert [row["issuer"] for row in latest.json()["records"]["corporate_finance_register"]] == ["АО Второй"]

    pinned = api.get(f"/api/v1/corporate-finance/overview?dataset_versions={first_id}", headers=UPLOADER)
    assert pinned.status_code == 200, pinned.text
    assert pinned.json()["pinned_dataset_types"] == ["corporate_finance_register"]
    assert [row["issuer"] for row in pinned.json()["records"]["corporate_finance_register"]] == ["АО Первый"]


def test_brokerage_workbook_version_keeps_all_child_datasets_together(api):
    """Selecting a brokerage workbook must never mix child datasets by date."""
    session = api.app.state.session_factory()
    try:
        first_upload = SourceUpload(
            source_sha256="1" * 64, original_filename="brokerage-june.xlsx", storage_key="brokerage-june.xlsx",
            file_format="xlsx", detected_source_type="client_brokerage", detection={}, uploader_id="uploader",
        )
        second_upload = SourceUpload(
            source_sha256="2" * 64, original_filename="brokerage-july.xlsx", storage_key="brokerage-july.xlsx",
            file_format="xlsx", detected_source_type="client_brokerage", detection={}, uploader_id="uploader",
        )
        session.add_all([first_upload, second_upload])
        session.flush()

        datasets: list[DatasetVersion] = []
        for upload, suffix, business_date in (
            (first_upload, "June", date(2026, 6, 30)),
            (second_upload, "July", date(2026, 7, 31)),
        ):
            ledger = DatasetVersion(
                source_upload=upload, dataset_type="brokerage_trade_ledger", detected_key="trades",
                scope_type="business_domain", scope_code="BROKERAGE", source_report_date=business_date,
                business_date=business_date, parser_version="test", version=1, status=ImportStatus.PUBLISHED,
                summary={}, uploader_id="uploader",
            )
            derivatives = DatasetVersion(
                source_upload=upload, dataset_type="derivatives_register", detected_key="derivatives",
                scope_type="business_domain", scope_code="BROKERAGE", source_report_date=business_date,
                business_date=business_date, parser_version="test", version=1, status=ImportStatus.PUBLISHED,
                summary={}, uploader_id="uploader",
            )
            session.add_all([ledger, derivatives])
            session.flush()
            session.add_all([
                DatasetRecord(dataset_id=ledger.id, record_type="trade", record_key=f"trade-{suffix}", payload={"trade_number": suffix}, source_ref={}, raw_values={}, formulas={}, cached_values={}),
                DatasetRecord(dataset_id=derivatives.id, record_type="derivative", record_key=f"derivative-{suffix}", payload={"instrument": suffix}, source_ref={}, raw_values={}, formulas={}, cached_values={}),
            ])
            datasets.extend([ledger, derivatives])
        session.commit()
        first_upload_id = str(first_upload.id)
    finally:
        session.close()

    response = api.get(f"/api/v1/brokerage/overview?source_upload_id={first_upload_id}", headers=UPLOADER)
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["selected_source_upload_id"] == first_upload_id
    assert {source["source_upload_id"] for source in payload["sources"]} == {first_upload_id}
    assert [row["trade_number"] for row in payload["records"]["brokerage_trade_ledger"]] == ["June"]
    assert [row["instrument"] for row in payload["records"]["derivatives_register"]] == ["June"]

    # The downloadable workbook follows the same selected physical package.
    # It must not export the newer July children when the June version bar is
    # selected in the brokerage UI.
    exported = api.get(f"/api/v1/brokerage/export?source_upload_id={first_upload_id}", headers=UPLOADER)
    assert exported.status_code == 200, exported.text
    assert_workbook_is_compliant(exported.content)
    workbook = load_workbook(BytesIO(exported.content), data_only=False)
    trades = workbook["Сделки"]
    trade_values = [trades.cell(row, 1).value for row in range(1, trades.max_row + 1)]
    assert "June" in trade_values
    assert "July" not in trade_values


def test_module_read_route_rejects_a_pinned_version_outside_the_actors_domain(api):
    session = api.app.state.session_factory()
    try:
        upload = SourceUpload(
            source_sha256="d" * 64, original_filename="client.xlsx", storage_key="client.xlsx",
            file_format="xlsx", detected_source_type="client_brokerage", detection={}, uploader_id="uploader",
        )
        dataset = DatasetVersion(
            source_upload=upload, dataset_type="brokerage_trade_ledger", detected_key="trades",
            scope_type="business_domain", scope_code="BROKERAGE", source_report_date=date(2026, 7, 20),
            business_date=date(2026, 7, 20), parser_version="test", version=1,
            status=ImportStatus.PUBLISHED, summary={}, uploader_id="uploader",
        )
        session.add(dataset)
        session.flush()
        brokerage_dataset_id = str(dataset.id)
        session.commit()
    finally:
        session.close()

    # An actor scoped to corpfin only (no client_ops access) must not be able
    # to pull a brokerage dataset into the corporate-finance page just by
    # passing its id - domain access is enforced per requested version, not
    # only for whatever the module would have picked on its own.
    corpfin_only_reader = {"X-Actor-Id": "reader", "X-Actor-Roles": "reader", "X-Actor-Domains": "corpfin"}
    response = api.get(f"/api/v1/corporate-finance/overview?dataset_versions={brokerage_dataset_id}", headers=corpfin_only_reader)
    assert response.status_code == 403, response.text


def test_real_source_precedes_demo_fixture_and_hides_demo_registry_rows(api):
    session = api.app.state.session_factory()
    try:
        demo_upload = SourceUpload(
            source_sha256="a" * 64,
            original_filename="demo_brokerage.xlsx",
            storage_key="demo.xlsx",
            file_format="xlsx",
            detected_source_type="client_brokerage",
            detection={},
            uploader_id="demo",
        )
        real_upload = SourceUpload(
            source_sha256="b" * 64,
            original_filename="Клиентский_дашборд.xlsx",
            storage_key="real.xlsx",
            file_format="xlsx",
            detected_source_type="client_brokerage",
            detection={},
            uploader_id="uploader",
        )
        session.add_all([demo_upload, real_upload])
        session.flush()
        for upload, filename, source_date in (
            (demo_upload, "demo_brokerage.xlsx", date(2026, 7, 20)),
            (real_upload, "Клиентский_дашборд.xlsx", date(2026, 7, 10)),
        ):
            session.add(
                DatasetVersion(
                    source_upload=upload,
                    dataset_type="brokerage_trade_ledger",
                    detected_key="trades",
                    scope_type="business_domain",
                    scope_code="BROKERAGE",
                    source_report_date=source_date,
                    business_date=source_date,
                    parser_version="test",
                    version=1,
                    status=ImportStatus.PUBLISHED,
                    summary={"source_filename": filename},
                    uploader_id="uploader",
                )
            )
        session.commit()
        selected = latest_published(session, "brokerage_trade_ledger", "BROKERAGE")
        assert selected is not None
        assert selected.source_upload.original_filename == "Клиентский_дашборд.xlsx"
        visible = list_datasets(session, dataset_type="brokerage_trade_ledger", scope_code="BROKERAGE")
        assert [item.source_upload.original_filename for item in visible] == ["Клиентский_дашборд.xlsx"]
    finally:
        session.close()


def test_latest_published_uses_workbook_report_date_for_operational_feeds(api):
    """A brokerage ledger's latest trade date must not select the older workbook."""
    session = api.app.state.session_factory()
    try:
        older_upload = SourceUpload(
            source_sha256="c" * 64, original_filename="brokerage-20.07.2026.xlsx",
            storage_key="brokerage-20.07.2026.xlsx", file_format="xlsx",
            detected_source_type="client_brokerage", detection={}, uploader_id="uploader",
        )
        newer_upload = SourceUpload(
            source_sha256="e" * 64, original_filename="brokerage-21.07.2026.xlsx",
            storage_key="brokerage-21.07.2026.xlsx", file_format="xlsx",
            detected_source_type="client_brokerage", detection={}, uploader_id="uploader",
        )
        session.add_all([older_upload, newer_upload])
        session.flush()
        session.add_all([
            DatasetVersion(
                source_upload=older_upload, dataset_type="brokerage_trade_ledger", detected_key="trades",
                scope_type="business_domain", scope_code="BROKERAGE", source_report_date=date(2026, 7, 20),
                business_date=date(2026, 7, 20), parser_version="test", version=1,
                status=ImportStatus.PUBLISHED, summary={"date_basis": "latest_transaction"}, uploader_id="uploader",
            ),
            DatasetVersion(
                source_upload=newer_upload, dataset_type="brokerage_trade_ledger", detected_key="trades",
                scope_type="business_domain", scope_code="BROKERAGE", source_report_date=date(2026, 7, 21),
                business_date=date(2026, 7, 10), parser_version="test", version=1,
                status=ImportStatus.PUBLISHED, summary={"date_basis": "latest_transaction"}, uploader_id="uploader",
            ),
        ])
        session.commit()
        selected = latest_published(session, "brokerage_trade_ledger", "BROKERAGE")
        assert selected is not None
        assert selected.source_upload.original_filename == "brokerage-21.07.2026.xlsx"
    finally:
        session.close()


def test_module_dates_ignore_operational_and_period_end_child_dates(api):
    session = api.app.state.session_factory()
    try:
        upload = SourceUpload(
            source_sha256="f" * 64, original_filename="Клиентский_дашборд_20.07.2026.xlsx",
            storage_key="client-dates.xlsx", file_format="xlsx", detected_source_type="client_brokerage",
            detection={}, uploader_id="uploader",
        )
        session.add(upload)
        session.flush()
        common = {
            "source_upload": upload, "scope_type": "business_domain", "scope_code": "BROKERAGE",
            "parser_version": "test", "version": 1, "status": ImportStatus.PUBLISHED, "uploader_id": "uploader",
        }
        session.add_all([
            DatasetVersion(
                **common, dataset_type="brokerage_trade_ledger", detected_key="trades",
                source_report_date=date(2026, 7, 20), business_date=date(2026, 7, 10),
                summary={"date_basis": "latest_transaction"},
            ),
            DatasetVersion(
                **common, dataset_type="derivatives_register", detected_key="derivatives",
                source_report_date=date(2026, 7, 20), business_date=date(2026, 7, 31),
                summary={"date_basis": "reporting_period_end"},
            ),
            DatasetVersion(
                **common, dataset_type="client_account_snapshot", detected_key="clients",
                source_report_date=date(2026, 7, 20), business_date=date(2026, 7, 20), summary={},
            ),
        ])
        session.commit()
        payload = module_payload(session, "brokerage")
        assert payload["report_date_mismatch"] is False
        assert payload["report_dates"] == ["2026-07-20"]
    finally:
        session.close()


def test_export_respects_a_pinned_dataset_version(api):
    """Regression: /accounting/export, /risk/export, and /funds/TABYS/export
    all ignored the domain-version bar's pinned selection and always
    regenerated from latest_published - pin an older workbook version on
    screen, hit Export, and the downloaded file silently showed the latest
    version's numbers instead of what was on screen. /brokerage/export was
    the only one that ever threaded a pin through. Fixed by giving every
    export route the same dataset_versions/source_upload_id handling the
    read endpoints already had (_resolve_version_overrides, shared with
    module_payload's own override-merging)."""
    def publish(issuer: str) -> str:
        workbook = load_workbook(BytesIO(_accounting_portfolio_workbook()))
        workbook["ОСИП_ПОРТФЕЛЬ"].cell(5, 10, issuer)
        content = BytesIO()
        workbook.save(content)
        uploaded = api.post("/api/v1/source-uploads", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, files={"file": ("portfolio.xlsx", content.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        assert uploaded.status_code == 201, uploaded.text
        materialized = api.post(f"/api/v1/source-uploads/{uploaded.json()['id']}/datasets", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, json={"datasets": [{"detected_key": "portfolio_detail", "scope_code": "ACCOUNTING"}]})
        assert materialized.status_code == 200, materialized.text
        dataset = materialized.json()["items"][0]
        approved = api.post(f"/api/v1/dataset-versions/{dataset['id']}/approve", headers={**UPLOADER, "X-Actor-Domains": "accounting"}, json={"comment": "Проверено", "acknowledged_dq_codes": []})
        assert approved.status_code == 200, approved.text
        published = api.post(f"/api/v1/dataset-versions/{dataset['id']}/publish", headers={**UPLOADER, "X-Actor-Domains": "accounting"})
        assert published.status_code == 200, published.text
        return dataset["id"]

    old_id = publish("Old Issuer LLP")
    publish("New Issuer LLP")

    headers = {**UPLOADER, "X-Actor-Domains": "accounting"}
    latest_export = api.get("/api/v1/accounting/export", headers=headers)
    assert latest_export.status_code == 200, latest_export.text
    latest_workbook = load_workbook(BytesIO(latest_export.content))
    latest_issuers = {cell.value for row in latest_workbook["Детализация портфеля"].iter_rows(min_row=9) for cell in row if cell.column == 2}
    assert "New Issuer LLP" in latest_issuers
    assert "Old Issuer LLP" not in latest_issuers

    pinned_export = api.get(f"/api/v1/accounting/export?dataset_versions={old_id}", headers=headers)
    assert pinned_export.status_code == 200, pinned_export.text
    pinned_workbook = load_workbook(BytesIO(pinned_export.content))
    pinned_issuers = {cell.value for row in pinned_workbook["Детализация портфеля"].iter_rows(min_row=9) for cell in row if cell.column == 2}
    assert "Old Issuer LLP" in pinned_issuers
    assert "New Issuer LLP" not in pinned_issuers
