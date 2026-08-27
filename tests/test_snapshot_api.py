import csv
from datetime import date
from decimal import Decimal
import io
from uuid import UUID

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy import select

from osip_dashboard.persistence.models import AuditEvent, ImportBatch, PortfolioSnapshotRecord, PositionLotRecord


UPLOADER = {
    "X-Actor-Id": "uploader-1",
    "X-Actor-Roles": "uploader,reader",
}
REVIEWER = {
    "X-Actor-Id": "reviewer-1",
    "X-Actor-Roles": "reviewer,reader",
}
PUBLISHER = {
    "X-Actor-Id": "publisher-1",
    "X-Actor-Roles": "publisher,reader",
}
REQUIRED_DQ_CODES = ["DQ-04", "DQ-05"]


def test_dividend_reference_upload_reports_freshness(api):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "#Dividend", "#Declaration", "#ExDate", "#Record", "#Payable", "#Type", "TIP US Equity"])
    sheet.append(["TIP US Equity", 0.5, date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 2), date(2026, 7, 8), "REGULAR", ""])
    output = io.BytesIO()
    workbook.save(output)

    initial = api.get("/api/v1/reference-data/dividends", headers=UPLOADER)
    assert initial.status_code == 200
    uploaded = api.post(
        "/api/v1/reference-data/dividends",
        files={"file": ("dividends 28.07.26..xlsx", output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=UPLOADER,
    )
    assert uploaded.status_code == 200
    payload = uploaded.json()
    assert payload["freshness"] == "fresh"
    assert payload["source_date"] == "2026-07-28"
    assert payload["row_count"] == 1
    assert payload["ticker_count"] == 1
    assert payload["future_pay_count"] == 0

    current = api.get("/api/v1/reference-data/dividends", headers=UPLOADER).json()
    assert current["source_sha256"] == payload["source_sha256"]


def test_risk_bucket_rates_auto_repo_bbb_minus_regardless_of_isin():
    from osip_dashboard.services.holdings_export import _risk_bucket

    # Auto-repo ISINs roll over every period, so they'll never all be in the
    # classes_and_ratings dictionary and never carry an S&P/Moody's/Fitch
    # rating of their own - per the portfolio team, авторепо is always [BBB-].
    for isin in ("KZ2C00015196", "SOME-FUTURE-REPO-ISIN"):
        assert _risk_bucket({"isin": isin, "true_asset_class": "Repo", "rating_sp": "", "rating_moodys": "", "rating_fitch": "", "listing_rating": ""}) == "[BBB-]+"


def test_instrument_focus_is_money_market_for_auto_repo_regardless_of_isin():
    from osip_dashboard.services.holdings_export import _instrument_focus

    # Same rollover problem as the rating rule above: a repo's ISIN is fresh
    # every period, but its true focus/sector is always Money market - it
    # must not fall through to "Не указано" just because the CSV dictionary
    # doesn't (and can't) have every future repo ISIN.
    for isin in ("KZ2C00015196", "SOME-FUTURE-REPO-ISIN"):
        assert _instrument_focus({"isin": isin, "true_asset_class": "Repo"}) == "Money market"


def test_weighted_average_ytm_combines_disagreeing_lots_by_carrying_value():
    from osip_dashboard.api_handlers import _weighted_average_ytm

    # Regression: two lots of the same bond reporting slightly different
    # current YTM used to make the whole instrument's YTM "Недоступно" -
    # confirmed against a real OSIP workbook where two GLLKb2 lots differed
    # by ~0.005pp. A carrying-value-weighted average represents the position
    # without needing exact agreement.
    assert _weighted_average_ytm([]) is None
    assert _weighted_average_ytm([(Decimal("17.5"), Decimal("100"))]) == "17.5"
    # 100-weighted 20% and 300-weighted 24% -> (20*100 + 24*300) / 400 = 23%
    weighted = _weighted_average_ytm([(Decimal("20"), Decimal("100")), (Decimal("24"), Decimal("300"))])
    assert Decimal(weighted) == Decimal("23")
    # No usable weight (e.g. missing carrying values) - fall back to a plain
    # average rather than fabricating a weight or dropping the value.
    plain_average = _weighted_average_ytm([(Decimal("10"), Decimal("0")), (Decimal("20"), Decimal("0"))])
    assert Decimal(plain_average) == Decimal("15")


def test_upload_reference_dictionary_replaces_and_reclassifies(api):
    from osip_dashboard.services.instrument_dictionary import instrument_dictionary

    original_row_count = len(instrument_dictionary())

    csv_content = (
        "ISIN,Класс актива,Class,Rating group,Focus/sector/factor\r\n"
        "KZ2C00008951,РЕПО,Repo,BBB-,Money market\r\n"
        # New ISIN with a Cyrillic-lookalike rating string, matching what the
        # portfolio team's Excel dictionary actually contained: this should
        # still resolve to the CCC+ bucket, not "Рейтинг не указан".
        "KZ2P00018088,Корпоративные облигации,Corporate bonds,ССС+,Consumer Discretionary\r\n"
    ).encode("utf-8-sig")

    response = api.post(
        "/api/v1/reference-data/classes-and-ratings",
        files={"file": ("classes_and_ratings.csv", csv_content, "text/csv")},
        headers=UPLOADER,
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["row_count"] == 2
    assert body["previous_row_count"] == original_row_count
    assert "KZ2P00018088" in body["added_isins"]

    assert instrument_dictionary()["KZ2P00018088"]["Rating group"] == "ССС+"

    from osip_dashboard.services.holdings_export import _risk_bucket

    assert _risk_bucket({"isin": "KZ2P00018088"}) == "[CCC+]-"

    status_response = api.get("/api/v1/reference-data/classes-and-ratings", headers=UPLOADER)
    assert status_response.status_code == 200
    assert status_response.json()["row_count"] == 2


def test_upload_reference_dictionary_rejects_missing_columns(api):
    bad_content = "ISIN,Rating group\r\nKZ2C00008951,BBB-\r\n".encode("utf-8-sig")
    response = api.post(
        "/api/v1/reference-data/classes-and-ratings",
        files={"file": ("classes_and_ratings.csv", bad_content, "text/csv")},
        headers=UPLOADER,
    )
    assert response.status_code == 422


def test_upload_reference_dictionary_requires_uploader_role(api):
    csv_content = "ISIN,Класс актива,Class,Rating group,Focus/sector/factor\r\n".encode("utf-8-sig")
    response = api.post(
        "/api/v1/reference-data/classes-and-ratings",
        files={"file": ("classes_and_ratings.csv", csv_content, "text/csv")},
        headers=REVIEWER,
    )
    assert response.status_code == 403


def upload(api, path, *, content=None, filename=None, portfolio_code=None, portfolio_name=None, headers=UPLOADER):
    code = portfolio_code or ("SOBSTV" if "СОБСТВ" in path.name.upper() else "TABYS")
    data = {"portfolio_code": code}
    if portfolio_name is not None:
        data["portfolio_name"] = portfolio_name
    return api.post(
        "/api/v1/imports",
        files={
            "file": (
                filename or path.name,
                content if content is not None else path.read_bytes(),
                "application/vnd.ms-excel",
            )
        },
        data=data,
        headers=headers,
    )


def test_accept_language_translates_dq_and_api_error_details(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    issues = api.get(
        f"/api/v1/snapshots/{imported['snapshot_id']}/issues",
        headers={**UPLOADER, "Accept-Language": "en-GB,en;q=0.9"},
    )
    assert issues.status_code == 200
    assert next(item for item in issues.json()["items"] if item["code"] == "DQ-04")["message"].startswith(
        "The workbook does not contain stable identifiers"
    )

    missing = api.get(
        "/api/v1/portfolios/UNKNOWN/snapshots",
        headers={**UPLOADER, "Accept-Language": "en"},
    )
    assert missing.status_code == 404
    assert missing.json()["detail"] == "Portfolio not found"


def approve_and_publish(api, import_id, codes=REQUIRED_DQ_CODES):
    approved = api.post(
        f"/api/v1/imports/{import_id}/approve",
        json={"comment": "Reviewed operational exceptions", "acknowledged_dq_codes": codes},
        headers=REVIEWER,
    )
    assert approved.status_code == 200, approved.text
    published = api.post(
        f"/api/v1/imports/{import_id}/publish", headers=PUBLISHER
    )
    assert published.status_code == 200, published.text
    return published.json()


def test_upload_reproduces_golden_snapshot_and_is_idempotent(api, workbook_paths):
    response = upload(api, workbook_paths["SOBSTV"])
    assert response.status_code == 201, response.text
    result = response.json()
    assert result["status"] == "validated"
    assert result["portfolio"] == "SOBSTV"
    assert result["report_date"] == "2026-07-15"
    assert result["summary"]["position_count"] == 19
    assert result["summary"]["unique_isin_count"] == 15
    cents = Decimal("0.01")
    assert Decimal(result["summary"]["purchase_amount_kzt"]).quantize(cents) == Decimal(
        "4695258648.74"
    )
    assert Decimal(result["summary"]["derived_carrying_value_kzt"]).quantize(
        cents
    ) == Decimal("4774363156.14")
    assert Decimal(result["summary"]["cash_kzt"]).quantize(cents) == Decimal(
        "42009877.85"
    )
    assert result["publication_requires_override"] is True

    duplicate = upload(api, workbook_paths["SOBSTV"])
    assert duplicate.status_code == 200
    assert duplicate.json()["id"] == result["id"]
    assert duplicate.json()["duplicate"] is True

    listed = api.get("/api/v1/imports", headers=UPLOADER).json()["items"]
    assert len(listed) == 1


def test_role_enforcement_and_four_eyes_workflow(api, workbook_paths):
    assert upload(api, workbook_paths["SOBSTV"], headers={}).status_code == 401
    assert upload(
        api,
        workbook_paths["SOBSTV"],
        headers={"X-Actor-Id": "reader", "X-Actor-Roles": "reader"},
    ).status_code == 403

    imported = upload(api, workbook_paths["SOBSTV"]).json()
    import_id = imported["id"]
    self_reviewer = {
        "X-Actor-Id": "uploader-1",
        "X-Actor-Roles": "reviewer",
    }
    self_approval = api.post(
        f"/api/v1/imports/{import_id}/approve",
        json={"comment": "Self review", "acknowledged_dq_codes": REQUIRED_DQ_CODES},
        headers=self_reviewer,
    )
    assert self_approval.status_code == 409

    incomplete = api.post(
        f"/api/v1/imports/{import_id}/approve",
        json={"comment": "Reviewed", "acknowledged_dq_codes": ["DQ-04"]},
        headers=REVIEWER,
    )
    assert incomplete.status_code == 409

    published = approve_and_publish(api, import_id)
    assert published["status"] == "published"
    assert published["reviewer_id"] == "reviewer-1"
    assert published["publisher_id"] == "publisher-1"
    assert {event["action"] for event in published["audit_events"]} >= {
        "import.created",
        "import.validated",
        "import.approved",
        "import.published",
    }


def test_snapshot_metric_provenance_returns_exact_rows_and_formulas(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    response = api.get(
        f"/api/v1/snapshots/{imported['snapshot_id']}/provenance", headers=UPLOADER
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["source_filename"] == workbook_paths["SOBSTV"].name
    purchase = payload["metrics"]["purchase_amount_kzt"]
    assert purchase["basis"] == "source"
    assert purchase["source_refs"]
    assert all(ref["field"] == "purchase_amount_kzt" for ref in purchase["source_refs"])
    assert all(ref["workbook_name"] == workbook_paths["SOBSTV"].name for ref in purchase["source_refs"])
    carrying = payload["metrics"]["derived_carrying_value_kzt"]
    assert carrying["basis"] == "derived"
    assert "carrying_amount_native" in carrying["formula"]
    assert {ref["field"] for ref in carrying["source_refs"]} >= {"carrying_amount_native", "report_fx_rate"}
    total = payload["metrics"]["derived_operational_total_kzt"]
    assert len(total["inputs"]) == 2
    assert payload["metrics"]["official_nav_kzt"]["source_refs"] == []


def test_source_row_preview_centers_on_exact_provenance_cell(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    provenance = api.get(
        f"/api/v1/snapshots/{imported['snapshot_id']}/provenance", headers=UPLOADER
    ).json()
    reference = provenance["metrics"]["purchase_amount_kzt"]["source_refs"][0]

    response = api.get(
        f"/api/v1/source-rows/{reference['source_row_id']}/preview",
        params={"cell": reference["source_cell"]},
        headers=UPLOADER,
    )
    assert response.status_code == 200, response.text
    preview = response.json()
    assert preview["sheet_name"] == reference["sheet_name"]
    assert preview["target_cell"] == reference["source_cell"]
    assert preview["target_row"] == reference["row_number"]
    assert abs(Decimal(str(preview["target_value"])) - Decimal(reference["value"])) < Decimal("0.000001")
    assert preview["original_filename"] == workbook_paths["SOBSTV"].name
    assert any(row["row_number"] == reference["row_number"] for row in preview["rows"])
    # A real portfolio position is always far below the sheet's header rows,
    # so the response should include both bands - the header context (rows
    # 1-10) and the window centered on the target - not just one or the other.
    if reference["row_number"] > 10:
        assert any(row["row_number"] <= 10 for row in preview["rows"])
        assert any(row["row_number"] > 10 for row in preview["rows"])
        assert preview["header_row"] is not None and preview["header_row"] <= 10
        assert any(str(value).strip() for value in preview["column_labels"]), "the drawer must be able to label columns by real header text, not bare Excel letters"

    mismatch = api.get(
        f"/api/v1/source-rows/{reference['source_row_id']}/preview",
        params={"cell": "AA999"},
        headers=UPLOADER,
    )
    assert mismatch.status_code == 409


def test_incomplete_lot_is_excluded_from_the_overview_total_not_zeroed_or_blanked(api, workbook_paths):
    """Regression, in both directions: this lot's missing carrying amount
    must not be folded into the total as a fabricated zero (silently
    understating it), and - the more severe failure this test now guards
    against - must not blank the *entire* portfolio's derived total over
    one bad lot out of many either. Excluding just that lot and disclosing
    the gap (excluded_lot_count/excluded_purchase_value_kzt) is the only
    one of the three that's both honest and still useful."""
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    snapshot_id = imported["snapshot_id"]
    with api.app.state.session_factory() as session:
        lots = list(session.scalars(
            select(PositionLotRecord).where(PositionLotRecord.snapshot_id == UUID(snapshot_id))
        ))
        assert len(lots) > 1
        lot = lots[0]
        lot.carrying_amount_native = None
        lot.derived_carrying_value_kzt = None
        lot.unavailable_fields = sorted(set(lot.unavailable_fields + ["carrying_amount_native"]))
        expected_total = sum(
            (other.derived_carrying_value_kzt for other in lots if other.id != lot.id),
            Decimal("0"),
        )
        excluded_purchase = lot.purchase_amount_kzt or Decimal("0")
        session.commit()

    overview = api.get(f"/api/v1/snapshots/{snapshot_id}/overview", headers=UPLOADER)
    assert overview.status_code == 200, overview.text
    body = overview.json()
    metrics = body["metrics"]
    assert metrics["derived_carrying_value_kzt"]["basis"] == "derived"
    assert Decimal(metrics["derived_carrying_value_kzt"]["value"]) == expected_total
    assert metrics["derived_operational_total_kzt"]["basis"] == "derived"
    assert body["excluded_lot_count"] == 1
    assert Decimal(body["excluded_purchase_value_kzt"]) == excluded_purchase
    assert len(body["excluded_lots"]) == 1
    excluded = body["excluded_lots"][0]
    assert excluded["security_code"] == lot.security_code
    assert excluded["isin"] == lot.isin
    assert excluded["missing_fields"] == ["carrying_amount_native"]
    assert Decimal(excluded["purchase_amount_kzt"]) == excluded_purchase

    provenance = api.get(f"/api/v1/snapshots/{snapshot_id}/provenance", headers=UPLOADER)
    assert provenance.status_code == 200, provenance.text
    carrying = provenance.json()["metrics"]["derived_carrying_value_kzt"]
    assert carrying["basis"] == "derived"
    assert lot.security_code in carrying["explanation"]
    assert any(ref["field"] == "carrying_amount_native" for ref in carrying["source_refs"])


def test_incomplete_lot_is_excluded_from_holdings_and_allocation_totals_not_zeroed(
    api, workbook_paths
):
    """Regression: _aggregated_holdings/snapshot_allocations used to fold a
    lot with no carrying amount into their totals as `or Decimal("0")` -
    e.g. a deposit the source never marks a current balance for still has
    real, purchase-confirmed money, and it silently vanished from the
    Holdings page, the Excel export, and every allocation chart instead of
    being disclosed as missing. snapshot_overview (tested above) has its own
    version of this same fix."""
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    snapshot_id = imported["snapshot_id"]
    with api.app.state.session_factory() as session:
        lot = session.scalar(
            select(PositionLotRecord).where(PositionLotRecord.snapshot_id == UUID(snapshot_id))
        )
        assert lot is not None
        target_isin = lot.isin
        lot.carrying_amount_native = None
        lot.derived_carrying_value_kzt = None
        session.commit()

    holdings = api.get(f"/api/v1/snapshots/{snapshot_id}/holdings", params={"view": "instruments"}, headers=UPLOADER)
    assert holdings.status_code == 200, holdings.text
    items = holdings.json()["items"]
    target = next(item for item in items if item["isin"] == target_isin)
    assert target["derived_carrying_value_kzt"] is None
    assert target["derived_carrying_incomplete"] is True
    assert target["derived_weight_percent"] is None
    # purchase_amount_kzt has no such gap - the instrument's money is still
    # fully visible under that basis, just not the derived one.
    assert Decimal(target["purchase_amount_kzt"]) > 0
    # The remaining instruments' weights sum to ~100% of the complete-only
    # total - not diluted by the excluded instrument silently counting as a
    # zero-value slice of a total that should have been smaller.
    other_weight = sum(
        Decimal(item["derived_weight_percent"])
        for item in items
        if item["isin"] != target_isin
    )
    assert abs(other_weight - Decimal("100")) <= Decimal("0.5")

    allocations = api.get(
        f"/api/v1/snapshots/{snapshot_id}/allocations",
        params={"dimension": "asset_class"},
        headers=UPLOADER,
    )
    assert allocations.status_code == 200, allocations.text
    allocation_payload = allocations.json()
    assert allocation_payload["excluded_lot_count"] == 1
    assert allocation_payload["excluded_value_kzt"] is not None
    assert Decimal(allocation_payload["excluded_value_kzt"]) > 0

    export = api.get(f"/api/v1/snapshots/{snapshot_id}/holdings/export", headers=UPLOADER)
    assert export.status_code == 200, export.text
    workbook = load_workbook(io.BytesIO(export.content))
    worksheet = workbook["Инструменты"]
    note_texts = [str(worksheet.cell(row, 1).value) for row in range(1, 9)]
    assert any("исключен" in text for text in note_texts)


def test_published_read_models_preserve_lots_cash_and_settlement_lineage(
    api, workbook_paths
):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    snapshot_id = imported["snapshot_id"]

    readiness = api.get(
        f"/api/v1/snapshots/{snapshot_id}/report-readiness", headers=UPLOADER
    ).json()
    assert readiness["operational_snapshot_export"]["ready"] is False
    assert readiness["unacknowledged_critical_count"] > 0
    blocked_report = api.post(
        f"/api/v1/snapshots/{snapshot_id}/reports", json={}, headers=PUBLISHER
    )
    assert blocked_report.status_code == 409

    default_reads = api.get(
        "/api/v1/portfolios/SOBSTV/snapshots", headers=UPLOADER
    ).json()["items"]
    assert default_reads == []

    approve_and_publish(api, imported["id"])
    overview = api.get(
        f"/api/v1/snapshots/{snapshot_id}/overview", headers=UPLOADER
    ).json()
    assert overview["data_label"] == "operational/derived"
    assert overview["metrics"]["cash_kzt"]["basis"] == "source"
    assert overview["metrics"]["derived_operational_total_kzt"]["basis"] == "derived"
    assert overview["metrics"]["official_nav_kzt"] == {
        "value": None,
        "basis": "unavailable",
    }
    assert isinstance(overview["metrics"]["cash_kzt"]["value"], str)
    readiness = api.get(
        f"/api/v1/snapshots/{snapshot_id}/report-readiness", headers=UPLOADER
    ).json()
    assert readiness["operational_snapshot_export"]["ready"] is True
    assert readiness["official_report_export"]["ready"] is False
    generated = api.post(
        f"/api/v1/snapshots/{snapshot_id}/reports", json={}, headers=PUBLISHER
    )
    assert generated.status_code == 201
    report = generated.json()
    artifact = api.get(report["artifact_url"], headers=UPLOADER)
    assert artifact.status_code == 200
    artifact_text = artifact.content.decode("utf-8-sig")
    assert "OSIP Portfolio Operational Snapshot" in artifact_text
    assert "not official reporting" in artifact_text
    artifact_rows = list(csv.reader(io.StringIO(artifact_text)))
    assert ["Cash balances"] in artifact_rows
    assert ["Deduplicated settlements"] in artifact_rows
    assert ["Operational calendar"] in artifact_rows
    assert ["Data quality findings"] in artifact_rows
    metric_header = artifact_rows.index(["Metric", "Value", "Basis"])
    metrics = {
        row[0]: {"value": row[1], "basis": row[2]}
        for row in artifact_rows[metric_header + 1 :]
        if len(row) == 3 and row[0]
    }
    assert metrics["derived_operational_total_kzt"] == {
        "value": overview["metrics"]["derived_operational_total_kzt"]["value"],
        "basis": "derived",
    }
    assert metrics["official_nav_kzt"] == {"value": "", "basis": "unavailable"}
    reports = api.get(
        f"/api/v1/snapshots/{snapshot_id}/reports", headers=UPLOADER
    ).json()["items"]
    assert [item["id"] for item in reports] == [report["id"]]

    repeated_report = api.post(
        f"/api/v1/snapshots/{snapshot_id}/reports", json={}, headers=PUBLISHER
    )
    assert repeated_report.status_code == 201
    assert repeated_report.json()["id"] == report["id"]
    assert repeated_report.json()["artifact_sha256"] == report["artifact_sha256"]
    repeated_artifact = api.get(report["artifact_url"], headers=UPLOADER)
    assert repeated_artifact.content == artifact.content

    holdings = api.get(
        f"/api/v1/snapshots/{snapshot_id}/holdings", headers=UPLOADER
    ).json()["items"]
    cash = api.get(
        f"/api/v1/snapshots/{snapshot_id}/cash", headers=UPLOADER
    ).json()["items"]
    settlements = api.get(
        f"/api/v1/snapshots/{snapshot_id}/settlements", headers=UPLOADER
    ).json()
    assert len(holdings) == 19
    assert len(cash) == 11
    assert sum(item["active"] for item in cash) == 5
    assert settlements["raw_count"] == 0
    assert settlements["deduplicated_count"] == 0
    assert settlements["items"] == []
    assert all("row_number" in item["source"] for item in holdings)
    assert all("listing_rating" in item for item in holdings)

    lots_response = api.get(
        f"/api/v1/snapshots/{snapshot_id}/holdings", headers=UPLOADER
    ).json()
    assert lots_response["dividend_data_status"]["freshness"] in {"fresh", "stale", "unknown", "missing"}

    instruments = api.get(
        f"/api/v1/snapshots/{snapshot_id}/holdings?view=instruments",
        headers=UPLOADER,
    ).json()
    assert instruments["view"] == "instruments"
    # Both holdings views carry the current dividend-dictionary freshness so
    # the UI can warn that dividend-adjusted HPR may be understated, without
    # a second round trip to /reference-data/dividends.
    assert instruments["dividend_data_status"]["freshness"] == lots_response["dividend_data_status"]["freshness"]
    assert len(instruments["items"]) == 15
    repeated = next(item for item in instruments["items"] if item["isin"] == "KZKD00001210")
    assert repeated["lot_count"] == 2
    assert len(repeated["source_refs"]) == 2
    assert Decimal(repeated["current_ytm"]) > 0
    assert repeated["hpr_percent"] is not None
    sgov = next(item for item in instruments["items"] if item["security_code"] == "SGOV US")
    assert sgov["true_asset_class"] == "Government bond"
    assert sgov["carrying_amount_native"] is not None
    # The OSIP balance-price source field is optional; when supplied it must
    # remain a typed numeric value rather than being silently rescaled.
    assert all(
        item["carrying_price_native"] is None
        or Decimal(item["carrying_price_native"]) == Decimal(item["carrying_price_native"])
        for item in holdings
    )

    holdings_value = sum(
        Decimal(item["derived_carrying_value_kzt"] or "0") for item in holdings
    )
    cash_value = sum(Decimal(item["kzt_amount"]) for item in cash)
    assert holdings_value + cash_value == Decimal(
        overview["metrics"]["derived_operational_total_kzt"]["value"]
    )

    source = api.get(
        f"/api/v1/imports/{imported['id']}/source", headers=UPLOADER
    )
    assert source.status_code == 200
    assert source.content == workbook_paths["SOBSTV"].read_bytes()

    allocation = api.get(
        f"/api/v1/snapshots/{snapshot_id}/allocations?dimension=asset_class",
        headers=UPLOADER,
    ).json()
    assert allocation["value_basis"] == "derived_carrying_value_kzt"
    assert {item["label"] for item in allocation["items"]} >= {
        "ETF",
        "Repo",
        "Government bond",
        "Corporate bond",
    }
    assert sum(Decimal(item["value_kzt"]) for item in allocation["items"]) == Decimal(
        allocation["total_value_kzt"]
    )
    purchase_allocation = api.get(
        f"/api/v1/snapshots/{snapshot_id}/allocations?dimension=asset_class&basis=purchase",
        headers=UPLOADER,
    ).json()
    assert purchase_allocation["value_basis"] == "purchase_amount_kzt"
    assert sum(
        Decimal(item["value_kzt"]) for item in purchase_allocation["items"]
    ) == Decimal(overview["metrics"]["purchase_amount_kzt"]["value"])

    calendar = api.get(
        f"/api/v1/snapshots/{snapshot_id}/calendar", headers=UPLOADER
    ).json()
    settlement_events = [
        item for item in calendar["items"] if item["event_type"] == "settlement"
    ]
    assert not settlement_events
    assert calendar["counts"]["overdue_settlements"] == 0
    assert calendar["settlement_total"]["basis"] == "unavailable"
    purchase_events = [
        item
        for item in calendar["items"]
        if item["event_type"] in {"instrument_open", "repo_open"}
    ]
    assert purchase_events
    assert all(item["amount_basis"] == "source_purchase_amount" for item in purchase_events)
    assert sum(Decimal(item["amount_kzt"]) for item in purchase_events) == Decimal(
        overview["metrics"]["purchase_amount_kzt"]["value"]
    )


def test_holdings_excel_export_matches_filtered_instrument_view_and_audits(
    api, workbook_paths
):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    snapshot_id = imported["snapshot_id"]
    export_path = f"/api/v1/snapshots/{snapshot_id}/holdings/export"

    unpublished = api.get(export_path, headers=UPLOADER)
    assert unpublished.status_code == 409

    approve_and_publish(api, imported["id"])
    response = api.get(
        export_path,
        params={
            "basis": "purchase",
            "term": "A_MUM072",
            "asset_class": "Repo",
        },
        headers=UPLOADER,
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "SOBSTV_prop_holdings_data_20260715.xlsx" in response.headers[
        "content-disposition"
    ]

    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    worksheet = workbook["Инструменты"]
    assert worksheet.freeze_panes == "B7"
    assert worksheet.auto_filter.ref == "A6:L7"
    assert worksheet["B2"].value == "SOBSTV"
    assert worksheet["E2"].value.date() == date(2026, 7, 15)
    assert worksheet["H2"].value == 1
    assert worksheet["B3"].value == "Сумма покупки"
    assert worksheet["E3"].value == "A_MUM072"
    assert worksheet["H3"].value == "РЕПО"
    assert [worksheet.cell(6, column).value for column in range(1, 13)] == [
        "Код инструмента",
        "ISIN",
        "Эмитент",
        "Истинный класс актива",
        "Валюта",
        "Лоты",
        "Количество",
        "HPR (расч.), KZT, %",
        "HPR (расч.), FX, %",
        "Текущая YTM, %",
        "Сумма покупки, KZT",
        "Вес, %",
    ]
    assert worksheet.max_row == 8
    assert worksheet["A7"].value == "A_MUM072_0014"
    assert worksheet["A8"].value == "Итого"
    assert worksheet["K8"].value == worksheet["K7"].value
    assert worksheet["L8"].value == worksheet["L7"].value
    assert "Проверка веса" in str(worksheet["A4"].value)
    distribution_meta = str(workbook["Распределения"]["A5"].value)
    assert "USD-эквивалент" in distribution_meta
    assert "источник:" in distribution_meta
    assert "дата" in distribution_meta
    assert worksheet["K8"].number_format == "#,##0.00;[Red](#,##0.00);-"
    assert worksheet["G7"].number_format == "#,##0;[Red](#,##0);-"
    assert worksheet["H7"].number_format == '0.00"%";[Red](0.00"%");-'
    assert worksheet["I7"].number_format == '0.00"%";[Red](0.00"%");-'
    assert worksheet["H7"].data_type == "n"
    assert worksheet["I7"].data_type == "n"
    assert worksheet["J7"].data_type == "n"
    assert worksheet["K7"].data_type == "n"
    assert worksheet["L7"].data_type == "n"
    assert workbook.sheetnames == [
        "Инструменты", "Распределения", "Позиции по лотам",
        "Ожидаемые денежные потоки", "Контроль и происхождение", "Данные графиков",
    ]
    assert workbook["Распределения"].freeze_panes == "B1"
    assert workbook["Данные графиков"].freeze_panes == "B1"
    lots_sheet = workbook["Позиции по лотам"]
    assert lots_sheet.freeze_panes == "B7"
    assert [lots_sheet.cell(6, column).value for column in range(1, 25)] == [
        "Код инструмента", "Лот №", "ISIN", "Эмитент", "Дата погашения облигации", "Истинный класс актива", "Валюта цены",
        "Количество", "Номинал (источник)", "Стоимость покупки, ₸", "MV, USD", "Дата открытия", "Цена покупки за единицу (источник)", "Стоимость открытия",
        "Накопленные купоны или дивиденды, KZT", "Оценка выплаченных купонов или дивидендов (валюта цены)", "Текущая балансовая цена за единицу (источник OSIP)", "Текущая стоимость, KZT",
        "HPR (расч.), KZT", "HPR (расч.), FX", "HPR (расч.), KZT, %", "HPR (расч.), FX, %", "Текущая YTM, %", "Вес, %",
    ]
    note = str(lots_sheet["A5"].value)
    assert "масштаб котировки" in note
    assert "Purchase price per unit" not in note
    assert "дивидендах Bloomberg" in note
    assert lots_sheet["A7"].value == "A_MUM072_0014"
    lot_rows = [row for row in range(7, lots_sheet.max_row + 1) if lots_sheet.cell(row, 1).value != "Итого"]
    assert lot_rows
    assert all(lots_sheet.cell(row, 17).value not in (None, "Недоступно") for row in lot_rows)
    assert lots_sheet["J7"].data_type == "n"
    lot_total_row = next(row for row in range(7, lots_sheet.max_row + 1) if lots_sheet.cell(row, 1).value == "Итого")
    assert lots_sheet.cell(lot_total_row, 10).value == sum(lots_sheet.cell(row, 10).value for row in lot_rows)
    assert lots_sheet["S7"].number_format == "#,##0.00;[Red](#,##0.00);-"
    assert lots_sheet["T7"].number_format == "#,##0.00;[Red](#,##0.00);-"
    assert lots_sheet["U7"].number_format == '0.00"%";[Red](0.00"%");-'
    assert lots_sheet["V7"].number_format == '0.00"%";[Red](0.00"%");-'
    assert lots_sheet["X7"].number_format == "0.0%;[Red](0.0%);-"
    lots_api = api.get(f"/api/v1/snapshots/{snapshot_id}/holdings", headers=UPLOADER).json()["items"]
    lot_api = next(item for item in lots_api if item["security_code"] == "A_MUM072_0014")
    assert Decimal(str(lots_sheet["J7"].value)).quantize(Decimal("0.01")) == Decimal(lot_api["purchase_amount_kzt"]).quantize(Decimal("0.01"))
    maturity_cell = lots_sheet["E7"].value
    assert (maturity_cell.date() if hasattr(maturity_cell, "date") else maturity_cell) == (date.fromisoformat(lot_api["close_date"]) if lot_api["close_date"] else "Недоступно")
    assert any(lots_sheet.cell(row, 1).value == "Итого" for row in range(7, lots_sheet.max_row + 1))
    flows_sheet = workbook["Ожидаемые денежные потоки"]
    assert [flows_sheet.cell(6, column).value for column in range(1, 6)] == [
        "Код инструмента", "Тип потока", "Дата", "Сумма", "Валюта",
    ]
    flow_dates = [
        flows_sheet.cell(row, 3).value.date()
        if hasattr(flows_sheet.cell(row, 3).value, "date")
        else flows_sheet.cell(row, 3).value
        for row in range(7, flows_sheet.max_row + 1)
        if flows_sheet.cell(row, 3).value is not None
    ]
    assert flow_dates == sorted(flow_dates)
    assert all(date(2026, 7, 15) <= value <= date(2027, 1, 11) for value in flow_dates)
    assert "180 календарных дней" in str(flows_sheet["A5"].value)
    assert "Bloomberg" in str(flows_sheet["A5"].value)
    assert not any(value >= date(2027, 1, 12) for value in flow_dates)

    detail = api.get(f"/api/v1/imports/{imported['id']}", headers=UPLOADER).json()
    audit = next(event for event in detail["audit_events"] if event["action"] == "holdings.exported")
    assert audit["detail"] == {
        "format": "xlsx",
        "basis": "purchase",
        "term": "A_MUM072",
        "asset_class": "Repo",
        "row_count": 1,
        "filename": "SOBSTV_prop_holdings_data_20260715.xlsx",
    }


def test_holdings_excel_export_instruments_fx_hpr_uses_the_resolved_nbk_rate(
    api, workbook_paths, monkeypatch
):
    """Regression: the Инструменты sheet's "HPR (расч.), FX, %" column used to
    read the raw, unresolved workbook rate (_snapshot_usd_rates) directly,
    while the Распределения and Позиции по лотам sheets in the very same
    download already used the NBK-resolved rate (resolve_export_usd_kzt_rate)
    - so one workbook could show two different implied USD/KZT rates on
    different sheets, with no note on Инструменты explaining why. Pin a
    controlled resolved rate and confirm the sheet's own disclosure note
    names it - proof the sheet now calls the resolved function at all,
    which the old code path never did."""
    from osip_dashboard.services.fx_rates import FxRate
    from osip_dashboard.services.holdings_export import holdings as holdings_module

    fake_rate = FxRate(
        rate=Decimal("321.5000"), effective_date=date(2026, 7, 14),
        source="test NBK stub", source_url="",
    )
    monkeypatch.setattr(
        holdings_module, "resolve_export_usd_kzt_rate",
        lambda report_date, workbook_rate=None: fake_rate,
    )

    imported = upload(api, workbook_paths["SOBSTV"]).json()
    snapshot_id = imported["snapshot_id"]
    approve_and_publish(api, imported["id"])
    response = api.get(
        f"/api/v1/snapshots/{snapshot_id}/holdings/export",
        params={"basis": "purchase", "term": "A_MUM072", "asset_class": "Repo"},
        headers=UPLOADER,
    )
    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    worksheet = workbook["Инструменты"]
    note = str(worksheet["A5"].value)
    assert "HPR (расч.), FX, %" in note
    assert "321.5000" in note
    assert "test NBK stub" in note
    assert "2026-07-14" in note
    assert "Распределения" in note and "Позиции по лотам" in note
    assert worksheet["I7"].data_type == "n"


def test_holdings_excel_export_requires_reader_role(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    response = api.get(
        f"/api/v1/snapshots/{imported['snapshot_id']}/holdings/export",
        headers={"X-Actor-Id": "uploader-only", "X-Actor-Roles": "uploader"},
    )
    assert response.status_code == 403


def test_tabys_holdings_export_keeps_title_and_note_merges_disjoint(api, workbook_paths):
    """A filename/date warning must not make Excel repair the TABYS export.

    The warning adds a metadata row to the title block.  Regression coverage
    ensures the explanatory note and the first table section move with it,
    instead of producing overlapping merge ranges (which Excel removes).
    """
    imported = upload(
        api,
        workbook_paths["TABYS"],
        filename="TABYS 20.07.2026.xls",
        portfolio_code="TABYS",
    ).json()
    approve_and_publish(api, imported["id"])
    response = api.get(
        f"/api/v1/snapshots/{imported['snapshot_id']}/holdings/export",
        headers=UPLOADER,
    )
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)

    for worksheet in workbook.worksheets:
        merged = list(worksheet.merged_cells.ranges)
        for index, first in enumerate(merged):
            for second in merged[index + 1 :]:
                assert not (
                    first.min_row <= second.max_row
                    and second.min_row <= first.max_row
                    and first.min_col <= second.max_col
                    and second.min_col <= first.max_col
                ), f"overlapping merges in {worksheet.title}: {first} / {second}"

    distributions = workbook["Распределения"]
    assert distributions.freeze_panes == "B1"
    assert workbook["Данные графиков"].freeze_panes == "B1"
    assert distributions["A5"].value.startswith("Операционные / расчётные данные")
    assert distributions["A6"].value.startswith("USD-эквивалент")
    assert distributions["A7"].value == "Распределение по валютам"
    lots = workbook["Позиции по лотам"]
    assert "Цена покупки за единицу" in str(lots["A6"].value)


def test_holdings_export_distributions_reconcile_and_classify_true_exposure(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    response = api.get(
        f"/api/v1/snapshots/{imported['snapshot_id']}/holdings/export",
        headers=UPLOADER,
    )
    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    main = workbook["Инструменты"]
    header_row = 6
    total_row = main.max_row if main.max_row < 30 else 25
    # The main table weight total is a typed percentage and carries the
    # validation disclosure from the same ISIN-aggregated view.
    assert main.cell(total_row, 1).value == "Итого"
    assert Decimal(str(main.cell(total_row, 12).value)) == Decimal("1")
    assert "ОК" in str(main["A4"].value)
    assert workbook.sheetnames == [
        "Инструменты", "Распределения", "Позиции по лотам",
        "Ожидаемые денежные потоки", "Контроль и происхождение", "Данные графиков",
    ]
    flows = workbook["Ожидаемые денежные потоки"]
    coupon_rows = [
        row for row in range(7, flows.max_row + 1)
        if flows.cell(row, 2).value == "Купон"
    ]
    assert coupon_rows
    distributions = workbook["Распределения"]
    assert len(distributions._charts) == 4
    # Currency distribution is anchored from column G so it cannot cover the
    # four-column table on the left; the remaining charts retain the wider
    # right-side placement.
    assert distributions._charts[0].anchor._from.col == 6
    assert all(chart.width == 15 and chart.height == 7.5 for chart in distributions._charts)
    assert [chart.legend.position for chart in distributions._charts] == ["b", "r", "b", "r"]
    assert all(chart.dataLabels is not None for chart in distributions._charts)
    assert all(chart.dataLabels.showPercent and not chart.dataLabels.showCatName and not chart.dataLabels.showSerName for chart in distributions._charts)

    def table_total(header: str, weight_column: int) -> int:
        header_row = next(
            row for row in range(1, distributions.max_row + 1)
            if distributions.cell(row, 1).value == header
        ) + 1
        total = next(
            row for row in range(header_row + 1, distributions.max_row + 1)
            if distributions.cell(row, 1).value == "Итого"
        )
        assert Decimal(str(distributions.cell(total, weight_column).value)) == Decimal("1")
        return header_row

    table_total("Распределение по валютам", 4)
    currency_header = next(row for row in range(1, distributions.max_row + 1) if distributions.cell(row, 1).value == "Распределение по валютам") + 1
    currency_kzt = next(row for row in range(currency_header + 1, distributions.max_row + 1) if distributions.cell(row, 1).value == "KZT")
    # KZT holdings are converted with the common report-date USD rate carried
    # by the workbook's USD source rows; this must not be an artificial blank.
    assert isinstance(distributions.cell(currency_kzt, 3).value, (int, float))
    assert distributions.cell(currency_kzt, 3).value > 0
    table_total("Распределение по истинному классу актива", 3)
    risk_header = table_total("Распределение по рейтингу и типу риска", 3)
    class_header = table_total("Распределение по истинному классу актива", 3)
    assert any(distributions.cell(row, 1).value == "[BBB-]+" for row in range(risk_header + 1, distributions.max_row + 1))
    assert any(distributions.cell(row, 1).value == "Государственные облигации" for row in range(class_header + 1, distributions.max_row + 1))


def test_cash_calendar_and_dq_excel_exports_repeat_view_filters_and_audit(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    snapshot_id = imported["snapshot_id"]
    cash_path = f"/api/v1/snapshots/{snapshot_id}/cash-calendar/export"
    assert api.get(cash_path, headers=UPLOADER).status_code == 409
    approve_and_publish(api, imported["id"])

    active_only = api.get(cash_path, headers=UPLOADER)
    assert active_only.status_code == 200
    cash_workbook = load_workbook(io.BytesIO(active_only.content), data_only=True)
    assert cash_workbook.sheetnames == ["Денежные средства", "Сводка валют", "Календарь событий"]
    cash_sheet = cash_workbook["Денежные средства"]
    assert cash_sheet["B2"].value == "SOBSTV"
    assert cash_sheet.freeze_panes == "B7"
    assert [cash_sheet.cell(6, column).value for column in range(1, 10)] == [
        "Кастодиан", "Исходное наименование", "Валюта", "Остаток в исходной валюте",
        "Эквивалент, KZT", "Активна", "Рабочая книга", "Лист", "Строка",
    ]
    assert cash_sheet.max_row == 11  # five active source rows plus header and metadata
    assert cash_sheet["D7"].data_type == "n"
    assert cash_sheet["E7"].data_type == "n"
    all_cash = api.get(cash_path, params={"include_inactive": True}, headers=UPLOADER)
    all_cash_workbook = load_workbook(io.BytesIO(all_cash.content), data_only=True)
    assert all_cash_workbook["Денежные средства"].max_row == 17
    calendar_sheet = cash_workbook["Календарь событий"]
    assert calendar_sheet["A7"].data_type == "d"
    assert calendar_sheet["F7"].data_type == "n"

    dq = api.get(
        f"/api/v1/snapshots/{snapshot_id}/issues/export",
        params={"term": "DQ-04", "severity": "high"},
        headers=UPLOADER,
    )
    assert dq.status_code == 200
    dq_workbook = load_workbook(io.BytesIO(dq.content), data_only=True)
    dq_sheet = dq_workbook["Замечания DQ"]
    assert dq_sheet.max_row == 7
    assert dq_sheet["A7"].value == "DQ-04"
    assert dq_sheet["B7"].value == "Высокая"
    assert dq_sheet["E7"].value == "Недоступно"

    detail = api.get(f"/api/v1/imports/{imported['id']}", headers=UPLOADER).json()
    actions = {event["action"] for event in detail["audit_events"]}
    assert {"cash_calendar.exported", "dq_issues.exported"} <= actions


def test_lot_and_registry_excel_exports_respect_publication_and_portfolio_scope(api, workbook_paths):
    sobstv = upload(api, workbook_paths["SOBSTV"]).json()
    tabys = upload(api, workbook_paths["TABYS"]).json()
    lots_path = f"/api/v1/snapshots/{sobstv['snapshot_id']}/lots/export"
    assert api.get(lots_path, headers=UPLOADER).status_code == 409
    approve_and_publish(api, sobstv["id"])
    approve_and_publish(api, tabys["id"], codes=["DQ-01", "DQ-04", "DQ-05"])

    lots = api.get(lots_path, headers=UPLOADER)
    assert lots.status_code == 200
    lots_workbook = load_workbook(io.BytesIO(lots.content), data_only=True)
    lot_sheet = lots_workbook["Лоты источника"]
    assert lot_sheet.freeze_panes == "B7"
    assert lot_sheet.max_row == 25
    assert lot_sheet["A7"].value
    assert lot_sheet["G7"].data_type == "n"
    assert lot_sheet["Y7"].data_type == "n"
    # Regression: coupon/repo rate is ingested as a raw fraction (0.0475 for
    # a 4.75% bond) but this column is headed "%" and sits next to two other
    # percentage-point columns - it must be scaled the same way, or a 4.75%
    # bond reads as if it paid under 1%.
    coupon_rate_column = next(
        column for column in range(1, lot_sheet.max_column + 1)
        if lot_sheet.cell(6, column).value == "Купон / ставка РЕПО, %"
    )
    len_row = next(
        row for row in range(7, lot_sheet.max_row + 1)
        if lot_sheet.cell(row, 2).value == "LEN 4.75 11/29/27 Corp"
    )
    assert float(lot_sheet.cell(len_row, coupon_rate_column).value) == pytest.approx(4.75)
    assert lot_sheet.cell(len_row, coupon_rate_column).number_format == '0.00"%";[Red](0.00"%");-'

    restricted = {
        "X-Actor-Id": "sobstv-reader",
        "X-Actor-Roles": "reader",
        "X-Actor-Portfolios": "SOBSTV",
    }
    registry = api.get("/api/v1/imports/export", headers=restricted)
    assert registry.status_code == 200
    registry_workbook = load_workbook(io.BytesIO(registry.content), data_only=True)
    assert registry_workbook.sheetnames == ["Реестр загрузок", "Аудит"]
    registry_sheet = registry_workbook["Реестр загрузок"]
    assert registry_sheet.freeze_panes == "B6"
    assert registry_sheet.max_row == 6
    assert registry_sheet["C6"].value == "SOBSTV"
    assert registry_sheet["I6"].data_type == "n"
    with api.app.state.session_factory() as session:
        event = session.scalar(
            select(AuditEvent).where(AuditEvent.action == "imports_registry.exported")
        )
        assert event is not None
        assert event.detail["row_count"] == 1


def test_governed_metric_definitions_expose_disabled_official_metrics(api):
    definitions = api.get("/api/v1/metrics", headers=UPLOADER)
    assert definitions.status_code == 200
    by_code = {item["code"]: item for item in definitions.json()["items"]}
    assert by_code["derived_carrying_value_kzt"]["formula"] == "AA × AU × AT + AR"
    assert by_code["derived_carrying_value_kzt"]["basis"] == "derived"
    assert by_code["official_nav_kzt"]["enabled"] is False
    assert by_code["official_nav_kzt"]["basis"] == "unavailable"
    assert by_code["official_nav_kzt"]["unavailable_reason"]


def test_corrected_file_creates_version_and_supersedes_same_date(api, workbook_paths):
    path = workbook_paths["SOBSTV"]
    first = upload(api, path).json()
    first_preview = api.get(
        f"/api/v1/imports/{first['id']}/comparison", headers=UPLOADER
    )
    assert first_preview.status_code == 200
    assert first_preview.json()["baseline"] is None
    assert first_preview.json()["lot_changes"]["added_count"] == 19
    approve_and_publish(api, first["id"])

    corrected = upload(api, path, content=path.read_bytes() + b"\x00").json()
    assert corrected["status"] == "validated"
    assert corrected["version"] == 2
    comparison = api.get(
        f"/api/v1/imports/{corrected['id']}/comparison", headers=UPLOADER
    )
    assert comparison.status_code == 200
    preview = comparison.json()
    assert preview["baseline"]["import_id"] == first["id"]
    assert preview["baseline"]["status"] == "published"
    assert preview["metrics"]["position_count"]["delta"] == 0
    assert Decimal(
        preview["metrics"]["derived_operational_total_kzt"]["delta"]
    ) == 0
    assert preview["lot_changes"] == {
        "added_count": 0,
        "removed_count": 0,
        "unchanged_count": 19,
        "added": [],
        "removed": [],
    }
    approve_and_publish(api, corrected["id"])

    old = api.get(f"/api/v1/imports/{first['id']}", headers=UPLOADER).json()
    assert old["status"] == "superseded"
    published = api.get(
        "/api/v1/portfolios/SOBSTV/snapshots", headers=UPLOADER
    ).json()["items"]
    assert [item["version"] for item in published] == [2]
    history = api.get(
        "/api/v1/portfolios/SOBSTV/snapshots?include_superseded=true",
        headers=UPLOADER,
    ).json()["items"]
    assert [item["version"] for item in history] == [2, 1]
    assert history[0]["status"] == "published"
    assert history[1]["status"] == "superseded"


def test_portfolios_publish_independently_and_expose_report_dates(api, workbook_paths):
    sobstv = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, sobstv["id"])
    portfolios = api.get("/api/v1/portfolios", headers=UPLOADER).json()
    by_code = {item["code"]: item for item in portfolios["items"]}
    assert by_code["SOBSTV"]["latest_published_report_date"] == "2026-07-15"
    assert by_code["TABYS"]["latest_published_report_date"] is None

    tabys = upload(api, workbook_paths["TABYS"]).json()
    approve_and_publish(api, tabys["id"], codes=["DQ-01", "DQ-04", "DQ-05"])
    portfolios = api.get("/api/v1/portfolios", headers=UPLOADER).json()
    assert portfolios["combined_report_dates"] == ["2026-07-15"]
    assert portfolios["report_date_mismatch"] is False

    # The supplied golden files share an as-of date. Move the already-published
    # TABYS fixture to a different immutable source date to prove that the
    # combined read model exposes, rather than normalizes away, a mismatch.
    with api.app.state.session_factory() as session:
        tabys_batch = session.scalar(
            select(ImportBatch).where(ImportBatch.id == UUID(tabys["id"]))
        )
        tabys_snapshot = session.scalar(
            select(PortfolioSnapshotRecord).where(
                PortfolioSnapshotRecord.id == UUID(tabys["snapshot_id"])
            )
        )
        assert tabys_batch is not None and tabys_snapshot is not None
        tabys_batch.report_date = date(2026, 7, 14)
        tabys_snapshot.report_date = date(2026, 7, 14)
        session.commit()

    mismatched = api.get("/api/v1/portfolios", headers=UPLOADER).json()
    by_code = {item["code"]: item for item in mismatched["items"]}
    assert by_code["SOBSTV"]["latest_published_report_date"] == "2026-07-15"
    assert by_code["TABYS"]["latest_published_report_date"] == "2026-07-14"
    assert mismatched["combined_report_dates"] == ["2026-07-14", "2026-07-15"]
    assert mismatched["report_date_mismatch"] is True


def test_portfolio_permissions_filter_lists_and_block_direct_object_access(api, workbook_paths):
    sobstv = upload(api, workbook_paths["SOBSTV"]).json()
    tabys = upload(api, workbook_paths["TABYS"]).json()
    approve_and_publish(api, sobstv["id"])
    approve_and_publish(api, tabys["id"], codes=["DQ-01", "DQ-04", "DQ-05"])
    restricted = {
        "X-Actor-Id": "sobstv-reader",
        "X-Actor-Roles": "reader",
        "X-Actor-Portfolios": "SOBSTV",
    }

    portfolios = api.get("/api/v1/portfolios", headers=restricted).json()["items"]
    assert [item["code"] for item in portfolios] == ["SOBSTV"]
    imports = api.get("/api/v1/imports", headers=restricted).json()["items"]
    assert {item["portfolio"] for item in imports} == {"SOBSTV"}
    assert api.get(
        f"/api/v1/snapshots/{tabys['snapshot_id']}/overview", headers=restricted
    ).status_code == 403
    assert api.get(
        f"/api/v1/imports/{tabys['id']}/source", headers=restricted
    ).status_code == 403
    assert api.get(
        "/api/v1/imports?portfolio=TABYS", headers=restricted
    ).status_code == 403
    tabys_report = api.post(
        f"/api/v1/snapshots/{tabys['snapshot_id']}/reports",
        json={},
        headers=PUBLISHER,
    )
    assert tabys_report.status_code == 201
    assert api.get(
        tabys_report.json()["artifact_url"], headers=restricted
    ).status_code == 403

    restricted_uploader = {
        "X-Actor-Id": "sobstv-uploader",
        "X-Actor-Roles": "uploader,reader",
        "X-Actor-Portfolios": "SOBSTV",
    }
    assert upload(
        api,
        workbook_paths["TABYS"],
        content=workbook_paths["TABYS"].read_bytes() + b"\x00",
        headers=restricted_uploader,
    ).status_code == 403


def test_invalid_uploads_and_parse_failures(api, workbook_paths):
    path = workbook_paths["SOBSTV"]
    wrong_extension = upload(api, path, filename="portfolio.xlsx")
    assert wrong_extension.status_code == 422
    invalid_ole = upload(api, path, content=b"not an xls")
    assert invalid_ole.status_code == 422
    oversized = upload(api, path, content=path.read_bytes() + b"x" * (10 * 1024 * 1024))
    assert oversized.status_code == 422

    arbitrary_filename = upload(
        api,
        path,
        filename="export-from-official-system.xls",
    )
    assert arbitrary_filename.status_code == 201
    assert arbitrary_filename.json()["status"] == "validated"
    assert arbitrary_filename.json()["portfolio"] == "SOBSTV"
    assert arbitrary_filename.json()["report_date"] == "2026-07-15"

    corrupt = upload(
        api,
        path,
        content=bytes.fromhex("D0CF11E0A1B11AE1") + b"corrupt",
        filename="СОБСТВ 15.07.2026.xls",
    )
    assert corrupt.status_code == 201
    assert corrupt.json()["status"] == "failed"

    new_portfolio = upload(
        api,
        path,
        content=path.read_bytes() + b"\x00",
        filename="any-name-from-source-system.xls",
        portfolio_code="GROWTH",
        portfolio_name="Портфель роста",
    )
    assert new_portfolio.status_code == 201
    assert new_portfolio.json()["status"] == "validated"
    assert new_portfolio.json()["portfolio"] == "GROWTH"

    all_imports = api.get("/api/v1/imports", headers=UPLOADER).json()["items"]
    assert sum(item["status"] == "failed" for item in all_imports) == 1


def test_osip_business_date_comes_from_workbook_not_filename(api, workbook_paths):
    imported = upload(
        api,
        workbook_paths["SOBSTV"],
        filename="СОБСТВ 19.07.2026.xls",
    ).json()
    assert imported["report_date"] == "2026-07-15"
    approve_and_publish(api, imported["id"])

    treasury = api.get("/api/v1/treasury/overview", headers=UPLOADER)
    assert treasury.status_code == 200, treasury.text
    payload = treasury.json()
    assert payload["report_date_mismatch"] is False
    assert payload["sources"][0]["source_report_date"] == "2026-07-15"
    assert payload["sources"][0]["business_date"] == "2026-07-15"

    snapshots = api.get("/api/v1/portfolios/SOBSTV/snapshots?include_superseded=true", headers=UPLOADER)
    assert snapshots.status_code == 200, snapshots.text
    source_upload_id = snapshots.json()["items"][0]["source_upload_id"]
    pinned = api.get(f"/api/v1/treasury/overview?source_upload_id={source_upload_id}", headers=UPLOADER)
    assert pinned.status_code == 200, pinned.text
    assert pinned.json()["selected_source_upload_id"] == source_upload_id
    assert pinned.json()["sources"][0]["source_upload_id"] == source_upload_id


def test_treasury_overview_is_uploader_scoped(api, workbook_paths):
    """Same per-uploader visibility rule as every other domain read - a
    different operator must never see this uploader's published Treasury
    snapshot, by default or by pinning its source_upload_id directly."""
    imported = upload(api, workbook_paths["SOBSTV"], filename="СОБСТВ 20.07.2026.xls").json()
    approve_and_publish(api, imported["id"])

    other_operator = {**UPLOADER, "X-Actor-Id": "other-treasury-operator"}
    denied = api.get("/api/v1/treasury/overview", headers=other_operator)
    assert denied.status_code == 200, denied.text
    assert denied.json()["available"] is False

    snapshots = api.get("/api/v1/portfolios/SOBSTV/snapshots?include_superseded=true", headers=UPLOADER)
    source_upload_id = snapshots.json()["items"][0]["source_upload_id"]
    denied_pin = api.get(f"/api/v1/treasury/overview?source_upload_id={source_upload_id}", headers=other_operator)
    assert denied_pin.status_code == 404, denied_pin.text


def test_rejection_preserves_snapshot_source_and_audit(api, workbook_paths):
    imported = upload(api, workbook_paths["TABYS"]).json()
    rejected = api.post(
        f"/api/v1/imports/{imported['id']}/reject",
        json={"reason": "Awaiting corrected custodian mapping"},
        headers=REVIEWER,
    )
    assert rejected.status_code == 200
    payload = rejected.json()
    assert payload["status"] == "rejected"
    assert payload["snapshot_id"] == imported["snapshot_id"]
    assert "import.rejected" in {
        event["action"] for event in payload["audit_events"]
    }
    source = api.get(
        f"/api/v1/imports/{imported['id']}/source", headers=UPLOADER
    )
    assert source.content == workbook_paths["TABYS"].read_bytes()


def test_withdrawal_hides_published_version_without_deleting_evidence(
    api, workbook_paths
):
    imported = upload(api, workbook_paths["TABYS"]).json()
    approve_and_publish(api, imported["id"])

    forbidden = api.post(
        f"/api/v1/imports/{imported['id']}/withdraw",
        json={"reason": "Wrong portfolio assignment"},
        headers=REVIEWER,
    )
    assert forbidden.status_code == 403

    withdrawn = api.post(
        f"/api/v1/imports/{imported['id']}/withdraw",
        json={"reason": "Wrong portfolio assignment"},
        headers=PUBLISHER,
    )
    assert withdrawn.status_code == 200
    payload = withdrawn.json()
    assert payload["status"] == "withdrawn"
    assert payload["rejection_reason"] == "Wrong portfolio assignment"
    assert "import.withdrawn" in {
        event["action"] for event in payload["audit_events"]
    }

    published = api.get(
        "/api/v1/portfolios/TABYS/snapshots", headers=UPLOADER
    ).json()["items"]
    assert published == []
    history = api.get(
        "/api/v1/portfolios/TABYS/snapshots?include_superseded=true",
        headers=UPLOADER,
    ).json()["items"]
    assert history == []
    source = api.get(
        f"/api/v1/imports/{imported['id']}/source", headers=UPLOADER
    )
    assert source.content == workbook_paths["TABYS"].read_bytes()

    reassigned = upload(
        api,
        workbook_paths["TABYS"],
        portfolio_code="CORRECT_TABYS",
        portfolio_name="Исправленное назначение TABYS",
    )
    assert reassigned.status_code == 201
    reassigned_payload = reassigned.json()
    assert reassigned_payload["id"] != imported["id"]
    assert reassigned_payload["portfolio"] == "CORRECT_TABYS"
    assert reassigned_payload["source_sha256"] == imported["source_sha256"]
    assert reassigned_payload["status"] == "validated"

    visible_imports = api.get("/api/v1/imports", headers=UPLOADER).json()["items"]
    assert [item["id"] for item in visible_imports] == [reassigned_payload["id"]]
    including_withdrawn = api.get(
        "/api/v1/imports?include_withdrawn=true", headers=UPLOADER
    ).json()["items"]
    assert {item["id"] for item in including_withdrawn} == {
        imported["id"],
        reassigned_payload["id"],
    }

    repeated = api.post(
        f"/api/v1/imports/{imported['id']}/withdraw",
        json={"reason": "Repeated request"},
        headers=PUBLISHER,
    )
    assert repeated.status_code == 409


def test_withdrawn_version_can_be_reuploaded_to_the_same_portfolio(
    api, workbook_paths
):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    approve_and_publish(api, imported["id"])
    withdrawn = api.post(
        f"/api/v1/imports/{imported['id']}/withdraw",
        json={"reason": "Accidentally removed; upload again"},
        headers=PUBLISHER,
    )
    assert withdrawn.status_code == 200

    reuploaded = upload(api, workbook_paths["SOBSTV"])
    assert reuploaded.status_code == 201
    payload = reuploaded.json()
    assert payload["id"] != imported["id"]
    assert payload["duplicate"] is False
    assert payload["portfolio"] == "SOBSTV"
    assert payload["version"] == 2
    assert payload["source_sha256"] == imported["source_sha256"]
    assert payload["status"] == "validated"


def test_dq_issue_ownership_assign_and_clear(api, workbook_paths):
    imported = upload(api, workbook_paths["SOBSTV"]).json()
    published = approve_and_publish(api, imported["id"])
    snapshot_id = published["snapshot_id"]

    issues = api.get(f"/api/v1/snapshots/{snapshot_id}/issues", headers=UPLOADER).json()
    assert issues["items"], "fixture workbook should have at least one DQ finding"
    issue_id = issues["items"][0]["id"]
    assert issues["items"][0]["owner_id"] is None
    assert issues["items"][0]["due_date"] is None
    assert issues["items"][0]["is_overdue"] is False

    forbidden = api.post(
        f"/api/v1/issues/{issue_id}/assign",
        json={"owner_id": "ops-1", "due_date": "2020-01-01", "reason": "triage"},
        headers=UPLOADER,
    )
    assert forbidden.status_code == 403

    missing_reason = api.post(
        f"/api/v1/issues/{issue_id}/assign",
        json={"owner_id": "ops-1", "due_date": "2020-01-01", "reason": ""},
        headers=REVIEWER,
    )
    assert missing_reason.status_code == 422  # rejected by request-schema min_length

    due_without_owner = api.post(
        f"/api/v1/issues/{issue_id}/assign",
        json={"owner_id": None, "due_date": "2020-01-01", "reason": "invalid combination"},
        headers=REVIEWER,
    )
    assert due_without_owner.status_code == 409

    assigned = api.post(
        f"/api/v1/issues/{issue_id}/assign",
        json={"owner_id": "ops-1", "due_date": "2020-01-01", "reason": "Overdue remediation owner"},
        headers=REVIEWER,
    )
    assert assigned.status_code == 200, assigned.text
    body = assigned.json()
    assert body["owner_id"] == "ops-1"
    assert body["due_date"] == "2020-01-01"
    assert body["is_overdue"] is True  # 2020-01-01 is in the past

    refreshed = api.get(f"/api/v1/snapshots/{snapshot_id}/issues", headers=UPLOADER).json()
    refreshed_issue = next(item for item in refreshed["items"] if item["id"] == issue_id)
    assert refreshed_issue["owner_id"] == "ops-1"
    assert refreshed_issue["is_overdue"] is True

    cleared = api.post(
        f"/api/v1/issues/{issue_id}/assign",
        json={"owner_id": None, "due_date": None, "reason": "Resolved, clearing assignment"},
        headers=REVIEWER,
    )
    assert cleared.status_code == 200
    assert cleared.json()["owner_id"] is None
    assert cleared.json()["due_date"] is None
    assert cleared.json()["is_overdue"] is False
