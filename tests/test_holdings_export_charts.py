from datetime import date
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import Workbook

from osip_dashboard.services.holdings_export import (
    _accrued_or_dividend_kzt,
    _coupon_or_dividend_native,
    _expected_cash_flow_rows,
    _derive_expected_coupon_native,
    estimated_coupon_income_native,
    estimated_paid_coupon_income_native,
    _write_distribution_block,
    create_cash_calendar_xlsx,
)
from osip_dashboard.services.dividends import DividendHistory, DividendRecord
from osip_dashboard.services.instrument_dictionary import true_asset_class


def test_lot_income_columns_combine_coupon_and_validated_dividend():
    lot = SimpleNamespace(
        raw_security_type="Корпоративные облигации",
        accrued_income_kzt=Decimal("100"),
    )
    dividend = SimpleNamespace(
        matched_count=1,
        kzt_amount=Decimal("15"),
        native_amount=Decimal("10"),
    )
    assert _accrued_or_dividend_kzt(lot, dividend) == Decimal("115")
    assert _coupon_or_dividend_native(Decimal("25"), dividend) == Decimal("35")


def test_estimated_coupon_income_uses_nominal_quantity_rate_and_360_day_holding_period():
    lot = SimpleNamespace(
        raw_security_type="Корпоративные облигации",
        instrument_currency="KZT",
        nominal_value=Decimal("1000"),
        quantity=Decimal("10"),
        coupon_or_repo_rate=Decimal("0.12"),
        purchase_date=date(2026, 1, 1),
        report_fx_rate=Decimal("500"),
        instrument=None,
    )
    assert estimated_coupon_income_native(lot, date(2026, 7, 1)) == Decimal("603.3333333333333333333333333")


def test_estimated_coupon_income_is_not_applied_to_repo_or_missing_inputs():
    repo = SimpleNamespace(
        raw_security_type="РЕПО",
        instrument_currency="KZT",
        nominal_value=Decimal("1000"),
        quantity=Decimal("10"),
        coupon_or_repo_rate=Decimal("0.12"),
        purchase_date=date(2026, 1, 1),
        report_fx_rate=Decimal("500"),
        instrument=None,
    )
    incomplete = SimpleNamespace(
        raw_security_type="Облигации",
        instrument_currency="KZT",
        nominal_value=Decimal("1000"),
        quantity=Decimal("10"),
        coupon_or_repo_rate=None,
        purchase_date=date(2026, 1, 1),
        report_fx_rate=Decimal("500"),
        instrument=None,
    )
    assert estimated_coupon_income_native(repo, date(2026, 7, 1)) is None
    assert estimated_coupon_income_native(incomplete, date(2026, 7, 1)) is None


def test_paid_coupon_estimate_excludes_current_accrued_coupon():
    lot = SimpleNamespace(
        raw_security_type="Корпоративные облигации",
        instrument_currency="KZT",
        nominal_value=Decimal("1000"),
        quantity=Decimal("10"),
        coupon_or_repo_rate=Decimal("0.12"),
        purchase_date=date(2026, 1, 1),
        previous_coupon_date=date(2026, 1, 1),
        accrued_income_kzt=Decimal("100"),
        report_fx_rate=Decimal("500"),
        instrument=None,
    )
    gross = estimated_coupon_income_native(lot, date(2026, 7, 1))
    paid_estimate = estimated_paid_coupon_income_native(lot, date(2026, 7, 1))
    assert gross == Decimal("603.3333333333333333333333333")
    assert paid_estimate == Decimal("503.3333333333333333333333333")


def _write_block(sheet, charts_sheet, rows, **kwargs):
    return _write_distribution_block(
        sheet, 1, "Focus", ["Focus", "Weight"], rows,
        widths=[30, 14], numeric_formats={2: "0.0%"}, chart_column=2, chart_title="Test",
        charts_sheet=charts_sheet, charts_row=1, table_name="ChartDataTest",
        **kwargs,
    )


def test_write_distribution_block_pushes_small_adjacent_labels_outside():
    workbook = Workbook()
    sheet = workbook.active
    charts_sheet = workbook.create_sheet("Данные графиков")
    # Only one slice (A) is under 2%, so no grouping - both A and B are
    # still small enough (<8%) and adjacent to need outside placement.
    rows = [["A", Decimal("1")], ["B", Decimal("4")], ["C", Decimal("35")], ["D", Decimal("60")]]
    _write_block(sheet, charts_sheet, rows)
    chart = sheet._charts[0]
    overridden = {label.idx: label.dLblPos for label in chart.dataLabels.dLbl}
    assert overridden == {0: "outEnd", 1: "outEnd"}
    assert chart.dataLabels.dLblPos == "inEnd"


def test_write_distribution_block_writes_a_visible_table_and_plots_it():
    # The chart's data source is now a real, visible Excel Table on the
    # "Данные графиков" sheet (not a hidden range), so
    # visible_cells_only stays at openpyxl's default (True) - unlike the
    # earlier hidden-helper-column approach, which had to disable it.
    workbook = Workbook()
    sheet = workbook.active
    charts_sheet = workbook.create_sheet("Данные графиков")
    rows = [["A", Decimal("1")], ["B", Decimal("99")]]
    _write_block(sheet, charts_sheet, rows)
    assert sheet._charts[0].visible_cells_only is True
    assert "ChartDataTest" in charts_sheet.tables
    assert charts_sheet["A2"].value == "A" or charts_sheet["A3"].value == "A"


def test_write_distribution_block_groups_and_nudges_the_real_scenario():
    workbook = Workbook()
    sheet = workbook.active
    charts_sheet = workbook.create_sheet("Данные графиков")
    rows = [
        ["Consumer Discretionary", Decimal("1.0")], ["Development Institution", Decimal("4.5")],
        ["Government", Decimal("35.1")], ["Industrials", Decimal("1.0")], ["Large-cap", Decimal("1.6")],
        ["Money market", Decimal("39.8")], ["Quasi-sovereign", Decimal("17.0")],
    ]
    _write_block(sheet, charts_sheet, rows)
    chart = sheet._charts[0]
    # Grouping already leaves only 5 slices (see test_excel_charts.py's
    # test_chart_series_groups_two_or_more_small_slices_into_other):
    # Development Institution (4.5%) and Прочее (3.6%) are adjacent (wrapping
    # around the circle) and both under the small-slice floor, so each still
    # gets pushed outside with a distinct nudge apart.
    by_index = {label.idx: label for label in chart.dataLabels.dLbl}
    assert set(by_index) == {0, 4}
    assert by_index[0].layout is not None
    assert by_index[4].layout is not None
    offset_0 = (by_index[0].layout.manualLayout.x, by_index[0].layout.manualLayout.y)
    offset_4 = (by_index[4].layout.manualLayout.x, by_index[4].layout.manualLayout.y)
    assert offset_0 != offset_4


def test_write_distribution_block_returns_next_free_charts_row():
    workbook = Workbook()
    sheet = workbook.active
    charts_sheet = workbook.create_sheet("Данные графиков")
    rows = [["A", Decimal("1")], ["B", Decimal("99")]]
    _, next_charts_row = _write_block(sheet, charts_sheet, rows)
    assert next_charts_row > 1


def test_expected_cash_flows_use_inclusive_180_day_window_and_date_order():
    snapshot = SimpleNamespace(report_date=date(2026, 1, 1))
    lots = [
        SimpleNamespace(
            isin="OUTSIDE-ISIN",
            security_code="OUTSIDE",
            close_date=date(2026, 7, 1),
            nominal_value=Decimal("10"),
            quantity=Decimal("1"),
            instrument_currency="KZT",
            next_coupon_date=None,
        ),
        SimpleNamespace(
            isin="BOUNDARY-ISIN",
            security_code="BOUNDARY",
            close_date=date(2026, 6, 30),
            nominal_value=Decimal("20"),
            quantity=Decimal("1"),
            instrument_currency="KZT",
            next_coupon_date=None,
        ),
        SimpleNamespace(
            isin="START-ISIN",
            security_code="START",
            close_date=date(2026, 1, 1),
            nominal_value=Decimal("30"),
            quantity=Decimal("1"),
            instrument_currency="KZT",
            next_coupon_date=None,
        ),
    ]

    rows = _expected_cash_flow_rows(snapshot, lots)

    assert [(row[0], row[2]) for row in rows] == [
        ("START", date(2026, 1, 1)),
        ("BOUNDARY", date(2026, 6, 30)),
    ]


def test_expected_cash_flows_group_same_instrument_lots_on_the_same_date():
    snapshot = SimpleNamespace(report_date=date(2026, 1, 1))
    # Two lots of the same instrument (same ISIN) closing on the same date
    # is one economic redemption, not two - they must combine into a single
    # row with the combined amount rather than appearing as two rows.
    lots = [
        SimpleNamespace(isin="SAME-ISIN", security_code="BOND", close_date=date(2026, 2, 1), nominal_value=Decimal("100"), quantity=Decimal("2"), instrument_currency="KZT", next_coupon_date=None),
        SimpleNamespace(isin="SAME-ISIN", security_code="BOND", close_date=date(2026, 2, 1), nominal_value=Decimal("100"), quantity=Decimal("3"), instrument_currency="KZT", next_coupon_date=None),
    ]
    rows = _expected_cash_flow_rows(snapshot, lots)
    assert rows == [["BOND", "Погашение", date(2026, 2, 1), Decimal("500"), "KZT"]]


def test_expected_cash_flows_include_bloomberg_dividends_by_pay_date_and_tax():
    snapshot = SimpleNamespace(report_date=date(2026, 7, 1))
    lot = SimpleNamespace(
        isin="ETF-ISIN", security_code="SPY US", purchase_date=date(2026, 1, 1),
        quantity=Decimal("10"), instrument_currency="USD", close_date=None,
        next_coupon_date=None,
    )
    history = DividendHistory([
        DividendRecord("SPY US Equity", Decimal("2"), None, date(2026, 6, 10), None, date(2026, 7, 15), "REGULAR"),
        # The lot was not held on this ex-date, so this payment is excluded.
        DividendRecord("SPY US Equity", Decimal("3"), None, date(2025, 12, 10), None, date(2026, 7, 20), "REGULAR"),
        # Outside the report-date + 180-day window.
        DividendRecord("SPY US Equity", Decimal("4"), None, date(2026, 7, 10), None, date(2027, 2, 1), "REGULAR"),
    ])
    rows = _expected_cash_flow_rows(snapshot, [lot], dividend_history=history)
    assert rows == [["SPY US", "Дивиденд (Bloomberg)", date(2026, 7, 15), Decimal("17.00"), "USD"]]


def test_expected_cash_flows_mark_group_unavailable_if_any_lot_lacks_an_amount():
    snapshot = SimpleNamespace(report_date=date(2026, 1, 1))
    lots = [
        SimpleNamespace(isin="SAME-ISIN", security_code="BOND", close_date=date(2026, 2, 1), nominal_value=Decimal("100"), quantity=Decimal("2"), instrument_currency="KZT", next_coupon_date=None),
        SimpleNamespace(isin="SAME-ISIN", security_code="BOND", close_date=date(2026, 2, 1), nominal_value=None, quantity=Decimal("3"), instrument_currency="KZT", next_coupon_date=None),
    ]
    rows = _expected_cash_flow_rows(snapshot, lots)
    assert rows == [["BOND", "Погашение", date(2026, 2, 1), "Недоступно", "KZT"]]


def test_expected_cash_flows_label_repo_closes_distinctly():
    snapshot = SimpleNamespace(report_date=date(2026, 1, 1))
    repo_lot = SimpleNamespace(
        isin="REPO-ISIN", security_code="A_KFUSb114", close_date=date(2026, 2, 1),
        nominal_value=Decimal("100"), quantity=Decimal("1"), instrument_currency="KZT", next_coupon_date=None,
        raw_security_type="авторепо", raw_sector="", instrument=SimpleNamespace(normalized_asset_class="Repo"),
    )
    rows = _expected_cash_flow_rows(snapshot, [repo_lot])
    assert rows == [["A_KFUSb114", "Закрытие репо", date(2026, 2, 1), Decimal("100"), "KZT"]]


def test_true_asset_class_repo_wins_over_a_stale_or_mismatched_dictionary_entry():
    # Repo ISINs roll over every period; even if a dictionary_class happens
    # to be supplied (e.g. a coincidental match to an old, unrelated entry),
    # the workbook's own "Repo" section label must win rather than the
    # dictionary silently reclassifying it as something else.
    assert true_asset_class("Repo", "авторепо", "", "A_KFUSb999", dictionary_class="Government bonds") == "Repo"
    assert true_asset_class("Repo", "авторепо", "", "A_KFUSb999", dictionary_class=None) == "Repo"


def test_write_distribution_block_total_is_unavailable_if_any_row_is():
    workbook = Workbook()
    sheet = workbook.active
    charts_sheet = workbook.create_sheet("Данные графиков")
    rows = [["KZT", Decimal("10"), Decimal("5")], ["EUR", Decimal("3"), "Недоступно"]]
    _write_distribution_block(
        sheet, 1, "Currencies", ["Currency", "MV KZT", "MV USD"], rows,
        widths=[18, 20, 20], numeric_formats={2: "0.00", 3: "0.00"}, chart_column=2, chart_title="Test",
        charts_sheet=charts_sheet, charts_row=1, table_name="ChartDataCurrencyTotal",
    )
    total_row_number = 1 + 1 + len(rows) + 1
    # Column 2 (MV KZT) has no gaps, so it still sums; column 3 (MV USD) has
    # one unresolved row, so the total must disclose that gap rather than
    # silently summing only the rows that happened to resolve.
    assert sheet.cell(total_row_number, 2).value == Decimal("13")
    assert sheet.cell(total_row_number, 3).value == "Недоступно"


def test_derive_expected_coupon_native_matches_real_osip_recalculation():
    # Regression fixture: values below are exactly what LibreOffice computes
    # when it recalculates OSIP's own (broken-cache) formula for these real
    # SOBSTV workbook rows - confirmed by converting the actual .xls with
    # headless LibreOffice and reading back column AS. The source's own
    # formula branches on the real "Купоннный период" (coupon period, in
    # days) column directly: /2 for a 180-day period, unchanged for 360,
    # /12 for 30, /4 for 90, and a pro-rata nominal*qty*rate/365*period
    # fallback for anything else (a "stub" period - see the stub-period
    # test below). An earlier version of this function *inferred* a
    # payment frequency from the calendar gap between coupon dates instead
    # of reading the period directly - that happened to match for these
    # four standard periods (rounding near 180/360/30/90 lands on the same
    # fraction either way) but silently diverged on a real stub period.
    annual = SimpleNamespace(
        coupon_or_repo_rate=Decimal("0.095"), nominal_value=Decimal("1000"), quantity=Decimal("416452"),
        coupon_period_days=Decimal("360"), coupon_indexation=Decimal("1"),
    )
    assert _derive_expected_coupon_native(annual) == Decimal("39562940.000")

    semi_annual = SimpleNamespace(
        coupon_or_repo_rate=Decimal("0.19"), nominal_value=Decimal("1000"), quantity=Decimal("400000"),
        coupon_period_days=Decimal("180"), coupon_indexation=Decimal("1"),
    )
    assert _derive_expected_coupon_native(semi_annual) == Decimal("38000000.0")

    quarterly = SimpleNamespace(
        coupon_or_repo_rate=Decimal("0.23"), nominal_value=Decimal("10000"), quantity=Decimal("9926"),
        coupon_period_days=Decimal("90"), coupon_indexation=Decimal("1"),
    )
    assert _derive_expected_coupon_native(quarterly) == Decimal("5707450.00")


def test_derive_expected_coupon_native_handles_a_real_stub_period():
    # Regression: ASDBe25, a real ADB bond in a live SOBSTV workbook, has a
    # 211-day first coupon period (neither 180/360/30/90). The old
    # date-gap-inferred-frequency version rounded 211 days to a semi-annual
    # (frequency=2) schedule and understated this lot's expected coupon by
    # 13.5% (3.49M KZT) versus what LibreOffice's live recalculation of
    # OSIP's own formula gives - confirmed on the real cell.
    stub = SimpleNamespace(
        coupon_or_repo_rate=Decimal("0.149"), nominal_value=Decimal("1"), quantity=Decimal("300000000"),
        coupon_period_days=Decimal("211"), coupon_indexation=Decimal("1"),
    )
    assert _derive_expected_coupon_native(stub) == Decimal("25840273.972602739726027397") \
        or abs(_derive_expected_coupon_native(stub) - Decimal("25840273.97")) < Decimal("0.01")


def test_derive_expected_coupon_native_requires_a_coupon_period():
    # A first coupon with no recorded period has no basis to derive from -
    # must not guess one rather than silently fabricating a number.
    lot = SimpleNamespace(
        coupon_or_repo_rate=Decimal("0.095"), nominal_value=Decimal("1000"), quantity=Decimal("2"),
        coupon_period_days=None, coupon_indexation=Decimal("1"),
    )
    assert _derive_expected_coupon_native(lot) is None


def test_expected_cash_flow_rows_fall_back_to_derived_coupon_when_source_blank():
    snapshot = SimpleNamespace(report_date=date(2026, 1, 1))
    lot = SimpleNamespace(
        isin="COUPON-ISIN", security_code="BOND", close_date=None,
        next_coupon_date=date(2026, 6, 1), instrument_currency="KZT",
        coupon_or_repo_rate=Decimal("0.1"), nominal_value=Decimal("1000"), quantity=Decimal("1"),
        coupon_period_days=Decimal("182"), coupon_indexation=Decimal("1"), source_row=None,
    )
    rows = _expected_cash_flow_rows(snapshot, [lot])
    assert len(rows) == 1
    assert rows[0][0] == "BOND"
    assert rows[0][1] == "Купон"
    assert isinstance(rows[0][3], Decimal) and rows[0][3] > 0


def test_cash_calendar_export_excludes_upcoming_settlements():
    # Regression: the sheet's own header claims "Предстоящие расчёты:
    # Исключены" but nothing filtered settlement-type events out - they
    # only ever "worked" by accident because no test snapshot had any.
    snapshot = SimpleNamespace(portfolio_code="SOBSTV", report_date=date(2026, 1, 1), version=1)
    calendar_items = [
        {"event_type": "settlement", "event_date": "2026-02-01", "status": "upcoming", "security_code": "X", "isin": "X1", "amount_native": None, "amount_kzt": None, "currency": "KZT", "amount_basis": "unavailable", "source_refs": []},
        {"event_type": "maturity", "event_date": "2026-03-01", "status": "upcoming", "security_code": "Y", "isin": "Y1", "amount_native": None, "amount_kzt": None, "currency": "KZT", "amount_basis": "unavailable", "source_refs": []},
    ]
    content = create_cash_calendar_xlsx(snapshot, [], calendar_items, include_inactive=False)
    from openpyxl import load_workbook
    import io
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook["Календарь событий"]
    security_codes = [sheet.cell(row, 3).value for row in range(7, sheet.max_row + 1) if sheet.cell(row, 3).value]
    assert "X" not in security_codes
    assert "Y" in security_codes
