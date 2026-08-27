from dataclasses import dataclass
from datetime import date
from decimal import Decimal

from openpyxl import Workbook

from osip_dashboard.services.holdings_export import _append_control_sheet


@dataclass
class _FakeSnapshot:
    portfolio_code: str
    report_date: date
    version: int


def _snapshot() -> _FakeSnapshot:
    return _FakeSnapshot(portfolio_code="SOBSTV", report_date=date(2026, 7, 20), version=1)


def _items() -> list[dict]:
    return [
        {
            "isin": "KZ2C00012904", "security_code": "NITCb1", "issuer": "АО Тест",
            "source_refs": [{"workbook_name": "test.xls", "sheet_name": "ОСИП_ПОРТФЕЛЬ", "row_number": 5}],
        },
        {
            "isin": "KZKD00001210", "security_code": "A_MUM072_0014", "issuer": "МФ РК",
            "source_refs": [
                {"workbook_name": "test.xls", "sheet_name": "ОСИП_ПОРТФЕЛЬ", "row_number": 6},
                {"workbook_name": "test.xls", "sheet_name": "ОСИП_ПОРТФЕЛЬ", "row_number": 7},
            ],
        },
    ]


def test_control_sheet_reports_the_same_totals_passed_in():
    workbook = Workbook()
    _append_control_sheet(
        workbook, _snapshot(), _items(),
        total_value=Decimal("1000000"), total_weight=Decimal("1"), weight_validation="ОК — 100.0%",
    )
    sheet = workbook["Контроль и происхождение"]
    values = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in range(1, 20) if sheet.cell(row, 1).value}
    assert values["Итоговая стоимость, KZT"] == Decimal("1000000")
    assert values["Итоговый вес"] == Decimal("1")
    assert values["Ожидаемый вес"] == Decimal("1")
    assert values["Разница"] == Decimal("0")
    assert values["Статус проверки"] == "ОК — 100.0%"
    assert values["Количество ISIN"] == 2


def test_control_sheet_lists_every_instrument_with_its_source():
    workbook = Workbook()
    _append_control_sheet(
        workbook, _snapshot(), _items(),
        total_value=Decimal("1000000"), total_weight=Decimal("1"), weight_validation="ОК — 100.0%",
    )
    sheet = workbook["Контроль и происхождение"]
    isins = {sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1)}
    assert "KZ2C00012904" in isins
    assert "KZKD00001210" in isins
    rows_by_isin = {sheet.cell(row, 1).value: row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value in {"KZ2C00012904", "KZKD00001210"}}
    multi_source_row = rows_by_isin["KZKD00001210"]
    assert "6" in str(sheet.cell(multi_source_row, 6).value) and "7" in str(sheet.cell(multi_source_row, 6).value)


def test_control_sheet_reports_a_nonzero_difference_when_weight_is_off():
    workbook = Workbook()
    _append_control_sheet(
        workbook, _snapshot(), _items(),
        total_value=Decimal("1000000"), total_weight=Decimal("0.97"), weight_validation="Фильтрованный вид — 97.0000%",
    )
    sheet = workbook["Контроль и происхождение"]
    values = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in range(1, 20) if sheet.cell(row, 1).value}
    assert values["Разница"] == Decimal("-0.03")
