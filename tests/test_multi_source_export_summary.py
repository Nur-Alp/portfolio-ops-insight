from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
import io
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from osip_dashboard.services import multi_source_export
from osip_dashboard.services.fx_rates import FxRate
from osip_dashboard.services.multi_source_export import _fields, _write_full_mix_table, _write_module_summary
from osip_dashboard.services.date_provenance import extract_filename_date, filename_date_mismatch


@dataclass
class _FakeRecord:
    payload: dict[str, Any]
    record_type: str = ""


@dataclass
class _FakeDataset:
    dataset_type: str
    summary: dict[str, Any] = field(default_factory=dict)
    records: list[_FakeRecord] = field(default_factory=list)
    business_date: date | None = None


def test_asset_management_summary_adds_unit_value_and_allocation_charts():
    unit_series = _FakeDataset("fund_unit_series", records=[
        _FakeRecord({"date": "2026-07-01", "unit_value_kzt": "100.0"}),
        _FakeRecord({"date": "2026-07-02", "unit_value_kzt": "101.5"}),
    ])
    holdings = _FakeDataset("fund_holdings", records=[
        _FakeRecord({"currency": "KZT", "purchase_value_kzt": "1000000"}),
        _FakeRecord({"currency": "USD", "purchase_value_kzt": "200000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_unit_series": unit_series, "fund_holdings": holdings})
    sheet = workbook["Сводка фонда"]
    assert len(sheet._charts) == 2


def test_asset_management_summary_is_skipped_when_no_relevant_datasets():
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"unrelated": _FakeDataset("unrelated")})
    assert "Сводка фонда" not in workbook.sheetnames


def test_unique_etf_positions_sheet_aggregates_repeat_lots_by_isin():
    holdings = _FakeDataset("fund_holdings", records=[
        _FakeRecord({"instrument": "SPTL US Equity", "isin": "US78464A6644", "currency": "USD", "quantity": "137", "purchase_value_kzt": "3531.86", "purchase_value_native": "3531.86", "purchase_date": "2023-11-16"}),
        _FakeRecord({"instrument": "SPTL US Equity", "isin": "US78464A6644", "currency": "USD", "quantity": "352", "purchase_value_kzt": "9539.20", "purchase_value_native": "9539.20", "purchase_date": "2023-12-01"}),
        _FakeRecord({"instrument": "QQQ US Equity", "isin": "US46090E1038", "currency": "USD", "quantity": "17", "purchase_value_kzt": "6676.92", "purchase_value_native": "6676.92", "purchase_date": "2023-12-04"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_holdings": holdings})
    sheet = workbook["Уникальные позиции ETF"]
    rows = {sheet.cell(r, 2).value: [sheet.cell(r, c).value for c in range(1, 9)] for r in range(6, sheet.max_row + 1) if sheet.cell(r, 2).value}
    sptl = rows["US78464A6644"]
    assert sptl[0] == "SPTL US Equity"
    assert sptl[3] == Decimal("489")  # 137 + 352
    assert sptl[4] == Decimal("13071.06")  # 3531.86 + 9539.20
    assert sptl[6] == 2  # lot count
    assert sptl[7] == date(2023, 11, 16)  # earliest purchase date, as a real date not a string
    qqq = rows["US46090E1038"]
    assert qqq[6] == 1
    assert len(rows) == 2
    sptl_row = next(r for r in range(6, sheet.max_row + 1) if sheet.cell(r, 2).value == "US78464A6644")
    # A whole-share quantity must render as a plain integer, not a mask
    # with optional decimal places - the latter left a dangling "300."
    # trailing decimal point in Excel even though the value is exact.
    assert sheet.cell(sptl_row, 4).number_format == "#,##0;[Red](#,##0);-"
    assert sheet.cell(sptl_row, 8).number_format == "dd.mm.yyyy"


def test_module_summary_sheets_follow_the_workbook_navigation_standard():
    # Dashboard summary sheets (stacked tables feeding charts) freeze
    # nothing - read chart-first, not scrolled as a long table; the ETF
    # sheet is a genuine continuous table and pins through its header row
    # instead. See the "Workbook navigation standard" section in
    # docs/export-column-audit.md.
    holdings = _FakeDataset("fund_holdings", records=[
        _FakeRecord({"instrument": "SPTL US Equity", "isin": "US78464A6644", "currency": "USD", "quantity": "137", "purchase_value_kzt": "3531.86", "purchase_value_native": "3531.86", "purchase_date": "2023-11-16"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_holdings": holdings})
    assert workbook["Сводка фонда"].freeze_panes is None
    etf_sheet = workbook["Уникальные позиции ETF"]
    assert etf_sheet.freeze_panes != "B1"
    assert etf_sheet.freeze_panes == "B6"


def test_unique_etf_positions_sheet_skipped_when_no_holdings():
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"unrelated": _FakeDataset("unrelated")})
    assert "Уникальные позиции ETF" not in workbook.sheetnames


def test_risk_export_gives_large_numeric_columns_enough_width_for_excel():
    dataset = SimpleNamespace(
        dataset_type="risk_limits_sobstv",
        summary={},
        records=[SimpleNamespace(
            payload={
                "portfolio_code": "SOBSTV",
                "dimension": "instrument_category",
                "label": "Финансовые инструменты",
                "limit_pct": "0.5",
                "limit_kzt": "4763236555.94",
                "actual_pct": "0.4051",
                "actual_kzt": "1929618863.37",
                "free_limit_kzt": "2833617692.57",
                "signal": "OK",
                "near_breach_policy_version": "utilization-ratio-v1",
            },
            record_type="risk_limit",
            source_ref={"sheet_name": "Лимиты", "row_number": 9},
        )],
        business_date=date(2026, 7, 1),
        source_report_date=date(2026, 7, 1),
        version=3,
        scope_code="SOBSTV",
        source_upload=SimpleNamespace(original_filename="risk-sobstv.xlsx"),
    )

    workbook = load_workbook(io.BytesIO(multi_source_export.create_module_xlsx("risk", [dataset])))
    # Raw risk data is split one sheet per dimension (see create_module_xlsx)
    # rather than one combined "Лимиты SOBSTV" sheet - this row's dimension
    # is instrument_category.
    sheet = workbook["SOBSTV · Классы инструментов"]

    assert sheet.freeze_panes == "B9"
    headers = {cell.value: cell.column for cell in sheet[8]}
    limit_kzt_column = headers["Лимит, KZT"]
    assert sheet.cell(9, limit_kzt_column).value == float(Decimal("4763236555.94"))
    assert sheet.cell(9, limit_kzt_column).number_format == "#,##0.00;[Red](#,##0.00);-"
    assert sheet.cell(9, headers["Лимит, %"]).number_format == '0.00"%";[Red](0.00"%");-'
    assert sheet.cell(9, headers["Факт, %"]).number_format == '0.00"%";[Red](0.00"%");-'
    assert sheet.column_dimensions[get_column_letter(limit_kzt_column)].width >= 20
    assert sheet.column_dimensions[get_column_letter(headers["Факт, KZT"])].width >= 20
    assert sheet.column_dimensions[get_column_letter(headers["Свободный лимит, KZT"])].width >= 22
    policy_column = headers["Версия политики порога близости к превышению"]
    assert sheet.cell(9, policy_column).value == "utilization-ratio-v1"


def test_accounting_summary_reports_reconciling_totals_and_charts():
    balance_sheet = _FakeDataset("accounting_balance_sheet", records=[
        _FakeRecord({"line_code": "25", "line_label": "Итого активы", "current_period_kzt": "4200000"}),
        _FakeRecord({"line_code": "42", "line_label": "Итого обязательства", "current_period_kzt": "200000"}),
        _FakeRecord({"line_code": "52", "line_label": "Итого капитал", "current_period_kzt": "4000000"}),
    ])
    income_statement = _FakeDataset("accounting_income_statement", records=[
        _FakeRecord({"line_code": "13", "line_label": "Итого доходов", "quarter_kzt": "287651"}),
        _FakeRecord({"line_code": "28", "line_label": "Итого расходов", "quarter_kzt": "184920"}),
        _FakeRecord({"line_code": "29", "line_label": "Чистая прибыль (убыток) до уплаты корпоративного подоходного налога", "quarter_kzt": "102731"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "accounting", {"accounting_balance_sheet": balance_sheet, "accounting_income_statement": income_statement})
    sheet = workbook["Сводка ФО"]
    values = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value}
    assert values["Активы = Обязательства + Капитал"] == "Сходится"
    assert values["Доходы − Расходы = Чистая прибыль"] == "Сходится"
    assert len(sheet._charts) == 2


def test_accounting_summary_flags_a_balance_sheet_mismatch():
    balance_sheet = _FakeDataset("accounting_balance_sheet", records=[
        _FakeRecord({"line_code": "25", "line_label": "Итого активы", "current_period_kzt": "9999999"}),
        _FakeRecord({"line_code": "42", "line_label": "Итого обязательства", "current_period_kzt": "200000"}),
        _FakeRecord({"line_code": "52", "line_label": "Итого капитал", "current_period_kzt": "4000000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "accounting", {"accounting_balance_sheet": balance_sheet})
    sheet = workbook["Сводка ФО"]
    values = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value}
    assert values["Активы = Обязательства + Капитал"] == "Расхождение"


def test_accounting_summary_adds_the_management_balance_sheet_rollup():
    # A minimal but real-shaped balance sheet: one line per category the
    # rollup groups (see _CONDENSED_BALANCE_SHEET_LINES in
    # services/multi_source.py) plus the three source totals, each tagged
    # record_type="balance_sheet_line" and a section - exactly what the real
    # ingestion parser (_parse_accounting_balance_sheet) produces.
    balance_sheet = _FakeDataset("accounting_balance_sheet", records=[
        _FakeRecord({"line_code": "1", "line_label": "Денежные средства", "section": "Активы", "current_period_kzt": "1000000", "prior_period_kzt": "900000"}, record_type="balance_sheet_line"),
        _FakeRecord({"line_code": "2", "line_label": "Ценные бумаги", "section": "Активы", "current_period_kzt": "3200000", "prior_period_kzt": "3100000"}, record_type="balance_sheet_line"),
        _FakeRecord({"line_code": "25", "line_label": "Итого активы", "section": "Активы", "current_period_kzt": "4200000", "prior_period_kzt": "4000000"}, record_type="balance_sheet_line"),
        _FakeRecord({"line_code": "26", "line_label": "Займы полученные", "section": "Обязательства", "current_period_kzt": "150000", "prior_period_kzt": "140000"}, record_type="balance_sheet_line"),
        _FakeRecord({"line_code": "42", "line_label": "Итого обязательства", "section": "Обязательства", "current_period_kzt": "200000", "prior_period_kzt": "190000"}, record_type="balance_sheet_line"),
        _FakeRecord({"line_code": "43", "line_label": "Уставный капитал", "section": "Собственный капитал", "current_period_kzt": "4000000", "prior_period_kzt": "3800000"}, record_type="balance_sheet_line"),
        _FakeRecord({"line_code": "52", "line_label": "Итого капитал", "section": "Собственный капитал", "current_period_kzt": "4000000", "prior_period_kzt": "3810000"}, record_type="balance_sheet_line"),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "accounting", {"accounting_balance_sheet": balance_sheet})
    sheet = workbook["Сводка ФО"]
    values = {sheet.cell(row, 1).value: sheet.cell(row, 2).value for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value}
    # Category rows are scaled to full KZT the same way the KPI totals above
    # are (source states figures in thousands) - 1,000,000 thousand -> 1e9.
    assert values["Денежные средства"] == Decimal("1000000000")
    assert values["ИТОГО АКТИВЫ"] == Decimal("4200000000")
    assert len(sheet._charts) == 2  # Состав баланса pie + the new rollup bar (no income statement in this fixture)


def test_accounting_summary_skips_the_management_rollup_when_no_lines_resolved():
    # No record here carries record_type="balance_sheet_line" (the fixture
    # only reads the KPI totals by line_label) - the rollup must not render
    # an all-zero 13-row table in that case.
    balance_sheet = _FakeDataset("accounting_balance_sheet", records=[
        _FakeRecord({"line_code": "25", "line_label": "Итого активы", "current_period_kzt": "4200000"}),
        _FakeRecord({"line_code": "42", "line_label": "Итого обязательства", "current_period_kzt": "200000"}),
        _FakeRecord({"line_code": "52", "line_label": "Итого капитал", "current_period_kzt": "4000000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "accounting", {"accounting_balance_sheet": balance_sheet})
    sheet = workbook["Сводка ФО"]
    assert "Управленческий баланс" not in {sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1)}


def test_risk_summary_adds_all_six_conditional_charts():
    def dataset(scope_code: str, records: list[dict[str, Any]]) -> SimpleNamespace:
        return SimpleNamespace(
            dataset_type=f"risk_limits_{scope_code.casefold()}",
            summary={},
            records=[SimpleNamespace(payload=payload, record_type="risk_limit", source_ref={"sheet_name": "Лимиты", "row_number": index + 9}) for index, payload in enumerate(records)],
            business_date=date(2026, 7, 1),
            source_report_date=date(2026, 7, 1),
            version=3,
            scope_code=scope_code,
            source_upload=SimpleNamespace(original_filename=f"risk-{scope_code.casefold()}.xlsx"),
        )

    sobstv = dataset("SOBSTV", [
        {"portfolio_code": "SOBSTV", "dimension": "country", "label": "Казахстан", "limit_pct": "1", "actual_pct": "0.8", "utilization": "0.8", "signal": "OK"},
        {"portfolio_code": "SOBSTV", "dimension": "exposure_detail", "label": "USD cash", "currency": "USD", "amount_native": "1000", "fx_rate": "500", "amount_kzt": "500000", "signal": "not_applicable"},
        {"portfolio_code": "SOBSTV", "dimension": "duration", "label": "Bond A", "issuer": "Issuer A", "isin": "KZ0001", "modified_duration": "4", "duration_limit": "5", "duration_headroom": "1", "signal": "OK"},
    ])
    tabys = dataset("TABYS", [
        {"portfolio_code": "TABYS", "dimension": "issuer", "label": "Issuer B", "limit_kzt": "100", "actual_kzt": "120", "signal": "breach"},
    ])

    workbook = load_workbook(io.BytesIO(multi_source_export.create_module_xlsx("risk", [sobstv, tabys])))
    sheet = workbook["Сводка по лимитам"]
    titles = [chart.title.tx.rich.p[0].r[0].t for chart in sheet._charts]

    assert len(titles) == 6
    assert "Статусы лимитов по измерению" in titles
    assert "Сравнение портфелей по статусам" in titles
    assert "Использование лимита, топ-10" in titles
    # SOBSTV's "country" row has utilization, so it gets a per-dimension
    # utilization chart; TABYS's "issuer" row here has none, so it doesn't.
    assert "Использование лимита: Страна · SOBSTV" in titles
    assert "Валютная экспозиция, KZT" in titles
    assert "Модифицированная дюрация против лимита" in titles
    status_chart = next(chart for chart in sheet._charts if chart.title.tx.rich.p[0].r[0].t == "Статусы лимитов по измерению")
    assert status_chart.type == "bar"
    assert status_chart.legend is not None
    assert status_chart.x_axis.majorGridlines is not None
    assert status_chart.visible_cells_only is True
    assert status_chart.y_axis.numFmt.formatCode == "#,##0"
    assert status_chart.x_axis.numFmt is None
    assert ":" in status_chart.series[0].cat.numRef.f
    status_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Статусы лимитов по измерению")
    portfolio_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Сравнение портфелей")
    status_values = [sheet.cell(row, 1).value for row in range(status_row + 2, portfolio_row) if sheet.cell(row, 2).value is not None]
    assert not any(str(value).startswith("Детализация") for value in status_values)
    utilization_chart = next(chart for chart in sheet._charts if chart.title.tx.rich.p[0].r[0].t == "Использование лимита, топ-10")
    assert utilization_chart.y_axis.numFmt.formatCode == "0%"
    assert portfolio_row - status_row >= 30


def test_unit_value_chart_uses_the_full_history_from_a_separate_charts_sheet():
    # Regression: the history used to be capped to the most recent 60 rows
    # and written inline on "Сводка фонда", pushing every later section down
    # by however many rows were kept. The full series must now live on
    # "Данные графиков" instead, uncapped.
    records = [_FakeRecord({"date": f"2026-01-{day:02d}", "unit_value_kzt": str(100 + day)}) for day in range(1, 32)] + \
        [_FakeRecord({"date": f"2026-02-{day:02d}", "unit_value_kzt": str(140 + day)}) for day in range(1, 29)]
    unit_series = _FakeDataset("fund_unit_series", records=records)
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_unit_series": unit_series})
    sheet = workbook["Сводка фонда"]
    charts_sheet = workbook["Данные графиков"]
    # No 60-row (or any-row) date table inline on the summary sheet itself.
    inline_dates = [sheet.cell(row, 1).value for row in range(1, sheet.max_row + 1) if isinstance(sheet.cell(row, 1).value, str) and sheet.cell(row, 1).value.startswith("2026-")]
    assert inline_dates == []
    charted_dates = [charts_sheet.cell(row, 1).value for row in range(1, charts_sheet.max_row + 1) if isinstance(charts_sheet.cell(row, 1).value, str) and charts_sheet.cell(row, 1).value.startswith("2026-")]
    assert len(charted_dates) == len(records)
    assert charted_dates[0] == "2026-01-01"
    assert charted_dates[-1] == "2026-02-28"


def test_unit_value_line_chart_has_visible_axis_labels():
    # Regression: the line chart never set axis delete/numFmt, so both axes
    # rendered with gridlines but no tick labels at all in real Excel - the
    # same fix write_bar_chart already applies.
    unit_series = _FakeDataset("fund_unit_series", records=[
        _FakeRecord({"date": "2026-07-01", "unit_value_kzt": "100.0"}),
        _FakeRecord({"date": "2026-07-02", "unit_value_kzt": "101.5"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_unit_series": unit_series})
    sheet = workbook["Сводка фонда"]
    chart = sheet._charts[0]
    assert chart.x_axis.delete is False
    assert chart.y_axis.delete is False
    assert chart.y_axis.numFmt.formatCode == "#,##0.00"
    # Regression: an unset chart.style faded the stroke from dark to pale
    # across the series in real Excel (a built-in style effect, not
    # something this code ever asked for) - the series color must be an
    # explicit solid fill instead.
    assert chart.series[0].graphicalProperties.line.solidFill.srgbClr == "9226A8"


def test_asset_management_summary_adds_a_nav_history_chart_when_nav_is_present():
    unit_series = _FakeDataset("fund_unit_series", records=[
        _FakeRecord({"date": "2026-07-01", "unit_value_kzt": "100.0", "nav_kzt": "5000000"}),
        _FakeRecord({"date": "2026-07-02", "unit_value_kzt": "101.5", "nav_kzt": "5075000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_unit_series": unit_series})
    sheet = workbook["Сводка фонда"]
    assert len(sheet._charts) == 2
    titles = [chart.title.tx.rich.p[0].r[0].t for chart in sheet._charts]
    assert "Стоимость пая во времени, KZT" in titles
    assert "СЧА фонда во времени, KZT" in titles


def test_asset_management_summary_adds_nav_composition_bar_chart_from_valuation():
    # Liabilities is a deduction from NAV, not a positive share of it, so
    # this must be a bar (magnitude comparison), never a pie (whole-to-part).
    valuation = _FakeDataset("fund_valuation", summary={
        "securities_value_kzt": "900000000", "cash_kzt": "50000000", "liabilities_kzt": "10000000", "nav_kzt": "940000000",
    })
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_valuation": valuation})
    sheet = workbook["Сводка фонда"]
    assert len(sheet._charts) == 1
    from openpyxl.chart import BarChart
    chart = sheet._charts[0]
    assert isinstance(chart, BarChart)
    # Securities routinely dominates cash/liabilities by 2+ orders of
    # magnitude - a linear axis flattened the smaller two to invisible
    # slivers (confirmed against a real export); log scale keeps all three
    # readable.
    assert chart.y_axis.scaling.logBase == 10.0


def test_brokerage_summary_never_combines_currencies_on_one_axis(monkeypatch):
    # No real network call in a unit test: KZT and USD both resolve to a
    # fixed test rate (real KZT-is-always-1 handling lives in fx_rates.py,
    # which this mock replaces entirely, so it's restated here).
    def _fake_resolve(currency: str, report_date, workbook_rate=None):
        rate = Decimal("1") if currency == "KZT" else Decimal("500") if currency == "USD" else None
        return FxRate(rate=rate, effective_date=report_date, source="test", source_url="") if rate is not None else None

    monkeypatch.setattr(multi_source_export, "resolve_export_fx_rate", _fake_resolve)
    ledger = _FakeDataset("brokerage_trade_ledger", business_date=date(2026, 7, 20), records=[
        _FakeRecord({"currency": "KZT", "amount": "1000000", "side": "Покупка", "venue": "KASE", "security_type": "Облигация", "execution_status": "Исполнена"}),
        _FakeRecord({"currency": "KZT", "amount": "800000", "side": "Продажа", "venue": "KASE", "security_type": "Облигация", "execution_status": "Исполнена"}),
        _FakeRecord({"currency": "USD", "amount": "5000", "side": "Покупка", "venue": "OTC", "security_type": "Акция", "execution_status": "Не исполнена"}),
        _FakeRecord({"currency": "USD", "amount": "3000", "side": "Продажа", "venue": "OTC", "security_type": "Облигация", "execution_status": "Исполнена"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "brokerage", {"brokerage_trade_ledger": ledger})
    sheet = workbook["Сводка брокерской деятельности"]
    # All native Excel charts are presentation objects on the summary sheet;
    # their editable source tables live separately on the chart-data sheet.
    assert len(sheet._charts) == 5
    assert len(workbook["Данные графиков"]._charts) == 0
    assert sheet._charts[0].anchor == "I9"  # column I, beside the turnover table (which itself ends at G)
    header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Валюта")
    header_row_values = [sheet.cell(header_row, column).value for column in range(1, 6)]
    # No "KZT" claim on the native-currency columns - amounts stay in each
    # trade's own currency (see the "Валюта" column); the KZT equivalent is
    # its own explicitly-labelled column.
    assert header_row_values == ["Валюта", "Покупка", "Продажа", "Покупка, ≈KZT", "Продажа, ≈KZT"]
    kzt_row = next(row for row in range(header_row + 1, sheet.max_row + 1) if sheet.cell(row, 1).value == "KZT")
    usd_row = next(row for row in range(header_row + 1, sheet.max_row + 1) if sheet.cell(row, 1).value == "USD")
    # KZT converts to itself; USD converts at the mocked 500 KZT rate.
    assert sheet.cell(kzt_row, 4).value == Decimal("1000000")
    assert sheet.cell(usd_row, 4).value == Decimal("2500000")
    # Stocks vs bonds: bonds = 1,000,000 + 800,000 KZT + 3,000 USD * 500 =
    # 3,300,000; stocks = 5,000 USD * 500 = 2,500,000 - same classification
    # and NBK-rate conversion as the web brokerage chart.
    stock_bond_title_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Акции против облигаций")
    stock_bond_data_row = stock_bond_title_row + 2
    assert sheet.cell(stock_bond_data_row, 1).value == "Облигации"
    assert sheet.cell(stock_bond_data_row, 2).value == Decimal("3300000")
    assert sheet.cell(stock_bond_data_row + 1, 1).value == "Акции"
    assert sheet.cell(stock_bond_data_row + 1, 2).value == Decimal("2500000")
    # Every brokerage chart must be backed by an editable table on the final
    # chart-data sheet, never by a visible operational summary table.
    assert all(
        "'Данные графиков'" in series.val.numRef.f
        for chart in sheet._charts
        for series in chart.ser
        if series.val is not None and series.val.numRef is not None
    )


def test_brokerage_export_fields_exclude_client_information():
    fields = _fields("brokerage_trade_ledger", [_FakeRecord({
        "trade_number": "D-001", "trade_date": "2026-07-20", "client_name": "Sensitive client",
        "account": "ACC-001", "iin": "000000000000", "counterparty": "Sensitive counterparty",
        "venue": "KASE", "instrument": "Bond", "security_type": "Облигация", "isin": "KZ0001",
        "amount": "100", "currency": "KZT", "execution_status": "Исполнено",
    })])
    assert "client_name" not in fields
    assert "account" not in fields
    assert "iin" not in fields
    assert "counterparty" not in fields
    assert "venue" in fields
    assert "security_type" in fields


def test_brokerage_export_adds_a_non_pii_six_month_maturity_calendar():
    class _DatasetRecord:
        def __init__(self, payload, row_number):
            self.payload = payload
            self.record_type = "maturity"
            self.source_ref = {"sheet_name": "Календарь погашения", "row_number": row_number}

    class _Dataset:
        def __init__(self, dataset_type, records, business_date):
            self.dataset_type = dataset_type
            self.summary = {}
            self.records = records
            self.business_date = business_date
            self.source_report_date = None
            self.version = 1
            self.scope_code = "BROKERAGE"
            self.source_upload = type("Upload", (), {"original_filename": "brokerage.xls"})()

    as_of = date(2026, 7, 20)
    maturity = _Dataset("client_maturity_calendar", [
        _DatasetRecord({"client_name": "Скрытый клиент", "manager": "Скрытый менеджер", "instrument": "Later", "isin": "ISIN-181", "maturity_date": "2027-01-17", "value_kzt": "181"}, 14),
        _DatasetRecord({"client_name": "Скрытый клиент", "manager": "Скрытый менеджер", "instrument": "At start", "isin": "ISIN-000", "maturity_date": "2026-07-20", "value_kzt": "1"}, 10),
        _DatasetRecord({"client_name": "Скрытый клиент", "manager": "Скрытый менеджер", "instrument": "At end", "isin": "ISIN-180", "maturity_date": "2027-01-16", "value_kzt": "180"}, 13),
        _DatasetRecord({"client_name": "Скрытый клиент", "manager": "Скрытый менеджер", "instrument": "Too early", "isin": "ISIN-OLD", "maturity_date": "2026-07-19", "value_kzt": "0"}, 9),
    ], as_of)
    workbook_bytes = multi_source_export.create_module_xlsx("brokerage", [maturity])
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
    sheet = workbook["Календарь погашения"]

    headers = [sheet.cell(8, column).value for column in range(1, 10)]
    assert headers[:2] == ["Инструмент", "ISIN"]
    assert "Клиент" not in headers
    assert "Менеджер" not in headers
    assert [sheet.cell(row, 1).value for row in range(9, 11)] == ["At start", "At end"]
    assert sheet.cell(9, 3).value.date() == date(2026, 7, 20)
    assert sheet.cell(10, 3).value.date() == date(2027, 1, 16)
    assert sheet.cell(7, 1).value == (
        "Период: с 20.07.2026 по 16.01.2027 включительно; "
        "на этом листе идентификаторы клиентов и менеджеров не показаны "
        "(на листах с клиентскими счетами и позициями эти данные присутствуют)."
    )


def test_fund_unit_series_fields_include_investors():
    # Regression: the preferred list used to say "investor_count", a field
    # that never existed in the payload (the real key is "investors") - the
    # column silently vanished from the export instead of erroring, since
    # `_fields` only filters a preferred list down, it doesn't validate it.
    fields = _fields("fund_unit_series", [_FakeRecord({
        "date": "2026-07-01", "nav_kzt": "100", "units": "1", "investors": "5", "unit_value_kzt": "100",
    })])
    assert "investors" in fields


def test_fields_drops_a_column_that_is_blank_on_every_record():
    # Regression: "available" only checked whether a payload KEY existed,
    # not whether it ever had a real value - ingestion always sets a
    # "branch" key on client_account_snapshot_client rows even when every
    # source cell was blank, so a fully-empty column rendered as a real
    # column full of "Недоступно" instead of being omitted.
    records = [
        _FakeRecord({"client_name": "ТОО Ромашка", "account": "AC1", "branch": ""}),
        _FakeRecord({"client_name": "ТОО Василёк", "account": "AC2", "branch": ""}),
    ]
    fields = _fields("client_account_snapshot_client", records)
    assert "branch" not in fields
    assert "client_name" in fields and "account" in fields


def test_client_open_dates_fields_match_the_real_payload_shape():
    # Regression: the preferred list named client_name/account/opening_date,
    # none of which this dataset's ingestion ever produces (it emits
    # source_name/normalized_name/open_date/match_status) - every preferred
    # field missed, so _fields fell through to an alphabetically-sorted,
    # unlabelled dump of the raw field names instead of a curated view.
    #
    # match_status is deliberately excluded from the export: it's an
    # internal reconciliation signal nothing downstream depends on, and
    # this is a local dashboard, not a formal reconciliation process - an
    # "unmatched" label with no action attached just raises questions.
    fields = _fields("client_open_dates", [_FakeRecord({
        "normalized_name": "IVANOV I", "source_name": "Иванов И.", "open_date": "2024-01-15", "match_status": "exact",
    })])
    assert fields == ["source_name", "normalized_name", "open_date"]


def test_client_account_snapshot_fields_surface_record_type_and_position_detail():
    # Regression: "record_type" is a DatasetRecord column, not a payload key,
    # so it never matched `available` and silently dropped; "instrument" was
    # never a real field name for this dataset (position rows use issuer/
    # security_type/isin instead), so it dropped too - the sheet had no way
    # to tell a client-summary row from a position line item.
    fields = _fields("client_account_snapshot", [_FakeRecord({
        "client_name": "ТОО Ромашка", "account": "ACC-1", "iin": "123", "branch": "Алматы",
        "manager": "Иванов", "cash_kzt": "100", "total_assets_kzt": "200",
    }), _FakeRecord({
        "client_name": "ТОО Ромашка", "account": "ACC-1", "issuer": "МФ РК", "security_type": "ГЦБ",
        "isin": "KZ0001", "quantity": "10", "market_value_kzt": "1000",
    })])
    assert "record_type" in fields
    assert "instrument" not in fields
    assert "issuer" in fields
    assert "security_type" in fields
    assert "isin" in fields
    assert "market_value_kzt" in fields


def test_field_value_resolves_record_type_from_the_record_not_the_payload():
    @dataclass
    class _FakeDatasetRecord:
        record_type: str
        payload: dict[str, Any]

    record = _FakeDatasetRecord(record_type="client_position", payload={"client_name": "ТОО Ромашка"})
    assert multi_source_export._field_value(record, "record_type") == "client_position"
    assert multi_source_export._field_value(record, "client_name") == "ТОО Ромашка"


def test_brokerage_turnover_shows_unavailable_when_nbk_has_no_rate(monkeypatch):
    monkeypatch.setattr(multi_source_export, "resolve_export_fx_rate", lambda currency, report_date, workbook_rate=None: None)
    ledger = _FakeDataset("brokerage_trade_ledger", business_date=date(2026, 7, 20), records=[
        _FakeRecord({"currency": "USD", "amount": "5000", "side": "Покупка", "venue": "OTC", "security_type": "Bond", "execution_status": "Исполнена"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "brokerage", {"brokerage_trade_ledger": ledger})
    sheet = workbook["Сводка брокерской деятельности"]
    header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Валюта")
    usd_row = header_row + 1
    assert sheet.cell(usd_row, 1).value == "USD"
    assert sheet.cell(usd_row, 4).value == "Недоступно"


def test_brokerage_summary_table_stays_ungrouped_while_the_chart_data_folds_the_tail():
    # One dominant instrument type (90 trades) plus 19 negligible ones (1
    # trade each, ~0.9% share) - each small type is well under the 2% "Прочее"
    # threshold and there are far more than 2 of them, so chart_series folds
    # them into a single row for the chart. The summary table itself must
    # keep every real category - only the separate chart-source sheet groups.
    records = [_FakeRecord({"security_type": "Bond", "execution_status": "Исполнена"}) for _ in range(90)]
    records += [_FakeRecord({"security_type": f"Type{index}", "execution_status": "Исполнена"}) for index in range(19)]
    ledger = _FakeDataset("brokerage_trade_ledger", records=records)
    workbook = Workbook()
    _write_module_summary(workbook, "brokerage", {"brokerage_trade_ledger": ledger})
    sheet = workbook["Сводка брокерской деятельности"]
    instrument_header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Состав по типам инструментов")
    category_rows = []
    row = instrument_header_row + 2
    while sheet.cell(row, 1).value:
        category_rows.append(sheet.cell(row, 1).value)
        row += 1
    assert category_rows == ["Bond"] + [f"Type{index}" for index in range(19)]

    charts_sheet = workbook["Данные графиков"]
    chart_instrument_header_row = next(row for row in range(1, charts_sheet.max_row + 1) if charts_sheet.cell(row, 1).value == "Состав по типам инструментов")
    chart_category_rows = []
    row = chart_instrument_header_row + 2
    while charts_sheet.cell(row, 1).value:
        chart_category_rows.append(charts_sheet.cell(row, 1).value)
        row += 1
    assert chart_category_rows == ["Bond", "Прочее"]

    # Every trade here shares one venue and one execution status, so those
    # two tables have a single category each and draw no chart (see
    # _write_full_mix_table) - only instrument_mix (20 categories) does, and
    # only a section that actually drew a chart needs the full 20-row gap
    # before the next section to avoid overlapping it.
    section_rows = [
        next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == title)
        for title in ("Площадки исполнения", "Состав по типам инструментов", "Статус исполнения сделок")
    ]
    assert section_rows[1] - section_rows[0] < 20
    assert section_rows[2] - section_rows[1] >= 20


def test_write_full_mix_table_skips_the_chart_for_a_single_category():
    workbook = Workbook()
    sheet = workbook.create_sheet("Summary")
    charts_sheet = workbook.create_sheet("Данные графиков")
    _write_full_mix_table(
        sheet, 2, "Статус исполнения сделок", {"исполненная": 8732},
        charts_sheet=charts_sheet, charts_row=4, table_name="ChartDataStatus",
        chart_title="Статус исполнения сделок",
    )
    assert len(sheet._charts) == 0
    assert len(charts_sheet._charts) == 0
    assert sheet.cell(3, 1).value == "Категория"
    assert sheet.cell(4, 1).value == "исполненная"
    assert sheet.cell(4, 2).value == Decimal("8732")


def test_write_full_mix_table_still_charts_two_or_more_categories():
    workbook = Workbook()
    sheet = workbook.create_sheet("Summary")
    charts_sheet = workbook.create_sheet("Данные графиков")
    _write_full_mix_table(
        sheet, 2, "Площадки исполнения", {"KASE": 5873, "OTC": 202},
        charts_sheet=charts_sheet, charts_row=4, table_name="ChartDataVenue",
        chart_title="Площадки исполнения",
    )
    assert len(sheet._charts) == 1


def test_clients_summary_uses_a_distinct_sheet_name_from_the_raw_dataset():
    # "Сводка клиентов" is already the display name for the raw
    # client_dashboard_snapshot dataset sheet (multi_source_export._sheet_name)
    # - the summary sheet must not collide with it in the same workbook.
    snapshot = _FakeDataset("client_account_snapshot", summary={
        "manager_mix": {"Иванов": {"total_assets_kzt": "500000"}, "Петров": {"total_assets_kzt": "300000"}},
        "cash_kzt": "200000", "securities_value_kzt": "600000",
    })
    workbook = Workbook()
    _write_module_summary(workbook, "clients", {"client_account_snapshot": snapshot})
    assert "Аналитика клиентов" in workbook.sheetnames
    assert "Сводка клиентов" not in workbook.sheetnames
    sheet = workbook["Аналитика клиентов"]
    assert len(sheet._charts) == 2


def test_clients_summary_prefers_dashboard_records_over_a_thin_account_summary():
    # client_account_snapshot.summary can be missing securities_value_kzt
    # entirely and collapse every manager into one "Не указано" bucket when
    # the source sheet's own manager column was blank - confirmed against
    # real data. client_dashboard_snapshot carries the same figures per real
    # client row, so it should be preferred whenever it's actually published.
    account = _FakeDataset("client_account_snapshot", summary={
        "manager_mix": {"Не указано": {"total_assets_kzt": "999999999"}},
        "cash_kzt": "111",
    })
    dashboard = _FakeDataset("client_dashboard_snapshot", records=[
        _FakeRecord({"manager": "Иванов", "total_assets_kzt": "500000", "cash_kzt": "100000", "securities_value_kzt": "400000"}),
        _FakeRecord({"manager": "Иванов", "total_assets_kzt": "250000", "cash_kzt": "50000", "securities_value_kzt": "200000"}),
        _FakeRecord({"manager": "Петров", "total_assets_kzt": "300000", "cash_kzt": "60000", "securities_value_kzt": "240000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "clients", {"client_account_snapshot": account, "client_dashboard_snapshot": dashboard})
    sheet = workbook["Аналитика клиентов"]
    manager_header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Активы по менеджерам")
    rows = []
    row = manager_header_row + 2
    while sheet.cell(row, 1).value:
        rows.append((sheet.cell(row, 1).value, sheet.cell(row, 2).value))
        row += 1
    assert rows == [("Иванов", Decimal("750000")), ("Петров", Decimal("300000"))]
    split_header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Состав активов: деньги против бумаг")
    assert sheet.cell(split_header_row + 2, 2).value == Decimal("210000")
    assert sheet.cell(split_header_row + 3, 2).value == Decimal("840000")
    assert "Данные графиков" in workbook.sheetnames


def test_asset_management_summary_currency_allocation_respects_the_search_term():
    # Regression: the currency-allocation table used to always sum every
    # holdings row regardless of the search term, so it silently disagreed
    # with the "Портфель фонда" detail sheet (which does filter by term)
    # whenever a term was applied to the same export.
    holdings = _FakeDataset("fund_holdings", records=[
        _FakeRecord({"instrument": "Bond A", "currency": "KZT", "purchase_value_kzt": "1000000"}),
        _FakeRecord({"instrument": "Bond B", "currency": "USD", "purchase_value_kzt": "200000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "asset-management", {"fund_holdings": holdings}, term="Bond A")
    sheet = workbook["Сводка фонда"]
    header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Распределение фонда по валютам")
    assert sheet.cell(header_row + 2, 1).value == "KZT"
    assert sheet.cell(header_row + 3, 1).value is None


def test_corporate_finance_summary_counts_and_charts_respect_the_search_term():
    # Regression: deal_count/active_count came from the ingestion-time
    # dataset.summary (computed over ALL records) regardless of the search
    # term, and the fee/issuer/ISIN charts iterated the unfiltered record
    # list too - so filtering to one issuer left the detail sheet filtered
    # but the summary sheet's counts and charts unchanged.
    register = _FakeDataset(
        "corporate_finance_register",
        summary={"deal_count": 2, "active_count": 2},
        records=[
            _FakeRecord({"issuer": "АО Первый", "active": True, "fee_received_kzt": "1000000", "isins": ["KZ001"]}),
            _FakeRecord({"issuer": "АО Второй", "active": True, "fee_received_kzt": "2000000", "isins": ["KZ002"]}),
        ],
    )
    workbook = Workbook()
    _write_module_summary(workbook, "corporate-finance", {"corporate_finance_register": register}, term="Первый")
    sheet = workbook["Сводка сделок"]
    assert sheet.cell(4, 2).value == 1
    assert sheet.cell(5, 2).value == 1
    issuer_header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Количество сделок по эмитентам")
    assert sheet.cell(issuer_header_row + 2, 1).value == "АО Первый"
    assert sheet.cell(issuer_header_row + 3, 1).value is None


def test_risk_summary_combines_both_portfolios_and_respects_the_search_term():
    sobstv = _FakeDataset("risk_limits_sobstv", records=[
        _FakeRecord({"portfolio_code": "SOBSTV", "dimension": "issuer", "label": 'АО "SAMPLE BANK"', "limit_kzt": "1073063250", "actual_kzt": "949184.22", "signal": "OK"}),
        _FakeRecord({"portfolio_code": "SOBSTV", "dimension": "country", "label": "ГЕРМАНИЯ", "limit_usd": "10000000", "actual_usd": "12000000", "signal": "breach"}),
    ])
    tabys = _FakeDataset("risk_limits_tabys", records=[
        _FakeRecord({"portfolio_code": "TABYS", "dimension": "issuer", "label": "SPDR", "limit_kzt": "12535913", "actual_kzt": "2587460", "signal": "OK"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "risk", {"risk_limits_sobstv": sobstv, "risk_limits_tabys": tabys})
    sheet = workbook["Сводка по лимитам"]
    assert sheet.cell(4, 2).value == 3
    assert sheet.cell(5, 2).value == 1
    breach_header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Превышенные лимиты")
    assert [sheet.cell(breach_header_row + 2, column).value for column in (1, 2, 3)] == ["SOBSTV", "Страна", "ГЕРМАНИЯ"]

    # Filtering to one issuer must narrow the summary counts too, not just
    # the raw dataset sheets - same "recompute, don't reuse the precomputed
    # summary" reasoning already applied to corporate finance.
    filtered_workbook = Workbook()
    _write_module_summary(filtered_workbook, "risk", {"risk_limits_sobstv": sobstv, "risk_limits_tabys": tabys}, term="SAMPLE BANK")
    filtered_sheet = filtered_workbook["Сводка по лимитам"]
    assert filtered_sheet.cell(4, 2).value == 1
    assert filtered_sheet.cell(5, 2).value == 0


def test_risk_duration_chart_uses_fifteen_rows_when_available():
    duration_records = [
        _FakeRecord({
            "portfolio_code": "SOBSTV",
            "dimension": "duration",
            "label": f"Bond {index}",
            "issuer": f"Issuer {index}",
            "modified_duration": str(index + 1),
            "duration_limit": "20",
            "duration_headroom": str(19 - index),
            "signal": "OK",
        })
        for index in range(15)
    ]
    workbook = Workbook()
    _write_module_summary(workbook, "risk", {
        "risk_limits_sobstv": _FakeDataset("risk_limits_sobstv", records=duration_records),
    })
    sheet = workbook["Сводка по лимитам"]
    title_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Контроли дюрации (топ-15 по использованию)")
    chart = next(chart for chart in sheet._charts if chart.title.tx.rich.p[0].r[0].t == "Модифицированная дюрация против лимита")
    category_reference = chart.series[0].cat.numRef.f if chart.series[0].cat.numRef else chart.series[0].cat.strRef.f
    assert f"$A${title_row + 2}:$A${title_row + 16}" in category_reference


def test_export_warns_when_filename_date_differs_from_internal_source_date():
    dataset = SimpleNamespace(
        dataset_type="risk_limits_tabys",
        summary={},
        records=[SimpleNamespace(
            payload={"portfolio_code": "TABYS", "dimension": "country", "label": "Казахстан", "signal": "OK"},
            record_type="risk_limit",
            source_ref={"sheet_name": "Пр2-16", "row_number": 12},
        )],
        business_date=date(2026, 6, 30),
        source_report_date=date(2026, 6, 30),
        version=1,
        scope_code="TABYS",
        source_upload=SimpleNamespace(original_filename="Риски_Tabys_Лимиты на 01.07.26.xls"),
    )

    workbook = load_workbook(io.BytesIO(multi_source_export.create_module_xlsx("risk", [dataset])))
    sheet = workbook["TABYS · Страны"]
    assert "01.07.2026" in sheet["D3"].value
    assert "30.06.2026" in sheet["D3"].value


def test_filename_date_provenance_supports_compact_dates():
    assert extract_filename_date("Направление_Корпфин_01072026.xlsx") == date(2026, 7, 1)
    assert extract_filename_date("source_2026-07-01.xlsx") == date(2026, 7, 1)
    assert filename_date_mismatch("Направление_Корпфин_01072026.xlsx", date(2026, 6, 30)) is True


def test_clients_summary_respects_the_search_term():
    dashboard = _FakeDataset("client_dashboard_snapshot", records=[
        _FakeRecord({"manager": "Иванов", "client_name": "ТОО Ромашка", "total_assets_kzt": "500000", "cash_kzt": "0", "securities_value_kzt": "500000"}),
        _FakeRecord({"manager": "Петров", "client_name": "ТОО Василёк", "total_assets_kzt": "300000", "cash_kzt": "0", "securities_value_kzt": "300000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "clients", {"client_dashboard_snapshot": dashboard}, term="Ромашка")
    sheet = workbook["Аналитика клиентов"]
    manager_header_row = next(row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "Активы по менеджерам")
    assert sheet.cell(manager_header_row + 2, 1).value == "Иванов"
    assert sheet.cell(manager_header_row + 3, 1).value is None


def test_brokerage_summary_bundles_the_clients_summary_onto_one_charts_sheet():
    # Brokerage export now carries the Clients domain too (routes/multi_source.py),
    # so both summaries land in the same workbook and must share one "Данные
    # графиков" sheet instead of openpyxl silently renaming the second one.
    ledger = _FakeDataset("brokerage_trade_ledger", records=[
        _FakeRecord({"currency": "KZT", "amount": "100", "side": "Покупка", "venue": "KASE", "security_type": "Bond", "execution_status": "Исполнена"}),
        _FakeRecord({"currency": "KZT", "amount": "50", "side": "Продажа", "venue": "OTC", "security_type": "Bond", "execution_status": "Исполнена"}),
    ])
    dashboard = _FakeDataset("client_dashboard_snapshot", records=[
        _FakeRecord({"manager": "Иванов", "total_assets_kzt": "500000", "cash_kzt": "100000", "securities_value_kzt": "400000"}),
        _FakeRecord({"manager": "Петров", "total_assets_kzt": "300000", "cash_kzt": "60000", "securities_value_kzt": "240000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "brokerage", {"brokerage_trade_ledger": ledger, "client_dashboard_snapshot": dashboard})
    assert workbook.sheetnames.count("Данные графиков") == 1
    assert "Данные графиков1" not in workbook.sheetnames
    charts_sheet = workbook["Данные графиков"]
    all_text = " ".join(str(charts_sheet.cell(row, 1).value) for row in range(1, charts_sheet.max_row + 1) if charts_sheet.cell(row, 1).value)
    assert "Сводка брокерской деятельности" in all_text
    assert "Аналитика клиентов" in all_text
    assert "Сводка брокерской деятельности" in workbook.sheetnames
    assert "Аналитика клиентов" in workbook.sheetnames


def test_corporate_finance_summary_explains_the_omitted_placement_demand_chart():
    register = _FakeDataset("corporate_finance_register", summary={"deal_count": 2, "active_count": 1}, records=[
        _FakeRecord({"issuer": "АО Тест", "fee_received_kzt": "1000000"}),
        _FakeRecord({"issuer": "БРК", "fee_received_kzt": "500000"}),
    ])
    workbook = Workbook()
    _write_module_summary(workbook, "corporate-finance", {"corporate_finance_register": register})
    sheet = workbook["Сводка сделок"]
    assert len(sheet._charts) == 4
    all_text = " ".join(str(sheet.cell(row, 1).value) for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value)
    assert "Диаграмма «Объём размещения / спрос по валютам»" not in all_text
    assert "не построена" not in all_text
    assert "Полнота данных источника" not in all_text
    assert [sheet.cell(row, 1).value for row in (9, 10)] == ["АО Тест", "БРК"]
    assert sheet.cell(8, 3).value is None
    assert sheet._charts[0].y_axis.scaling.logBase == 10.0
    assert sheet._charts[0].y_axis.scaling.min == 100000.0
    assert len(sheet._charts[0]._charts) == 1  # remuneration bars only
    assert sheet.cell(28, 1).value == "Статус сделок"
    assert sheet.cell(46, 1).value == "Количество сделок по эмитентам"
    assert sheet.cell(64, 1).value == "ISIN по сделкам"
