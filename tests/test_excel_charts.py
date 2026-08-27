from decimal import Decimal

from openpyxl import Workbook

from osip_dashboard.services.excel_charts import chart_series, cluster_nudges, small_adjacent_slice_indices, write_bar_chart


def test_flags_two_adjacent_small_slices_for_outside_placement():
    assert small_adjacent_slice_indices([Decimal("0.01"), Decimal("0.04"), Decimal("0.35"), Decimal("0.60")]) == {0, 1}


def test_does_not_flag_an_isolated_small_slice():
    assert small_adjacent_slice_indices([Decimal("0.23"), Decimal("0.02"), Decimal("0.35"), Decimal("0.40")]) == set()


def test_flags_small_slices_that_are_adjacent_by_wrapping_around():
    assert small_adjacent_slice_indices([Decimal("0.03"), Decimal("0.60"), Decimal("0.35"), Decimal("0.02")]) == {0, 3}


def test_single_slice_is_never_flagged():
    assert small_adjacent_slice_indices([Decimal("1")]) == set()


def test_cluster_nudges_gives_a_lone_flagged_slice_no_nudge():
    assert cluster_nudges({2}, count=5) == {}


def test_cluster_nudges_fans_a_pair_in_opposite_directions():
    nudges = cluster_nudges({3, 4}, count=7)
    assert set(nudges) == {3, 4}
    assert nudges[3] == -nudges[4]
    assert nudges[3] < 0 < nudges[4]


def test_cluster_nudges_wraps_a_run_spanning_the_last_and_first_index():
    nudges = cluster_nudges({0, 6}, count=7)
    assert set(nudges) == {0, 6}
    assert nudges[6] < nudges[0]


def test_chart_series_leaves_a_single_small_slice_alone():
    rows = [["A", Decimal("1.0")], ["B", Decimal("4.5")], ["C", Decimal("94.5")]]
    labels, values = chart_series(rows, chart_column=2)
    assert labels == ["A", "B", "C"]
    assert values == [Decimal("1.0"), Decimal("4.5"), Decimal("94.5")]


def test_chart_series_groups_two_or_more_small_slices_into_other():
    # The real case that prompted this: three slices under 2% (Consumer
    # Discretionary, Industrials, Large-cap) fold into one "Прочее" wedge;
    # Development Institution stays separate since 4.5% is above the floor.
    rows = [
        ["Consumer Discretionary", Decimal("1.0")], ["Development Institution", Decimal("4.5")],
        ["Government", Decimal("35.1")], ["Industrials", Decimal("1.0")], ["Large-cap", Decimal("1.6")],
        ["Money market", Decimal("39.8")], ["Quasi-sovereign", Decimal("17.0")],
    ]
    labels, values = chart_series(rows, chart_column=2)
    assert labels == ["Development Institution", "Government", "Money market", "Quasi-sovereign", "Прочее"]
    assert values == [Decimal("4.5"), Decimal("35.1"), Decimal("39.8"), Decimal("17.0"), Decimal("3.6")]


def test_write_bar_chart_log_scale_sets_log_base_on_the_value_axis():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Категория", "Количество"])
    sheet.append(["A", 5873])
    sheet.append(["B", 202])
    write_bar_chart(
        sheet, label_col=1, value_col_first=2, value_col_last=2, header_row=1,
        first_row=2, last_row=3, title="t", anchor="D1", log_scale=True,
    )
    chart = sheet._charts[0]
    assert chart.y_axis.scaling.logBase == 10
    assert chart.y_axis.scaling.min == 100.0


def test_write_bar_chart_log_scale_ignores_zero_when_finding_lower_bound():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Категория", "Количество"])
    sheet.append(["A", 0])
    sheet.append(["B", 46])
    write_bar_chart(
        sheet, label_col=1, value_col_first=2, value_col_last=2, header_row=1,
        first_row=2, last_row=3, title="t", anchor="D1", log_scale=True,
    )
    chart = sheet._charts[0]
    assert chart.y_axis.scaling.min == 10.0


def test_write_bar_chart_defaults_to_a_linear_axis():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Категория", "Количество"])
    sheet.append(["A", 5873])
    sheet.append(["B", 202])
    write_bar_chart(
        sheet, label_col=1, value_col_first=2, value_col_last=2, header_row=1,
        first_row=2, last_row=3, title="t", anchor="D1",
    )
    chart = sheet._charts[0]
    assert chart.y_axis.scaling.logBase is None


def test_write_bar_chart_accepts_a_custom_axis_number_format():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Позиция", "Использование"])
    sheet.append(["A", 0.8])
    write_bar_chart(
        sheet, label_col=1, value_col_first=2, value_col_last=2, header_row=1,
        first_row=2, last_row=2, title="t", anchor="D1", y_axis_number_format="0%",
    )
    assert sheet._charts[0].y_axis.numFmt.formatCode == "0%"


def test_write_bar_chart_supports_horizontal_categories_for_long_labels():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Позиция", "Использование"])
    sheet.append(["SOBSTV · Очень длинное наименование инструмента", 3.7])
    write_bar_chart(
        sheet, label_col=1, value_col_first=2, value_col_last=2, header_row=1,
        first_row=2, last_row=2, title="t", anchor="D1", y_axis_number_format="0%", horizontal=True,
    )
    chart = sheet._charts[0]
    assert chart.type == "bar"
    assert chart.width == 24
    assert chart.height == 14.0
    assert chart.overlap == 25
    assert chart.gapWidth == 250
    assert chart.legend is None
    assert chart.series[0].graphicalProperties.line.solidFill.srgbClr == "000000"
    assert chart.series[0].graphicalProperties.line.width == 8000
    assert chart.x_axis.majorGridlines is None
    assert chart.y_axis.majorGridlines is not None
    assert chart.y_axis.majorGridlines.spPr.ln.solidFill.srgbClr == "D9DEE7"
    assert chart.y_axis.numFmt.formatCode == "0%"
    assert chart.x_axis.numFmt is None
