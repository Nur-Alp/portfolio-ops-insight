from datetime import date, timedelta
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import Workbook

from osip_dashboard.services.dividends import (
    configure_dividend_data_root,
    DividendHistory,
    DividendRecord,
    dividend_data_status,
    is_us_listed,
    load_dividend_history,
    lot_dividend_contribution,
    replace_dividend_history,
)
from osip_dashboard.services.hpr import hpr_amount, hpr_percent
from osip_dashboard.services.holdings_export import _hpr_usd_percent_for_item


def test_dividend_history_applies_strict_ex_and_pay_date_rules_and_normalizes_equity_suffix():
    history = DividendHistory([
        DividendRecord(
            ticker="SPY US Equity",
            dividend=Decimal("2"),
            declaration_date=None,
            ex_date=date(2026, 6, 10),
            record_date=None,
            pay_date=date(2026, 6, 30),
            dividend_type="REGULAR",
        ),
        DividendRecord(
            ticker="SPY US Equity",
            dividend=Decimal("3"),
            declaration_date=None,
            ex_date=date(2026, 7, 1),
            record_date=None,
            pay_date=date(2026, 7, 28),
            dividend_type="REGULAR",
        ),
    ])

    received = history.received(
        "SPY US",
        purchase_date=date(2026, 1, 1),
        current_date=date(2026, 7, 28),
    )
    assert len(received) == 1
    assert received[0].dividend == Decimal("2")
    assert is_us_listed("SPY US")
    assert not is_us_listed("ZPRX GY")


def test_us_dividend_is_reduced_by_fifteen_percent_and_converted_to_kzt():
    history = DividendHistory([
        DividendRecord("SPY US Equity", Decimal("2"), None, date(2026, 6, 10), None, date(2026, 6, 30), "REGULAR"),
    ])
    lot = SimpleNamespace(
        security_code="SPY US",
        instrument_currency="USD",
        quantity=Decimal("10"),
        purchase_date=date(2026, 1, 1),
        report_fx_rate=Decimal("500"),
    )

    contribution = lot_dividend_contribution(
        lot, history=history, current_date=date(2026, 7, 28)
    )
    assert contribution.native_amount == Decimal("17.00")
    assert contribution.kzt_amount == Decimal("8500.00")
    assert contribution.matched_count == 1
    assert not contribution.unavailable


def test_non_us_dividend_is_not_taxed_and_missing_fx_is_explicitly_unavailable():
    history = DividendHistory([
        DividendRecord("ZPRX GY", Decimal("1.25"), None, date(2026, 6, 10), None, date(2026, 6, 30), "REGULAR"),
    ])
    lot = SimpleNamespace(
        security_code="ZPRX GY",
        instrument_currency="EUR",
        quantity=Decimal("4"),
        purchase_date=date(2026, 1, 1),
        report_fx_rate=None,
    )

    contribution = lot_dividend_contribution(
        lot, history=history, current_date=date(2026, 7, 28)
    )
    assert contribution.native_amount == Decimal("5.00")
    assert contribution.kzt_amount is None
    assert contribution.unavailable


def test_hpr_adjustment_is_exact_decimal_and_unavailable_for_missing_or_zero_inputs():
    history = DividendHistory([
        DividendRecord("SPY US Equity", Decimal("0.3333"), None, date(2026, 6, 10), None, date(2026, 6, 30), "REGULAR"),
    ])
    lot = SimpleNamespace(
        security_code="SPY US",
        instrument_currency="USD",
        quantity=Decimal("7"),
        purchase_date=date(2026, 1, 1),
        report_fx_rate=Decimal("500"),
    )
    dividend = lot_dividend_contribution(lot, history=history, current_date=date(2026, 7, 28))
    purchase_kzt = Decimal("1000.00")
    carrying_kzt = Decimal("1100.00")
    hpr_kzt = hpr_amount(purchase_kzt, carrying_kzt, dividend.kzt_amount)
    assert dividend.kzt_amount == Decimal("991.567500")
    assert hpr_kzt == Decimal("1091.567500")
    assert hpr_percent(purchase_kzt, carrying_kzt, dividend.kzt_amount) == Decimal("109.15675")
    assert hpr_amount(None, carrying_kzt, dividend.kzt_amount) is None
    assert hpr_amount(Decimal("0"), carrying_kzt, dividend.kzt_amount) is None
    assert hpr_amount(purchase_kzt, None, dividend.kzt_amount) is None
    assert hpr_amount(purchase_kzt, carrying_kzt, None) is None


def test_hpr_usd_percent_is_explicit_and_uses_the_same_native_scale():
    item = {
        "instrument_currency": "USD",
        "purchase_amount_native": Decimal("1000"),
        "carrying_amount_native": Decimal("1100"),
        "dividend_income_native": Decimal("17"),
        "dividend_unavailable": False,
    }
    assert _hpr_usd_percent_for_item(item, usd_rate=Decimal("500")) == Decimal("11.7")

    kzt_item = {
        "instrument_currency": "KZT",
        "purchase_amount_kzt": Decimal("500000"),
        "derived_carrying_value_kzt": Decimal("550000"),
        "dividend_income_kzt": Decimal("8500"),
        "dividend_unavailable": False,
    }
    assert _hpr_usd_percent_for_item(kzt_item, usd_rate=Decimal("500")) == Decimal("11.7")
    assert _hpr_usd_percent_for_item(kzt_item, usd_rate=None) == Decimal("11.7")
    kzt_item["dividend_unavailable"] = True
    assert _hpr_usd_percent_for_item(kzt_item, usd_rate=Decimal("500")) is None


def test_supplied_workbook_is_loaded_with_bloomberg_headers(tmp_path):
    path = tmp_path / "dividends.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ticker", "dividend", "declaration_date", "ex_date", "record_date", "pay_date", "dividend_type", "tickers_list"])
    sheet.append(["TIP US Equity", 0.5, date(2026, 1, 1), date(2026, 2, 1), date(2026, 2, 2), date(2026, 2, 5), "REGULAR", "TIP US"])
    workbook.save(path)

    history = load_dividend_history(str(path))
    assert len(history.records) == 1
    assert history.for_ticker("TIP US") == history.records


def test_uploaded_bloomberg_history_reports_freshness_and_coverage(tmp_path):
    path = tmp_path / "dividends 28.07.26..xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "#Dividend", "#Declaration", "#ExDate", "#Record", "#Payable", "#Type", "TIP US Equity"])
    sheet.append(["TIP US Equity", 0.5, date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 2), date(2026, 7, 8), "REGULAR", ""])
    workbook.save(path)

    try:
        configure_dividend_data_root(tmp_path / "runtime")
        status = replace_dividend_history(filename=path.name, content=path.read_bytes())
        assert status.freshness == "fresh"
        assert status.source_date == date(2026, 7, 28)
        assert status.latest_ex_date == date(2026, 7, 2)
        assert status.latest_pay_date == date(2026, 7, 8)
        assert status.future_pay_count == 0
        assert status.row_count == 1
        assert status.ticker_count == 1
        assert dividend_data_status(as_of=date(2026, 9, 2)).freshness == "stale"
    finally:
        configure_dividend_data_root(None)


def test_dividend_status_counts_future_unpaid_pay_dates(tmp_path):
    # Dates are relative to today rather than hardcoded absolutes - a fixed
    # 2026 date silently stops being "in the future" once the calendar
    # catches up to it, which previously made this test fail for a reason
    # that has nothing to do with the behavior it's meant to check.
    today = date.today()
    declaration = today - timedelta(days=5)
    ex_date = today - timedelta(days=4)
    pay_date = today + timedelta(days=10)
    path = tmp_path / f"dividends {today:%d.%m.%y}..xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["ID", "#Dividend", "#Declaration", "#ExDate", "#Record", "#Payable", "#Type", "TIP US Equity"])
    sheet.append(["TIP US Equity", 0.5, declaration, ex_date, ex_date, pay_date, "REGULAR", ""])
    workbook.save(path)

    try:
        configure_dividend_data_root(tmp_path / "runtime")
        status = replace_dividend_history(filename=path.name, content=path.read_bytes())
        assert status.future_pay_count == 1
        assert dividend_data_status(as_of=pay_date + timedelta(days=1)).future_pay_count == 0
    finally:
        configure_dividend_data_root(None)


def test_export_note_discloses_future_unpaid_dividends():
    from osip_dashboard.services.holdings_export import _dividend_freshness_note

    status = SimpleNamespace(
        freshness="fresh",
        source_date=date(2026, 7, 28),
        stale_after_days=35,
        future_pay_count=2,
    )
    note = _dividend_freshness_note(status)
    assert "актуальны на 28.07.2026" in note
    assert "2 будущие выплаты" in note
    assert "не включены в HPR" in note
