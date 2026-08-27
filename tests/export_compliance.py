"""Shared structural checks for generated Excel exports.

Encodes the automatable parts of the checklist in
``docs/excel-export-design-standards.md`` §7 so every export generator is
checked the same way instead of relying on someone remembering to look.
This intentionally covers only what can be verified generically, without
per-sheet knowledge of expected columns/values - those remain the job of
each export's own dedicated tests.

Known, deliberate gap: a literal "###" column-too-narrow display artifact
cannot be detected from the saved file. openpyxl's cell values and number
formats are unaffected by "###" - it is purely a rendering decision Excel
makes at open time based on actual column pixel width and font metrics,
neither of which round-trips through this checker. Column widths are
still asserted to be reasonably wide by each export's own dedicated tests.
"""

from __future__ import annotations

from datetime import date, datetime
import io
import re
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


_ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}
_PLACEHOLDER_VALUES = {"Недоступно", "Неприменимо", "Да", "Нет", None}
_PLAIN_NUMBER = re.compile(r"^-?[\d\s]+([.,]\d+)?%?$")
_PLAIN_DATE = re.compile(r"^\d{2}\.\d{2}\.\d{4}(\s\d{2}:\d{2})?$")

# ``Сделки`` contains several independent filterable tables. Excel stores
# only the last auto_filter range, while the generator deliberately freezes
# the first table's header. This is the one documented header-row exception
# in docs/export-column-audit.md; all continuous tables must freeze exactly
# below their detected header.
_HEADER_FREEZE_EXCEPTIONS = {"Сделки"}

# Module summary/dashboard sheets (stacked tables feeding charts) and the
# module-level chart-helper-data sheet freeze nothing at all, per the
# "Workbook navigation standard" exception in docs/export-column-audit.md -
# they're read chart-first, not scrolled as a long table, so a frozen
# column A has nothing to do there. Named explicitly (like the exception
# above) rather than inferred, so a genuinely broken continuous table can't
# accidentally slip through by resembling a dashboard sheet.
_UNFROZEN_DASHBOARD_SHEETS = {
    "Данные графиков",
    "Сводка фонда",
    "Сводка брокерской деятельности",
    "Аналитика клиентов",
    "Сводка по лимитам",
    "Сводка ФО",
    "Сводка сделок",
}


def assert_workbook_is_compliant(content: bytes, *, exempt_sheets: set[str] = frozenset()) -> None:
    """Assert an exported workbook meets the reusable structural standard.

    ``exempt_sheets`` names sheets that are intentionally trivial (for
    example a single "no data" placeholder cell) and therefore have nothing
    to freeze or type-check.
    """
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        broken = archive.testzip()
        assert broken is None, f"Corrupt XLSX archive member: {broken}"

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    for sheet in workbook.worksheets:
        exempt = sheet.title in exempt_sheets
        header_row = _header_row(sheet)
        _assert_freezes_column_a_and_header(sheet, header_row, exempt=exempt)
        _assert_no_error_cells(sheet)
        _assert_no_overlapping_merges(sheet)
        _assert_no_bare_empty_strings(sheet)
        _assert_column_typing_is_consistent(sheet, header_row)
        _assert_charts_do_not_overlap_the_table(sheet, header_row)


def _header_row(sheet) -> int | None:
    """The first row of the sheet's primary filterable table, if any.

    Every continuous-table generator in this codebase sets ``auto_filter``
    starting at its header row - reusing that existing, deliberately-placed
    marker is more robust than re-guessing "where the header is" from cell
    styling, and avoids the checker silently drifting from what the
    generator itself considers the header.
    """
    ref = sheet.auto_filter.ref
    if not ref:
        return None
    _, min_row, _, _ = range_boundaries(ref)
    return min_row


def _assert_freezes_column_a_and_header(sheet, header_row: int | None, *, exempt: bool) -> None:
    trivial = sheet.max_row <= 1 and sheet.max_column <= 1
    if exempt or trivial or sheet.title in _UNFROZEN_DASHBOARD_SHEETS:
        return
    freeze = sheet.freeze_panes
    assert freeze is not None, (
        f'Sheet "{sheet.title}" has no freeze_panes - column A (and, for a '
        "continuous table, the header row) must stay visible while "
        "scrolling. See docs/excel-export-design-standards.md §5."
    )
    assert freeze[0] == "B", (
        f'Sheet "{sheet.title}" freezes at {freeze!r}, which does not pin '
        'column A (expected "B<row>"). See docs/excel-export-design-standards.md §5.'
    )
    if header_row is not None and sheet.title not in _HEADER_FREEZE_EXCEPTIONS:
        frozen_row = int(freeze[1:])
        assert frozen_row == header_row + 1, (
            f'Sheet "{sheet.title}" has a filterable table with header row '
            f"{header_row} but freezes at {freeze!r}; continuous tables must "
            "freeze exactly below the header so it remains visible. See "
            "docs/export-column-audit.md for deliberate exceptions."
        )


def _assert_no_error_cells(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "e" or cell.value in _ERROR_VALUES:
                raise AssertionError(
                    f'Sheet "{sheet.title}" cell {cell.coordinate} contains an '
                    f"Excel error value: {cell.value!r}."
                )


def _assert_no_overlapping_merges(sheet) -> None:
    ranges = list(sheet.merged_cells.ranges)
    for index, left in enumerate(ranges):
        for right in ranges[index + 1 :]:
            assert left.isdisjoint(right), (
                f'Sheet "{sheet.title}" has overlapping merged ranges '
                f"{left.coord} and {right.coord}."
            )


def _assert_no_bare_empty_strings(sheet) -> None:
    """A missing value must be an explicit disclosure, never a silent "".

    docs/excel-export-design-standards.md §3: absent inputs are written as
    "Недоступно"/"Неприменимо", not left as a blank-looking empty string
    that is indistinguishable from a genuinely unset cell.
    """
    for row in sheet.iter_rows():
        for cell in row:
            assert cell.value != "", (
                f'Sheet "{sheet.title}" cell {cell.coordinate} is an empty '
                'string - use None (genuinely not applicable) or an explicit '
                '"Недоступно"/"Неприменимо" label instead.'
            )


def _assert_column_typing_is_consistent(sheet, header_row: int | None) -> None:
    """Flag a column that mixes real numbers/dates with number-shaped text.

    This does not require per-sheet column knowledge: it only compares a
    column against itself. A column that is genuinely textual (an ISIN, a
    trade number kept as an identifier) never trips this, because none of
    its other rows are typed as numbers/dates - there is nothing to be
    inconsistent with. It only fires when the same column has some cells
    openpyxl typed as numeric/date and other cells holding a plain
    number/date shape as a string, which is the actual signature of a
    formatting slip rather than an intentional identifier column.
    """
    start_row = (header_row or 0) + 1
    if start_row > sheet.max_row:
        return
    columns: dict[int, set[str]] = {}
    for row in sheet.iter_rows(min_row=start_row):
        for cell in row:
            if cell.value in _PLACEHOLDER_VALUES:
                continue
            kind = _cell_kind(cell)
            if kind is None:
                continue
            columns.setdefault(cell.column, set()).add(kind)
    for column, kinds in columns.items():
        if {"typed", "text-shaped-number"} <= kinds:
            raise AssertionError(
                f'Sheet "{sheet.title}" column {column} mixes typed numeric/date '
                "cells with number- or date-shaped text in other rows - a value "
                "in this column was likely written as a string by mistake."
            )


def _cell_kind(cell) -> str | None:
    if isinstance(cell.value, (int, float, date, datetime)):
        return "typed"
    if isinstance(cell.value, str):
        if _PLAIN_NUMBER.match(cell.value) or _PLAIN_DATE.match(cell.value):
            return "text-shaped-number"
    return None


def _assert_charts_do_not_overlap_the_table(sheet, header_row: int | None) -> None:
    """A chart must not be anchored on top of the sheet's primary table.

    Known limitation: this checks the *declared* anchor cell against the
    table's cell range, not actual rendered pixel overlap (which also
    depends on chart size and column widths/row heights) - the same
    limitation noted for the "###" display-only case above.
    """
    if header_row is None or not sheet.auto_filter.ref:
        return
    min_col, min_row, max_col, max_row = range_boundaries(sheet.auto_filter.ref)
    for chart in sheet._charts:
        anchor = _chart_anchor_cell(chart)
        if anchor is None:
            continue
        column, row = anchor
        if min_col <= column <= max_col and min_row <= row <= max_row:
            raise AssertionError(
                f'Sheet "{sheet.title}" has a chart anchored at column {column}, '
                f"row {row}, inside the primary table range {sheet.auto_filter.ref} "
                "- charts must be placed clear of the source table, not on top of it."
            )


def _chart_anchor_cell(chart) -> tuple[int, int] | None:
    anchor = chart.anchor
    if isinstance(anchor, str):
        min_col, min_row, _, _ = range_boundaries(f"{anchor}:{anchor}")
        return min_col, min_row
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    # AnchorMarker columns/rows are zero-based; range_boundaries is one-based.
    return marker.col + 1, marker.row + 1
