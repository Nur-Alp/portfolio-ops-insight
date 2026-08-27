import csv
import io

from openpyxl import Workbook, load_workbook

from osip_dashboard.services.excel_safety import SafeCsvWriter, neutralize_formulas


def test_neutralize_formulas_defuses_leading_equals_without_altering_the_text():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = '=HYPERLINK("http://evil.example/","click")'
    sheet["A2"] = "Обычный текст"
    assert sheet["A1"].data_type == "f"

    neutralize_formulas(workbook)

    buffer = io.BytesIO()
    workbook.save(buffer)
    reloaded = load_workbook(io.BytesIO(buffer.getvalue()))["Sheet"]
    assert reloaded["A1"].data_type == "s"
    assert reloaded["A1"].value == '=HYPERLINK("http://evil.example/","click")'
    assert reloaded["A2"].data_type == "s"
    assert reloaded["A2"].value == "Обычный текст"


def test_safe_csv_writer_quotes_formula_triggers_but_preserves_plain_numbers():
    buffer = io.StringIO()
    writer = SafeCsvWriter(buffer)
    writer.writerow([
        "=HYPERLINK(\"http://evil.example/\")",
        "-2+3+cmd|'/c calc'!A1",
        "@SUM(A1:A9)",
        "-100.50",
        "+42",
        "Обычный эмитент",
    ])

    reloaded = next(csv.reader(io.StringIO(buffer.getvalue())))
    assert reloaded[0] == "'=HYPERLINK(\"http://evil.example/\")"
    assert reloaded[1] == "'-2+3+cmd|'/c calc'!A1"
    assert reloaded[2] == "'@SUM(A1:A9)"
    assert reloaded[3] == "-100.50"
    assert reloaded[4] == "+42"
    assert reloaded[5] == "Обычный эмитент"
