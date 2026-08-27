from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from osip_dashboard.persistence import Base
from osip_dashboard.persistence.database import create_database_engine
from osip_dashboard.persistence.models import DatasetRecord, DatasetVersion, ImportStatus, ReconciliationResult, SourceUpload
from osip_dashboard.services.demo_multi_source import seed_multi_source_demo
from osip_dashboard.services.multi_source import publish_dataset, reconcile_fund
from osip_dashboard.storage import LocalBlobStore


def test_every_demo_record_source_ref_points_at_a_real_sheet_in_its_blob(tmp_path):
    """Regression: every DatasetRecord.source_ref (and SourceUpload.detection)
    hardcodes sheet_name="Demo", but the placeholder workbook _upload()
    generates used to keep openpyxl's default sheet title "Sheet" - the
    source-cell preview endpoint opens the real blob and checks the claimed
    sheet name against it, so every demo row 404'd with "workbook sheet
    unavailable" the moment someone actually clicked one (found via E2E,
    not this suite - hence this direct check now)."""
    engine = create_database_engine(f"sqlite:///{tmp_path / 'demo.sqlite3'}")
    Base.metadata.create_all(engine)
    blob_store = LocalBlobStore(tmp_path / "blobs")
    with Session(engine) as session:
        seed_multi_source_demo(session, blob_store)
        session.commit()
        uploads = session.scalars(select(SourceUpload)).all()
        assert uploads
        for upload in uploads:
            workbook = load_workbook(blob_store.path_for(upload.storage_key))
            record = session.scalar(select(DatasetRecord).join(DatasetVersion).where(DatasetVersion.source_upload_id == upload.id))
            if record is None:
                continue
            claimed_sheet = record.source_ref["sheet_name"]
            assert claimed_sheet in workbook.sheetnames, (
                f"{upload.original_filename}: source_ref claims sheet {claimed_sheet!r}, "
                f"but the actual blob only has {workbook.sheetnames!r}"
            )
    engine.dispose()


def test_fund_reconciliation_is_idempotent_and_cleans_duplicate_results(tmp_path):
    engine = create_database_engine(f"sqlite:///{tmp_path / 'demo.sqlite3'}")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        seed_multi_source_demo(session, LocalBlobStore(tmp_path / "blobs"))
        session.commit()
        reconcile_fund(session, "TABYS")
        reconcile_fund(session, "TABYS")
        session.commit()

        grouped = session.execute(
            select(ReconciliationResult.rule_code, func.count(ReconciliationResult.id))
            .group_by(ReconciliationResult.rule_code)
        ).all()
        assert grouped == [("FUND-NAV-UNIT-SERIES", 1)]
    engine.dispose()


def test_fund_reconciliation_survives_a_same_date_republish(tmp_path):
    """A corrected re-upload for the same business date must replace the
    prior reconciliation row, not accumulate alongside it - the underlying
    dataset gets a new id even though the (rule, scope, date) it evaluates
    is unchanged."""
    engine = create_database_engine(f"sqlite:///{tmp_path / 'demo.sqlite3'}")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        seed_multi_source_demo(session, LocalBlobStore(tmp_path / "blobs"))
        session.commit()
        reconcile_fund(session, "TABYS")
        session.commit()

        original = session.scalar(select(DatasetVersion).where(
            DatasetVersion.dataset_type == "fund_valuation",
            DatasetVersion.status == ImportStatus.PUBLISHED,
        ))
        republished = DatasetVersion(
            source_upload_id=original.source_upload_id, dataset_type=original.dataset_type,
            detected_key=original.detected_key, scope_type=original.scope_type, scope_code=original.scope_code,
            source_report_date=original.source_report_date, business_date=original.business_date,
            parser_version=original.parser_version, version=original.version + 1,
            status=ImportStatus.VALIDATED, summary=original.summary, uploader_id="republish-test",
        )
        session.add(republished)
        session.flush()
        publish_dataset(session, republished, "republish-test", source_first=True)
        session.commit()
        reconcile_fund(session, "TABYS")
        session.commit()

        grouped = session.execute(
            select(ReconciliationResult.rule_code, func.count(ReconciliationResult.id))
            .group_by(ReconciliationResult.rule_code)
        ).all()
        assert grouped == [("FUND-NAV-UNIT-SERIES", 1)]
    engine.dispose()
